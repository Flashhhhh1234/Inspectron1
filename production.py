""" Modern Production Tool - HIGHLIGHTER STYLE - Complete Integration
Full conversion from box selection to highlighter annotations
AUTO-OPENS PRODUCTION MODE when cabinet is loaded from queue
UPDATED: Proper highlighter display, box annotations removed
UPDATED: Touch-friendly pen tool (throttled drag, OS gesture suppression)
         and on-screen keyboard relaunch on every text field focus
"""
import tkinter as tk
from tkinter import messagebox, simpledialog, Menu, ttk
from PIL import Image, ImageTk, ImageDraw, ImageFont
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import os
import sys
import json
import getpass
import re
import pg_sqlite_compat as sqlite3
import numpy as np
from handover_database import HandoverDB
from database_manager import DatabaseManager
from path_policy import to_relative_path, to_relative_storage_location
import sys
import ctypes
from ctypes import wintypes
import shutil
import time
import subprocess

User = sys.argv[1] if len(sys.argv) > 1 else None
Name = sys.argv[2] if len(sys.argv) > 2 else None

print(f"✓ Production Tool started by: {Name} (username: {User})")


def disable_touch_gestures(hwnd):
    """
    Disable Windows' built-in touch-to-gesture translation (pan/scroll/zoom)
    for the given window handle, so touch input is delivered as raw mouse
    events instead of being intercepted as a system pan gesture.
    FUNCTIONAL USE: Fixes touchscreen drag being swallowed by Windows' pan
    gesture instead of reaching Tkinter's Button-1/B1-Motion handlers.
    Args: hwnd - Native window handle (int), e.g. from widget.winfo_id()
    Returns: bool - True if successfully applied, False otherwise
    """
    if os.name != 'nt':
        return False

    try:
        user32 = ctypes.windll.user32

        # Ensure this HWND is registered for direct touch input rather than
        # the legacy gesture-recognizer path. TWF_WANTPALM=0x00000002 asks
        # for palm/finger contacts to be delivered rather than filtered.
        TWF_WANTPALM = 0x00000002
        RegisterTouchWindow = user32.RegisterTouchWindow
        RegisterTouchWindow.restype = wintypes.BOOL
        RegisterTouchWindow.argtypes = [wintypes.HWND, wintypes.ULONG]
        registered = RegisterTouchWindow(hwnd, TWF_WANTPALM)

        # Disable the OS-level press-and-hold / pan feedback UI which is
        # what causes the "drag becomes scroll" hijack on some drivers.
        try:
            SetWindowFeedbackSetting = user32.SetWindowFeedbackSetting
            FEEDBACK_TOUCH_CONTACTVISUALIZATION = 3
            FEEDBACK_PEN_BARRELVISUALIZATION = 4
            FEEDBACK_TOUCH_TAP = 5
            FEEDBACK_TOUCH_DOUBLETAP = 6
            FEEDBACK_TOUCH_PRESSANDHOLD = 7
            FEEDBACK_TOUCH_RIGHTTAP = 8
            FEEDBACK_TOUCH_PRESSANDTAP = 9
            FEEDBACK_GESTURE_PRESSANDTAP = 10

            feedback_off = ctypes.c_int(0)
            for feedback_id in (
                FEEDBACK_TOUCH_CONTACTVISUALIZATION,
                FEEDBACK_TOUCH_TAP,
                FEEDBACK_TOUCH_DOUBLETAP,
                FEEDBACK_TOUCH_PRESSANDHOLD,
                FEEDBACK_TOUCH_RIGHTTAP,
                FEEDBACK_TOUCH_PRESSANDTAP,
                FEEDBACK_GESTURE_PRESSANDTAP,
            ):
                SetWindowFeedbackSetting(
                    hwnd, feedback_id, 0,
                    ctypes.sizeof(feedback_off), ctypes.byref(feedback_off)
                )
        except Exception as e:
            print(f"[WARN] SetWindowFeedbackSetting skipped: {e}")

        if registered:
            print(f"[INFO] Touch gestures disabled for hwnd={hwnd}")
        else:
            print(f"[WARN] RegisterTouchWindow failed for hwnd={hwnd} "
                  f"(error={ctypes.get_last_error()})")

        return bool(registered)

    except Exception as e:
        print(f"[WARN] Could not disable touch gestures: {e}")
        return False


def configure_touch_feedback(hwnd):
    """Keep touch drawing responsive while suppressing Windows one-finger pan.

    Windows may interpret a mostly vertical finger movement as a scroll/pan
    gesture before Tk receives B1-Motion. Block only GID_PAN on the canvas,
    keep GID_ZOOM enabled, and disable visual feedback. No Python WNDPROC or
    raw WM_TOUCH callback is installed, so this remains safe on Python 3.14.
    FUNCTIONAL USE: Allows the OS to keep delivering two-finger pinch as a
    native zoom gesture (consumed via the Tk <<TouchpadPinch>> virtual
    event) while stopping it from hijacking one-finger drawing/pan strokes.
    """
    if os.name != 'nt':
        return False
    try:
        user32 = ctypes.windll.user32

        class GESTURECONFIG(ctypes.Structure):
            _fields_ = [
                ('dwID', wintypes.DWORD),
                ('dwWant', wintypes.DWORD),
                ('dwBlock', wintypes.DWORD),
            ]

        GID_ZOOM = 3
        GID_PAN = 4
        GC_ZOOM = 0x00000001
        GC_PAN = 0x00000001

        set_gesture_config = user32.SetGestureConfig
        set_gesture_config.argtypes = [
            wintypes.HWND, wintypes.DWORD, wintypes.UINT,
            ctypes.POINTER(GESTURECONFIG), wintypes.UINT
        ]
        set_gesture_config.restype = wintypes.BOOL

        # Allow two-finger zoom but block Windows from stealing a one-finger
        # vertical or horizontal stroke as a native pan/scroll gesture.
        configs = (GESTURECONFIG * 2)(
            GESTURECONFIG(GID_ZOOM, GC_ZOOM, 0),
            GESTURECONFIG(GID_PAN, 0, GC_PAN),
        )
        gesture_ok = bool(set_gesture_config(
            hwnd, 0, len(configs), configs, ctypes.sizeof(GESTURECONFIG)
        ))
        if not gesture_ok:
            print(f'[WARN] SetGestureConfig failed: {ctypes.get_last_error()}')

        feedback_ok = False
        try:
            setting = user32.SetWindowFeedbackSetting
            setting.restype = wintypes.BOOL
            setting.argtypes = [
                wintypes.HWND, wintypes.DWORD, wintypes.DWORD,
                wintypes.UINT, ctypes.c_void_p
            ]
            feedback_off = wintypes.BOOL(False)
            for feedback_id in (3, 5, 6, 7, 8, 9, 10):
                if setting(hwnd, feedback_id, 0, ctypes.sizeof(feedback_off),
                           ctypes.byref(feedback_off)):
                    feedback_ok = True
        except Exception as exc:
            print(f'[WARN] Touch feedback configuration skipped: {exc}')

        if gesture_ok:
            print('[INFO] One-finger Windows pan blocked; tool strokes enabled')
        return gesture_ok or feedback_ok
    except Exception as exc:
        print(f'[WARN] Touch gesture configuration skipped: {exc}')
        return False


def show_onscreen_keyboard():
    """
    Launch Windows on-screen keyboard (touch keyboard preferred, falls back to osk.exe).
    FUNCTIONAL USE: Pops up a virtual keyboard whenever a text entry field gains focus.
    Uses ShellExecuteW instead of subprocess.Popen because both TabTip.exe and osk.exe
    are shell-integrated components that Windows expects to be launched via the shell
    activation path — invoking them through a raw CreateProcess (which subprocess.Popen
    does) commonly fails with WinError 740 "requires elevation" even without any actual
    admin requirement.
    """
    if os.name != 'nt':
        return

    try:
        tabtip_path = r"C:\Program Files\Common Files\Microsoft Shared\ink\TabTip.exe"
        if os.path.exists(tabtip_path):
            result = ctypes.windll.shell32.ShellExecuteW(
                None, "open", tabtip_path, None, None, 1  # SW_SHOWNORMAL
            )
            # ShellExecuteW returns a value > 32 on success
            if result > 32:
                return
            print(f"[WARN] TabTip ShellExecute returned {result}")
    except Exception as e:
        print(f"[WARN] TabTip launch failed: {e}")

    try:
        osk_path = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "System32", "osk.exe")
        result = ctypes.windll.shell32.ShellExecuteW(
            None, "open", osk_path, None, None, 1
        )
        if result > 32:
            return
        print(f"[WARN] osk.exe ShellExecute returned {result}")
    except Exception as e:
        print(f"[WARN] osk.exe launch failed: {e}")


def hide_onscreen_keyboard():
    """
    Attempt to close the on-screen keyboard windows (osk.exe / TabTip).
    FUNCTIONAL USE: Called when a text field loses focus so the keyboard
    doesn't stay on screen unnecessarily, and right before relaunching it
    on a fresh focus-in so a collapsed/minimized keyboard reliably reopens.
    """
    if os.name != 'nt':
        return
    try:
        subprocess.run(["taskkill", "/IM", "osk.exe", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass
    try:
        subprocess.run(["taskkill", "/IM", "TabTip.exe", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


def getbase():
    """
    Returns the base directory path where the application is running.
    FUNCTIONAL USE: Determines if app is frozen (compiled) or running from source code.
    Used to construct absolute paths for config files, databases, and resources.
    Returns: Directory path string (either compiled executable dir or script dir)
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def asset_path(filename):
    """Resolve asset paths for source runs and PyInstaller bundles."""
    bundle_dir = getattr(sys, "_MEIPASS", "")
    if bundle_dir:
        bundled_path = os.path.join(bundle_dir, "assets", filename)
        if os.path.exists(bundled_path):
            return bundled_path

    if getattr(sys, 'frozen', False):
        return os.path.join(getbase(), "assets", filename)

    return os.path.join(getbase(), "assets", filename)


class ManagerDB:
    """Manager database integration for status tracking"""
    
    def __init__(self, db_path):
        self.db_path = db_path
    
    def updcab(self, cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                      total_punches, open_punches, implemented_punches, closed_punches, status,
                      storage_location=None, excel_path=None):
        """
        Insert or replace complete cabinet record with all statistics and metadata.
        FUNCTIONAL USE: Updates manager dashboard with cabinet progress: punch counts, implementation status,
        storage location, and associated Excel file path. Creates record if new, updates if exists.
        Used by production module to sync work progress with quality management system.
        """
        try:
            storage_location_db = to_relative_storage_location(storage_location)
            excel_path_db = to_relative_path(excel_path)

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 storage_location, excel_path, created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                storage_location_db, excel_path_db,
                  cabinet_id, datetime.now().isoformat(),
                  datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            print(f"✓ Manager DB: Updated {cabinet_id} - Status: {status}")
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def updstats(self, cabinet_id, status):
        """
        Update cabinet status field and last_updated timestamp.
        FUNCTIONAL USE: Lightweight status-only update for handover transitions between quality/production.
        Updates database with current date/time to track workflow progress.
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE cabinets
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            print(f"✓ Manager DB: Status updated for {cabinet_id} → {status}")
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False


class ProductionTool:
    def __init__(self, root):
        self.root = root
        self.logged_in_username = User
        self.logged_in_fullname = Name
        self.root.title("Production Tool - Highlighter Mode")
        self.root.geometry("1400x900")
        # Bind window close event to auto-save
        self.root.protocol("WM_DELETE_WINDOW", self.closing)
        
        # Data / files
        self.pdf_document = None
        self.current_pdf_path = None
        self.current_session_path = None
        self.current_page = 0
        self.project_name = ""
        self.sales_order_no = ""
        self.cabinet_id = ""
        self.storage_location = ""
        self.annotations = []
        
        self.handover_db = HandoverDB("handover_db")
        self.db = DatabaseManager("inspection_tool")
        self.manager_db = ManagerDB("manager")
        
        self.excel_file = None
        self.working_excel_path = None
        self.zoom_level = 1.0
        self.ZOOM_MIN = 0.5
        self.ZOOM_MAX = 3.0
        self.ZOOM_STEP = 0.1
        self._zoom_render_after_id = None
        self.zoom_slider_frame = None
        self._zoom_slider_pct_var = None
        self._zoom_slider_generation = 0
        self._zoom_root_release_funcid = None
        self._zoom_is_dragging = False
        self._render_cache = {}
        self.current_sr_no = 1
        self.current_page_image = None
        self.session_refs = set()
        
        # Visual navigation for production mode
        self.production_highlight_tags = []
        self.production_dialog_open = False
        
        # Highlighter colors with RGBA for semi-transparency
        self.highlighter_colors = {
            'yellow': {'rgb': (255, 255, 0), 'rgba': (255, 255, 0, 100)},
            'green': {'rgb': (0, 255, 0), 'rgba': (0, 255, 0, 100)},
            'blue': {'rgb': (0, 191, 255), 'rgba': (0, 191, 255, 100)},
            'pink': {'rgb': (255, 105, 180), 'rgba': (255, 105, 180, 100)},
            'orange': {'rgb': (255, 165, 0), 'rgba': (255, 165, 0, 100)}
        }
        
        # Column mapping
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
        
        self.interphase_sheet_name = 'Interphase'
        self.interphase_cols = {
            'ref_no': 'B',
            'description': 'C',
            'status': 'D',
        }
        
        self.header_cells = {
            "Interphase": {
                "project_name": "C4",
                "sales_order": "C6",
                "cabinet_id": "F6"
            },
            "Punch Sheet": {
                "project_name": "C2",
                "sales_order": "C4",
                "cabinet_id": "H4"
            }
        }
        
        # Highlighter drawing state - NO BOX SELECTION
        self.drawing = False
        self.highlighter_start_x = None
        self.highlighter_start_y = None
        self.temp_highlight_id = None
        self.selected_annotation = None
        
        # Tool modes (pen, text)
        self.current_tool = None  # None, 'pen', 'text'
        self.tool_mode = None  # Alias for current_tool
        self.pen_points = []
        self.temp_pen_line = None
        self.temp_line_ids = []  # Store temporary drawing line IDs
        self.drawing_type = None  # 'pen', 'text'
        self.text_pos_x = None
        self.text_pos_y = None

        # NEW: touch-friendly drag throttling state for the pen tool
        self._pending_drag_event = None
        self._drag_frame_scheduled = False
        self._panning = False
        self._touch_scroll_lock_until = 0.0

        # Pinch-to-zoom state (ported from Quality, "as it is") - Tk-managed
        # <<TouchpadPinch>> handling with proportional zoom tracking, a
        # throttled live preview, and a watchdog-finalized full-quality
        # render when the fingers are released.
        self._native_pinch_start_distance = None
        self._native_pinch_start_zoom = None
        self._native_pinch_last_zoom = None
        self._native_pinch_render_after_id = None
        self._native_pinch_preview_after_id = None
        self._native_pinch_preview_pending = False
        self._native_pinch_preview_interval_ms = 16
        self._native_pinch_finish_delay_ms = 180
        self._safe_pinch_last_raw = None
        self._safe_pinch_last_time = 0.0
        self._safe_pinch_accumulator = 1.0
        self._safe_pinch_active = False
        self._safe_pinch_release_funcid = None
        self._safe_pinch_watchdog_ms = 1200
        self._safe_pinch_bound = False
        # Cached source pixmap for the page being pinch-zoomed, captured once
        # at gesture start. Live pinch frames resize this in-memory PIL image
        # instead of re-rasterizing the PDF on every touch event, which is
        # what made pinch feel laggy - PyMuPDF get_pixmap() is comparatively
        # expensive and the digitizer can fire far faster than it can keep up.
        self._pinch_base_image = None
        self._pinch_base_scale = None
        self._pinch_base_page = None
        self._pinch_frame_pending = False
        self._pinch_frame_min_interval = 1.0 / 60.0  # cap live redraws to ~60fps
        self._pinch_last_frame_time = 0.0
        # Microsoft-style text box state
        self._text_box_start = None
        self._text_box_preview_id = None
        self._text_editor = None
        self._text_editor_window_id = None
        self._text_edit_annotation = None
        self._text_transform_mode = None
        self._text_transform_start = None
        self._text_transform_original_bbox = None
        self._text_selection_ids = []
        
        # Highlighter state
        self.active_highlighter = False
        
        # Undo stack
        self.undo_stack = []
        self.max_undo = 50

        # Debounced annotation save. Pen strokes are persisted shortly after
        # release without blocking every raw drawing event.
        self._annotation_save_after_id = None
        self._session_dirty = False
        
        self.uisetup()
        self.bind_global_keyboard_popup()   # NEW: relaunch on-screen keyboard every text focus
        self.current_sr_no = self.getnextsr()

    # ================================================================
    # TOUCH SUPPORT - GESTURE SUPPRESSION + ON-SCREEN KEYBOARD
    # ================================================================

    def _apply_touch_gesture_fix(self):
        """
        Apply the Windows touch-gesture suppression to the canvas HWND.
        FUNCTIONAL USE: Called shortly after canvas creation once the widget
        has a valid native window handle. Retries briefly if the handle
        isn't ready yet. This is what makes touch drags land as normal
        Button-1/B1-Motion events instead of being intercepted as a
        pan/scroll gesture by Windows.
        """
        if os.name != 'nt':
            return

        try:
            hwnd = self.canvas.winfo_id()
            if not hwnd:
                self.root.after(100, self._apply_touch_gesture_fix)
                return

            success = disable_touch_gestures(hwnd)
            self._touch_gesture_fix_applied = success

            # Quality-style gesture config: explicitly allow GID_ZOOM (so
            # two-finger pinch keeps reaching Tk as <<TouchpadPinch>>) while
            # blocking GID_PAN so a one-finger stroke isn't hijacked.
            configure_touch_feedback(hwnd)
            self.setup_canvas_pinch_zoom()

            if not success:
                print("[WARN] Touch gesture fix did not apply cleanly; "
                    "touch drags may still be intercepted as scroll.")
        except Exception as e:
            print(f"[WARN] _apply_touch_gesture_fix error: {e}")

    def bind_global_keyboard_popup(self):
        """
        Globally bind FocusIn/FocusOut on all Entry and Text widgets (including
        those inside dialogs like simpledialog) to show/hide the on-screen
        keyboard. Relaunches the keyboard EVERY time a text field is focused,
        not just the first time - handles the case where TabTip was manually
        collapsed by the user in between text fields.
        FUNCTIONAL USE: Ensures the popup keyboard appears automatically
        whenever the user taps into ANY text box anywhere in the app.
        Call this once in __init__ after uisetup().
        """
        def on_focus_in(event):
            widget = event.widget
            if isinstance(widget, (tk.Entry, tk.Text)):
                # Force a fresh relaunch every time: if TabTip/osk was
                # collapsed by the user, ShellExecuteW("open", ...) alone
                # won't reliably bring it back - killing it first
                # guarantees a real relaunch.
                hide_onscreen_keyboard()
                self.root.after(80, show_onscreen_keyboard)

        def on_focus_out(event):
            widget = event.widget
            if isinstance(widget, (tk.Entry, tk.Text)):
                # Small delay avoids flicker when focus moves between two text fields
                self.root.after(150, self._maybe_hide_keyboard)

        self.root.bind_class("Entry", "<FocusIn>", on_focus_in, add="+")
        self.root.bind_class("Text", "<FocusIn>", on_focus_in, add="+")

        self.root.bind_class("Entry", "<FocusOut>", on_focus_out, add="+")
        self.root.bind_class("Text", "<FocusOut>", on_focus_out, add="+")

    def _maybe_hide_keyboard(self):
        """
        Hide the on-screen keyboard only if focus has actually left a
        text-entry widget.
        FUNCTIONAL USE: Prevents keyboard flicker when tabbing between
        input fields.
        """
        try:
            focused = self.root.focus_get()
        except Exception:
            focused = None

        if not isinstance(focused, (tk.Entry, tk.Text)):
            hide_onscreen_keyboard()

    # ================================================================
    # MANAGER SYNC - PRODUCTION SPECIFIC
    # ================================================================
    
    # ----------------------------------------------------------------
    # LOADING OVERLAY - copied from Quality's busy()/unbusy() design
    # (full-window dark scrim + centered card with a spinning glyph),
    # kept behind the same show_loading()/update_loading()/hide_loading()
    # names and nested-call depth counting the rest of Production already
    # calls, so no other call site needs to change.
    # ----------------------------------------------------------------

    def show_loading(self, message="Working...", detail="Please wait while the operation completes."):
        """Show one reusable modal loading overlay; nested calls only update its text."""
        self._loading_depth = getattr(self, '_loading_depth', 0) + 1
        display_message = f"{message}\n{detail}" if detail else message
        try:
            existing = getattr(self, '_loading_window', None)
            if existing is not None and existing.winfo_exists():
                self._loading_message_var.set(display_message)
                existing.lift()
                self._spin_loading_overlay()
                self.root.update_idletasks()
                return

            overlay = tk.Toplevel(self.root)
            overlay.overrideredirect(True)
            overlay.attributes('-topmost', True)
            try:
                overlay.attributes('-alpha', 0.97)
            except tk.TclError:
                pass
            overlay.configure(bg='#0f172a')
            self._loading_window = overlay

            # Cover the whole main window so nothing underneath is clickable.
            self.root.update_idletasks()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            w = self.root.winfo_width()
            h = self.root.winfo_height()
            overlay.geometry(f"{max(w, 1)}x{max(h, 1)}+{x}+{y}")

            # Semi-dark full-window scrim so the frozen-looking background
            # doesn't show through, plus a centered card with a spinner.
            scrim = tk.Frame(overlay, bg='#0f172a')
            scrim.place(relx=0, rely=0, relwidth=1, relheight=1)

            card = tk.Frame(scrim, bg='#1e293b', highlightthickness=1,
                             highlightbackground='#334155')
            card.place(relx=0.5, rely=0.5, anchor='center')

            self._loading_spinner_var = tk.StringVar(value='◐')
            tk.Label(card, textvariable=self._loading_spinner_var, bg='#1e293b',
                     fg='#60a5fa', font=('Segoe UI', 26)).pack(padx=36, pady=(28, 6))

            self._loading_message_var = tk.StringVar(value=display_message)
            tk.Label(card, textvariable=self._loading_message_var, bg='#1e293b',
                     fg='white', font=('Segoe UI', 11, 'bold'), wraplength=360,
                     justify='center').pack(padx=36, pady=(0, 26))

            self._loading_spin_frames = ['◐', '◓', '◑', '◒']
            self._loading_spin_index = 0
            self._loading_spin_after_id = None

            overlay.lift()
            overlay.focus_force()
            self._spin_loading_overlay()

            # Force Tk to actually draw the overlay right now, before the
            # caller goes on to do the expensive blocking work - otherwise
            # the overlay would just sit in the event queue, unseen, while
            # the freeze happens exactly as before.
            self.root.update_idletasks()
            self.root.update()
        except tk.TclError:
            self._loading_window = None

    def _spin_loading_overlay(self):
        """Advance the loading-overlay spinner glyph. Re-arms itself while the overlay exists."""
        overlay = getattr(self, '_loading_window', None)
        if overlay is None or not overlay.winfo_exists():
            return
        try:
            self._loading_spin_index = (self._loading_spin_index + 1) % len(self._loading_spin_frames)
            self._loading_spinner_var.set(self._loading_spin_frames[self._loading_spin_index])
        except (tk.TclError, AttributeError):
            return
        self._loading_spin_after_id = self.root.after(160, self._spin_loading_overlay)

    def update_loading(self, message=None, detail=None):
        """Update the active loading overlay without creating another window."""
        try:
            if message is not None or detail is not None:
                if message is not None and detail is not None:
                    self._loading_message_var.set(f"{message}\n{detail}")
                elif message is not None:
                    self._loading_message_var.set(message)
                else:
                    self._loading_message_var.set(detail)
            self.root.update_idletasks()
        except Exception:
            pass

    def hide_loading(self, force=False):
        """Close the reusable loading overlay after the outermost heavy operation."""
        depth = getattr(self, '_loading_depth', 0)
        self._loading_depth = 0 if force else max(0, depth - 1)
        if self._loading_depth:
            return
        after_id = getattr(self, '_loading_spin_after_id', None)
        if after_id is not None:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass
            self._loading_spin_after_id = None
        overlay = getattr(self, '_loading_window', None)
        self._loading_window = None
        if overlay is not None:
            try:
                overlay.destroy()
            except tk.TclError:
                pass
        self.root.update_idletasks()

    def syncmgrstats(self):
        """
        Calculate current punch statistics from Excel and sync to manager database.
        FUNCTIONAL USE: Counts open/implemented/closed punches from Excel Punch Sheet (rows 9+).
        Syncs cabinet status and statistics to the manager PostgreSQL schema for dashboard visibility.
        Called during production work to update manager on progress.
        """
        if not self.cabinet_id:
            return
        
        try:
            # Count from Excel - start from row 9
            implemented_punches = 0
            closed_punches = 0
            total_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 9  # Start from row 9
                    while row <= ws.max_row + 5:
                        checked = self.read_cell(ws, row, 'E')
                        if not checked:
                            row += 1
                            if row > 2000:
                                break
                            continue
                        
                        total_punches += 1
                        implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                        closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                        
                        if closed:
                            closed_punches += 1
                        elif implemented:
                            implemented_punches += 1
                        
                        row += 1
                        if row > 2000:
                            break
                    
                    wb.close()
                except Exception as e:
                    print(f"Excel read error: {e}")
            
            open_punches = total_punches - implemented_punches - closed_punches
            
            self.manager_db.updcab(
                self.cabinet_id,
                self.project_name,
                self.sales_order_no,
                0,
                0,
                total_punches,
                open_punches,
                implemented_punches,
                closed_punches,
                'in_progress',
                storage_location=getattr(self, 'storage_location', None),
                excel_path=self.excel_file
            )
        
        except Exception as e:
            print(f"Manager sync error: {e}")
            import traceback
            traceback.print_exc()
    
    def syncmgrstatsonly(self):
        """Lightweight sync without full recount - for display updates"""
        # Only sync if we have the necessary data loaded
        if self.cabinet_id and self.excel_file:
            self.syncmgrstats()

    # ================================================================
    # CELL HELPERS
    # ================================================================
    
    def split_cell(self, cell_ref):
        """
        Parse Excel cell reference (e.g., 'A1', 'B42') into row and column components.
        FUNCTIONAL USE: Splits Excel notation into numeric row and string column for openpyxl operations.
        Args: cell_ref - Cell reference string (e.g., 'B5', 'H10')
        Returns: Tuple of (row_number, column_letter)
        """
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col
    
    def _resolve_merged_target(self, ws, row, col_idx):
        """
        Find actual cell coordinates when target cell is part of a merged cell range.
        FUNCTIONAL USE: Handles merged cells in Excel by returning the top-left cell of merge range.
        Ensures writes/reads go to correct cell even when targeting merged area.
        Args: ws - Worksheet, row - row number, col_idx - column index
        Returns: Tuple of (actual_row, actual_col) accounting for merges
        """
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx
    
    def write_cell(self, ws, row, col, value):
        """
        Write value to Excel cell, handling merged cells and column format conversion.
        FUNCTIONAL USE: Unified write interface that accepts column as letter ('A') or number (1).
        Automatically routes to correct cell if target is part of merged range.
        Args: ws - Worksheet, row - row number, col - column (letter or index), value - data to write
        """
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        ws.cell(row=target_row, column=target_col).value = value
    
    def read_cell(self, ws, row, col):
        """
        Read value from Excel cell, handling merged cells and column format conversion.
        FUNCTIONAL USE: Unified read interface that accepts column as letter ('A') or number (1).
        Automatically finds actual cell if target is part of merged range.
        Args: ws - Worksheet, row - row number, col - column (letter or index)
        Returns: Cell value (string, number, date, etc.)
        """
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value

    # ================================================================
    # MODERN UI SETUP
    # ================================================================
    
    def uisetup(self):
        """
        Create complete user interface with toolbar, menu, canvas, and status bar.
        FUNCTIONAL USE: Builds UI components including file menu, tools menu, navigation buttons,
        zoom controls, pen/text tool buttons, canvas for PDF display, and keyboard shortcuts.
        Sets up all event bindings for mouse and keyboard interactions.
        """
        # Main toolbar
        toolbar = tk.Frame(self.root, bg='#1e293b', height=70)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        # Enhanced Menu Bar
        menubar = Menu(self.root, bg='#1e293b', fg='white', activebackground='#3b82f6')
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="📁 File", menu=file_menu)
        file_menu.add_command(label="Load from Production Queue", command=self.loadfrmhandover, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools Menu
        tools_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="🛠️ Tools", menu=tools_menu)
        tools_menu.add_command(label="🏭 Production Mode", command=self.prodmode, accelerator="Ctrl+P")
        tools_menu.add_separator()
        tools_menu.add_command(label="✅ Complete & Handback", command=self.compreworkhndbck, accelerator="Ctrl+H")
        
        # View Menu
        view_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="👁️ View", menu=view_menu)
        
        # Keyboard shortcuts
        self.root.bind_all("<Control-o>", lambda e: self.loadfrmhandover())
        self.root.bind_all("<Control-p>", lambda e: self.prodmode())
        self.root.bind_all("<Control-h>", lambda e: self.compreworkhndbck())
        self.root.bind_all("<Control-z>", lambda e: self.undolast())
        self.root.bind_all("<Escape>", lambda e: self.deactivate_all())
        
        # Left section - Load operations
        left_frame = tk.Frame(toolbar, bg='#1e293b')
        left_frame.pack(side=tk.LEFT, padx=10, pady=10)
        
        tk.Button(left_frame, text="📦 Load from Queue", command=self.loadfrmhandover,
                 bg='#8b5cf6', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # Center section - Navigation
        center_frame = tk.Frame(toolbar, bg='#1e293b')
        center_frame.pack(side=tk.LEFT, padx=20)
        
        self.page_label = tk.Label(center_frame, text="Page: 0/0", bg='#1e293b', fg='white',
                                   font=('Segoe UI', 10, 'bold'))
        self.page_label.pack(side=tk.LEFT, padx=10)
        
        nav_btn_style = {
            'bg': '#64748b',
            'fg': 'white',
            'font': ('Segoe UI', 9, 'bold'),
            'relief': tk.FLAT,
            'cursor': 'hand2'
        }
        
        tk.Button(center_frame, text="◀", command=self.prev, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(center_frame, text="▶", command=self.next, width=3,
                 **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        
        # Tool section - Pen, Text, Undo
        tool_frame = tk.Frame(toolbar, bg='#1e293b')
        tool_frame.pack(side=tk.LEFT, padx=10)

        tk.Label(tool_frame, text="Tools:", bg='#1e293b', fg='#94a3b8', 
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 8))

        # Load icons or use fallback
        self.pen_btn = None
        self.text_btn = None
        
        try:
            icon_size = (44, 44)
            
            pen_icon_path = asset_path("pen_icon.png")
            pen_img = Image.open(pen_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.pen_icon = ImageTk.PhotoImage(pen_img)
            
            text_icon_path = asset_path("text_icon.png")
            text_img = Image.open(text_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.text_icon = ImageTk.PhotoImage(text_img)
            
            undo_icon_path = asset_path("undo_icon.png")
            undo_img = Image.open(undo_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.undo_icon = ImageTk.PhotoImage(undo_img)
            
            self.pen_btn = tk.Button(tool_frame, image=self.pen_icon, 
                                     command=lambda: self.settlmd("pen"),
                                     bg='#334155', width=48, height=48, 
                                     relief=tk.FLAT, cursor='hand2')
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, image=self.text_icon, 
                                      command=lambda: self.settlmd("text"),
                                      bg='#334155', width=48, height=48, 
                                      relief=tk.FLAT, cursor='hand2')
            self.text_btn.pack(side=tk.LEFT, padx=2)
            
            self.undo_btn = tk.Button(tool_frame, image=self.undo_icon,
                                      command=self.undolast,
                                      bg='#334155', width=48, height=48, 
                                      relief=tk.FLAT, cursor='hand2')
            self.undo_btn.pack(side=tk.LEFT, padx=2)
            
        except Exception as e:
            print(f"Could not load tool icons: {e}")
            # Fallback to text buttons
            self.pen_btn = tk.Button(tool_frame, text="✏️ Pen", 
                     command=lambda: self.settlmd("pen"),
                     bg='#334155', fg='white', padx=10, pady=8,
                     font=('Segoe UI', 9, 'bold'), relief=tk.FLAT,
                     cursor='hand2')
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, text="🅰️ Text", 
                     command=lambda: self.settlmd("text"),
                     bg='#334155', fg='white', padx=10, pady=8,
                     font=('Segoe UI', 9, 'bold'), relief=tk.FLAT,
                     cursor='hand2')
            self.text_btn.pack(side=tk.LEFT, padx=2)
            
            tk.Button(tool_frame, text="↶ Undo",
                     command=self.undolast,
                     bg='#334155', fg='white', padx=10, pady=8,
                     font=('Segoe UI', 9, 'bold'), relief=tk.FLAT,
                     cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        # Right section - Action buttons
        right_frame = tk.Frame(toolbar, bg='#1e293b')
        right_frame.pack(side=tk.RIGHT, padx=10, pady=10)
        
        tk.Button(right_frame, text="🏭 Production Mode", command=self.prodmode,
                 bg='#f59e0b', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        tk.Button(right_frame, text="✅ Handback to Quality", command=self.compreworkhndbck,
                 bg='#10b981', fg='white', padx=15, pady=10,
                 font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, borderwidth=0,
                 cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # Canvas with scrollbars
        canvas_frame = tk.Frame(self.root, bg='#f1f5f9')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        v_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(canvas_frame, bg='#f8fafc',
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.setup_quality_zoom_controls()
        
        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)

        # NEW: disable OS-level touch gesture translation on this canvas so
        # touch drags reach leftclick/leftdrag/leftrls as raw mouse-like
        # events instead of being intercepted as a pan/scroll gesture by
        # Windows. Mirrors the same fix used in the Quality tool.
        self.canvas.update_idletasks()  # ensure winfo_id() is valid
        self.root.after(50, self._apply_touch_gesture_fix)
        
        # Bind mouse events - CRITICAL FOR PEN AND TEXT TOOLS
        self.canvas.bind("<ButtonPress-1>", self.leftclick)
        self.canvas.bind("<B1-Motion>", self.leftdrag)
        self.canvas.bind("<ButtonRelease-1>", self.leftrls)
        self.canvas.bind("<Double-Button-1>", self.doubleclick)
        self.canvas.bind("<Double-Button-3>", self.doubleright)
        self._bind_display_mouse_controls()
        
        # Modern status bar
        status_bar = tk.Frame(self.root, bg='#334155', height=40)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        instructions_text = "Pen: Freehand | Text: Drag a box | No tool: Drag to pan | Ctrl+Wheel: Zoom | Double-click: Zoom slider"
        tk.Label(status_bar, text=instructions_text, bg='#334155', fg='#e2e8f0',
                font=('Segoe UI', 9), pady=10).pack()

    def _bind_display_mouse_controls(self):
        """Register wheel scrolling handlers for the PDF display canvas."""
        if getattr(self, '_display_mouse_controls_bound', False):
            return

        self.root.bind_all("<MouseWheel>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-MouseWheel>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-Button-4>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-Button-5>", self._on_display_mousewheel, add="+")

        self._display_mouse_controls_bound = True

    def _is_pointer_over_canvas(self):
        """Return True when the mouse pointer is over the display canvas."""
        if not hasattr(self, 'canvas') or not self.canvas or not self.canvas.winfo_exists():
            return False

        try:
            widget = self.root.winfo_containing(self.root.winfo_pointerx(), self.root.winfo_pointery())
        except tk.TclError:
            return False

        while widget is not None:
            if widget == self.canvas:
                return True
            widget = widget.master
        return False

    def _on_display_mousewheel(self, event):
        """Scroll the display canvas with the mouse wheel when hovering over it."""
        if not self._is_pointer_over_canvas():
            return

        delta = 0
        if getattr(event, 'num', None) == 4:
            delta = 1
        elif getattr(event, 'num', None) == 5:
            delta = -1
        elif getattr(event, 'delta', 0):
            delta = 1 if event.delta > 0 else -1

        if delta == 0:
            return

        horizontal = bool(getattr(event, 'state', 0) & 0x0001)
        try:
            if horizontal:
                self.canvas.xview_scroll(-delta, "units")
            else:
                self.canvas.yview_scroll(-delta, "units")
        except tk.TclError:
            return

        return "break"
    
    # ================================================================
    # LOAD FROM HANDOVER QUEUE - WITH AUTO-OPEN PRODUCTION MODE
    # ================================================================
    
    def _keep_latest_queue_versions(self, items):
        """Return one queue entry per cabinet, keeping the latest handover.

        Entries are grouped by normalized cabinet_id. The newest valid
        handed_over_date wins. If dates are missing or invalid, the later
        entry returned by the database wins, which provides deterministic
        behaviour without dropping a cabinet from the queue.
        """
        latest_by_cabinet = {}

        def parse_handover_date(value):
            if isinstance(value, datetime):
                return value
            if value is None:
                return None

            value = str(value).strip()
            if not value:
                return None

            # Accept standard ISO timestamps, including a trailing Z.
            try:
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            except (TypeError, ValueError):
                return None

        for position, item in enumerate(items or []):
            if not isinstance(item, dict):
                continue

            cabinet_id = str(item.get("cabinet_id") or "").strip()
            # Do not merge unrelated records that have no cabinet ID.
            key = cabinet_id.casefold() if cabinet_id else f"__missing_{position}"
            candidate_date = parse_handover_date(item.get("handed_over_date"))

            existing = latest_by_cabinet.get(key)
            if existing is None:
                latest_by_cabinet[key] = (item, candidate_date, position)
                continue

            _, existing_date, existing_position = existing

            if candidate_date is not None and existing_date is not None:
                try:
                    candidate_is_newer = candidate_date > existing_date
                except TypeError:
                    # Handles comparison between timezone-aware and naive values.
                    candidate_is_newer = candidate_date.replace(tzinfo=None) > existing_date.replace(tzinfo=None)
            elif candidate_date is not None:
                candidate_is_newer = True
            elif existing_date is None:
                candidate_is_newer = position > existing_position
            else:
                candidate_is_newer = False

            if candidate_is_newer:
                latest_by_cabinet[key] = (item, candidate_date, position)

        # Show latest handovers first. Missing dates are placed last.
        deduplicated = list(latest_by_cabinet.values())
        deduplicated.sort(
            key=lambda record: (
                record[1] is not None,
                record[1].replace(tzinfo=None) if record[1] is not None else datetime.min,
                record[2],
            ),
            reverse=True,
        )

        removed_count = len(items or []) - len(deduplicated)
        if removed_count:
            print(f"[INFO] Production queue: removed {removed_count} redundant entr{'y' if removed_count == 1 else 'ies'}.")

        return [record[0] for record in deduplicated]

    def loadfrmhandover(self):
        """Show the production queue using the Projects & Cabinets dialog design."""
        pending_items = self.handover_db.get_pending_production_items()
        pending_items = self._keep_latest_queue_versions(pending_items)

        if not pending_items:
            messagebox.showinfo(
                "No Items",
                "No items in the production queue. All items have been processed!",
                icon='info'
            )
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Production Queue")
        dlg.geometry("1000x620")
        dlg.minsize(780, 500)
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg='#1e293b', height=58)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header, text="Load from Production Queue", bg='#1e293b', fg='white',
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=14)

        search_frame = tk.Frame(dlg, bg='#f8fafc')
        search_frame.pack(fill=tk.X, padx=18, pady=(16, 8))
        tk.Label(
            search_frame, text="Search:", bg='#f8fafc', fg='#334155',
            font=('Segoe UI', 10, 'bold')
        ).pack(side=tk.LEFT, padx=(0, 8))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=('Segoe UI', 11))
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        count_var = tk.StringVar()
        tk.Label(
            search_frame, textvariable=count_var, bg='#f8fafc', fg='#64748b',
            font=('Segoe UI', 9)
        ).pack(side=tk.RIGHT, padx=(12, 0))

        body = tk.Frame(dlg, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=8)

        columns = ('cabinet', 'project', 'punches', 'handover_by', 'date', 'status')
        tree = ttk.Treeview(body, columns=columns, show='headings', selectmode='browse')
        tree.heading('cabinet', text='Cabinet ID')
        tree.heading('project', text='Project')
        tree.heading('punches', text='Open Punches')
        tree.heading('handover_by', text='Handed Over By')
        tree.heading('date', text='Date')
        tree.heading('status', text='Status')
        tree.column('cabinet', width=150, minwidth=120)
        tree.column('project', width=250, minwidth=180)
        tree.column('punches', width=100, anchor='center', stretch=False)
        tree.column('handover_by', width=160, minwidth=120)
        tree.column('date', width=105, anchor='center', stretch=False)
        tree.column('status', width=110, anchor='center', stretch=False)

        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        row_items = {}

        def populate(*args):
            query = search_var.get().strip().casefold()
            tree.delete(*tree.get_children())
            row_items.clear()
            visible = 0
            for item in pending_items:
                searchable = ' '.join(str(item.get(key, '')) for key in (
                    'cabinet_id', 'project_name', 'handed_over_by', 'handed_over_date', 'status'
                )).casefold()
                if query and query not in searchable:
                    continue
                date_text = str(item.get('handed_over_date') or '')[:10]
                status_text = str(item.get('status') or '').replace('_', ' ').title()
                row_id = tree.insert('', tk.END, values=(
                    item.get('cabinet_id', ''),
                    item.get('project_name', ''),
                    item.get('open_punches', 0),
                    item.get('handed_over_by', ''),
                    date_text,
                    status_text,
                ))
                row_items[row_id] = item
                visible += 1
            count_var.set(f"{visible} of {len(pending_items)} item(s)")
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])

        def load_selected(event=None):
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Please select an item first.", parent=dlg)
                return
            item = row_items.get(selected[0])
            if not item:
                return
            dlg.grab_release()
            dlg.destroy()
            self.loadhndovritm(item)

        search_var.trace_add('write', populate)
        tree.bind('<Double-Button-1>', load_selected)
        tree.bind('<Return>', load_selected)

        buttons = tk.Frame(dlg, bg='#f8fafc')
        buttons.pack(fill=tk.X, padx=18, pady=(4, 16))
        tk.Button(
            buttons, text="Cancel", command=dlg.destroy, bg='#64748b', fg='white',
            font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=20, pady=9,
            cursor='hand2'
        ).pack(side=tk.LEFT)
        tk.Button(
            buttons, text="Load Selected", command=load_selected, bg='#3b82f6', fg='white',
            font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=26, pady=9,
            cursor='hand2'
        ).pack(side=tk.RIGHT)

        populate()
        search_entry.focus_set()

    def loadhndovritm(self, item):
        """
        Load PDF, Excel, and session data for a quality-handover item into production workspace.
        FUNCTIONAL USE: Initializes ProductionTool workspace with cabinet info, loads PDF document,
        Excel punch sheet, and previous session annotations. Auto-opens production mode dialog.
        Args: item - Dictionary with cabinet_id, pdf_path, excel_path, storage_location, project info
        """
        try:
            # Verify files exist
            if not item.get('pdf_path') or not os.path.exists(item['pdf_path']):
                messagebox.showerror("File Not Found", 
                                   f"PDF not found:\n{item['pdf_path']}")
                return
            
            if not item.get('excel_path') or not os.path.exists(item['excel_path']):
                messagebox.showerror("File Not Found", 
                                   f"Excel not found:\n{item['excel_path']}")
                return
            
            # Get project from database
            project_data = self.db.get_project(item['cabinet_id'])
            if not project_data:
                messagebox.showerror("Error", "Project not found in database")
                return
            
            # Load PDF
            self.pdf_document = fitz.open(item['pdf_path'])
            self.current_pdf_path = item['pdf_path']
            self.current_page = 0
            self.zoom_level = 1.0
            
            # Set project details
            self.cabinet_id = item['cabinet_id']
            self.project_name = item['project_name']
            self.sales_order_no = item['sales_order_no']
            self.storage_location = project_data['storage_location']
            
            # Set Excel
            self.excel_file = item['excel_path']
            self.working_excel_path = item['excel_path']
            
            # Load session if available
            print(f"\n{'='*60}")
            print(f"Loading handover item: {self.cabinet_id}")
            print(f"PDF: {item['pdf_path']}")
            print(f"Excel: {item['excel_path']}")
            print(f"Session path from item: {item.get('session_path')}")
            
            if item.get('session_path') and os.path.exists(item['session_path']):
                self.current_session_path = os.path.abspath(item['session_path'])
                print(f"✓ Session file exists, loading...")
                self.loadsessfrompath(self.current_session_path)
                print(f"After loading: {len(self.annotations)} annotations loaded")
                
                # Debug: Show what's in annotations
                highlight_count = sum(1 for a in self.annotations if a.get('type') == 'highlight')
                error_count = sum(1 for a in self.annotations if a.get('type') == 'error')
                print(f"  Highlights: {highlight_count}, Errors: {error_count}")
                
                for i, ann in enumerate(self.annotations[:3]):  # First 3 only
                    print(f"  Annotation {i}: type={ann.get('type')}, "
                          f"page={ann.get('page')}, "
                          f"has_points_page={'points_page' in ann}, "
                          f"has_bbox_page={'bbox_page' in ann}, "
                          f"sr_no={ann.get('sr_no')}")
            else:
                print(f"⚠️ No session file found")
                self.current_session_path = None
                self.annotations = []
                self.session_refs.clear()
            
            print(f"{'='*60}\n")
            
            # Mark as in progress
            username = self.logged_in_fullname or "Unknown User"
            
            self.handover_db.update_production_status(
                item['cabinet_id'],
                status='in_progress',
                user=username
            )
            
            # Update manager status
            self.manager_db.updstats(self.cabinet_id, 'in_progress')
            self.syncmgrstats()
            
            self.display()
            
            
            # AUTO-OPEN PRODUCTION MODE
            self.root.after(500, self.prodmode)
        
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load item:\n{e}")
            import traceback
            traceback.print_exc()
    def schedule_annotation_save(self, delay_ms=250):
        """Mark the in-memory session dirty; persistence occurs at close or handback."""
        self._session_dirty = True

    def _flush_annotation_save(self):
        """Compatibility no-op. Final session saving is intentionally deferred."""
        self._annotation_save_after_id = None
        self._session_dirty = True

    def closing(self):
        """
        Save current session and close application gracefully.
        FUNCTIONAL USE: Auto-saves all annotations and work to session file before exit.
        Ensures no unsaved production work is lost.
        """
        if self._text_editor is not None:
            self._commit_text_editor()
        if self._annotation_save_after_id is not None:
            try:
                self.root.after_cancel(self._annotation_save_after_id)
            except Exception:
                pass
            self._annotation_save_after_id = None
        if self.pdf_document and self.storage_location:
            try:
                print("\n🔄 Auto-saving before closing...")
                self.savesess()
                print("✓ Session auto-saved successfully")
                
                # Sync stats one last time. This is intentionally outside display().
                self.syncmgrstatsonly()
                print("✓ Statistics synced")
                
            except Exception as e:
                print(f"⚠️ Auto-save on close failed: {e}")
                # Ask user if they want to close anyway
                proceed = messagebox.askyesno(
                    "Save Failed",
                    f"Failed to auto-save:\n{e}\n\nClose anyway?",
                    icon='warning'
                )
                if not proceed:
                    return  # Don't close the application
        
        # Close the application
        self.root.destroy()
    # ================================================================
    # COMPLETE REWORK & HANDBACK - CHECK IMPLEMENTED COLUMN
    # ================================================================
    
    def compreworkhndbck(self):
        """
        Finalize production work and return cabinet to quality for verification.
        FUNCTIONAL USE: Validates all punches have implementation status, auto-saves session,
        creates handback record in database for quality module to receive and review.
        Updates manager database with completion status.
        """
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("No Item Loaded", 
                                 "Please load an item from the production queue first.")
            return
        
        item = self.handover_db.get_item_by_cabinet_id(self.cabinet_id, "quality_to_production")
        if not item:
            messagebox.showwarning("Not from Queue", 
                                 "This item was not loaded from the handover queue.")
            return
        
        # Check for punches without implementation
        not_implemented = self.findnotimplemented()
        if not_implemented:
            self.shownotimplemented(not_implemented)
            return
        
        # AUTO-SAVE SESSION BEFORE HANDBACK
        print("Auto-saving session before handback...")
        try:
            self.savesess()
            print("✓ Session auto-saved successfully")
        except Exception as e:
            print(f"⚠️ Session auto-save failed: {e}")
            # Continue anyway - not critical
        remarks=None
        
        username = self.logged_in_fullname or "Unknown User"
        
        handback_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": self.getsesspathforpdf(),
            "rework_completed_by": username,
            "rework_completed_date": datetime.now().isoformat(),
            "production_remarks": remarks or "No remarks"
        }
        
        success = self.handover_db.add_production_handback(handback_data)
        
        if success:
            self.syncmgrstats()
            self.manager_db.updstats(self.cabinet_id, 'being_closed_by_quality')
            
            
            # Clear current work
            self.pdf_document = None
            self.current_pdf_path = None
            self.excel_file = None
            self.annotations = []
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            self.root.title("Production Tool - Highlighter Mode")
        else:
            messagebox.showerror("Error", "Failed to handback item to Quality.")
    
    def findnotimplemented(self):
        """
        Scan Excel Punch Sheet and identify punches lacking implementation.
        FUNCTIONAL USE: Checks 'Implemented By' column for each punch from row 9 onwards.
        Returns list of unimplemented punches to prevent premature handback to quality.
        """
        not_implemented = []
        
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return not_implemented
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            row = 9
            while row <= ws.max_row + 5:
                checked = self.read_cell(ws, row, 'E')
                if not checked:
                    row += 1
                    if row > 2000:
                        break
                    continue
                
                closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
                if closed:
                    row += 1
                    continue
                
                implemented = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                if not implemented:
                    sr_no = self.read_cell(ws, row, self.punch_cols['sr_no'])
                    ref_no = self.read_cell(ws, row, self.punch_cols['ref_no'])
                    desc = self.read_cell(ws, row, self.punch_cols['desc'])
                    category = self.read_cell(ws, row, self.punch_cols['category'])
                    
                    not_implemented.append({
                        'row': row,
                        'sr_no': sr_no,
                        'ref_no': ref_no,
                        'description': desc,
                        'category': category
                    })
                
                row += 1
                if row > 2000:
                    break
            
            wb.close()
            return not_implemented
        
        except Exception as e:
            print(f"Error checking implementation: {e}")
            return []
    
    def shownotimplemented(self, not_implemented):
        """
        Display warning dialog with list of punches needing implementation.
        FUNCTIONAL USE: Visual feedback preventing handback with incomplete work.
        Shows punch details and requires user acknowledgment.
        Args: not_implemented - List of punch records with incomplete implementation
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("⚠️ Implementation Required")
        dlg.geometry("800x600")
        dlg.configure(bg='#fef3c7')
        dlg.transient(self.root)
        dlg.grab_set()
        
        header_frame = tk.Frame(dlg, bg='#0f172a', height=82)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="⚠️ IMPLEMENTATION REQUIRED",
                bg='#f59e0b', fg='white',
                font=('Segoe UI', 14, 'bold')).pack(pady=15)
        
        info_frame = tk.Frame(dlg, bg='#fef3c7')
        info_frame.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(info_frame, 
                text=f"The following {len(not_implemented)} punch(es) have not been marked as 'Implemented'.\n"
                     "Please complete implementation before handing back to Quality.",
                font=('Segoe UI', 11), bg='#fef3c7', fg='#78350f',
                justify='left').pack(anchor='w')
        
        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(list_frame, text="Punches requiring implementation:",
                font=('Segoe UI', 10, 'bold'), bg='white', fg='#1e293b').pack(anchor='w', padx=10, pady=(10, 5))
        
        scroll_frame = tk.Frame(list_frame, bg='white')
        scroll_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(scroll_frame, wrap=tk.WORD, font=('Courier New', 9),
                            yscrollcommand=scrollbar.set, bg='#f8fafc', relief=tk.FLAT,
                            padx=10, pady=10)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        for idx, punch in enumerate(not_implemented, 1):
            text_widget.insert(tk.END, f"\n{'='*70}\n")
            text_widget.insert(tk.END, f"#{idx} - SR No: {punch['sr_no']} | Ref: {punch['ref_no']}\n")
            text_widget.insert(tk.END, f"Category: {punch['category']}\n")
            text_widget.insert(tk.END, f"\nDescription:\n{punch['description']}\n")
        
        text_widget.config(state=tk.DISABLED)
        
        tk.Button(dlg, text="OK - I'll Complete Implementation First",
                 command=dlg.destroy, bg='#f59e0b', fg='white',
                 font=('Segoe UI', 10, 'bold'), padx=20, pady=12,
                 relief=tk.FLAT, cursor='hand2').pack(pady=20)

    # ================================================================
    # ENHANCED PRODUCTION MODE WITH HIGHLIGHTER NAVIGATION
    # ================================================================
    
    def prodmode(self):
        """Open the redesigned production rework workspace."""
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("No cabinet loaded", "Load a cabinet from the production queue first.")
            return

        punches = self.openpunches()
        if not punches:
            messagebox.showinfo("Production complete", "All punches are closed. The cabinet can be handed back to Quality.")
            return
        punches.sort(key=lambda p: (p['implemented'], p['sr_no']))

        colors = {
            'window': '#eef2f7', 'nav': '#0f172a', 'nav_card': '#172033',
            'card': '#ffffff', 'muted': '#64748b', 'text': '#0f172a',
            'line': '#e2e8f0', 'primary': '#2563eb', 'success': '#059669',
            'warning': '#d97706', 'soft_blue': '#eff6ff', 'soft_green': '#ecfdf5'
        }
        dlg = tk.Toplevel(self.root)
        dlg.title("Production Rework Workspace")
        dlg.geometry("1050x650")
        dlg.minsize(920, 580)
        dlg.configure(bg=colors['window'])
        dlg.transient(self.root)
        dlg.grab_set()
        self.production_dialog_open = True

        header = tk.Frame(dlg, bg=colors['nav'], height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        heading = tk.Frame(header, bg=colors['nav'])
        heading.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        tk.Label(heading, text="Production Rework", bg=colors['nav'], fg='white',
                 font=('Segoe UI Semibold', 17, 'bold')).pack(anchor='w', pady=(9, 0))
        tk.Label(heading, text=f"{self.cabinet_id}  •  {self.project_name}", bg=colors['nav'],
                 fg='#94a3b8', font=('Segoe UI', 10)).pack(anchor='w')
        summary = tk.Frame(header, bg=colors['nav'])
        summary.pack(side=tk.RIGHT, fill=tk.Y, padx=20)
        implemented_count = sum(1 for item in punches if item['implemented'])
        tk.Label(summary, text=f"{implemented_count} / {len(punches)} implemented", bg=colors['nav'],
                 fg='#bfdbfe', font=('Segoe UI Semibold', 10, 'bold')).pack(anchor='e', pady=(15, 1))
        tk.Label(summary, text="Select a punch, review Quality's note, then record your action.",
                 bg=colors['nav'], fg='#94a3b8', font=('Segoe UI', 9)).pack(anchor='e')

        body = tk.Frame(dlg, bg=colors['window'])
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)

        sidebar = tk.Frame(body, bg=colors['card'], width=250, highlightthickness=1,
                           highlightbackground=colors['line'])
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)
        tk.Label(sidebar, text="PUNCH LIST", bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=14, pady=(12, 6))
        filter_bar = tk.Frame(sidebar, bg=colors['soft_blue'])
        filter_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(filter_bar, text=f"{len(punches)} open punches", bg=colors['soft_blue'],
                 fg='#1d4ed8', font=('Segoe UI Semibold', 9, 'bold')).pack(anchor='w', padx=8, pady=6)
        list_frame = tk.Frame(sidebar, bg=colors['card'])
        list_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        punch_list = tk.Listbox(list_frame, activestyle='none', selectmode=tk.SINGLE,
                                font=('Segoe UI', 9), bg=colors['card'], fg=colors['text'],
                                selectbackground='#dbeafe', selectforeground='#1e3a8a',
                                relief=tk.FLAT, borderwidth=0, highlightthickness=0,
                                yscrollcommand=scrollbar.set)
        punch_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=punch_list.yview)

        content = tk.Frame(body, bg=colors['window'])
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        meta = tk.Frame(content, bg=colors['window'])
        meta.pack(fill=tk.X)
        sr_value = tk.StringVar(); ref_value = tk.StringVar(); status_value = tk.StringVar()
        def metric(parent, label, variable, tint, value_color):
            card = tk.Frame(parent, bg=tint, highlightthickness=1, highlightbackground=colors['line'])
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            tk.Label(card, text=label, bg=tint, fg=colors['muted'],
                     font=('Segoe UI Semibold', 7, 'bold')).pack(anchor='w', padx=10, pady=(7, 1))
            tk.Label(card, textvariable=variable, bg=tint, fg=value_color,
                     font=('Segoe UI Semibold', 11, 'bold')).pack(anchor='w', padx=10, pady=(0, 8))
        metric(meta, "SR NUMBER", sr_value, '#ffffff', colors['text'])
        metric(meta, "REFERENCE", ref_value, '#ffffff', colors['text'])
        metric(meta, "STATUS", status_value, '#ffffff', colors['warning'])

        detail_card = tk.Frame(content, bg=colors['card'], highlightthickness=1,
                               highlightbackground=colors['line'])
        detail_card.pack(fill=tk.BOTH, expand=True, pady=(9, 0))
        detail_head = tk.Frame(detail_card, bg=colors['card'])
        detail_head.pack(fill=tk.X, padx=14, pady=(11, 6))
        title_var = tk.StringVar(value="Punch details")
        tk.Label(detail_head, textvariable=title_var, bg=colors['card'], fg=colors['text'],
                 font=('Segoe UI Semibold', 13, 'bold')).pack(side=tk.LEFT)
        position_var = tk.StringVar()
        tk.Label(detail_head, textvariable=position_var, bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI', 9)).pack(side=tk.RIGHT)

        description = tk.Text(detail_card, height=5, wrap=tk.WORD, bg='#f8fafc', fg=colors['text'],
                              relief=tk.FLAT, borderwidth=0, padx=10, pady=8,
                              font=('Segoe UI', 10), cursor='arrow')
        description.pack(fill=tk.X, padx=14)
        description.config(state=tk.DISABLED)

        remarks = tk.Frame(detail_card, bg=colors['card'])
        remarks.pack(fill=tk.BOTH, expand=True, padx=14, pady=9)
        q_col = tk.Frame(remarks, bg=colors['soft_blue'], highlightthickness=1,
                         highlightbackground='#bfdbfe')
        q_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        tk.Label(q_col, text="QUALITY REMARK", bg=colors['soft_blue'], fg='#1d4ed8',
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=10, pady=(8, 3))
        quality_text = tk.Text(q_col, height=5, wrap=tk.WORD, bg=colors['soft_blue'], fg=colors['text'],
                               relief=tk.FLAT, padx=8, pady=6, font=('Segoe UI', 9), cursor='arrow')
        quality_text.pack(fill=tk.BOTH, expand=True, padx=3, pady=(0, 5))
        quality_text.config(state=tk.DISABLED)

        p_col = tk.Frame(remarks, bg=colors['soft_green'], highlightthickness=1,
                         highlightbackground='#a7f3d0')
        p_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))
        tk.Label(p_col, text="PRODUCTION ACTION / REMARK", bg=colors['soft_green'], fg='#047857',
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=10, pady=(8, 3))
        action_text = tk.Text(p_col, height=7, wrap=tk.WORD, bg='white', fg=colors['text'],
                              relief=tk.FLAT, highlightthickness=1, highlightbackground='#a7f3d0',
                              highlightcolor=colors['success'], padx=8, pady=6, font=('Segoe UI', 9))
        action_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        footer = tk.Frame(dlg, bg=colors['card'], height=60, highlightthickness=1,
                          highlightbackground=colors['line'])
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)
        left_actions = tk.Frame(footer, bg=colors['card'])
        left_actions.pack(side=tk.LEFT, padx=14, pady=9)
        right_actions = tk.Frame(footer, bg=colors['card'])
        right_actions.pack(side=tk.RIGHT, padx=14, pady=9)

        def button(parent, text, command, bg, fg='white', width=16):
            return tk.Button(parent, text=text, command=command, bg=bg, fg=fg,
                             activebackground=bg, activeforeground=fg, relief=tk.FLAT,
                             borderwidth=0, cursor='hand2', font=('Segoe UI Semibold', 9, 'bold'),
                             padx=10, pady=8, width=width)

        current = {'index': 0, 'refreshing': False}
        def find_ann(punch, create=False):
            sr_key = str(punch.get('sr_no', '')).strip()
            row_key = str(punch.get('row', '')).strip()
            matches = [a for a in self.annotations
                       if (str(a.get('sr_no', '')).strip() == sr_key and sr_key)
                       or (str(a.get('excel_row', '')).strip() == row_key and row_key)]
            # Prefer the object already carrying the shared remarks.
            ann = next((a for a in matches if a.get('implementation_remark') is not None
                        or a.get('quality_remark') is not None), matches[0] if matches else None)
            if ann is None and create:
                ann = {
                    'type': 'punch_meta', 'page': None,
                    'sr_no': punch.get('sr_no'), 'excel_row': punch.get('row'),
                    'ref_no': punch.get('ref_no'), 'punch_text': punch.get('punch_text'),
                    'category': punch.get('category'),
                    'created_by_role': 'production',
                    'timestamp': datetime.now().isoformat()
                }
                self.annotations.append(ann)
            return ann

        def refresh_list():
            selected = current['index']
            current['refreshing'] = True
            punch_list.delete(0, tk.END)
            for item in punches:
                marker = "DONE" if item['implemented'] else "OPEN"
                punch_list.insert(tk.END, f"  {marker:<5}   SR {item['sr_no']}   •   {item['ref_no']}")
            punch_list.selection_clear(0, tk.END)
            punch_list.selection_set(selected)
            punch_list.activate(selected)
            punch_list.see(selected)
            current['refreshing'] = False

        def show_item(index=None):
            if index is not None:
                current['index'] = max(0, min(len(punches) - 1, index))
            p = punches[current['index']]
            ann = find_ann(p)
            sr_value.set(str(p['sr_no']))
            ref_value.set(str(p['ref_no']))
            status_value.set("Implemented" if p['implemented'] else "Action required")
            title_var.set(str(p.get('category') or 'Punch details'))
            position_var.set(f"Punch {current['index'] + 1} of {len(punches)}")
            description.config(state=tk.NORMAL)
            description.delete('1.0', tk.END)
            description.insert('1.0', p.get('punch_text') or 'No description available.')
            description.config(state=tk.DISABLED)
            quality_text.config(state=tk.NORMAL)
            quality_text.delete('1.0', tk.END)
            quality_text.insert('1.0', (ann or {}).get('quality_remark') or 'No quality remark for this punch.')
            quality_text.config(state=tk.DISABLED)
            action_text.delete('1.0', tk.END)
            action_text.insert('1.0', (ann or {}).get('implementation_remark') or '')
            status_value.set("Implemented" if p['implemented'] else "Action required")
            refresh_list()
            self.navtopunch(p['sr_no'], p.get('punch_text'))
            action_text.focus_set()

        def save_remark_draft():
            if not punches:
                return False
            p = punches[current['index']]
            remark = action_text.get('1.0', 'end-1c').strip()
            ann = find_ann(p, create=bool(remark))
            if ann is None:
                return False
            if str(ann.get('implementation_remark') or '') == remark:
                return False
            ann['implementation_remark'] = remark
            ann['implemented'] = bool(p.get('implemented'))
            self._session_dirty = True
            return True

        def select_from_list(event=None):
            if current.get('refreshing'):
                return
            save_remark_draft()
            selection = punch_list.curselection()
            if selection:
                show_item(selection[0])

        def go(delta):
            save_remark_draft()
            show_item(current['index'] + delta)

        def mark_implemented():
            p = punches[current['index']]
            name = self.logged_in_fullname or "Unknown User"
            remark = action_text.get('1.0', 'end-1c').strip()
            self.show_loading("Marking punch implemented", f"Updating SR {p.get('sr_no')} in the Punch Sheet.")
            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]
                self.write_cell(ws, p['row'], self.punch_cols['implemented_name'], name)
                self.write_cell(ws, p['row'], self.punch_cols['implemented_date'],
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                wb.save(self.excel_file)
                wb.close()
            except PermissionError:
                self.hide_loading(force=True)
                messagebox.showerror("Excel file is open", "Close the Excel workbook and try again.", parent=dlg)
                return
            except Exception as exc:
                self.hide_loading(force=True)
                messagebox.showerror("Could not update punch", str(exc), parent=dlg)
                return
            ann = find_ann(p, create=True)
            if ann:
                ann['implemented'] = True
                ann['implemented_name'] = name
                ann['implemented_date'] = datetime.now().isoformat()
                ann['implementation_remark'] = remark
            p['implemented'] = True
            self._session_dirty = True
            self.hide_loading(force=True)
            refresh_list()
            status_value.set("Implemented")
            next_index = current['index'] + 1
            if next_index < len(punches):
                self.root.after_idle(lambda: show_item(next_index))
            else:
                messagebox.showinfo("Review complete", "All punches in this list have been reviewed.", parent=dlg)
            # Manager aggregation is not required to navigate to the next punch.
            self.root.after(250, self.syncmgrstatsonly)

        def on_close():
            save_remark_draft()
            self.clrborderhighlight()
            self.production_dialog_open = False
            dlg.destroy()

        punch_list.bind('<<ListboxSelect>>', select_from_list)
        button(left_actions, "Previous", lambda: go(-1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT, padx=(0, 8))
        button(left_actions, "Next", lambda: go(1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT)
        button(right_actions, "Close", on_close, '#475569', width=10).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "Mark Implemented", mark_implemented, colors['success'], width=18).pack(side=tk.RIGHT, padx=(8, 0))
        button(right_actions, "Save Remark", save_remark_draft, colors['primary'], width=12).pack(side=tk.RIGHT)
        dlg.protocol("WM_DELETE_WINDOW", on_close)
        dlg.bind('<Control-Return>', lambda event: mark_implemented())
        refresh_list()
        show_item(0)
    
    def navtopunch(self, sr_no, punch_text):
        """Navigate to highlighter annotation and highlight it - UPDATED FOR HIGHLIGHTER

        A punch can have several highlights attached to it (e.g. via
        Quality's multi-mark mode). Navigation only ever jumps to and
        highlights the FIRST highlight linked to this punch; any additional
        highlights for the same SR No are intentionally ignored here.
        """
        target_ann = None

        # Try SR No match - looking for 'error' type annotations (which are
        # highlighter marks). Compare SR numbers as normalized strings so a
        # str/int mismatch between the punch and the annotation can't cause
        # us to skip the true first highlight. Collect every match first so
        # "first" is well-defined instead of relying on incidental list
        # order breaking early.
        sr_key = str(sr_no).strip() if sr_no is not None else ''
        sr_matches = [
            ann for ann in self.annotations
            if sr_key and str(ann.get('sr_no', '')).strip() == sr_key
            and ann.get('type') in ('error', 'highlight')
        ]
        if sr_matches:
            target_ann = sr_matches[0]
            if len(sr_matches) > 1:
                print(f"ℹ️ SR {sr_no} has {len(sr_matches)} highlights - "
                      f"navigating to the first one only, ignoring the rest")
            print(f"✓ Found annotation by SR No: {sr_no}, type: {target_ann.get('type')}")

        # Fuzzy text match if no direct SR match
        if not target_ann:
            best_match = None
            best_score = 0
            
            for ann in self.annotations:
                if ann.get('type') in ('error', 'highlight') and ann.get('punch_text'):
                    ann_text = str(ann['punch_text']).lower()
                    search_text = str(punch_text).lower()
                    
                    if search_text in ann_text or ann_text in search_text:
                        score = len(set(search_text.split()) & set(ann_text.split()))
                        if score > best_score:
                            best_score = score
                            best_match = ann
            
            if best_match:
                target_ann = best_match
                print(f"✓ Found annotation by text match, SR: {best_match.get('sr_no')}")
        
        self.clrborderhighlight()
        
        if target_ann:
            print(f"Navigating to annotation:")
            print(f"  Type: {target_ann.get('type')}")
            print(f"  SR No: {target_ann.get('sr_no')}")
            print(f"  Has points_page: {'points_page' in target_ann}")
            print(f"  Has bbox_page: {'bbox_page' in target_ann}")
            
            target_page = target_ann.get('page')
            if target_page is not None and target_page != self.current_page:
                self.current_page = target_page
                self.display()
            
            # Highlight the annotation visually
            if 'points_page' in target_ann or 'bbox_page' in target_ann:
                self.highlightannonvisual(target_ann)
                self._last_highlighted_ann = target_ann
        else:
            # Punches added directly from Add Punch intentionally have no drawing object.
            self._last_highlighted_ann = None
    
    def highlightannonvisual(self, annotation):
        """Draw visual indicators for highlighter annotation - UPDATED"""
        # Calculate bounding box from points_page or use bbox_page
        if 'points_page' in annotation and annotation['points_page']:
            # Calculate bbox from highlighter points
            points = annotation['points_page']
            xs = [p[0] for p in points]
            ys = [p[1] for p in points]
            bbox_page = (min(xs), min(ys), max(xs), max(ys))
            bbox_display = self.bbox_page_to_display(bbox_page)
            print(f"  Using points_page to calculate bbox: {bbox_page}")
        elif 'bbox_page' in annotation:
            bbox_display = self.bbox_page_to_display(annotation['bbox_page'])
            print(f"  Using bbox_page: {annotation['bbox_page']}")
        else:
            print("⚠️ Annotation has no points_page or bbox_page - cannot highlight")
            return
        
        x1, y1, x2, y2 = bbox_display
        
        # Calculate center
        cx = (x1 + x2) / 2
        cy = (y1 + y2) / 2
        
        padding = 15
        
        # Glow layers - pulsing effect
        for i in range(3):
            glow_padding = padding + (i * 5)
            
            glow_id = self.canvas.create_rectangle(
                x1 - glow_padding, y1 - glow_padding,
                x2 + glow_padding, y2 + glow_padding,
                outline='#fbbf24', width=2, dash=(8, 4),
                tags='production_highlight'
            )
            self.production_highlight_tags.append(glow_id)
        
        # Main highlight border - bright orange
        main_id = self.canvas.create_rectangle(
            x1 - padding, y1 - padding,
            x2 + padding, y2 + padding,
            outline='#f59e0b', width=4, dash=(10, 5),
            tags='production_highlight'
        )
        self.production_highlight_tags.append(main_id)
        
        # Arrow pointing to the annotation
        arrow_start_x = cx - 120
        arrow_start_y = cy - 120
        
        # Arrow background (shadow)
        arrow_bg = self.canvas.create_line(
            arrow_start_x, arrow_start_y,
            cx - 20, cy - 20,
            arrow=tk.LAST, fill='#fbbf24', width=6,
            tags='production_highlight'
        )
        self.production_highlight_tags.append(arrow_bg)
        
        # Arrow foreground
        arrow_fg = self.canvas.create_line(
            arrow_start_x, arrow_start_y,
            cx - 20, cy - 20,
            arrow=tk.LAST, fill='#f59e0b', width=3,
            tags='production_highlight'
        )
        self.production_highlight_tags.append(arrow_fg)
        
        # Label background
        label_bg = self.canvas.create_rectangle(
            arrow_start_x - 60, arrow_start_y - 35,
            arrow_start_x + 10, arrow_start_y - 5,
            fill='#fef3c7', outline='#f59e0b', width=2,
            tags='production_highlight'
        )
        self.production_highlight_tags.append(label_bg)
        
        # Label text
        label_text = f"SR {annotation.get('sr_no', '?')}"
        label_txt = self.canvas.create_text(
            arrow_start_x - 25, arrow_start_y - 20,
            text=label_text,
            fill='#92400e',
            font=('Segoe UI', 12, 'bold'),
            tags='production_highlight'
        )
        self.production_highlight_tags.append(label_txt)
        
        # Scroll to make visible
        bbox_all = self.canvas.bbox("all")
        if bbox_all:
            self.canvas.yview_moveto(max(0, (y1 - 150) / max(1, bbox_all[3])))
            self.canvas.xview_moveto(max(0, (x1 - 150) / max(1, bbox_all[2])))
        
        print(f"✓ Visual highlight added at display coords: {bbox_display}")
    
    def clrborderhighlight(self):
        """Clear production mode visual indicators"""
        self.canvas.delete('production_highlight')
        self.production_highlight_tags.clear()
    
    def openpunches(self):
        """
        Extract list of open (non-closed) punches from Excel Punch Sheet.
        FUNCTIONAL USE: Reads from row 9 onwards, identifies punches without 'Closed By' entry.
        Returns punch details for production mode navigation and implementation tracking.
        """
        punches = []
        
        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches
        
        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
        
        row = 9
        while True:
            checked = self.read_cell(ws, row, 'E')
            if not checked:
                row += 1
                if row > 2000:
                    break
                continue
            
            closed = self.read_cell(ws, row, self.punch_cols['closed_name'])
            if closed:
                row += 1
                continue
            
            implemented = bool(self.read_cell(ws, row, self.punch_cols['implemented_name']))
            sr = self.read_cell(ws, row, self.punch_cols['sr_no'])
            
            punches.append({
                'sr_no': sr,
                'row': row,
                'ref_no': self.read_cell(ws, row, self.punch_cols['ref_no']),
                'punch_text': self.read_cell(ws, row, self.punch_cols['desc']),
                'category': self.read_cell(ws, row, self.punch_cols['category']),
                'implemented': implemented
            })
            
            row += 1
            if row > 2000:
                break
        
        wb.close()
        return punches

    # ================================================================
    # TOOL MODES - PEN, TEXT, UNDO
    # ================================================================
    
    def settlmd(self, mode):
        """
        Activate tool mode: pen for freehand drawing or text for text annotations.
        FUNCTIONAL USE: Sets current drawing tool (None, 'pen', 'text') for annotation workflow.
        Updates UI button states to reflect active tool.
        Args: mode - String ('pen' or 'text') or None to deactivate
        """
        # Deactivate highlighter if active (not applicable in production tool, but kept for consistency)
        if hasattr(self, 'active_highlighter') and self.active_highlighter:
            self.active_highlighter = None
        
        # Toggle tool mode
        if self.tool_mode == mode:
            self.tool_mode = None
            if mode == "pen":
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
        else:
            self.tool_mode = mode
            if mode == "pen":
                self.pen_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
        
        print(f"Tool mode: {self.tool_mode}")
    
    def deactivate_all(self):
        """
        Disable all active drawing tools and highlighter.
        FUNCTIONAL USE: Clears tool mode, stops active highlighting/drawing, resets canvas.
        Bound to Escape key for quick tool deactivation.
        """
        if self._text_editor is not None:
            self._commit_text_editor()
        if self.tool_mode:
            self.settlmd(self.tool_mode)
        self._clear_text_selection()
        self._panning = False
        self.drawing = False
        self.drawing_type = None
        self.pen_points = []
        self.temp_line_ids = []
        self.display()
    
    def updtoolpane(self):
        """Update annotation statistics - placeholder"""
        pass
    
    def _flash_status(self, message, bg='#10b981'):
        """
        Display temporary status message in status bar with color indication.
        FUNCTIONAL USE: Provides visual feedback for user actions (success, warning, info).
        Message auto-clears after timeout.
        Args: message - Text to display, bg - background color (green for success, orange for warning)
        """
        status_label = tk.Label(
            self.root, 
            text=message, 
            bg=bg, 
            fg='white', 
            font=('Segoe UI', 10, 'bold'),
            padx=25, 
            pady=12,
            relief=tk.FLAT
        )
        status_label.place(relx=0.5, rely=0.08, anchor='center')
        self.root.after(1500, lambda: status_label.destroy())
    
    def clear_temp_drawings(self):
        """
        Delete temporary preview drawings from canvas.
        FUNCTIONAL USE: Clears incomplete pen strokes and text previews when user cancels or switches tools.
        """
        for line_id in self.temp_line_ids:
            try:
                self.canvas.delete(line_id)
            except:
                pass
        self.temp_line_ids.clear()
    
    # ================================================================
    # UNDO FUNCTIONALITY
    # ================================================================
    
    def addtoundostck(self, action_type, annotation):
        """
        Push annotation action onto undo stack for later reversal.
        FUNCTIONAL USE: Maintains undo history limited to 50 most recent actions.
        Allows user to revert mistakes with Ctrl+Z.
        Args: action_type - String ('add', 'delete', 'modify'), annotation - Annotation data
        """
        self.undo_stack.append({
            'type': action_type,
            'annotation': annotation.copy()
        })
        
        if len(self.undo_stack) > self.max_undo:
            self.undo_stack.pop(0)
    
    def undolast(self):
        """
        Reverse most recent annotation change from undo stack.
        FUNCTIONAL USE: Removes last action and redraws canvas to show previous state.
        Bound to Ctrl+Z for quick access.
        """
        if not self.undo_stack:
            messagebox.showinfo("Nothing to Undo", "No actions to undo.", icon='info')
            return
        
        last_action = self.undo_stack.pop()
        
        if last_action['type'] == 'add_annotation':
            annotation = last_action['annotation']
            if annotation in self.annotations:
                self.annotations.remove(annotation)
                self.schedule_annotation_save()
                self.display()
                self._flash_status("✓ Annotation removed", bg='#10b981')
        
        self.updtoolpane()
    
    # ================================================================
    # MOUSE EVENT HANDLERS - PEN AND TEXT (TOUCH-FRIENDLY)
    # ================================================================
    
    def leftclick(self, event):
        """Start pen/text interaction, select a text box, or pan with no tool."""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return "break"
        if self._text_editor is not None:
            self._commit_text_editor()
        x = self.canvas.canvasx(event.x); y = self.canvas.canvasy(event.y)
        if self.tool_mode != 'pen':
            hit, mode = self._hit_test_text_box(x, y)
            if hit is not None:
                self.selected_annotation = hit
                self._text_transform_mode = mode
                self._text_transform_start = (x, y)
                self._text_transform_original_bbox = tuple(hit.get('bbox_page'))
                self.drawing = True; self.drawing_type = 'text_transform'
                self._draw_text_selection(hit)
                return "break"
        if self.tool_mode == 'pen':
            self.drawing = True; self.drawing_type = 'pen'
            self.pen_points = [(x, y)]; self.clear_temp_drawings()
            self._touch_scroll_lock_until = time.monotonic() + 5.0
            return "break"
        if self.tool_mode == 'text':
            self.drawing = True; self.drawing_type = 'text_box'
            self._text_box_start = (x, y)
            self._touch_scroll_lock_until = time.monotonic() + 5.0
            return "break"
        self._clear_text_selection()
        self.canvas.scan_mark(event.x, event.y)
        self._panning = True
        return "break"

    def leftdrag(self, event):
        if self._panning and not self.drawing:
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            return "break"
        if not self.drawing: return "break"
        self._pending_drag_event = event
        if not self._drag_frame_scheduled:
            self._drag_frame_scheduled = True
            self.root.after(16, self._process_pen_drag_frame)
        return "break"

    def _process_pen_drag_frame(self):
        self._drag_frame_scheduled = False
        event = self._pending_drag_event; self._pending_drag_event = None
        if event is None or not self.drawing: return
        x = self.canvas.canvasx(event.x); y = self.canvas.canvasy(event.y)
        if self.drawing_type == 'pen':
            if self.pen_points:
                lx, ly = self.pen_points[-1]
                if (x-lx)**2 + (y-ly)**2 < 36: return
            self.pen_points.append((x, y))
            if self.temp_line_ids:
                self.canvas.coords(self.temp_line_ids[0], *[v for pt in self.pen_points for v in pt])
            elif len(self.pen_points) >= 2:
                self.temp_line_ids.append(self.canvas.create_line(
                    *[v for pt in self.pen_points for v in pt], fill='red', width=3,
                    capstyle=tk.ROUND, joinstyle=tk.ROUND, smooth=True))
        elif self.drawing_type == 'text_box':
            sx, sy = self._text_box_start
            if self._text_box_preview_id is None:
                self._text_box_preview_id = self.canvas.create_rectangle(
                    sx, sy, x, y, outline='#2563eb', width=2, dash=(5,3), tags=('text_ui',))
            else: self.canvas.coords(self._text_box_preview_id, sx, sy, x, y)
        elif self.drawing_type == 'text_transform':
            self._update_text_transform(x, y)

    def leftrls(self, event):
        if self._panning:
            self._panning = False
            if not self.drawing: return "break"
        if not self.pdf_document or not self.drawing: return "break"
        if self.drawing_type == 'pen':
            if len(self.pen_points) >= 2:
                ann={'type':'pen','page':self.current_page,
                     'points':self.display_to_page_coords(self.pen_points),
                     'timestamp':datetime.now().isoformat(), 'created_by_role':'production'}
                self.annotations.append(ann); self.addtoundostck('add_annotation',ann)
                self.schedule_annotation_save()
            self.pen_points=[]; self.clear_temp_drawings(); self.drawing=False; self.drawing_type=None
            self.display(); return "break"
        if self.drawing_type == 'text_box':
            x=self.canvas.canvasx(event.x); y=self.canvas.canvasy(event.y)
            sx,sy=self._text_box_start; x1,x2=sorted((sx,x)); y1,y2=sorted((sy,y))
            if self._text_box_preview_id is not None:
                self.canvas.delete(self._text_box_preview_id); self._text_box_preview_id=None
            if x2-x1 < 80: x2=x1+220
            if y2-y1 < 35: y2=y1+90
            ann={'type':'text','page':self.current_page,
                 'bbox_page':self.bbox_display_to_page((x1,y1,x2,y2)),
                 'pos_page':self.display_to_page_coords((x1,y1)), 'text':'',
                 'font_size':12,'timestamp':datetime.now().isoformat(), 'created_by_role':'production'}
            self.annotations.append(ann); self.addtoundostck('add_annotation',ann)
            self.drawing=False; self.drawing_type=None; self.display(); self._open_text_editor(ann,False)
            return "break"
        if self.drawing_type == 'text_transform':
            self.drawing=False; self.drawing_type=None; self._text_transform_mode=None
            self.schedule_annotation_save()
            self.display(); self._draw_text_selection(self.selected_annotation); return "break"
        return "break"

    def _text_bbox_display(self, ann):
        bbox=ann.get('bbox_page')
        if not bbox:
            x,y=ann.get('pos_page',(0,0)); bbox=(x,y,x+140,y+55); ann['bbox_page']=bbox
        return self.bbox_page_to_display(bbox)

    def _hit_test_text_box(self,x,y):
        r=12
        for ann in reversed(self.annotations):
            if ann.get('type')!='text' or ann.get('page')!=self.current_page: continue
            x1,y1,x2,y2=self._text_bbox_display(ann)
            hs={'resize_nw':(x1,y1),'resize_n':((x1+x2)/2,y1),'resize_ne':(x2,y1),
                'resize_e':(x2,(y1+y2)/2),'resize_se':(x2,y2),'resize_s':((x1+x2)/2,y2),
                'resize_sw':(x1,y2),'resize_w':(x1,(y1+y2)/2)}
            for mode,(hx,hy) in hs.items():
                if abs(x-hx)<=r and abs(y-hy)<=r: return ann,mode
            if x1<=x<=x2 and y1<=y<=y2: return ann,'move'
        return None,None

    def _clear_text_selection(self):
        for i in self._text_selection_ids:
            try:self.canvas.delete(i)
            except Exception:pass
        self._text_selection_ids=[]; self.selected_annotation=None

    def _draw_text_selection(self,ann):
        self._clear_text_selection()
        if not ann:return
        self.selected_annotation=ann; x1,y1,x2,y2=self._text_bbox_display(ann)
        ids=[self.canvas.create_rectangle(x1,y1,x2,y2,outline='#2563eb',width=2,dash=(5,3),tags=('text_ui',))]
        for hx,hy in [(x1,y1),((x1+x2)/2,y1),(x2,y1),(x2,(y1+y2)/2),(x2,y2),((x1+x2)/2,y2),(x1,y2),(x1,(y1+y2)/2)]:
            ids.append(self.canvas.create_rectangle(hx-5,hy-5,hx+5,hy+5,fill='white',outline='#2563eb',width=2,tags=('text_ui',)))
        self._text_selection_ids=ids

    def _update_text_transform(self,x,y):
        ann=self.selected_annotation
        if not ann:return
        sx,sy=self._text_transform_start; dx=(x-sx)/self.page_to_display_scale(); dy=(y-sy)/self.page_to_display_scale()
        x1,y1,x2,y2=self._text_transform_original_bbox; m=self._text_transform_mode
        if m=='move': x1,y1,x2,y2=x1+dx,y1+dy,x2+dx,y2+dy
        else:
            if m in ('resize_nw','resize_w','resize_sw'):x1+=dx
            if m in ('resize_ne','resize_e','resize_se'):x2+=dx
            if m in ('resize_nw','resize_n','resize_ne'):y1+=dy
            if m in ('resize_sw','resize_s','resize_se'):y2+=dy
            if x2-x1<30: x2=x1+30
            if y2-y1<18: y2=y1+18
        ann['bbox_page']=(x1,y1,x2,y2); ann['pos_page']=(x1,y1); self._draw_text_selection(ann)

    def _open_text_editor(self,ann,select_all=True):
        self._commit_text_editor(); self.display(); x1,y1,x2,y2=self._text_bbox_display(ann)
        ed=tk.Text(self.canvas,wrap=tk.WORD,undo=True,relief=tk.SOLID,borderwidth=2,
                   highlightthickness=1,highlightbackground='#2563eb',highlightcolor='#2563eb',
                   font=('Segoe UI',max(8,int(ann.get('font_size',12)*self.zoom_level))))
        ed.insert('1.0',ann.get('text','')); self._text_editor=ed; self._text_edit_annotation=ann
        self._text_editor_window_id=self.canvas.create_window(x1,y1,anchor=tk.NW,window=ed,width=max(80,x2-x1),height=max(35,y2-y1),tags=('text_ui',))
        ed.focus_set()
        if select_all:ed.tag_add(tk.SEL,'1.0','end-1c')
        ed.bind('<Control-Return>',lambda e:(self._commit_text_editor(),'break'))
        ed.bind('<Escape>',lambda e:(self._cancel_text_editor(),'break'))
        ed.bind('<FocusOut>',lambda e:self.root.after_idle(self._commit_text_editor))

    def _commit_text_editor(self):
        ed=self._text_editor; ann=self._text_edit_annotation
        if ed is None or ann is None:return
        try:value=ed.get('1.0','end-1c').rstrip()
        except tk.TclError:value=ann.get('text','')
        ann['text']=value
        if not value and ann in self.annotations:self.annotations.remove(ann)
        self._destroy_text_editor()
        self.schedule_annotation_save()
        self.display()
        if value:self._draw_text_selection(ann)

    def _cancel_text_editor(self):
        ann=self._text_edit_annotation
        removed = ann is not None and not ann.get('text') and ann in self.annotations
        if removed:self.annotations.remove(ann)
        self._destroy_text_editor()
        if removed:self.schedule_annotation_save()
        self.display()

    def _destroy_text_editor(self):
        if self._text_editor_window_id is not None:
            try:self.canvas.delete(self._text_editor_window_id)
            except Exception:pass
        if self._text_editor is not None:
            try:self._text_editor.destroy()
            except Exception:pass
        self._text_editor=None; self._text_editor_window_id=None; self._text_edit_annotation=None

    def _wrap_text_for_box(self,draw,text,font,width):
        lines=[]
        for para in str(text).splitlines() or ['']:
            words=para.split()
            if not words:lines.append('');continue
            line=words[0]
            for word in words[1:]:
                trial=line+' '+word
                if draw.textlength(trial,font=font)<=width:line=trial
                else:lines.append(line);line=word
            lines.append(line)
        return '\n'.join(lines)

    # ================================================================
    # DISPLAY PAGE - HIGHLIGHTER RENDERING ONLY (NO BOXES)
    # ================================================================
    
    def display(self):
        """
        Render current PDF page on canvas with all annotations (highlighters, pen, text).
        FUNCTIONAL USE: Converts PDF page to image, scales per zoom level, draws all stored annotations.
        Updates page label and redraws complete view after changes.
        """
        if not self.pdf_document:
            self.canvas.delete("all")
            self.page_label.config(text="Page: 0/0")
            return

        try:
            old_x = self.canvas.xview()[0] if self.canvas.bbox("all") else 0.0
            old_y = self.canvas.yview()[0] if self.canvas.bbox("all") else 0.0
            page = self.pdf_document[self.current_page]
            mat = fitz.Matrix(self.page_to_display_scale(), self.page_to_display_scale())
            pix = page.get_pixmap(matrix=mat, annots=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.current_page_image = np.array(img)
            draw = ImageDraw.Draw(img, 'RGBA')

            # Try to load a font for text
            try:
                font_size = max(12, int(14 * self.zoom_level))
                font = ImageFont.truetype("arial.ttf", font_size)
            except:
                font = ImageFont.load_default()

            # Count annotations by type for debugging
            page_annotations = [ann for ann in self.annotations if ann.get('page') == self.current_page]
            
            highlight_count = 0
            error_count = 0
            pen_count = 0
            text_count = 0
            box_count = 0

            for ann in self.annotations:
                if ann.get('page') != self.current_page:
                    continue

                ann_type = ann.get('type')

                # -------- HIGHLIGHTER STROKES (type='highlight' or type='error') --------
                if ann_type in ('highlight', 'error') and 'points_page' in ann:
                    points_page = ann['points_page']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        color_key = ann.get('color', 'yellow')
                        rgba = self.highlighter_colors.get(color_key, self.highlighter_colors['yellow'])['rgba']
                        
                        # Draw thick semi-transparent strokes
                        stroke_width = max(15, int(15 * self.zoom_level))
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill=rgba, width=stroke_width)
                        
                        # Add closed indicator if applicable
                        if ann.get('closed_by'):
                            # Calculate bbox if not present
                            if 'bbox_page' in ann:
                                bbox_display = self.bbox_page_to_display(ann['bbox_page'])
                            else:
                                xs = [p[0] for p in points_page]
                                ys = [p[1] for p in points_page]
                                bbox_page = (min(xs), min(ys), max(xs), max(ys))
                                bbox_display = self.bbox_page_to_display(bbox_page)
                            
                            cx = bbox_display[0] + 8
                            cy = bbox_display[1] + 8
                            draw.ellipse([cx - 6, cy - 6, cx + 6, cy + 6], fill=(0, 128, 0, 200))
                        
                        if ann_type == 'highlight':
                            highlight_count += 1
                        else:
                            error_count += 1

                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points_page = ann['points']
                    if len(points_page) >= 2:
                        points_display = self.page_to_display_coords(points_page)
                        stroke_width = max(2, int(3 * self.zoom_level))
                        for i in range(len(points_display) - 1):
                            x1, y1 = points_display[i]
                            x2, y2 = points_display[i + 1]
                            draw.line([x1, y1, x2, y2], fill='red', width=stroke_width)
                        pen_count += 1

                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text':
                    text = ann.get('text', '')
                    bbox = ann.get('bbox_page')
                    if not bbox:
                        px, py = ann.get('pos_page', (0, 0)); bbox = (px, py, px+140, py+55); ann['bbox_page'] = bbox
                    x1, y1, x2, y2 = self.bbox_page_to_display(bbox)
                    if text:
                        wrapped = self._wrap_text_for_box(draw, text, font, max(20, x2-x1-8))
                        draw.rectangle([x1,y1,x2,y2], fill=(255,255,255,220))
                        draw.multiline_text((x1+4,y1+3), wrapped, fill='red', font=font, spacing=3)
                        text_count += 1
                
                # -------- BOX ANNOTATIONS - REMOVED (counting for debugging only) --------
                elif ann_type == 'box':
                    box_count += 1
                    print(f"  ⚠️ Skipping box annotation (boxes are disabled)")

            self.photo = ImageTk.PhotoImage(img)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo, tags=("pdf_page",))
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            self.canvas.xview_moveto(old_x)
            self.canvas.yview_moveto(old_y)
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {e}")
            import traceback
            traceback.print_exc()

    # ================================================================
    # COORDINATE CONVERSION HELPERS
    # ================================================================
    
    def getnextsr(self):
        """
        Calculate next available punch serial number from Excel Punch Sheet.
        FUNCTIONAL USE: Scans rows 9+ to find highest SR No, returns next sequential number.
        Used when creating new punch entries.
        """
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 1
            
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            last_sr_no = 0
            row_num = 9
            
            while row_num <= ws.max_row + 5:
                val = self.read_cell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                try:
                    last_sr_no = int(val)
                except:
                    pass
                row_num += 1
            
            wb.close()
            return last_sr_no + 1
        except Exception:
            return 1
    
    def page_to_display_scale(self):
        """
        Calculate scaling factor from PDF page coordinates to display canvas.
        FUNCTIONAL USE: Accounts for zoom level (default 2x magnification plus user zoom).
        Used to convert between PDF space and canvas rendering space.
        """
        return 2.0 * self.zoom_level
    
    def display_to_page_coords(self, pts):
        """
        Convert canvas display coordinates back to PDF page space.
        FUNCTIONAL USE: Reverses scaling from display_scale to find annotation position in PDF.
        Handles single point tuple or list of points.
        Args: pts - Single (x, y) tuple or list of [(x1, y1), (x2, y2), ...]
        Returns: Same structure but with page-space coordinates
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] / scale, pts[1] / scale)
        
        # Handle list of points
        return [(x / scale, y / scale) for x, y in pts]
    
    def page_to_display_coords(self, pts):
        """
        Convert PDF page coordinates to canvas display space.
        FUNCTIONAL USE: Scales from PDF space to display space for rendering annotations on canvas.
        Handles single point tuple or list of points.
        Args: pts - Single (x, y) tuple or list of [(x1, y1), (x2, y2), ...]
        Returns: Same structure but with display-space coordinates
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] * scale, pts[1] * scale)
        
        # Handle list of points
        return [(x * scale, y * scale) for x, y in pts]
    
    def bbox_page_to_display(self, bbox_page):
        """
        Convert bounding box from PDF coordinates to display coordinates.
        FUNCTIONAL USE: Scales rectangle coordinates for rendering on canvas.
        Args: bbox_page - Tuple (x1, y1, x2, y2) in PDF space
        Returns: Tuple (x1, y1, x2, y2) in display space
        """
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_page
        return (x1 * scale, y1 * scale, x2 * scale, y2 * scale)
    
    def bbox_display_to_page(self, bbox_display):
        """
        Convert bounding box from display coordinates to PDF coordinates.
        FUNCTIONAL USE: Reverses scaling to find annotation position in original PDF.
        Args: bbox_display - Tuple (x1, y1, x2, y2) in display space
        Returns: Tuple (x1, y1, x2, y2) in PDF space
        """
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_display
        return (x1 / scale, y1 / scale, x2 / scale, y2 / scale)
    
    # ================================================================
    # ROTATION TRANSFORMATION METHODS FOR PDF EXPORT
    # ================================================================
    
    def transform_bbox_for_rotation(self, rect, page):
        """
        Adjust annotation bbox when PDF page has rotation metadata.
        FUNCTIONAL USE: Handles PDFs with /Rotate property by transforming coordinates accordingly.
        Ensures annotations align correctly on rotated pages.
        Args: rect - Bounding box (x1, y1, x2, y2), page - PyMuPDF page object
        Returns: Transformed bounding box adjusted for page rotation
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x1, y1, x2, y2 = rect

        if r == 0:
            return fitz.Rect(x1, y1, x2, y2)
        if r == 90:
            return fitz.Rect(y1, w - x2, y2, w - x1)
        if r == 180:
            return fitz.Rect(w - x2, h - y2, w - x1, h - y1)
        if r == 270:
            return fitz.Rect(h - y2, x1, h - y1, x2)

        return fitz.Rect(x1, y1, x2, y2)

    def transform_point_for_rotation(self, point, page):
        """
        Adjust single point coordinates when PDF page has rotation metadata.
        FUNCTIONAL USE: Handles /Rotate metadata on rotated PDF pages.
        Ensures text annotations and marks position correctly.
        Args: point - (x, y) coordinate, page - PyMuPDF page object
        Returns: Transformed (x, y) adjusted for page rotation
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x, y = point

        if r == 0:
            return fitz.Point(x, y)
        elif r == 90:
            return fitz.Point(y, w - x)
        elif r == 180:
            return fitz.Point(w - x, h - y)
        elif r == 270:
            return fitz.Point(h - y, x)
        
        return fitz.Point(x, y)

    def transform_highlight_points_for_rotation(self, points, page):
        """
        Adjust list of points for highlighter stroke when page has rotation.
        FUNCTIONAL USE: Handles /Rotate metadata for multi-point annotations like pen strokes.
        Transforms entire stroke to align with rotated page.
        Args: points - List of (x, y) tuples, page - PyMuPDF page object
        Returns: List of transformed (x, y) tuples
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        
        transformed_points = []
        
        for point in points:
            x, y = point
            
            if r == 0:
                transformed_points.append(fitz.Point(x, y))
            elif r == 90:
                transformed_points.append(fitz.Point(y, w - x))
            elif r == 180:
                transformed_points.append(fitz.Point(w - x, h - y))
            elif r == 270:
                transformed_points.append(fitz.Point(h - y, x))
            else:
                transformed_points.append(fitz.Point(x, y))
        
        return transformed_points
    
    def setup_quality_zoom_controls(self):
        """Install the same non-toolbar zoom controls used by Quality."""
        self.root.bind_all("<Control-MouseWheel>", self._on_ctrl_scroll_zoom, add="+")
        self.root.bind_all("<Control-Button-4>", self._on_ctrl_scroll_zoom, add="+")
        self.root.bind_all("<Control-Button-5>", self._on_ctrl_scroll_zoom, add="+")
        # Pinch-to-zoom itself is wired up in setup_canvas_pinch_zoom(), called
        # from _apply_touch_gesture_fix() once the canvas HWND is ready -
        # matching Quality exactly.

    def _update_zoom_toolbar_label(self):
        """Production has no docked zoom label; keep floating percentage in sync."""
        var = getattr(self, '_zoom_slider_pct_var', None)
        if var is not None:
            try: var.set(f"{int(round(self.zoom_level * 100))}%")
            except tk.TclError: pass

    def step_zoom(self, delta):
        value = round((self.zoom_level + delta) / self.ZOOM_STEP) * self.ZOOM_STEP
        self.set_zoom_level(value, immediate=True)

    def set_zoom_level(self, new_zoom, immediate=False, low_res=False, full_render=False):
        """Quality-style zoom update with immediate partial rendering."""
        new_zoom = max(self.ZOOM_MIN, min(self.ZOOM_MAX, float(new_zoom)))
        if abs(new_zoom - self.zoom_level) < 0.001 and not full_render:
            return
        self.zoom_level = new_zoom
        self._update_zoom_toolbar_label()
        if not self.pdf_document:
            return
        if immediate:
            if self._zoom_render_after_id is not None:
                try: self.root.after_cancel(self._zoom_render_after_id)
                except Exception: pass
                self._zoom_render_after_id = None
            self._do_zoom_render(current_page_only=not full_render, low_res=low_res)
        else:
            if self._zoom_render_after_id is not None:
                try: self.root.after_cancel(self._zoom_render_after_id)
                except Exception: pass
            self._zoom_render_after_id = self.root.after(
                33, lambda: self._do_zoom_render(current_page_only=True, low_res=low_res))

    def _do_zoom_render(self, current_page_only=False, low_res=False):
        self._zoom_render_after_id = None
        if current_page_only:
            self._render_current_page_only(low_res=low_res)
        else:
            self.display()

    def _render_current_page_only(self, low_res=False):
        """Fast live zoom preview. Full annotations return on final render.

        Latency fix (ported from Quality): for the low_res live-preview path
        used during an active pinch gesture, a single base pixmap is
        rasterized once per gesture (at gesture-start zoom) and cached in
        self._pinch_base_image; every subsequent live frame just resizes
        that cached PIL image with PIL's cheap NEAREST resize instead of
        calling PyMuPDF's get_pixmap() again. Repeated PDF rasterization was
        the main per-frame cost that made live pinch feel like it was
        lagging behind the fingers.
        """
        if not self.pdf_document:
            return
        try:
            old_x = self.canvas.xview()[0] if self.canvas.bbox("all") else 0.0
            old_y = self.canvas.yview()[0] if self.canvas.bbox("all") else 0.0
            page = self.pdf_document[self.current_page]
            scale = self.page_to_display_scale()

            if low_res:
                # Reuse the cached base image for this gesture if we already
                # captured one for this page; otherwise capture it now. The
                # base is rasterized at a fixed, modest internal resolution
                # (independent of the live target zoom) so it stays valid -
                # and cheap to resize - for the whole gesture even as zoom
                # keeps changing.
                if (self._pinch_base_image is None or
                        self._pinch_base_page != self.current_page):
                    base_scale = scale * 0.5
                    mat = fitz.Matrix(base_scale, base_scale)
                    pix = page.get_pixmap(matrix=mat, alpha=False, annots=False)
                    self._pinch_base_image = Image.frombytes(
                        "RGB", [pix.width, pix.height], pix.samples
                    )
                    self._pinch_base_scale = base_scale
                    self._pinch_base_page = self.current_page

                # Scale the cached base image (cheap, no PDF rasterization)
                # to match how far the live zoom has moved since capture.
                ratio = scale / self._pinch_base_scale if self._pinch_base_scale else 1.0
                base_w, base_h = self._pinch_base_image.size
                target_w = max(1, int(base_w * ratio))
                target_h = max(1, int(base_h * ratio))
                img = self._pinch_base_image.resize((target_w, target_h), Image.Resampling.NEAREST)
            else:
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat, alpha=False, annots=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            photo = ImageTk.PhotoImage(img)
            self.canvas.delete("pdf_page")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=photo, tags=("pdf_page",))
            self.photo = photo
            self.canvas.config(scrollregion=(0, 0, img.width, img.height))
            self.canvas.xview_moveto(old_x)
            self.canvas.yview_moveto(old_y)
            self.page_label.config(text=f"Page: {self.current_page + 1}/{len(self.pdf_document)}")
        except Exception as e:
            print(f"[WARN] Partial zoom render failed: {e}")

    def _on_ctrl_scroll_zoom(self, event):
        if not self.pdf_document or not self._is_pointer_over_canvas():
            return
        delta = 0
        if getattr(event, 'num', None) == 4: delta = 1
        elif getattr(event, 'num', None) == 5: delta = -1
        elif getattr(event, 'delta', 0): delta = 1 if event.delta > 0 else -1
        if delta == 0: return "break"
        self.step_zoom(delta * self.ZOOM_STEP)
        return "break"

    # ================================================================
    # PINCH-TO-ZOOM (ported from Quality "as it is")
    # ================================================================

    def setup_canvas_pinch_zoom(self):
        """Bind Tk-managed pinch events without installing a native WNDPROC.

        Tk dispatches this callback on its own UI thread, avoiding the Python
        3.14 GIL crash caused by calling Python from a Windows window-procedure
        callback. Unsupported Tk builds simply ignore the virtual event.
        """
        if getattr(self, '_safe_pinch_bound', False):
            return
        try:
            self.canvas.bind('<<TouchpadPinch>>', self._on_pinch_zoom, add='+')
            # Most Windows Tk touch drivers promote the final finger-up to a
            # ButtonRelease event. Bind with add='+' so normal drawing/panning
            # release handlers remain intact.
            self.canvas.bind('<ButtonRelease-1>', self._on_safe_pinch_release, add='+')
            self._safe_pinch_bound = True
        except tk.TclError as exc:
            self._safe_pinch_bound = False
            print(f'[WARN] Tk pinch event unavailable: {exc}')

    def _on_pinch_zoom(self, event):
        """Track Tk pinch magnitude proportionally with immediate visual feedback.

        Latency fix: the zoom-level math still runs on every event (cheap),
        but the actual canvas redraw is throttled to a ~60fps budget and
        deferred via after_idle instead of calling update_idletasks()
        synchronously on every single touch event. Forcing a full Tk flush
        per-event was the main source of pinch lag on fast digitizers.
        """
        if (not self.pdf_document or self.active_highlighter or
                self.tool_mode is not None or self.drawing):
            return 'break'
        try:
            now = time.monotonic()
            raw = float(getattr(event, 'delta', 0.0))
            if raw == 0.0:
                return 'break'

            # A pause means a new physical pinch gesture. Reset only the raw
            # sample history, not the current zoom level.
            if now - self._safe_pinch_last_time > 0.25:
                self._safe_pinch_last_raw = None
                self._safe_pinch_accumulator = 1.0
                self._pinch_base_image = None  # force a fresh base capture
            self._safe_pinch_last_time = now

            # Tk builds expose pinch values differently. If consecutive values
            # look cumulative, use their ratio/difference. Otherwise treat the
            # value as an incremental gesture delta. Exponential scaling makes
            # equal finger movement produce equal proportional zoom movement.
            previous = self._safe_pinch_last_raw
            self._safe_pinch_last_raw = raw
            if previous is not None and raw > 0 and previous > 0:
                ratio = raw / previous
                if 0.70 <= ratio <= 1.40 and abs(ratio - 1.0) > 0.0005:
                    factor = ratio
                else:
                    factor = pow(1.0018, raw)
            else:
                if abs(raw) >= 10.0:
                    factor = pow(1.0018, raw)
                elif abs(raw) > 1.0:
                    factor = pow(1.018, raw)
                else:
                    factor = pow(2.0, raw)

            # Reject only impossible driver spikes, while retaining the actual
            # pinch amount for normal events.
            factor = max(0.70, min(1.40, factor))
            target = max(
                self.ZOOM_MIN,
                min(self.ZOOM_MAX, self.zoom_level * factor)
            )
            if abs(target - self.zoom_level) < 0.0005:
                return 'break'

            self.zoom_level = target
            self._native_pinch_last_zoom = target
            self._safe_pinch_active = True
            self._update_zoom_toolbar_label()
            self._panning = False

            # Throttle the actual redraw to a real frame budget instead of
            # painting on every raw event - the digitizer can fire well past
            # 100Hz, but repainting that often buys nothing visually and is
            # exactly what caused the lag. Only schedule a frame if one isn't
            # already pending, and skip if we're still inside this frame's
            # minimum interval.
            if not self._pinch_frame_pending:
                elapsed = now - self._pinch_last_frame_time
                delay_ms = 0 if elapsed >= self._pinch_frame_min_interval else \
                    int((self._pinch_frame_min_interval - elapsed) * 1000)
                self._pinch_frame_pending = True
                self.root.after(delay_ms, self._render_pinch_frame)

            # Final quality is normally applied by _on_safe_pinch_release.
            # This long timer is only a driver fallback when no release event
            # is exposed by Tk; it is deliberately not the normal finish path.
            if self._native_pinch_render_after_id is not None:
                try:
                    self.root.after_cancel(self._native_pinch_render_after_id)
                except Exception:
                    pass
            self._native_pinch_render_after_id = self.root.after(
                self._safe_pinch_watchdog_ms,
                self._finish_safe_pinch_render
            )
        except Exception as exc:
            print(f'[WARN] Tk pinch event failed: {exc}')
        return 'break'

    def _render_pinch_frame(self):
        """Paint exactly one throttled live-preview frame during an active pinch."""
        self._pinch_frame_pending = False
        if not self._safe_pinch_active:
            return
        self._pinch_last_frame_time = time.monotonic()
        self._render_current_page_only(low_res=True)

    def _on_safe_pinch_release(self, event=None):
        """Finalize full-quality zoom when the fingers are released."""
        if not self._safe_pinch_active:
            return
        self._finish_safe_pinch_render()
        return 'break'

    def _finish_safe_pinch_render(self):
        """Finalize one active pinch with a single full-quality render."""
        pending_id = self._native_pinch_render_after_id
        self._native_pinch_render_after_id = None
        if pending_id is not None:
            try:
                self.root.after_cancel(pending_id)
            except Exception:
                pass

        if not self._safe_pinch_active:
            return
        self._safe_pinch_active = False

        target = self._native_pinch_last_zoom
        self._native_pinch_last_zoom = None
        self._safe_pinch_last_raw = None
        self._safe_pinch_accumulator = 1.0
        self._pinch_base_image = None  # drop the cached preview source, gesture is over
        self._pinch_base_scale = None
        self._pinch_base_page = None
        if self.pdf_document and target is not None:
            self.set_zoom_level(target, immediate=True, full_render=True)

    def doubleclick(self, event):
        if not self.pdf_document: return "break"
        x = self.canvas.canvasx(event.x); y = self.canvas.canvasy(event.y)
        hit, _ = self._hit_test_text_box(x, y)
        if hit is not None:
            self._open_text_editor(hit, False)
            return "break"
        if self.tool_mode in ("pen", "text"):
            return "break"
        self.drawing = False
        self.clear_temp_drawings()
        self.show_zoom_slider(event)
        return "break"

    def doubleright(self, event):
        return "break"

    def show_zoom_slider(self, event=None):
        if getattr(self, 'zoom_slider_frame', None) is not None:
            try:
                self.zoom_slider_frame.lift()
                return
            except tk.TclError:
                self.zoom_slider_frame = None
        self._unbind_zoom_root_release()
        MIN_ZOOM, MAX_ZOOM = self.ZOOM_MIN, self.ZOOM_MAX
        overlay_width, overlay_height = 70, 260
        if event is not None:
            pos_x = min(max(event.x_root-self.root.winfo_rootx()-overlay_width//2, 10),
                        self.root.winfo_width()-overlay_width-10)
            pos_y = min(max(event.y_root-self.root.winfo_rooty()-overlay_height//2, 10),
                        self.root.winfo_height()-overlay_height-10)
        else:
            pos_x, pos_y = 40, 40
        frame = tk.Frame(self.root, bg='#1e293b', bd=2, relief=tk.RIDGE)
        frame.place(x=pos_x, y=pos_y, width=overlay_width, height=overlay_height)
        self.zoom_slider_frame = frame
        close_btn = tk.Button(frame, text="✕", font=('Segoe UI',9,'bold'), bg='#ef4444',
                              fg='white', relief=tk.FLAT, bd=0, width=2, height=1,
                              cursor='hand2', command=self.close_zoom_slider)
        close_btn.place(x=overlay_width-26, y=4)
        close_btn.bind("<ButtonPress-1>", lambda e: (self.close_zoom_slider(), "break"))
        pct_var = tk.StringVar(value=f"{int(self.zoom_level*100)}%")
        self._zoom_slider_pct_var = pct_var
        tk.Label(frame, textvariable=pct_var, bg='#1e293b', fg='white',
                 font=('Segoe UI',10,'bold')).place(x=6,y=30)
        track_top, track_bottom = 60, overlay_height-20
        track_height = track_bottom-track_top
        track_x_center = overlay_width//2
        track = tk.Canvas(frame, bg='#1e293b', width=overlay_width,
                          height=track_height+20, highlightthickness=0)
        track.place(x=0,y=track_top-10)
        track.create_line(track_x_center,10,track_x_center,track_height+10,
                          fill='#475569',width=4,capstyle=tk.ROUND)
        def zoom_to_y(z): return 10+(1-(z-MIN_ZOOM)/(MAX_ZOOM-MIN_ZOOM))*track_height
        def y_to_zoom(y): return MAX_ZOOM-max(0,min(1,(y-10)/track_height))*(MAX_ZOOM-MIN_ZOOM)
        radius=9; hy=zoom_to_y(self.zoom_level)
        handle=track.create_oval(track_x_center-radius,hy-radius,track_x_center+radius,hy+radius,
                                 fill='#3b82f6',outline='white',width=2)
        self._zoom_slider_generation += 1
        generation=self._zoom_slider_generation
        self._zoom_is_dragging=False
        state={'event':None,'scheduled':False}
        def current(): return self._zoom_slider_generation==generation and self.zoom_slider_frame is frame
        def process():
            state['scheduled']=False
            if not current(): return
            ev=state['event']; state['event']=None
            if ev is None: return
            y=max(10,min(track_height+10,ev.y))
            z=round(y_to_zoom(y)/0.05)*0.05
            if abs(z-self.zoom_level)<0.001: return
            track.coords(handle,track_x_center-radius,y-radius,track_x_center+radius,y+radius)
            # Immediate, Quality-identical low-resolution preview on every frame.
            self.zoom_level=z; pct_var.set(f"{int(z*100)}%")
            self._render_current_page_only(low_res=True)
        def drag(ev):
            if not current(): return
            self._zoom_is_dragging=True; state['event']=ev
            if not state['scheduled']:
                state['scheduled']=True
                self.root.after(33,process)
        def release(ev):
            if not current() or not self._zoom_is_dragging: return
            self._zoom_is_dragging=False
            if self._zoom_render_after_id is not None:
                try:self.root.after_cancel(self._zoom_render_after_id)
                except Exception:pass
                self._zoom_render_after_id=None
            self._do_zoom_render(current_page_only=False)
        track.tag_bind(handle,"<B1-Motion>",drag)
        track.bind("<Button-1>",drag)
        track.bind("<B1-Motion>",drag)
        track.bind("<ButtonRelease-1>",release)
        self._zoom_root_release_funcid=self.root.bind("<ButtonRelease-1>",release,add="+")

    def _unbind_zoom_root_release(self):
        fid=getattr(self,'_zoom_root_release_funcid',None)
        if fid:
            try:self.root.unbind("<ButtonRelease-1>",fid)
            except Exception:pass
            self._zoom_root_release_funcid=None

    def close_zoom_slider(self):
        if self._zoom_render_after_id is not None:
            try:self.root.after_cancel(self._zoom_render_after_id)
            except Exception:pass
            self._zoom_render_after_id=None
        self._zoom_is_dragging=False
        self._zoom_slider_generation += 1
        self._unbind_zoom_root_release()
        frame=getattr(self,'zoom_slider_frame',None)
        if frame is not None:
            try:frame.destroy()
            except tk.TclError:pass
        self.zoom_slider_frame=None; self._zoom_slider_pct_var=None
        self.display()

    def prev(self):
        """
        Navigate to previous page in PDF.
        FUNCTIONAL USE: Decrements current_page and redraws display.
        Bound to arrow button in toolbar.
        """
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display()
    
    def next(self):
        """
        Navigate to next page in PDF.
        FUNCTIONAL USE: Increments current_page and redraws display.
        Bound to arrow button in toolbar.
        """
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display()
    
    # ================================================================
    # SESSION MANAGEMENT - HIGHLIGHTER COMPATIBLE
    # ================================================================
    
    def getsesspathforpdf(self):
        """
        Generate session file path for current PDF.
        FUNCTIONAL USE: Creates .json path in cabinet session directory for storing annotations.
        Used for saving/loading work between sessions.
        """
        if not self.current_pdf_path or not self.cabinet_id:
            return None
        if self.current_session_path and os.path.exists(self.current_session_path):
            return self.current_session_path
        
        if hasattr(self, 'storage_location') and self.storage_location:
            project_folder = os.path.join(
                self.storage_location,
                self.project_name.replace(' ', '_')
            )
            cabinet_root = os.path.join(
                project_folder,
                self.cabinet_id.replace(' ', '_')
            )
            session_path = os.path.join(
                cabinet_root,
                "Sessions",
                f"{self.cabinet_id}_annotations.json"
            )
            
            return session_path if os.path.exists(session_path) else None
        
        return None
    
    def sync_production_annotations_to_base_pdf(self):
        """Embed production-created pen and text markups in the loaded base PDF."""
        if not self.pdf_document or not self.current_pdf_path or not os.path.exists(self.current_pdf_path):
            return False
        marker = "ProductionTool"
        try:
            for page in self.pdf_document:
                annot = page.first_annot
                while annot:
                    next_annot = annot.next
                    if str((annot.info or {}).get('subject', '')).startswith(marker):
                        page.delete_annot(annot)
                    annot = next_annot

            saved_count = 0
            for ann in self.annotations:
                if ann.get('created_by_role') != 'production':
                    continue
                page_index = ann.get('page')
                if not isinstance(page_index, int) or not 0 <= page_index < len(self.pdf_document):
                    continue
                page = self.pdf_document[page_index]
                if ann.get('type') == 'pen' and len(ann.get('points', [])) >= 2:
                    stroke = [[tuple(map(float, point)) for point in ann['points']]]
                    pdf_annot = page.add_ink_annot(stroke)
                    pdf_annot.set_colors(stroke=(1, 0, 0))
                    pdf_annot.set_border(width=2)
                    pdf_annot.set_info(title=self.logged_in_fullname or "Production",
                                       subject=f"{marker}:pen",
                                       content="Production pen annotation")
                    pdf_annot.update()
                    saved_count += 1
                elif ann.get('type') == 'text' and ann.get('text'):
                    bbox = ann.get('bbox_page')
                    if not bbox:
                        x, y = ann.get('pos_page', (0, 0)); bbox = (x, y, x + 140, y + 55)
                    pdf_annot = page.add_freetext_annot(
                        fitz.Rect(*map(float, bbox)), str(ann['text']),
                        fontsize=max(8, float(ann.get('font_size', 12))),
                        text_color=(1, 0, 0), fill_color=(1, 1, 1),
                        border_width=1, align=0)
                    pdf_annot.set_opacity(0.92)
                    pdf_annot.set_info(title=self.logged_in_fullname or "Production",
                                       subject=f"{marker}:text",
                                       content=str(ann['text']))
                    pdf_annot.update()
                    saved_count += 1

            if not self.pdf_document.can_save_incrementally():
                print("[WARN] Base PDF cannot be saved incrementally; JSON session is still safe.")
                return False
            self.pdf_document.saveIncr()
            print(f"[INFO] Embedded {saved_count} production annotation(s) in base PDF")
            return True
        except Exception as exc:
            print(f"[WARN] Base PDF annotation save failed: {exc}")
            return False

    def savesess(self, sync_pdf=True):
        """
        Serialize all current annotations to session JSON file.
        FUNCTIONAL USE: Writes annotations list, page references, and metadata to file.
        Enables resuming work across production module instances.
        """
        loading_here = bool(sync_pdf and getattr(self, '_loading_depth', 0) == 0)
        if loading_here:
            self.show_loading("Saving production work", "Updating the shared session and base PDF annotations.")
        if self._text_editor is not None:
            self._commit_text_editor()
        if not self.pdf_document:
            if loading_here:
                self.hide_loading(force=True)
            print("⚠️ No PDF loaded - cannot save session")
            return
        
        if not hasattr(self, 'storage_location') or not self.storage_location:
            if loading_here:
                self.hide_loading(force=True)
            print("⚠️ Storage location not set - cannot save session")
            return
        
        # Determine save path
        project_folder = os.path.join(
            self.storage_location,
            self.project_name.replace(' ', '_')
        )
        cabinet_root = os.path.join(
            project_folder,
            self.cabinet_id.replace(' ', '_')
        )
        sessions_dir = os.path.join(cabinet_root, "Sessions")
        
        # Ensure sessions directory exists
        os.makedirs(sessions_dir, exist_ok=True)
        
        canonical_path = os.path.join(
            sessions_dir,
            f"{self.cabinet_id}_annotations.json"
        )
        save_path = self.current_session_path or canonical_path
        save_path = os.path.abspath(save_path)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        self.current_session_path = save_path
        
        data = {
            'project_name': self.project_name,
            'sales_order_no': self.sales_order_no,
            'cabinet_id': self.cabinet_id,
            'pdf_path': self.current_pdf_path,
            'current_page': self.current_page,
            'zoom_level': self.zoom_level,
            'current_sr_no': self.current_sr_no,
            'session_refs': list(self.session_refs),
            'annotations': [],
            'undo_stack_size': len(self.undo_stack) if hasattr(self, 'undo_stack') else 0,
            'save_timestamp': datetime.now().isoformat()
        }
        
        # Process all annotation types
        for ann in self.annotations:
            entry = ann.copy()
            
            # ===== HIGHLIGHTER ANNOTATIONS - Convert tuples to lists =====
            if 'points_page' in entry:
                entry['points_page'] = [[float(x), float(y)] for x, y in entry['points_page']]
            
            # ===== BBOX for highlights =====
            if 'bbox_page' in entry:
                entry['bbox_page'] = [float(x) for x in entry['bbox_page']]
            
            # ===== PEN STROKES - Convert tuples to lists =====
            if 'points' in entry:
                entry['points'] = [[float(x), float(y)] for x, y in entry['points']]
            
            # ===== TEXT ANNOTATIONS - Convert tuple to list =====
            if 'pos_page' in entry:
                pos = entry['pos_page']
                entry['pos_page'] = [float(pos[0]), float(pos[1])]
            
            # Ensure text content is saved
            if 'text' in entry:
                entry['text'] = str(entry['text'])
            
            data['annotations'].append(entry)
        
        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

            if sync_pdf:
                self.sync_production_annotations_to_base_pdf()
            
            # Count annotation types for feedback
            highlight_count = len([a for a in self.annotations if a.get('type') == 'highlight'])
            error_count = len([a for a in self.annotations if a.get('type') == 'error'])
            pen_count = len([a for a in self.annotations if a.get('type') == 'pen'])
            text_count = len([a for a in self.annotations if a.get('type') == 'text'])
            
            print(f"\n✓ Session saved to: {save_path}")
            print(f"Total annotations: {len(self.annotations)}")
            if highlight_count > 0:
                print(f"  Highlights: {highlight_count}")
            if error_count > 0:
                print(f"  Errors: {error_count}")
            if pen_count > 0:
                print(f"  Pen strokes: {pen_count}")
            if text_count > 0:
                print(f"  Text annotations: {text_count}")
            self._session_dirty = False
            if loading_here:
                self.hide_loading()
            
        except Exception as e:
            if loading_here:
                self.hide_loading(force=True)
            print(f"❌ Failed to save session: {e}")
            import traceback
            traceback.print_exc()
    
    def loadsessfrompath(self, path):
        """Load annotation session - FULL HIGHLIGHTER SUPPORT"""
        self.current_session_path = os.path.abspath(path)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Session Load Error", f"Failed to load session:\n{e}")
            return
        
        self.project_name = data.get('project_name', self.project_name)
        self.sales_order_no = data.get('sales_order_no', self.sales_order_no)
        self.cabinet_id = data.get('cabinet_id', getattr(self, "cabinet_id", ""))
        self.current_page = data.get('current_page', 0)
        self.zoom_level = data.get('zoom_level', 1.0)
        self.current_sr_no = data.get('current_sr_no', self.current_sr_no)
        
        # Restore session refs
        self.annotations = []
        self.session_refs = set(data.get('session_refs', []))
        
        highlight_count = 0
        error_count = 0
        pen_count = 0
        text_count = 0
        box_count = 0
        
        for entry in data.get('annotations', []):
            ann = entry.copy()
            ann_type = ann.get('type')
            
            # ===== HIGHLIGHTER ANNOTATIONS - points_page =====
            if 'points_page' in ann:
                ann['points_page'] = [(float(p[0]), float(p[1])) for p in ann['points_page']]
                if ann_type == 'highlight':
                    highlight_count += 1
                elif ann_type == 'error':
                    error_count += 1
            
            # ===== BBOX - Convert list to tuple =====
            if 'bbox_page' in ann:
                ann['bbox_page'] = tuple(float(x) for x in ann['bbox_page'])
            
            # ===== PEN STROKES - points =====
            if 'points' in ann:
                ann['points'] = [(float(p[0]), float(p[1])) for p in ann['points']]
                pen_count += 1
            
            # ===== TEXT ANNOTATIONS - pos_page =====
            if 'pos_page' in ann:
                pos = ann['pos_page']
                ann['pos_page'] = (float(pos[0]), float(pos[1]))
                text_count += 1
            
            # ===== BOX ANNOTATIONS - Count but skip =====
            if ann_type == 'box':
                box_count += 1
                print(f"  ⚠️ Skipping box annotation (type='box') - boxes are disabled")
                continue  # Skip box annotations
            
            # Ensure text content is restored
            if 'text' in ann:
                ann['text'] = str(ann['text'])
            
            self.annotations.append(ann)
            
            # Add ref_no to session refs
            if ann.get('ref_no'):
                self.session_refs.add(str(ann['ref_no']).strip())
        
        self.display()
        
        print(f"\n✓ Session loaded from: {path}")
        print(f"Total annotations loaded: {len(self.annotations)}")
        print(f"  🅰️ Text annotations: {text_count}")
        if box_count > 0:
            print(f"  📦 Box annotations (skipped): {box_count}")
        
        types_loaded = {}
        for ann in self.annotations:
            ann_type = ann.get('type', 'unknown')
            types_loaded[ann_type] = types_loaded.get(ann_type, 0) + 1
        print(f"Annotation types loaded: {types_loaded}\n")


def main():
    root = tk.Tk()
    app = ProductionTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()