import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from PIL import Image, ImageTk
import json
import os
import sys
import subprocess
from datetime import datetime, timedelta
from collections import defaultdict
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import calendar


def get_app_base_dir():
    """
    Resolve the base directory of the application.

    Handles both normal Python execution and frozen executables
    (e.g., PyInstaller builds) to ensure all file paths
    are resolved relative to the correct runtime location.
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_financial_year():
    """
    Determine the current financial year.

    Financial year is assumed to start on October 1st.
    Returns a string in the format 'YYYY-YY' (e.g., '2026-27'),
    which is used across dashboards and analytics.
    """
    today = datetime.now()
    if today.month >= 10:  # October onwards
        return f"{today.year}-{str(today.year + 1)[-2:]}"
    else:
        return f"{today.year - 1}-{str(today.year)[-2:]}"


def get_week_number():
    """
    Return the ISO calendar week number for the current date.

    Used for weekly dashboard statistics and reporting labels.
    """
    today = datetime.now()
    return today.isocalendar()[1]


class ManagerDatabase:
    """
    Database access and business logic layer for the Manager Dashboard.

    Responsible for:
    - SQLite database initialization and queries
    - Reading punch data from Excel files
    - Status determination from Interphase sheets
    - Aggregated statistics for dashboard and analytics views
    """
    def __init__(self, db_path):
        """
        Initialize database connection and Excel column mappings.
    
        Also ensures required tables exist and configures
        punch-sheet column references to stay consistent
        with the Quality Inspection tool.
        """
        self.db_path = db_path
        self.init_database()
        
        # Excel column mapping (same as Quality Inspection tool)
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'checked_name': 'E',
            'checked_date': 'F',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
    
    def init_database(self):
        """
        Create required SQLite tables if they do not already exist.
    
        Tables:
        - cabinets: cabinet-level tracking and status
        - category_occurrences: defect category logging for analytics
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS cabinets (
            cabinet_id TEXT PRIMARY KEY,
            project_name TEXT,
            sales_order_no TEXT,
            total_pages INTEGER DEFAULT 0,
            annotated_pages INTEGER DEFAULT 0,
            total_punches INTEGER DEFAULT 0,
            open_punches INTEGER DEFAULT 0,
            implemented_punches INTEGER DEFAULT 0,
            closed_punches INTEGER DEFAULT 0,
            status TEXT DEFAULT 'quality_inspection',
            created_date TEXT,
            last_updated TEXT,
            storage_location TEXT,
            excel_path TEXT)''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS category_occurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cabinet_id TEXT,
            project_name TEXT,
            category TEXT,
            subcategory TEXT,
            occurrence_date TEXT)''')
        
        conn.commit()
        conn.close()
    
    def split_cell(self, cell_ref):
        """
        Split an Excel-style cell reference into row and column.
    
        Example:
            'F6' → (6, 'F')
    
        Used to safely parse dynamic Excel references.
        """
        import re
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col
    
    def _resolve_merged_target(self, ws, row, col_idx):
        """Handle merged cells"""
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx
    
    def read_cell(self, ws, row, col):
        """Read cell value handling merged cells"""
        from openpyxl.utils import column_index_from_string
        
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self._resolve_merged_target(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value
    
    def count_punches_from_excel(self, excel_path):
        """
        Count punch statistics directly from the Punch Sheet Excel file.
    
        Returns:
            (total_punches, implemented_punches, closed_punches)
    
        This ensures dashboard data always reflects the latest Excel state,
        not stale database values.
        """
        if not excel_path or not os.path.exists(excel_path):
            return (0, 0, 0)
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            if self.punch_sheet_name not in wb.sheetnames:
                wb.close()
                return (0, 0, 0)
            
            ws = wb[self.punch_sheet_name]
            
            total = 0
            implemented = 0
            closed = 0
            
            row = 9  # Start from row 9
            while row <= ws.max_row + 5:
                # Check if this row has a punch (has checked_name)
                checked = self.read_cell(ws, row, self.punch_cols['checked_name'])
                
                if checked:  # This is a logged punch
                    total += 1
                    
                    # Check if implemented
                    impl = self.read_cell(ws, row, self.punch_cols['implemented_name'])
                    if impl:
                        implemented += 1
                    
                    # Check if closed
                    closed_val = self.read_cell(ws, row, self.punch_cols['closed_name'])
                    if closed_val:
                        closed += 1
                
                row += 1
                
                # Safety limit
                if row > 2000:
                    break
            
            wb.close()
            return (total, implemented, closed)
            
        except Exception as e:
            print(f"Error counting punches from Excel: {e}")
            return (0, 0, 0)
    
    def get_status_from_interphase(self, excel_path):
        """
        Determine cabinet workflow status from the Interphase worksheet.
    
        Uses the lowest populated reference number to infer
        the current project phase (assembly, documentation, etc.).
    
        Returns:
            status string or None if unavailable
        """

        if not excel_path or not os.path.exists(excel_path):
            return None
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            # Check if Interphase worksheet exists
            if 'Interphase' not in wb.sheetnames:
                wb.close()
                return None
            
            ws = wb['Interphase']
            
            # Find the lowest filled status cell in column D
            lowest_status_row = None
            lowest_ref_no = None
            
            # Start from row 2 (assuming row 1 is header)
            for row in range(2, ws.max_row + 1):
                status_cell = self.read_cell(ws, row, 'D')
                
                # If status cell has content, check the reference number
                if status_cell:
                    ref_no_cell = self.read_cell(ws, row, 'B')
                    
                    if ref_no_cell:
                        lowest_status_row = row
                        lowest_ref_no = str(ref_no_cell).strip()
            
            wb.close()
            
            # If we found a reference number, determine the status
            if lowest_ref_no:
                try:
                    # Handle range formats like "1-2" or single numbers like "5"
                    if '-' in lowest_ref_no:
                        # Get the first number in the range
                        ref_num = int(lowest_ref_no.split('-')[0])
                    else:
                        ref_num = int(lowest_ref_no)
                    
                    # Determine status based on reference number
                    if 1 <= ref_num <= 2:
                        return 'project_info_sheet'
                    elif 3 <= ref_num <= 9:
                        return 'mechanical_assembly'
                    elif 10 <= ref_num <= 18:
                        return 'component_assembly'
                    elif 19 <= ref_num <= 26:
                        return 'final_assembly'
                    elif 27 <= ref_num <= 31:
                        return 'final_documentation'
                
                except (ValueError, IndexError):
                    # If we can't parse the reference number, return None
                    pass
            
            return None
            
        except Exception as e:
            print(f"Error reading Interphase worksheet: {e}")
            return None
    
    def get_all_projects(self):
        """
        Retrieve all projects with cabinet counts and last update timestamps.
    
        Used to populate the main dashboard project overview.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''SELECT project_name, COUNT(DISTINCT cabinet_id) as count,
                          MAX(last_updated) as updated
                          FROM cabinets
                          GROUP BY project_name
                          ORDER BY updated DESC''')
        projects = [{'project_name': r[0], 'cabinet_count': r[1], 'last_updated': r[2]} 
                   for r in cursor.fetchall()]
        conn.close()
        return projects
    
    def get_cabinets_by_project(self, project_name):
        """
        Retrieve all cabinets for a project with real-time metrics.
    
        Combines:
        - Database metadata
        - Live Excel punch counts
        - Interphase-based status resolution
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''SELECT cabinet_id, project_name, total_pages, annotated_pages,
                          status, excel_path, storage_location
                          FROM cabinets
                          WHERE project_name = ?
                          ORDER BY last_updated DESC''', (project_name,))
        
        cabinets = []
        for row in cursor.fetchall():
            cabinet_id, project_name, total_pages, annotated_pages, db_status, excel_path, storage_location = row
            
            # Get real counts from Excel
            total_punches, implemented_punches, closed_punches = self.count_punches_from_excel(excel_path)
            
            # Try to get status from Interphase worksheet first
            interphase_status = self.get_status_from_interphase(excel_path)
            
            # Use Interphase status if available, otherwise use database status
            # Only override if the database doesn't have a status set by production/quality code
            if interphase_status and db_status in ['quality_inspection', 'project_info_sheet', 
                                                     'mechanical_assembly', 'component_assembly', 
                                                     'final_assembly', 'final_documentation']:
                final_status = interphase_status
            else:
                final_status = db_status
            
            cabinets.append({
                'cabinet_id': cabinet_id,
                'project_name': project_name,
                'total_pages': total_pages or 0,
                'annotated_pages': annotated_pages or 0,
                'total_punches': total_punches,
                'implemented_punches': implemented_punches,
                'closed_punches': closed_punches,
                'status': final_status,
                'excel_path': excel_path,
                'storage_location': storage_location
            })
        
        conn.close()
        return cabinets
    
    def search_projects(self, search_term):
        """
        Search projects by name using partial matching.
    
        Used by the dashboard search bar for instant filtering.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''SELECT project_name, COUNT(DISTINCT cabinet_id) as count,
                          MAX(last_updated) as updated
                          FROM cabinets
                          WHERE project_name LIKE ?
                          GROUP BY project_name
                          ORDER BY updated DESC''', (f'%{search_term}%',))
        projects = [{'project_name': r[0], 'cabinet_count': r[1], 'last_updated': r[2]} 
                   for r in cursor.fetchall()]
        conn.close()
        return projects
    
    def get_all_project_names(self):
        """
        Return a list of unique project names.
    
        Used for analytics auto-suggestions and search hints.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT project_name FROM cabinets ORDER BY project_name')
        projects = [row[0] for row in cursor.fetchall()]
        conn.close()
        return projects
    
    def get_cabinet_statistics(self):
        """
        Calculate cabinet counts across time periods.
    
        Periods:
        - Daily
        - Weekly
        - Monthly
        - Financial Year
    
        Used for top dashboard statistic cards.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        
        # Financial year starts on October 1st
        current_year = today.year
        if today.month >= 10:
            fy_start = datetime(current_year, 10, 1).date()
        else:
            fy_start = datetime(current_year - 1, 10, 1).date()
        
        stats = {}
        
        # Daily count
        cursor.execute('''SELECT COUNT(DISTINCT cabinet_id) FROM cabinets 
                         WHERE DATE(created_date) = ?''', (today.isoformat(),))
        stats['daily'] = cursor.fetchone()[0]
        
        # Weekly count
        cursor.execute('''SELECT COUNT(DISTINCT cabinet_id) FROM cabinets 
                         WHERE DATE(created_date) >= ?''', (week_start.isoformat(),))
        stats['weekly'] = cursor.fetchone()[0]
        
        # Monthly count
        cursor.execute('''SELECT COUNT(DISTINCT cabinet_id) FROM cabinets 
                         WHERE DATE(created_date) >= ?''', (month_start.isoformat(),))
        stats['monthly'] = cursor.fetchone()[0]
        
        # Financial Yearly count
        cursor.execute('''SELECT COUNT(DISTINCT cabinet_id) FROM cabinets 
                         WHERE DATE(created_date) >= ?''', (fy_start.isoformat(),))
        stats['yearly'] = cursor.fetchone()[0]
        
        conn.close()
        return stats
    
    def get_category_stats(self, start_date=None, end_date=None, project_name=None):
        """
        Retrieve aggregated defect category statistics.
    
        Supports optional filtering by:
        - Date range
        - Project name
    
        Forms the backend for Pareto charts and exports.
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = 'SELECT category, subcategory, COUNT(*) as count FROM category_occurrences WHERE 1=1'
        params = []
        
        if start_date:
            query += ' AND occurrence_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND occurrence_date <= ?'
            params.append(end_date)
        
        if project_name:
            query += ' AND project_name = ?'
            params.append(project_name)
        
        query += ' GROUP BY category, subcategory ORDER BY count DESC'
        cursor.execute(query, params)
        stats = [{'category': r[0], 'subcategory': r[1], 'count': r[2]} 
                for r in cursor.fetchall()]
        conn.close()
        return stats


class ManagerUI:
    """
        Main Tkinter UI controller for the Manager Dashboard.
    
        Responsible for:
        - Navigation and layout
        - Dashboard rendering
        - Analytics and charts
        - Defect Library management
        - Template Excel management
    """
    def __init__(self, root):
        """
        Initialize the application UI, database, and configuration files.
    
        Loads:
        - SQLite database
        - Defect categories JSON
        - Excel template paths
        """
        self.root = root
        self.root.title("Manager Dashboard")
        self.root.geometry("1600x900")
        
        base_dir = get_app_base_dir()
        self.db = ManagerDatabase(os.path.join(base_dir, "manager.db"))
        self.category_file = os.path.join(os.path.dirname(base_dir), "assets", "categories.json")
        self.template_excel_file = os.path.join(base_dir, "Emerson.xlsx")
        self.categories = self.load_categories()
        
        self.setup_ui()
        self.show_dashboard()
    
    def load_categories(self):
        """
        Load defect categories from JSON storage.
    
        Returns an empty list if the file does not exist
        or cannot be parsed.
        """
        try:
            if os.path.exists(self.category_file):
                with open(self.category_file, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                    return loaded
        except Exception:
            pass
        return []
    
    def save_categories(self):
        """
        Persist defect categories to disk in JSON format.
    
        Ensures directory creation and safe file writing.
        """
        try:
            os.makedirs(os.path.dirname(self.category_file), exist_ok=True)
            with open(self.category_file, "w", encoding="utf-8") as f:
                json.dump(self.categories, f, indent=2)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")
    
    def setup_ui(self):
        """
        Construct the main application layout.
    
        Includes:
        - Top navigation bar
        - Content container
        - Navigation buttons
        """
        
        # Navigation
        nav = tk.Frame(self.root, bg='#1e293b', height=70)
        nav.pack(side=tk.TOP, fill=tk.X)
        nav.pack_propagate(False)
        
        tk.Label(nav, text="📊 Manager Dashboard", bg='#1e293b', fg='white',
                font=('Segoe UI', 18, 'bold')).pack(side=tk.LEFT, padx=30, pady=15)
        
        btn_style = {'font': ('Segoe UI', 11, 'bold'), 'relief': tk.FLAT,
                    'cursor': 'hand2', 'padx': 25, 'pady': 12}
        
        self.nav_btns = {}
        self.nav_btns['dashboard'] = tk.Button(nav, text="Dashboard",
                                               command=self.show_dashboard,
                                               bg='#3b82f6', fg='white', **btn_style)
        self.nav_btns['dashboard'].pack(side=tk.LEFT, padx=5)
        
        self.nav_btns['analytics'] = tk.Button(nav, text="Analytics",
                                               command=self.show_analytics,
                                               bg='#334155', fg='white', **btn_style)
        self.nav_btns['analytics'].pack(side=tk.LEFT, padx=5)
        
        # RENAMED: Categories -> Defect Library
        self.nav_btns['defect_library'] = tk.Button(nav, text="Defect Library",
                                                command=self.show_defect_library,
                                                bg='#334155', fg='white', **btn_style)
        self.nav_btns['defect_library'].pack(side=tk.LEFT, padx=5)
        
        # NEW: Template Excel Editor
        self.nav_btns['template_editor'] = tk.Button(nav, text="Template Excel",
                                                command=self.show_template_editor,
                                                bg='#334155', fg='white', **btn_style)
        self.nav_btns['template_editor'].pack(side=tk.LEFT, padx=5)
        
        # Content frame
        self.content = tk.Frame(self.root, bg='#f8fafc')
        self.content.pack(fill=tk.BOTH, expand=True)
    
    def set_active_nav(self, key):
        """
        Highlight the active navigation button.
    
        Provides visual feedback for the current view.
        """
        for k, btn in self.nav_btns.items():
            btn.config(bg='#3b82f6' if k == key else '#334155')
    
    def clear_content(self):
        """
        Remove all widgets from the content area.
    
        Used when switching between views.
        """

        for w in self.content.winfo_children():
            w.destroy()
    
    # ============ DASHBOARD - WITH PROPER DATE DISPLAYS AND SEARCH ============
    def show_dashboard(self):
        """
        Render the main dashboard view.
    
        Displays:
        - Time-based cabinet statistics
        - Project list with expandable cabinet details
        - Search functionality
        """
        self.set_active_nav('dashboard')
        self.clear_content()
        
        # Centered container with 70% width
        center_container = tk.Frame(self.content, bg='#f8fafc')
        center_container.place(relx=0.5, rely=0, anchor='n', relwidth=0.7, relheight=1.0)
        
        # Statistics Cards at the top
        stats_frame = tk.Frame(center_container, bg='#f8fafc')
        stats_frame.pack(fill=tk.X, padx=30, pady=(20, 10))
        
        stats = self.db.get_cabinet_statistics()
        today = datetime.now()
        
        # Create 4 stat cards with proper labels
        stat_cards = [
            (today.strftime("%B %d"), stats['daily'], "#3b82f6"),  # December 31
            (f"Week {get_week_number()}", stats['weekly'], "#8b5cf6"),  # Week 52
            (today.strftime("%B"), stats['monthly'], "#10b981"),  # December
            (f"FY {get_financial_year()}", stats['yearly'], "#f59e0b")  # FY 2024-25
        ]
        
        for label, count, color in stat_cards:
            card = tk.Frame(stats_frame, bg='white', relief=tk.SOLID, borderwidth=1)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
            
            tk.Label(card, text=label, font=('Segoe UI', 11, 'bold'), 
                    bg='white', fg='#64748b').pack(pady=(15, 5))
            tk.Label(card, text=str(count), font=('Segoe UI', 28, 'bold'),
                    bg='white', fg=color).pack(pady=(0, 5))
            tk.Label(card, text="Cabinets", font=('Segoe UI', 9),
                    bg='white', fg='#94a3b8').pack(pady=(0, 15))
        
        
        
        projects = self.db.get_all_projects()
        
        if not projects:
            empty_container = tk.Frame(center_container, bg='#f8fafc')
            empty_container.pack(expand=True, fill=tk.BOTH)
            center_frame = tk.Frame(empty_container, bg='#f8fafc')
            center_frame.place(relx=0.5, rely=0.5, anchor='center')
            
            tk.Label(center_frame, text="No projects found", 
                    font=('Segoe UI', 16, 'bold'), fg='#1e293b', bg='#f8fafc').pack(pady=10)
            tk.Label(center_frame, text="Projects will appear here once Quality Inspection tool syncs data.",
                    font=('Segoe UI', 11), fg='#64748b', bg='#f8fafc').pack(pady=5)
            return
        
        # Scrollable container
        canvas_container = tk.Frame(center_container, bg='#f8fafc')
        canvas_container.pack(expand=True, fill=tk.BOTH, padx=30, pady=(0, 20))
        
        canvas = tk.Canvas(canvas_container, bg='#f8fafc', highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='#f8fafc')
        
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        scroll_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        
        canvas_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Header with search bar
        header = tk.Frame(center_container, bg='#f8fafc')
        header.pack(fill=tk.X, padx=30, pady=(10, 10))
        header.pack_forget()
        header.pack(fill=tk.X, padx=30, pady=(10, 10), before=canvas_container)
        
        tk.Label(header, text="Projects Overview", font=('Segoe UI', 16, 'bold'),
                bg='#f8fafc').pack(side=tk.LEFT)
        
        # Search bar on the right
        search_frame = tk.Frame(header, bg='white', relief=tk.SOLID, borderwidth=1)
        search_frame.pack(side=tk.RIGHT, padx=10)
        
        tk.Label(search_frame, text="🔍", bg='white', font=('Segoe UI', 12)).pack(side=tk.LEFT, padx=(10, 5))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, width=30,
                               font=('Segoe UI', 10), relief=tk.FLAT, bg='white')
        search_entry.pack(side=tk.LEFT, padx=(0, 10), pady=8)
        
        def on_search(*args):
            search_term = search_var.get().strip()
            if search_term and search_term != "Search projects...":
                filtered_projects = self.db.search_projects(search_term)
            else:
                filtered_projects = self.db.get_all_projects()
            self.update_project_list(scroll_frame, filtered_projects)
        
        search_var.trace('w', on_search)
        search_entry.insert(0, "Search projects...")
        search_entry.config(fg='#94a3b8')
        
        def on_focus_in(event):
            if search_entry.get() == "Search projects...":
                search_entry.delete(0, tk.END)
                search_entry.config(fg='#1e293b')
        
        def on_focus_out(event):
            if not search_entry.get():
                search_entry.insert(0, "Search projects...")
                search_entry.config(fg='#94a3b8')
        
        search_entry.bind("<FocusIn>", on_focus_in)
        search_entry.bind("<FocusOut>", on_focus_out)
        
        # Store references for search updates
        self.dashboard_scroll_frame = scroll_frame
        self.update_project_list(scroll_frame, projects)
    
    def update_project_list(self, scroll_frame, projects):
        """
        Refresh the project cards displayed on the dashboard.

        Called after searches or data updates.
        """
        for w in scroll_frame.winfo_children():
            w.destroy()
        
        if not projects:
            tk.Label(scroll_frame, text="No matching projects found",
                    font=('Segoe UI', 12), fg='#64748b', bg='#f8fafc').pack(pady=50)
            return
        
        for proj in projects:
            self.create_project_card(scroll_frame, proj)
    
    def create_project_card(self, parent, project):
        """
        Create a collapsible UI card for a single project.
    
        Expands to show cabinet-level information on click.
        """
        card = tk.Frame(parent, bg='white', relief=tk.SOLID, borderwidth=1)
        card.pack(fill=tk.X, pady=10, padx=5)
        
        header = tk.Frame(card, bg='#eff6ff', cursor='hand2')
        header.pack(fill=tk.X)
        
        expand_var = tk.BooleanVar(value=False)
        
        indicator = tk.Label(header, text="▶", font=('Segoe UI', 12, 'bold'),
                           bg='#eff6ff', fg='#3b82f6', width=3)
        indicator.pack(side=tk.LEFT)
        
        tk.Label(header, text=project['project_name'], font=('Segoe UI', 13, 'bold'),
                bg='#eff6ff').pack(side=tk.LEFT, pady=15, padx=10)
        
        # Cabinet count on the right
        tk.Label(header, text=f"📦 {project['cabinet_count']} Cabinet(s)",
                font=('Segoe UI', 11, 'bold'), bg='#eff6ff', fg='#3b82f6').pack(side=tk.RIGHT, padx=20)
        
        dropdown = tk.Frame(card, bg='white')
        
        def toggle():
            if expand_var.get():
                dropdown.pack_forget()
                indicator.config(text="▶")
                expand_var.set(False)
            else:
                self.populate_cabinets(dropdown, project['project_name'])
                dropdown.pack(fill=tk.BOTH, padx=15, pady=10)
                indicator.config(text="▼")
                expand_var.set(True)
        
        header.bind("<Button-1>", lambda e: toggle())
        indicator.bind("<Button-1>", lambda e: toggle())
    
    def populate_cabinets(self, parent, project_name):
        """
        Populate cabinet rows for a selected project.
    
        Includes punch counts, status labels, and Excel links.
        """
        for w in parent.winfo_children():
            w.destroy()
        
        cabinets = self.db.get_cabinets_by_project(project_name)
        
        if not cabinets:
            tk.Label(parent, text="No cabinets", bg='white').pack(pady=20)
            return
        
        # Header - REMOVED Drawing % and Debug columns
        hdr = tk.Frame(parent, bg='#f1f5f9')
        hdr.pack(fill=tk.X, pady=5)
        
        headers = [
            ("Cabinet", 20), ("Total Punches", 12),
            ("Implemented", 12), ("Closed", 10), ("Status", 30)
        ]
        
        for text, w in headers:
            tk.Label(hdr, text=text, font=('Segoe UI', 9, 'bold'),
                    bg='#f1f5f9', width=w, anchor='w').pack(side=tk.LEFT, padx=3)
        
        # Rows
        for cab in cabinets:
            row = tk.Frame(parent, bg='white')
            row.pack(fill=tk.X, pady=2)
            
            # Cabinet ID - CLICKABLE
            cabinet_label = tk.Label(row, text=cab['cabinet_id'], font=('Segoe UI', 9, 'bold'),
                    bg='white', fg='#3b82f6', width=20, anchor='w', cursor='hand2')
            cabinet_label.pack(side=tk.LEFT, padx=3)
            
            # Make it clickable to open Excel
            def open_excel(excel_path=cab.get('excel_path')):
                self.open_excel_file(excel_path)
            
            cabinet_label.bind('<Button-1>', lambda e, ep=cab.get('excel_path'): self.open_excel_file(ep))
            
            # Add hover effect
            def on_enter(e, lbl=cabinet_label):
                lbl.config(fg='#1e40af', font=('Segoe UI', 9, 'bold', 'underline'))
            
            def on_leave(e, lbl=cabinet_label):
                lbl.config(fg='#3b82f6', font=('Segoe UI', 9, 'bold'))
            
            cabinet_label.bind('<Enter>', on_enter)
            cabinet_label.bind('<Leave>', on_leave)
            
            # REMOVED Drawing percentage section
            
            # Total Punches
            tk.Label(row, text=str(cab['total_punches']), font=('Segoe UI', 9),
                    bg='white', width=12, anchor='center').pack(side=tk.LEFT, padx=3)
            
            # Implemented
            tk.Label(row, text=str(cab['implemented_punches']), font=('Segoe UI', 9),
                    bg='white', width=12, anchor='center').pack(side=tk.LEFT, padx=3)
            
            # Closed
            tk.Label(row, text=str(cab['closed_punches']), font=('Segoe UI', 9),
                    bg='white', width=10, anchor='center').pack(side=tk.LEFT, padx=3)
            
            # Status - UPDATED status map
            status_map = {
                'project_info_sheet': (' Project Info Sheet', '#3b82f6'),
                'mechanical_assembly': (' Mechanical Assembly', '#8b5cf6'),
                'component_assembly': (' Component Assembly', '#f59e0b'),
                'final_assembly': (' Final Assembly', '#10b981'),
                'final_documentation': (' Final Documentation', '#64748b'),
                'handed_to_production': (' Handed to Production', '#8b5cf6'),
                'in_progress': ('Production Rework', '#f59e0b'),
                'being_closed_by_quality': (' Being Closed', '#10b981'),
                'closed': ('✓ Closed', '#64748b')
            }
            
            status_text, status_color = status_map.get(
                cab['status'],
                (cab['status'].replace('_', ' ').title(), '#64748b')
            )
            
            status_label = tk.Label(row, text=status_text, font=('Segoe UI', 9, 'bold'),
                                   bg=status_color, fg='white', padx=10, pady=4,
                                   anchor='w', width=30)
            status_label.pack(side=tk.LEFT, padx=3)
            
            # REMOVED Debug button
    
    def open_excel_file(self, excel_path):
        
        """Open Excel file in default application"""
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showwarning("File Not Found", 
                                 f"Excel file not found:\n{excel_path or 'No path specified'}")
            return
        
        try:
            if sys.platform == 'win32':
                os.startfile(excel_path)
            elif sys.platform == 'darwin':  # macOS
                subprocess.Popen(['open', excel_path])
            else:  # linux
                subprocess.Popen(['xdg-open', excel_path])
            
            messagebox.showinfo("Opening Excel", 
                              f"Opening:\n{os.path.basename(excel_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file:\n{e}")
    
    # ============ ANALYTICS - INTEGRATED SEARCH WITH FILTERS ============
    def show_analytics(self):
        """
        Display the Category Analytics view.
    
        Includes:
        - Search with suggestions
        - Date and level filters
        - Pareto chart visualization
        """
        self.set_active_nav('analytics')
        self.clear_content()
        
        # Header
        header = tk.Frame(self.content, bg='#f8fafc')
        header.pack(fill=tk.X, padx=30, pady=(20, 10))
        
        tk.Label(header, text="Category Analytics", font=('Segoe UI', 16, 'bold'),
                bg='#f8fafc').pack(side=tk.LEFT)
        
        # Integrated Search Bar with Filters
        search_control_frame = tk.Frame(self.content, bg='white', relief=tk.SOLID, borderwidth=1)
        search_control_frame.pack(fill=tk.X, padx=30, pady=(0, 10))
        
        # Main search bar
        search_bar_frame = tk.Frame(search_control_frame, bg='white')
        search_bar_frame.pack(fill=tk.X, padx=20, pady=(15, 10))
        
        tk.Label(search_bar_frame, text="🔍", bg='white', 
                font=('Segoe UI', 14)).pack(side=tk.LEFT, padx=(5, 10))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_bar_frame, textvariable=search_var, width=50,
                               font=('Segoe UI', 11), relief=tk.FLAT, bg='#f8fafc')
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0, 10))
        
        # Suggestion dropdown
        suggestion_frame = tk.Frame(search_control_frame, bg='white')
        suggestion_listbox = tk.Listbox(suggestion_frame, height=5, font=('Segoe UI', 10),
                                       relief=tk.FLAT, bg='#f8fafc', borderwidth=0)
        
        all_projects = self.db.get_all_project_names()
        
        def update_suggestions(*args):
            
            search_text = search_var.get().lower()
            suggestion_listbox.delete(0, tk.END)
            
            if search_text:
                matches = [p for p in all_projects if search_text in p.lower()]
                if matches:
                    for match in matches[:5]:
                        suggestion_listbox.insert(tk.END, match)
                    suggestion_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
                else:
                    suggestion_frame.pack_forget()
            else:
                suggestion_frame.pack_forget()
        
        def select_suggestion(event):
            if suggestion_listbox.curselection():
                selected = suggestion_listbox.get(suggestion_listbox.curselection())
                search_var.set(selected)
                suggestion_frame.pack_forget()
                apply_filters()
        
        search_var.trace('w', update_suggestions)
        suggestion_listbox.pack(fill=tk.X, padx=10, pady=5)
        suggestion_listbox.bind('<<ListboxSelect>>', select_suggestion)
        
        # Filter buttons
        filter_frame = tk.Frame(search_control_frame, bg='white')
        filter_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        tk.Label(filter_frame, text="Filter by:", font=('Segoe UI', 10, 'bold'),
                bg='white').pack(side=tk.LEFT, padx=(0, 10))
        
        # Date filter options
        date_filter_var = tk.StringVar(value="all")
        
        filter_buttons = [
            ("All Time", "all"),
            ("Today", "today"),
            ("This Month", "month"),
            ("This Quarter", "quarter"),
            ("This Year", "year"),
            ("Custom Date", "custom")
        ]
        
        for text, value in filter_buttons:
            btn = tk.Radiobutton(filter_frame, text=text, variable=date_filter_var,
                               value=value, bg='white', font=('Segoe UI', 9),
                               indicatoron=False, padx=15, pady=5,
                               selectcolor='#3b82f6', fg='#1e293b',
                               activebackground='#3b82f6', activeforeground='white',
                               relief=tk.FLAT, cursor='hand2')
            btn.pack(side=tk.LEFT, padx=2)
        
        # Level selection (Category/Subcategory)
        tk.Label(filter_frame, text=" | View:", font=('Segoe UI', 10, 'bold'),
                bg='white').pack(side=tk.LEFT, padx=(20, 10))
        
        level_var = tk.StringVar(value="category")
        tk.Radiobutton(filter_frame, text="Category", variable=level_var,
                      value="category", bg='white', font=('Segoe UI', 9),
                      indicatoron=False, padx=15, pady=5,
                      selectcolor='#10b981', fg='#1e293b',
                      activebackground='#10b981', activeforeground='white',
                      relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Radiobutton(filter_frame, text="Subcategory", variable=level_var,
                      value="subcategory", bg='white', font=('Segoe UI', 9),
                      indicatoron=False, padx=15, pady=5,
                      selectcolor='#10b981', fg='#1e293b',
                      activebackground='#10b981', activeforeground='white',
                      relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        # Problematic (80%) filter checkbox
        problematic_var = tk.BooleanVar(value=False)
        tk.Checkbutton(filter_frame, text=" Show Only Problematic (80%)",
                      variable=problematic_var,
                      bg='white', fg='#ef4444', 
                      font=('Segoe UI', 9, 'bold'),
                      selectcolor='white',
                      activebackground='white',
                      activeforeground='#dc2626',
                      cursor='hand2').pack(side=tk.LEFT, padx=20)
        
        
        # Custom date picker frame (hidden by default)
        custom_date_frame = tk.Frame(search_control_frame, bg='#f8fafc')
        
        tk.Label(custom_date_frame, text="From:", bg='#f8fafc',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(20, 5))
        start_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        start_date_entry = tk.Entry(custom_date_frame, textvariable=start_date_var,
                                    width=12, font=('Segoe UI', 9))
        start_date_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Label(custom_date_frame, text="To:", bg='#f8fafc',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(20, 5))
        end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        end_date_entry = tk.Entry(custom_date_frame, textvariable=end_date_var,
                                  width=12, font=('Segoe UI', 9))
        end_date_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Button(custom_date_frame, text="Apply", command=lambda: apply_filters(),
                 bg='#3b82f6', fg='white', font=('Segoe UI', 9, 'bold'),
                 padx=10, pady=3, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=10)
        
        def show_custom_date(*args):
            if date_filter_var.get() == "custom":
                custom_date_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
            else:
                custom_date_frame.pack_forget()
                apply_filters()
        
        date_filter_var.trace('w', show_custom_date)
        level_var.trace('w', lambda *args: apply_filters())
        problematic_var.trace('w', lambda *args: apply_filters())
        
        # Chart frame
        self.chart_frame = tk.Frame(self.content, bg='white')
        self.chart_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 20))
        
        # Store variables for export
        self.analytics_search_var = search_var
        self.analytics_date_filter = date_filter_var
        self.analytics_level = level_var
        self.analytics_start_date = start_date_var
        self.analytics_end_date = end_date_var
        self.analytics_problematic = problematic_var
        
        def apply_filters():
            project_filter = search_var.get().strip() if search_var.get() != "Search projects or select filters..." else None
            date_filter = date_filter_var.get()
            level = level_var.get()
            show_problematic_only = problematic_var.get()
            
            # Calculate date range
            start_date = None
            end_date = None
            
            if date_filter == "today":
                start_date = datetime.now().date().isoformat()
                end_date = start_date
            elif date_filter == "month":
                today = datetime.now().date()
                start_date = today.replace(day=1).isoformat()
                end_date = today.isoformat()
            elif date_filter == "quarter":
                today = datetime.now().date()
                # Calculate current quarter (Q1: Jan-Mar, Q2: Apr-Jun, Q3: Jul-Sep, Q4: Oct-Dec)
                quarter = (today.month - 1) // 3
                start_month = quarter * 3 + 1
                start_date = datetime(today.year, start_month, 1).date().isoformat()
                end_date = today.isoformat()
            elif date_filter == "year":
                today = datetime.now().date()
                # Financial year starts October 1st
                if today.month >= 10:
                    start_date = datetime(today.year, 10, 1).date().isoformat()
                else:
                    start_date = datetime(today.year - 1, 10, 1).date().isoformat()
                end_date = today.isoformat()
            elif date_filter == "custom":
                start_date = start_date_var.get()
                end_date = end_date_var.get()
            
            self.update_chart_with_filters(start_date, end_date, project_filter, level, show_problematic_only)
        
        # Initial load
        apply_filters()
        
        # Placeholder behavior
        search_entry.insert(0, "Search projects or select filters...")
        search_entry.config(fg='#94a3b8')
        
        def on_focus_in(event):
            if search_entry.get() == "Search projects or select filters...":
                search_entry.delete(0, tk.END)
                search_entry.config(fg='#1e293b')
        
        def on_focus_out(event):
            if not search_entry.get():
                search_entry.insert(0, "Search projects or select filters...")
                search_entry.config(fg='#94a3b8')
        
        search_entry.bind("<FocusIn>", on_focus_in)
        search_entry.bind("<FocusOut>", on_focus_out)
        search_entry.bind("<Return>", lambda e: apply_filters())

    def update_chart_with_filters(self, start_date, end_date, project, level, show_problematic_only=False):
        """
        Generate and render a Pareto chart based on active filters.
    
        Highlights the top 80% contributing categories
        using cumulative frequency analysis.
        """
        # Clear previous chart
        for w in self.chart_frame.winfo_children():
            w.destroy()
        plt.close('all')
        
        stats = self.db.get_category_stats(start_date, end_date, project)
        
        if not stats:
            empty_frame = tk.Frame(self.chart_frame, bg='white')
            empty_frame.place(relx=0.5, rely=0.5, anchor='center')
            
            tk.Label(empty_frame, text="No data available for the selected filters.",
                    font=('Segoe UI', 12), fg='#64748b', bg='white').pack(pady=5)
            tk.Label(empty_frame, 
                    text="Category data will appear once Quality Inspection logs punches.",
                    font=('Segoe UI', 10), fg='#94a3b8', bg='white').pack(pady=5)
            return
        
        counts = defaultdict(int)
        if level == "category":
            for item in stats:
                counts[item['category']] += item['count']
        else:
            # For subcategory, treat each unique subcategory independently
            for item in stats:
                key = f"{item['category']} → {item['subcategory'] or 'N/A'}"
                counts[key] += item['count']
        
        # Sort ALL items by count in descending order
        sorted_items = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:15]
        labels = [item[0] for item in sorted_items]
        values = [item[1] for item in sorted_items]
        
        # Calculate total from ALL sorted items
        total = sum(values)
        
        # Calculate cumulative percentage for these sorted items
        cumulative = []
        cum = 0
        for v in values:
            cum += v
            cumulative.append((cum/total)*100)
        
        # Calculate 80% threshold index - find where cumulative first reaches or exceeds 80%
        # This is based on the SORTED (descending) order, regardless of category grouping
        threshold_80_idx = None
        for i, cum_pct in enumerate(cumulative):
            if cum_pct >= 80:
                threshold_80_idx = i
                break
        
        # If no item reaches 80%, then all items are problematic
        if threshold_80_idx is None and len(cumulative) > 0:
            threshold_80_idx = len(cumulative) - 1
        
        # Filter to show only problematic if checkbox is checked
        if show_problematic_only and threshold_80_idx is not None:
            labels = labels[:threshold_80_idx + 1]
            values = values[:threshold_80_idx + 1]
            # Recalculate cumulative for filtered data
            cumulative = []
            cum = 0
            filtered_total = sum(values)
            for v in values:
                cum += v
                cumulative.append((cum/filtered_total)*100)
            # Update threshold for filtered data
            threshold_80_idx = None
            for i, cum_pct in enumerate(cumulative):
                if cum_pct >= 80:
                    threshold_80_idx = i
                    break
            if threshold_80_idx is None and len(cumulative) > 0:
                threshold_80_idx = len(cumulative) - 1
        
        fig = Figure(figsize=(14, 7), facecolor='white')
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twinx()
        
        # Color bars: red for problematic (up to and including 80% threshold), blue for rest
        # This applies INDEPENDENTLY to each item based on its position in descending order
        bar_colors = []
        for i in range(len(labels)):
            if threshold_80_idx is not None and i <= threshold_80_idx:
                bar_colors.append('#ef4444')  # Red for problematic (within 80% cumulative)
            else:
                bar_colors.append('#3b82f6')  # Blue for non-problematic (beyond 80%)
        
        bars = ax1.bar(range(len(labels)), values, color=bar_colors, alpha=0.7, edgecolor='black', linewidth=0.5)
        line = ax2.plot(range(len(labels)), cumulative, color='#f59e0b',
                       marker='o', linewidth=2, markersize=6, label='Cumulative %')
        ax2.axhline(y=80, color='#10b981', linestyle='--', linewidth=1.5, alpha=0.7, label='80% threshold')
        
        ax1.set_xlabel('Category', fontsize=11, fontweight='bold')
        ax1.set_ylabel('Frequency', fontsize=11, fontweight='bold', color='#1e293b')
        ax2.set_ylabel('Cumulative %', fontsize=11, fontweight='bold', color='#f59e0b')
        
        # Add filter info to title
        filter_text = f"{level.title()} Analysis"
        if project:
            filter_text += f" - {project}"
        
        # Add problematic count to title
        problematic_count = (threshold_80_idx + 1) if threshold_80_idx is not None else 0
        total_count = len(labels)
        ax1.set_title(f'Pareto Chart - {filter_text}\n'
                     f'({problematic_count}/{total_count} categories represent 80% of issues)',
                     fontsize=14, fontweight='bold')
        
        ax1.set_xticks(range(len(labels)))
        ax1.set_xticklabels(labels, rotation=45, ha='right', fontsize=9)
        ax1.tick_params(axis='y', labelcolor='#1e293b')
        ax2.tick_params(axis='y', labelcolor='#f59e0b')
        ax2.set_ylim(0, 105)
        ax2.legend(loc='lower right')
        ax1.grid(axis='y', alpha=0.3, linestyle='--')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        

        
        # Store current chart data for export
        self.current_chart_data = {
            'labels': labels,
            'values': values,
            'cumulative': cumulative,
            'threshold_80_idx': threshold_80_idx,
            'level': level,
            'total': total
        }

    

    def export_excel_filtered(self):
        """
        Export analytics data to Excel based on active filters.
    
        Automatically selects export format:
        - Standard
        - Project-wise
        - Month-wise
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Get current filter values
            project_filter = self.analytics_search_var.get().strip() if self.analytics_search_var.get() != "Search projects or select filters..." else None
            date_filter = self.analytics_date_filter.get()
            start_date = None
            end_date = None
            
            # Calculate date range based on filter
            if date_filter == "today":
                start_date = datetime.now().date().isoformat()
                end_date = start_date
            elif date_filter == "month":
                today = datetime.now().date()
                start_date = today.replace(day=1).isoformat()
                end_date = today.isoformat()
            elif date_filter == "quarter":
                today = datetime.now().date()
                quarter = (today.month - 1) // 3
                start_month = quarter * 3 + 1
                start_date = datetime(today.year, start_month, 1).date().isoformat()
                end_date = today.isoformat()
            elif date_filter == "year":
                today = datetime.now().date()
                if today.month >= 10:
                    start_date = datetime(today.year, 10, 1).date().isoformat()
                else:
                    start_date = datetime(today.year - 1, 10, 1).date().isoformat()
                end_date = today.isoformat()
            elif date_filter == "custom":
                start_date = self.analytics_start_date.get()
                end_date = self.analytics_end_date.get()
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Styling
            header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            problematic_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Export based on date filter
            if date_filter in ["month", "quarter"]:
                # Project-wise data for month and quarter
                self._export_project_wise(wb, start_date, end_date, date_filter, header_fill, header_font, problematic_fill, border)
            elif date_filter == "year":
                # Month-wise data for year
                self._export_month_wise(wb, start_date, end_date, header_fill, header_font, problematic_fill, border)
            else:
                # Standard export (Category and Subcategory sheets)
                self._export_standard(wb, start_date, end_date, project_filter, header_fill, header_font, problematic_fill, border)
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Category_Analytics_{date_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Analytics exported successfully to:\n{file_path}")
        
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export analytics:\n{str(e)}")

    def _export_standard(self, wb, start_date, end_date, project_filter, header_fill, header_font, problematic_fill, border):
        """Standard export with Category and Subcategory sheets"""
        stats = self.db.get_category_stats(start_date, end_date, project_filter)
        
        if not stats:
            messagebox.showwarning("No Data", "No data available for the selected filters.")
            return
        
        # CATEGORY SHEET
        ws_cat = wb.create_sheet("Category Analysis")
        ws_cat.append(["Rank", "Category", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Aggregate by category
        cat_counts = defaultdict(int)
        for item in stats:
            cat_counts[item['category']] += item['count']
        
        sorted_cats = sorted(cat_counts.items(), key=lambda x: x[1], reverse=True)
        total_cat = sum([v for _, v in sorted_cats])
        
        cumulative = 0
        threshold_idx = None
        for i, (cat, count) in enumerate(sorted_cats, 1):
            percentage = (count / total_cat) * 100
            cumulative += percentage
            status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
            
            if threshold_idx is None and cumulative >= 80:
                threshold_idx = i
            
            ws_cat.append([i, cat, count, round(percentage, 2), round(cumulative, 2), status])
            
            # Apply problematic highlighting
            row = ws_cat.max_row
            if status == "Problematic (80%)":
                for cell in ws_cat[row]:
                    cell.fill = problematic_fill
            
            # Apply borders and alignment
            for cell in ws_cat[row]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_cat.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_cat.column_dimensions[col_letter].width = adjusted_width
        
        # SUBCATEGORY SHEET
        ws_sub = wb.create_sheet("Subcategory Analysis")
        ws_sub.append(["Rank", "Category", "Subcategory", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_sub[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Aggregate by subcategory
        sub_counts = defaultdict(int)
        sub_mapping = {}
        for item in stats:
            key = f"{item['category']} → {item['subcategory'] or 'N/A'}"
            sub_counts[key] += item['count']
            sub_mapping[key] = (item['category'], item['subcategory'] or 'N/A')
        
        sorted_subs = sorted(sub_counts.items(), key=lambda x: x[1], reverse=True)
        total_sub = sum([v for _, v in sorted_subs])
        
        cumulative = 0
        threshold_idx = None
        for i, (key, count) in enumerate(sorted_subs, 1):
            cat, sub = sub_mapping[key]
            percentage = (count / total_sub) * 100
            cumulative += percentage
            status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
            
            if threshold_idx is None and cumulative >= 80:
                threshold_idx = i
            
            ws_sub.append([i, cat, sub, count, round(percentage, 2), round(cumulative, 2), status])
            
            # Apply problematic highlighting
            row = ws_sub.max_row
            if status == "Problematic (80%)":
                for cell in ws_sub[row]:
                    cell.fill = problematic_fill
            
            # Apply borders and alignment
            for cell in ws_sub[row]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_sub.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sub.column_dimensions[col_letter].width = adjusted_width

    def _export_project_wise(self, wb, start_date, end_date, date_filter, header_fill, header_font, problematic_fill, border):
        """Export project-wise data for month and quarter filters"""
        # Get all projects in the date range
        all_stats = self.db.get_category_stats(start_date, end_date, None)
        
        if not all_stats:
            messagebox.showwarning("No Data", "No data available for the selected filters.")
            return
        
        # Get unique projects
        projects = set()
        project_data = defaultdict(list)
        for item in all_stats:
            project = item.get('project_name', 'Unknown')
            projects.add(project)
            project_data[project].append(item)
        
        # CATEGORY SHEET - Project-wise
        ws_cat = wb.create_sheet("Category Analysis (Project-wise)")
        ws_cat.append(["Project", "Rank", "Category", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for project in sorted(projects):
            stats = project_data[project]
            
            # Aggregate by category
            cat_counts = defaultdict(int)
            for item in stats:
                cat_counts[item['category']] += item['count']
            
            sorted_cats = sorted(cat_counts.items(), key=lambda x: x[1], reverse=True)
            total_cat = sum([v for _, v in sorted_cats])
            
            if total_cat == 0:
                continue
            
            cumulative = 0
            threshold_idx = None
            for i, (cat, count) in enumerate(sorted_cats, 1):
                percentage = (count / total_cat) * 100
                cumulative += percentage
                status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
                
                if threshold_idx is None and cumulative >= 80:
                    threshold_idx = i
                
                ws_cat.append([project, i, cat, count, round(percentage, 2), round(cumulative, 2), status])
                
                # Apply problematic highlighting
                row = ws_cat.max_row
                if status == "Problematic (80%)":
                    for cell in ws_cat[row]:
                        cell.fill = problematic_fill
                
                # Apply borders
                for cell in ws_cat[row]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_cat.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_cat.column_dimensions[col_letter].width = adjusted_width
        
        # SUBCATEGORY SHEET - Project-wise
        ws_sub = wb.create_sheet("Subcategory Analysis (Project-wise)")
        ws_sub.append(["Project", "Rank", "Category", "Subcategory", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_sub[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for project in sorted(projects):
            stats = project_data[project]
            
            # Aggregate by subcategory
            sub_counts = defaultdict(int)
            sub_mapping = {}
            for item in stats:
                key = f"{item['category']} → {item['subcategory'] or 'N/A'}"
                sub_counts[key] += item['count']
                sub_mapping[key] = (item['category'], item['subcategory'] or 'N/A')
            
            sorted_subs = sorted(sub_counts.items(), key=lambda x: x[1], reverse=True)
            total_sub = sum([v for _, v in sorted_subs])
            
            if total_sub == 0:
                continue
            
            cumulative = 0
            threshold_idx = None
            for i, (key, count) in enumerate(sorted_subs, 1):
                cat, sub = sub_mapping[key]
                percentage = (count / total_sub) * 100
                cumulative += percentage
                status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
                
                if threshold_idx is None and cumulative >= 80:
                    threshold_idx = i
                
                ws_sub.append([project, i, cat, sub, count, round(percentage, 2), round(cumulative, 2), status])
                
                # Apply problematic highlighting
                row = ws_sub.max_row
                if status == "Problematic (80%)":
                    for cell in ws_sub[row]:
                        cell.fill = problematic_fill
                
                # Apply borders
                for cell in ws_sub[row]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_sub.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sub.column_dimensions[col_letter].width = adjusted_width

    def _export_month_wise(self, wb, start_date, end_date, header_fill, header_font, problematic_fill, border):
        """Export month-wise data for year filter"""
        # Parse start and end dates
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Generate list of months in the range
        months = []
        current = start.replace(day=1)
        while current <= end:
            months.append(current)
            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
        
        # CATEGORY SHEET - Month-wise
        ws_cat = wb.create_sheet("Category Analysis (Month-wise)")
        ws_cat.append(["Month", "Rank", "Category", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for month_date in months:
            # Calculate month range
            month_start = month_date.replace(day=1).date().isoformat()
            if month_date.month == 12:
                month_end = month_date.replace(year=month_date.year + 1, month=1, day=1)
            else:
                month_end = month_date.replace(month=month_date.month + 1, day=1)
            month_end = (month_end.date() - timedelta(days=1)).isoformat()
            
            # Get stats for this month
            stats = self.db.get_category_stats(month_start, month_end, None)
            
            if not stats:
                continue
            
            month_name = month_date.strftime('%B %Y')
            
            # Aggregate by category
            cat_counts = defaultdict(int)
            for item in stats:
                cat_counts[item['category']] += item['count']
            
            sorted_cats = sorted(cat_counts.items(), key=lambda x: x[1], reverse=True)
            total_cat = sum([v for _, v in sorted_cats])
            
            if total_cat == 0:
                continue
            
            cumulative = 0
            threshold_idx = None
            for i, (cat, count) in enumerate(sorted_cats, 1):
                percentage = (count / total_cat) * 100
                cumulative += percentage
                status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
                
                if threshold_idx is None and cumulative >= 80:
                    threshold_idx = i
                
                ws_cat.append([month_name, i, cat, count, round(percentage, 2), round(cumulative, 2), status])
                
                # Apply problematic highlighting
                row = ws_cat.max_row
                if status == "Problematic (80%)":
                    for cell in ws_cat[row]:
                        cell.fill = problematic_fill
                
                # Apply borders
                for cell in ws_cat[row]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_cat.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_cat.column_dimensions[col_letter].width = adjusted_width
        
        # SUBCATEGORY SHEET - Month-wise
        ws_sub = wb.create_sheet("Subcategory Analysis (Month-wise)")
        ws_sub.append(["Month", "Rank", "Category", "Subcategory", "Count", "Percentage (%)", "Cumulative (%)", "Status"])
        
        # Apply header styling
        for cell in ws_sub[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for month_date in months:
            # Calculate month range
            month_start = month_date.replace(day=1).date().isoformat()
            if month_date.month == 12:
                month_end = month_date.replace(year=month_date.year + 1, month=1, day=1)
            else:
                month_end = month_date.replace(month=month_date.month + 1, day=1)
            month_end = (month_end.date() - timedelta(days=1)).isoformat()
            
            # Get stats for this month
            stats = self.db.get_category_stats(month_start, month_end, None)
            
            if not stats:
                continue
            
            month_name = month_date.strftime('%B %Y')
            
            # Aggregate by subcategory
            sub_counts = defaultdict(int)
            sub_mapping = {}
            for item in stats:
                key = f"{item['category']} → {item['subcategory'] or 'N/A'}"
                sub_counts[key] += item['count']
                sub_mapping[key] = (item['category'], item['subcategory'] or 'N/A')
            
            sorted_subs = sorted(sub_counts.items(), key=lambda x: x[1], reverse=True)
            total_sub = sum([v for _, v in sorted_subs])
            
            if total_sub == 0:
                continue
            
            cumulative = 0
            threshold_idx = None
            for i, (key, count) in enumerate(sorted_subs, 1):
                cat, sub = sub_mapping[key]
                percentage = (count / total_sub) * 100
                cumulative += percentage
                status = "Problematic (80%)" if cumulative <= 80 or threshold_idx is None else "Normal"
                
                if threshold_idx is None and cumulative >= 80:
                    threshold_idx = i
                
                ws_sub.append([month_name, i, cat, sub, count, round(percentage, 2), round(cumulative, 2), status])
                
                # Apply problematic highlighting
                row = ws_sub.max_row
                if status == "Problematic (80%)":
                    for cell in ws_sub[row]:
                        cell.fill = problematic_fill
                
                # Apply borders
                for cell in ws_sub[row]:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for col in ws_sub.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sub.column_dimensions[col_letter].width = adjusted_width
    # ============ DEFECT LIBRARY (RENAMED FROM CATEGORIES) ============
    def show_defect_library(self):
        """
        Display the Defect Library management interface.
    
        Allows creation, editing, deletion, and testing
        of defect categories and subcategories.
        """
        self.set_active_nav('defect_library')
        self.clear_content()
        
        # Centered container
        center_container = tk.Frame(self.content, bg='#f8fafc')
        center_container.place(relx=0.5, rely=0, anchor='n', relwidth=0.7, relheight=1.0)
        
        # Header
        header = tk.Frame(center_container, bg='#f8fafc')
        header.pack(fill=tk.X, padx=30, pady=(20, 10))
        
        tk.Label(header, text="Defect Library Management", font=('Segoe UI', 16, 'bold'),
                bg='#f8fafc').pack(side=tk.LEFT)
        
        tk.Button(header, text="➕ Add Defect Type", command=self.add_category,
                 bg='#10b981', fg='white', font=('Segoe UI', 10, 'bold'),
                 padx=15, pady=8).pack(side=tk.RIGHT)
        
        if not self.categories:
            empty_container = tk.Frame(center_container, bg='#f8fafc')
            empty_container.pack(expand=True, fill=tk.BOTH)
            center_frame = tk.Frame(empty_container, bg='#f8fafc')
            center_frame.place(relx=0.5, rely=0.5, anchor='center')
            
            tk.Label(center_frame, text="No defect types defined",
                    font=('Segoe UI', 16, 'bold'), fg='#1e293b', bg='#f8fafc').pack(pady=10)
            tk.Label(center_frame, text="Click 'Add Defect Type' to create your first defect category.",
                    font=('Segoe UI', 11), fg='#64748b', bg='#f8fafc').pack(pady=5)
            return
        
        # Scrollable container
        canvas_container = tk.Frame(center_container, bg='#f8fafc')
        canvas_container.pack(expand=True, fill=tk.BOTH, padx=30, pady=(0, 20))
        
        canvas = tk.Canvas(canvas_container, bg='#f8fafc', highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='#f8fafc')
        
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        scroll_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        
        canvas_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        for cat in self.categories:
            self.create_category_card(scroll_frame, cat)

    def create_category_card(self, parent, category):
        card = tk.Frame(parent, bg='white', relief=tk.SOLID, borderwidth=1)
        card.pack(fill=tk.X, pady=8, padx=5)
        
        header = tk.Frame(card, bg='#dbeafe')
        header.pack(fill=tk.X)
        
        # Category name and ref number
        left_frame = tk.Frame(header, bg='#dbeafe')
        left_frame.pack(side=tk.LEFT, padx=15, pady=10)
        
        tk.Label(left_frame, text=category['name'], font=('Segoe UI', 12, 'bold'),
                bg='#dbeafe', fg='#1e40af').pack(side=tk.LEFT)
        
        ref_num = category.get('ref_number', '')
        if ref_num:
            tk.Label(left_frame, text=f" (Ref: {ref_num})", font=('Segoe UI', 9),
                    bg='#dbeafe', fg='#64748b').pack(side=tk.LEFT, padx=5)
        
        # Determine mode
        mode = category.get('mode')
        if not mode:
            mode = 'parent' if category.get('subcategories') else 'template'
        
        mode_text = {
            'template': '📝 Template',
            'parent': '📁 Parent',
            'wiring_selector': '⚡ Wiring Selector'
        }.get(mode, mode)
        
        tk.Label(header, text=mode_text, font=('Segoe UI', 9),
                bg='#dbeafe', fg='#64748b').pack(side=tk.LEFT, padx=10)
        
        btn_frame = tk.Frame(header, bg='#dbeafe')
        btn_frame.pack(side=tk.RIGHT, padx=10)
        
        tk.Button(btn_frame, text="Edit", command=lambda: self.edit_category(category),
                 bg='#3b82f6', fg='white', font=('Segoe UI', 9, 'bold'),
                 padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        tk.Button(btn_frame, text="Delete", command=lambda: self.delete_category(category),
                 bg='#ef4444', fg='white', font=('Segoe UI', 9, 'bold'),
                 padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        if mode == 'parent':
            tk.Button(btn_frame, text="Add Sub", command=lambda: self.add_subcategory(category),
                     bg='#10b981', fg='white', font=('Segoe UI', 9, 'bold'),
                     padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
        elif mode == 'wiring_selector':
            tk.Button(btn_frame, text="Add Wiring Sub", command=lambda: self.add_wiring_subcategory(category),
                     bg='#10b981', fg='white', font=('Segoe UI', 9, 'bold'),
                     padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
            # Add special subcategory button
            if 'special_subcategories' in category:
                tk.Button(btn_frame, text="Add Special Sub", command=lambda: self.add_special_subcategory(category),
                         bg='#8b5cf6', fg='white', font=('Segoe UI', 9, 'bold'),
                         padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
        elif mode == 'template':
            tk.Button(btn_frame, text="🧪 Test", command=lambda: self.handle_template_category(category),
                     bg='#8b5cf6', fg='white', font=('Segoe UI', 9, 'bold'),
                     padx=12, pady=6, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # Display wiring types for wiring_selector mode
        if mode == 'wiring_selector':
            self.display_wiring_types(card, category)
        # Display regular subcategories for parent mode
        elif category.get('subcategories'):
            sub_frame = tk.Frame(card, bg='white')
            sub_frame.pack(fill=tk.X, padx=20, pady=10)
            
            for sub in category['subcategories']:
                self.display_subcategory_row(sub_frame, category, sub)

    def display_wiring_types(self, card, category):
        """Display wiring types and their subcategories"""
        wiring_frame = tk.Frame(card, bg='white')
        wiring_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Display wiring types
        if category.get('wiring_types'):
            for wiring_type in category['wiring_types']:
                type_card = tk.Frame(wiring_frame, bg='#f1f5f9', relief=tk.SOLID, borderwidth=1)
                type_card.pack(fill=tk.X, pady=5, padx=10)
                
                # Wiring type header
                type_header = tk.Frame(type_card, bg='#e0e7ff')
                type_header.pack(fill=tk.X)
                
                type_label_frame = tk.Frame(type_header, bg='#e0e7ff')
                type_label_frame.pack(side=tk.LEFT, padx=10, pady=5)
                
                tk.Label(type_label_frame, text=f"⚡ {wiring_type['type']}", 
                        font=('Segoe UI', 10, 'bold'),
                        bg='#e0e7ff', fg='#3730a3').pack(side=tk.LEFT)
                
                type_ref = wiring_type.get('ref_number', '')
                if type_ref:
                    tk.Label(type_label_frame, text=f" (Ref: {type_ref})", 
                            font=('Segoe UI', 8),
                            bg='#e0e7ff', fg='#64748b').pack(side=tk.LEFT, padx=5)
                
                # Subcategories under this wiring type
                if wiring_type.get('subcategories'):
                    sub_container = tk.Frame(type_card, bg='#f1f5f9')
                    sub_container.pack(fill=tk.X, padx=10, pady=5)
                    
                    for sub in wiring_type['subcategories']:
                        self.display_wiring_subcategory_row(sub_container, category, wiring_type, sub)
        
        # Display special subcategories
        if category.get('special_subcategories'):
            special_frame = tk.Frame(wiring_frame, bg='#fef3c7', relief=tk.SOLID, borderwidth=1)
            special_frame.pack(fill=tk.X, pady=5, padx=10)
            
            tk.Label(special_frame, text="Special Subcategories", 
                    font=('Segoe UI', 10, 'bold'),
                    bg='#fef3c7', fg='#92400e').pack(anchor='w', padx=10, pady=5)
            
            for sub in category['special_subcategories']:
                self.display_special_subcategory_row(special_frame, category, sub)

    def display_subcategory_row(self, parent, category, sub):
        """Display a regular subcategory row"""
        sub_row = tk.Frame(parent, bg='#f8fafc')
        sub_row.pack(fill=tk.X, pady=3)
        
        label_frame = tk.Frame(sub_row, bg='#f8fafc')
        label_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10, pady=8)
        
        tk.Label(label_frame, text=f" ↳ {sub['name']}", font=('Segoe UI', 10),
                bg='#f8fafc', anchor='w').pack(side=tk.LEFT)
        
        ref_num = sub.get('ref_number', '')
        if ref_num:
            tk.Label(label_frame, text=f" (Ref: {ref_num})", font=('Segoe UI', 8),
                    bg='#f8fafc', fg='#64748b', anchor='w').pack(side=tk.LEFT, padx=5)
        
        sub_btn_frame = tk.Frame(sub_row, bg='#f8fafc')
        sub_btn_frame.pack(side=tk.RIGHT, padx=10)
        
        tk.Button(sub_btn_frame, text="Test",
                 command=lambda: self.handle_subcategory(category, sub),
                 bg='#8b5cf6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=10, pady=5, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Edit",
                 command=lambda: self.edit_subcategory(category, sub),
                 bg='#3b82f6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=10, pady=5, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Delete",
                 command=lambda: self.delete_subcategory(category, sub),
                 bg='#ef4444', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=10, pady=5, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)

    def display_wiring_subcategory_row(self, parent, category, wiring_type, sub):
        """Display a wiring subcategory row"""
        sub_row = tk.Frame(parent, bg='#f8fafc')
        sub_row.pack(fill=tk.X, pady=2)
        
        label_frame = tk.Frame(sub_row, bg='#f8fafc')
        label_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10, pady=6)
        
        tk.Label(label_frame, text=f"  → {sub['name']}", font=('Segoe UI', 9),
                bg='#f8fafc', anchor='w').pack(side=tk.LEFT)
        
        ref_num = sub.get('ref_number', '')
        if ref_num:
            tk.Label(label_frame, text=f" (Ref: {ref_num})", font=('Segoe UI', 8),
                    bg='#f8fafc', fg='#64748b', anchor='w').pack(side=tk.LEFT, padx=5)
        
        sub_btn_frame = tk.Frame(sub_row, bg='#f8fafc')
        sub_btn_frame.pack(side=tk.RIGHT, padx=10)
        
        tk.Button(sub_btn_frame, text="Test",
                 command=lambda: self.handle_subcategory(category, sub),
                 bg='#8b5cf6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Edit",
                 command=lambda: self.edit_wiring_subcategory(category, wiring_type, sub),
                 bg='#3b82f6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Delete",
                 command=lambda: self.delete_wiring_subcategory(category, wiring_type, sub),
                 bg='#ef4444', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)

    def display_special_subcategory_row(self, parent, category, sub):
        """Display a special subcategory row"""
        sub_row = tk.Frame(parent, bg='#fffbeb')
        sub_row.pack(fill=tk.X, pady=2)
        
        label_frame = tk.Frame(sub_row, bg='#fffbeb')
        label_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10, pady=6)
        
        tk.Label(label_frame, text=f" ⭐ {sub['name']}", font=('Segoe UI', 9),
                bg='#fffbeb', anchor='w').pack(side=tk.LEFT)
        
        ref_num = sub.get('ref_number', '')
        if ref_num:
            tk.Label(label_frame, text=f" (Ref: {ref_num})", font=('Segoe UI', 8),
                    bg='#fffbeb', fg='#64748b', anchor='w').pack(side=tk.LEFT, padx=5)
        
        sub_btn_frame = tk.Frame(sub_row, bg='#fffbeb')
        sub_btn_frame.pack(side=tk.RIGHT, padx=10)
        
        tk.Button(sub_btn_frame, text="Test",
                 command=lambda: self.handle_subcategory(category, sub),
                 bg='#8b5cf6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Edit",
                 command=lambda: self.edit_special_subcategory(category, sub),
                 bg='#3b82f6', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)
        
        tk.Button(sub_btn_frame, text="Delete",
                 command=lambda: self.delete_special_subcategory(category, sub),
                 bg='#ef4444', fg='white', font=('Segoe UI', 8, 'bold'),
                 padx=8, pady=4, relief=tk.FLAT, cursor='hand2').pack(side=tk.LEFT, padx=2)

    # Category management methods
    def collect_template_data(self, mandatory=True, existing=None, include_ref=False):
        """
        Collect structured input data for template-based categories.
    
        Supports:
        - Dynamic inputs
        - Text templates
        - Optional reference numbers
        """
        min_inputs = 1 if mandatory else 0
        default_inputs = len(existing.get("inputs", [])) if existing else min_inputs
        num_inputs = simpledialog.askinteger(
            "Expected Inputs",
            "How many inputs are required?",
            parent=self.root,
            minvalue=min_inputs,
            maxvalue=10,
            initialvalue=default_inputs
        )
        if num_inputs is None:
            return None
        
        inputs = []
        for i in range(num_inputs):
            default = existing["inputs"][i] if existing and i < len(existing.get("inputs", [])) else None
            name = simpledialog.askstring(
                "Input Name",
                f"Internal name for input #{i+1}",
                parent=self.root,
                initialvalue=default.get("name") if default else ""
            )
            if not name:
                return None
            
            label = simpledialog.askstring(
                "Input Label",
                f"Question asked to user for '{name}'",
                parent=self.root,
                initialvalue=default.get("label") if default else ""
            )
            if not label:
                return None
            
            inputs.append({"name": name.strip(), "label": label.strip()})
        
        placeholder_text = ", ".join(f"{{{i['name']}}}" for i in inputs)
        default_template = existing.get("template") if existing else ""
        template = simpledialog.askstring(
            "Punch Text Template",
            f"Enter punch text template.\nAvailable placeholders:\n{placeholder_text}",
            parent=self.root,
            initialvalue=default_template
        )
        
        if mandatory and not template:
            messagebox.showerror("Required", "Template is mandatory")
            return None
        
        result = {"inputs": inputs, "template": template.strip() if template else None}
        
        # Add reference number if requested
        if include_ref:
            ref_number = simpledialog.askstring(
                "Reference Number",
                "Enter reference number (optional):",
                parent=self.root,
                initialvalue=existing.get("ref_number", "") if existing else ""
            )
            result["ref_number"] = ref_number.strip() if ref_number else ""
        
        return result

    def create_category(self):
        name = simpledialog.askstring("New Category", "Enter category name:", parent=self.root)
        if not name:
            return None
        
        # Ask for reference number
        ref_number = simpledialog.askstring(
            "Reference Number",
            "Enter reference number (optional):",
            parent=self.root
        )
        
        category = {
            "name": name.strip(),
            "ref_number": ref_number.strip() if ref_number else "",
            "mode": None,
            "inputs": [],
            "template": None,
            "subcategories": []
        }
        
        # Ask for category type
        choice = messagebox.askquestion(
            "Category Type",
            "What type of category?\n\nYES → Template category (generates punch text directly)\nNO → Choose parent or wiring selector",
            parent=self.root
        )
        
        if choice == 'yes':
            category["mode"] = "template"
            data = self.collect_template_data(mandatory=False, include_ref=False)
            if data:
                category.update(data)
        else:
            # Ask if wiring selector
            is_wiring = messagebox.askyesno(
                "Wiring Selector?",
                "Is this a wiring selector category (Power/I/O/Ground)?",
                parent=self.root
            )
            if is_wiring:
                category["mode"] = "wiring_selector"
                category["wiring_types"] = []
                category["special_subcategories"] = []
            else:
                category["mode"] = "parent"
        
        return category

    def add_category(self):
        cat = self.create_category()
        if not cat:
            return
        
        if any(c["name"].lower() == cat["name"].lower() for c in self.categories):
            messagebox.showwarning("Duplicate", "Category already exists")
            return
        
        self.categories.append(cat)
        self.save_categories()
        self.show_defect_library()

    def edit_category(self, category):
        mode = category.get('mode')
        
        if not mode:
            mode = 'parent' if category.get('subcategories') else 'template'
            category['mode'] = mode
        
        # Edit name
        new_name = simpledialog.askstring(
            "Edit Category",
            "Enter new category name:",
            initialvalue=category["name"],
            parent=self.root
        )
        if not new_name:
            return
        
        # Edit reference number
        new_ref = simpledialog.askstring(
            "Edit Reference Number",
            "Enter reference number (optional):",
            initialvalue=category.get("ref_number", ""),
            parent=self.root
        )
        
        category["name"] = new_name.strip()
        category["ref_number"] = new_ref.strip() if new_ref else ""
        
        # Template-specific editing
        if mode == 'template':
            if category.get('inputs'):
                updated = self.collect_template_data(mandatory=False, existing=category, include_ref=False)
                if updated:
                    category["inputs"] = updated["inputs"]
                    category["template"] = updated["template"]
            else:
                new_template = simpledialog.askstring(
                    "Edit Template",
                    "Enter punch text template:",
                    initialvalue=category.get("template", ""),
                    parent=self.root
                )
                if new_template is not None:
                    category["template"] = new_template.strip()
        
        self.save_categories()
        self.show_defect_library()

    def delete_category(self, category):
        if not messagebox.askyesno("Confirm", f"Delete category '{category['name']}'?"):
            return
        self.categories.remove(category)
        self.save_categories()
        self.show_defect_library()

    # Wiring category methods
    def add_wiring_subcategory(self, category):
        """Add a subcategory to a wiring type"""
        # Ask which wiring types to add to
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Wiring Types")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Select wiring types to add this subcategory to:",
                font=('Segoe UI', 11, 'bold')).pack(pady=10)
        
        wiring_types = ["Power Wiring", "I/O Wiring", "Ground Wiring"]
        selections = {}
        
        for wtype in wiring_types:
            var = tk.BooleanVar()
            tk.Checkbutton(dialog, text=wtype, variable=var,
                          font=('Segoe UI', 10)).pack(anchor='w', padx=20, pady=5)
            selections[wtype] = var
        
        result = {'confirmed': False, 'types': []}
        
        def confirm():
            selected = [wtype for wtype, var in selections.items() if var.get()]
            if not selected:
                messagebox.showwarning("No Selection", "Please select at least one wiring type")
                return
            result['confirmed'] = True
            result['types'] = selected
            dialog.destroy()
        
        tk.Button(dialog, text="Continue", command=confirm,
                 bg='#10b981', fg='white', font=('Segoe UI', 10, 'bold'),
                 padx=20, pady=8).pack(pady=20)
        
        dialog.wait_window()
        
        if not result['confirmed']:
            return
        
        # Get subcategory details
        name = simpledialog.askstring("New Subcategory", "Enter subcategory name:", parent=self.root)
        if not name:
            return
        
        data = self.collect_template_data(mandatory=True, include_ref=True)
        if not data:
            return
        
        # Add to selected wiring types
        if 'wiring_types' not in category:
            category['wiring_types'] = []
        
        for selected_type in result['types']:
            # Find or create wiring type
            wiring_type = None
            for wt in category['wiring_types']:
                if wt['type'] == selected_type:
                    wiring_type = wt
                    break
            
            if not wiring_type:
                # Create new wiring type
                wiring_type = {
                    'type': selected_type,
                    'ref_number': category.get('ref_number', ''),
                    'subcategories': []
                }
                category['wiring_types'].append(wiring_type)
            
            # Add subcategory
            if 'subcategories' not in wiring_type:
                wiring_type['subcategories'] = []
            
            wiring_type['subcategories'].append({
                "name": name.strip(),
                **data
            })
        
        self.save_categories()
        self.show_defect_library()

    def add_special_subcategory(self, category):
        """Add a special subcategory (not tied to wiring type)"""
        name = simpledialog.askstring("New Special Subcategory", "Enter subcategory name:", parent=self.root)
        if not name:
            return
        
        data = self.collect_template_data(mandatory=True, include_ref=True)
        if not data:
            return
        
        if 'special_subcategories' not in category:
            category['special_subcategories'] = []
        
        category['special_subcategories'].append({
            "name": name.strip(),
            **data
        })
        
        self.save_categories()
        self.show_defect_library()

    def edit_wiring_subcategory(self, category, wiring_type, subcategory):
        """Edit a wiring subcategory"""
        updated = self.collect_template_data(mandatory=True, existing=subcategory, include_ref=True)
        if not updated:
            return
        
        new_name = simpledialog.askstring(
            "Edit Subcategory",
            "Enter new name:",
            initialvalue=subcategory['name'],
            parent=self.root
        )
        if not new_name:
            return
        
        subcategory.clear()
        subcategory.update(updated)
        subcategory["name"] = new_name.strip()
        
        self.save_categories()
        self.show_defect_library()

    def edit_special_subcategory(self, category, subcategory):
        """Edit a special subcategory"""
        updated = self.collect_template_data(mandatory=True, existing=subcategory, include_ref=True)
        if not updated:
            return
        
        new_name = simpledialog.askstring(
            "Edit Special Subcategory",
            "Enter new name:",
            initialvalue=subcategory['name'],
            parent=self.root
        )
        if not new_name:
            return
        
        subcategory.clear()
        subcategory.update(updated)
        subcategory["name"] = new_name.strip()
        
        self.save_categories()
        self.show_defect_library()

    def delete_wiring_subcategory(self, category, wiring_type, sub):
        if not messagebox.askyesno("Confirm", f"Delete subcategory '{sub['name']}'?"):
            return
        
        if 'subcategories' in wiring_type:
            wiring_type['subcategories'].remove(sub)
            self.save_categories()
            self.show_defect_library()

    def delete_special_subcategory(self, category, sub):
        if not messagebox.askyesno("Confirm", f"Delete special subcategory '{sub['name']}'?"):
            return
        
        if 'special_subcategories' in category:
            category['special_subcategories'].remove(sub)
            self.save_categories()
            self.show_defect_library()

    # Regular subcategory methods
    def add_subcategory(self, category):
        """Add a regular subcategory to a parent category"""
        name = simpledialog.askstring("New Subcategory", "Enter subcategory name:", parent=self.root)
        if not name:
            return
        
        data = self.collect_template_data(mandatory=True, include_ref=True)
        if not data:
            return
        
        if 'subcategories' not in category:
            category['subcategories'] = []
        
        category["subcategories"].append({"name": name.strip(), **data})
        self.save_categories()
        self.show_defect_library()

    def edit_subcategory(self, category, subcategory):
        """Edit a regular subcategory"""
        if subcategory.get('inputs'):
            updated = self.collect_template_data(mandatory=True, existing=subcategory, include_ref=True)
            if not updated:
                return
            
            new_name = simpledialog.askstring(
                "Edit Subcategory",
                "Enter new name:",
                initialvalue=subcategory['name'],
                parent=self.root
            )
            if not new_name:
                return
            
            subcategory.clear()
            subcategory.update(updated)
            subcategory["name"] = new_name.strip()
        else:
            # Old style
            new_name = simpledialog.askstring(
                "Edit Subcategory",
                "Enter new name:",
                initialvalue=subcategory['name'],
                parent=self.root
            )
            if not new_name:
                return
            
            new_ref = simpledialog.askstring(
                "Edit Reference Number",
                "Enter reference number (optional):",
                initialvalue=subcategory.get('ref_number', ''),
                parent=self.root
            )
            
            new_template = simpledialog.askstring(
                "Edit Template",
                "Enter new template:",
                initialvalue=subcategory.get('template', ''),
                parent=self.root
            )
            if new_template is None:
                return
            
            subcategory['name'] = new_name.strip()
            subcategory['ref_number'] = new_ref.strip() if new_ref else ""
            subcategory['template'] = new_template.strip()
        
        self.save_categories()
        self.show_defect_library()

    def delete_subcategory(self, category, sub):
        if not messagebox.askyesno("Confirm", f"Delete subcategory '{sub['name']}'?"):
            return
        
        if 'subcategories' in category:
            category['subcategories'].remove(sub)
            self.save_categories()
            self.show_defect_library()

    def run_template(self, template_def, tag_name=None):
        """
        Execute a punch text template by prompting the user for inputs.
    
        Returns the fully formatted punch text.
        """
        values = {}
        if tag_name:
            values["tag"] = tag_name
        
        for inp in template_def.get("inputs", []):
            val = simpledialog.askstring(
                "Input Required",
                inp["label"],
                parent=self.root
            )
            if not val:
                return None
            values[inp["name"]] = val.strip()
        
        try:
            return template_def["template"].format(**values)
        except KeyError as e:
            messagebox.showerror("Template Error", f"Missing placeholder: {e}")
            return None

    def handle_template_category(self, category, bbox_page=None):
        """Handle template category execution"""
        if category.get('inputs'):
            punch_text = self.run_template(category, tag_name=None)
            if not punch_text:
                return
        else:
            punch_text = category.get('template', 'No template defined')
        
        messagebox.showinfo("Generated Punch Text", 
                          f"Category: {category['name']}\n\nPunch Text:\n{punch_text}")

    def handle_subcategory(self, category, subcategory, bbox_page=None):
        """Handle subcategory execution"""
        if subcategory.get('inputs'):
            punch_text = self.run_template(subcategory, tag_name=None)
            if not punch_text:
                return
        else:
            punch_text = subcategory.get('template', 'No template defined')
        
        messagebox.showinfo("Generated Punch Text",
                          f"Category: {category['name']}\nSubcategory: {subcategory['name']}\n\nPunch Text:\n{punch_text}")
    # ============ NEW: TEMPLATE EXCEL EDITOR ============
    def show_template_editor(self):
        """
        Display the Template Excel Editor interface.
    
        Allows:
        - Opening the template
        - Replacing the template
        - Exporting copies
        - Verifying structure integrity
        """
        self.set_active_nav('template_editor')
        self.clear_content()
        
        # Centered container
        center_container = tk.Frame(self.content, bg='#f8fafc')
        center_container.place(relx=0.5, rely=0, anchor='n', relwidth=0.7, relheight=1.0)
        
        # Header
        header = tk.Frame(center_container, bg='#f8fafc')
        header.pack(fill=tk.X, padx=30, pady=(20, 10))
        
        tk.Label(header, text="📝 Template Excel Editor", font=('Segoe UI', 16, 'bold'),
                bg='#f8fafc').pack(side=tk.LEFT)
        
        # Info card
        info_card = tk.Frame(center_container, bg='#eff6ff', relief=tk.SOLID, borderwidth=1)
        info_card.pack(fill=tk.X, padx=30, pady=10)
        
        info_text = f"""Current Template: Emerson.xlsx

This template is used by both Quality Inspection and Production tools.
Any changes made here will affect all new projects.

Template Location: {self.template_excel_file}
"""
        
        tk.Label(info_card, text=info_text, font=('Segoe UI', 10),
                bg='#eff6ff', fg='#1e40af', justify='left').pack(padx=20, pady=15)
        
        # Action buttons
        action_frame = tk.Frame(center_container, bg='white', relief=tk.SOLID, borderwidth=1)
        action_frame.pack(fill=tk.X, padx=30, pady=10)
        
        tk.Label(action_frame, text="Template Actions", font=('Segoe UI', 12, 'bold'),
                bg='white', fg='#1e293b').pack(anchor='w', padx=20, pady=(15, 10))
        
        btn_style = {
            'font': ('Segoe UI', 10, 'bold'),
            'relief': tk.FLAT,
            'cursor': 'hand2',
            'padx': 20,
            'pady': 12,
            'width': 30
        }
        
        # Open template button
        tk.Button(action_frame, text="📂 Open Template Excel",
                 command=self.open_template_excel,
                 bg='#3b82f6', fg='white', **btn_style).pack(padx=20, pady=(0, 10))
        
        # Replace template button
        tk.Button(action_frame, text="🔄 Replace Template File",
                 command=self.replace_template_excel,
                 bg='#f59e0b', fg='white', **btn_style).pack(padx=20, pady=(0, 10))
        
        # Export template button
        tk.Button(action_frame, text="💾 Export Template Copy",
                 command=self.export_template_copy,
                 bg='#10b981', fg='white', **btn_style).pack(padx=20, pady=(0, 15))
        
        # Template structure info
        structure_frame = tk.Frame(center_container, bg='white', relief=tk.SOLID, borderwidth=1)
        structure_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=10)
        
        tk.Label(structure_frame, text="📋 Template Structure", font=('Segoe UI', 12, 'bold'),
                bg='white', fg='#1e293b').pack(anchor='w', padx=20, pady=(15, 10))
        
        structure_text = """Required Sheets:
• Interphase - Project checklist and status tracking
• Punch Sheet - Defect and punch list management

The template must maintain:
✓ Correct sheet names (case-sensitive)
✓ Header structure in rows 1-7 (Punch Sheet) and 1-10 (Interphase)
✓ Column mapping for automated data entry
✓ Merged cells for project information

Warning: Modifying the template structure may cause errors in Quality and Production tools.
"""
        
        tk.Label(structure_frame, text=structure_text, font=('Segoe UI', 9),
                bg='white', fg='#64748b', justify='left').pack(anchor='w', padx=40, pady=(0, 15))
        
        # Check template button
        tk.Button(structure_frame, text="✓ Verify Template Structure",
                 command=self.verify_template_structure,
                 bg='#8b5cf6', fg='white', font=('Segoe UI', 10, 'bold'),
                 padx=20, pady=10, relief=tk.FLAT, cursor='hand2').pack(pady=(0, 20))
    
    def open_template_excel(self):
        """Open template Excel file in default application"""
        if not os.path.exists(self.template_excel_file):
            messagebox.showerror("Template Not Found", 
                               f"Template file not found:\n{self.template_excel_file}")
            return
        
        try:
            if sys.platform == 'win32':
                os.startfile(self.template_excel_file)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', self.template_excel_file])
            else:
                subprocess.Popen(['xdg-open', self.template_excel_file])
            
            messagebox.showinfo("Template Opened", 
                              "Template Excel file opened.\n\n"
                              "⚠️ Important:\n"
                              "• Do not modify sheet names\n"
                              "• Do not change header structure\n"
                              "• Save changes before closing\n\n"
                              "Changes will affect all new projects.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open template:\n{e}")
    
    def replace_template_excel(self):
        """Replace template Excel with a new file"""
        confirm = messagebox.askyesno(
            "Replace Template",
            "⚠️ WARNING\n\n"
            "This will replace the current template file.\n"
            "All new projects will use the new template.\n\n"
            "Existing projects will NOT be affected.\n\n"
            "Continue?",
            icon='warning'
        )
        
        if not confirm:
            return
        
        # Select new template file
        new_template = filedialog.askopenfilename(
            title="Select New Template Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not new_template:
            return
        
        try:
            # Verify it's a valid Excel file
            wb = load_workbook(new_template, data_only=True)
            
            # Check for required sheets
            required_sheets = ['Interphase', 'Punch Sheet']
            missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
            
            if missing_sheets:
                wb.close()
                messagebox.showerror("Invalid Template", 
                                   f"Template is missing required sheets:\n" + 
                                   "\n".join(missing_sheets))
                return
            
            wb.close()
            
            # Backup current template
            backup_path = self.template_excel_file + ".backup"
            if os.path.exists(self.template_excel_file):
                import shutil
                shutil.copy2(self.template_excel_file, backup_path)
            
            # Replace template
            import shutil
            shutil.copy2(new_template, self.template_excel_file)
            
            messagebox.showinfo("Template Replaced", 
                              f"✓ Template successfully replaced!\n\n"
                              f"New template: {os.path.basename(new_template)}\n"
                              f"Backup saved: {os.path.basename(backup_path)}\n\n"
                              "All new projects will use this template.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to replace template:\n{e}")
    
    def export_template_copy(self):
        """Export a copy of the template"""
        save_path = filedialog.asksaveasfilename(
            title="Save Template Copy As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Emerson_Template_Copy.xlsx"
        )
        
        if not save_path:
            return
        
        try:
            import shutil
            shutil.copy2(self.template_excel_file, save_path)
            
            messagebox.showinfo("Template Exported", 
                              f"✓ Template copy saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export template:\n{e}")
    
    def verify_template_structure(self):
        """Verify template Excel structure"""
        if not os.path.exists(self.template_excel_file):
            messagebox.showerror("Template Not Found", 
                               "Template file not found!")
            return
        
        try:
            wb = load_workbook(self.template_excel_file, data_only=True)
            
            issues = []
            warnings = []
            
            # Check required sheets
            required_sheets = ['Interphase', 'Punch Sheet']
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    issues.append(f"✗ Missing required sheet: {sheet_name}")
                else:
                    warnings.append(f"✓ Sheet found: {sheet_name}")
            
            # Check Punch Sheet structure
            if 'Punch Sheet' in wb.sheetnames:
                ws = wb['Punch Sheet']
                
                # Check for expected columns
                expected_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
                for col in expected_cols:
                    if ws[f'{col}8'].value is None and ws[f'{col}7'].value is None:
                        warnings.append(f"⚠ Column {col} header might be missing")
            
            # Check Interphase structure
            if 'Interphase' in wb.sheetnames:
                ws = wb['Interphase']
                
                # Check for key cells
                if ws['C4'].value is None:
                    warnings.append("⚠ Project Name cell (C4) is empty")
                if ws['C6'].value is None:
                    warnings.append("⚠ Sales Order cell (C6) is empty")
            
            wb.close()
            
            # Show results
            result_text = "Template Structure Verification\n\n"
            
            if issues:
                result_text += "❌ CRITICAL ISSUES:\n"
                result_text += "\n".join(issues)
                result_text += "\n\n"
            
            if warnings:
                result_text += "ℹ️ Information:\n"
                result_text += "\n".join(warnings[:10])  # Show first 10
                if len(warnings) > 10:
                    result_text += f"\n... and {len(warnings) - 10} more"
            
            if not issues:
                result_text += "\n\n✓ Template structure appears valid!"
                messagebox.showinfo("Verification Complete", result_text)
            else:
                messagebox.showwarning("Verification Issues", result_text)
            
        except Exception as e:
            messagebox.showerror("Verification Error", 
                               f"Failed to verify template:\n{e}")
    
    # ============ REPORT GENERATOR ============


def main():
    root = tk.Tk()
    app = ManagerUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
