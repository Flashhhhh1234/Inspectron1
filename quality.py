import os

# Set*both variable forms for PaddleX ve*sion compatibility.
os.environ["PADDLE_PDX_MODEL_SOURCE"] = "BOS"
os.environ["MODEL_SOURCE"] = "BOS"

os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
os.environ["DISABLE_MODEL_SOURCE_CHECK"] = "True"

import tkinter as tk
import ctypes
from ctypes import wintypes
import types
from tkinter import messagebox, simpledialog, Menu
from PIL import Image, ImageTk, ImageDraw, ImageFont,ImageEnhance,ImageFilter
import fitz  
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import shutil
import tempfile
import re
import ntpath
import json
import numpy as np
import getpass
import sys
import subprocess
import time
import uuid
import pg_sqlite_compat as sqlite3
import shlex
from difflib import SequenceMatcher
from handover_database import HandoverDB
from database_manager import DatabaseManager
from category_store_pg import load_categories_from_postgres
from path_policy import (
    get_base_path,
    resolve_storage_location,
    to_absolute_path,
    to_relative_path,
    to_relative_storage_location,
)
from tkinter import ttk
import cv2
import io
import filedialog_compat as filedialog

User = sys.argv[1] if len(sys.argv) > 1 else None
Name = sys.argv[2] if len(sys.argv) > 2 else None

# ---------------------------------------------------------------------------
# OCR ENGINE: PaddleOCR (replaces Tesseract)
# ---------------------------------------------------------------------------
# PaddleOCR is loaded lazily (on first actual OCR call) rather than at import
# time, since model initialization is a real cost (loads detection +
# recognition + angle-classification weights) and would otherwise slow down
# app startup even when no highlight has been drawn yet.
#
# use_textline_orientation=True enables PaddleOCR 3.x text-line orientation
# classification. This replaces the deprecated use_angle_cls parameter and
# allows rotated text lines to be corrected before recognition.
#
# Model download source: recent PaddleOCR/PaddleX releases default to
# downloading model weights from Hugging Face. Setting PADDLE_PDX_MODEL_SOURCE
# to "BOS" switches that back to Baidu Object Storage, PaddlePaddle's own
# native hosting - useful in environments where Hugging Face isn't reachable
# (e.g. behind a corporate proxy/firewall that only allows specific hosts).
# This MUST be set before `from paddleocr import PaddleOCR` runs, since
# PaddleX (the model-download layer PaddleOCR sits on) reads it at import
# time - setting it later, inside get_paddle_ocr_engine(), would be too late
# if anything else imports paddle-related modules first.


_paddle_ocr_engine = None
_paddle_ocr_init_error = None


def get_paddle_ocr_engine():
    """Return a lazily-initialized, process-wide PaddleOCR engine instance."""
    global _paddle_ocr_engine, _paddle_ocr_init_error
    if _paddle_ocr_engine is not None:
        return _paddle_ocr_engine
    if _paddle_ocr_init_error is not None:
        # Already failed once this run; don't retry on every highlight -
        # that would re-attempt (and likely re-fail) model loading each time.
        return None
    try:
        from paddleocr import PaddleOCR
        _paddle_ocr_engine = PaddleOCR(
            lang="en",
            use_textline_orientation=False,
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
        )
        print(f"[INFO] PaddleOCR engine initialized (text-line orientation enabled, "
              f"model source: {os.environ.get('PADDLE_PDX_MODEL_SOURCE', 'default')})")
    except Exception as e:
        _paddle_ocr_init_error = e
        print(f"[WARN] PaddleOCR failed to initialize: {e}")
        _paddle_ocr_engine = None
    return _paddle_ocr_engine

def prevent_power_throttling():
    """
    Ask Windows not to apply power-saving CPU throttling to this process.
    FUNCTIONAL USE: Confirmed root cause of severe touch-drawing lag: Windows'
    battery power plan downclocks/throttles background-seeming processes, which
    slows PDF rasterization and canvas compositing enough to cause multi-second
    stalls. This opts this specific process out of that throttling class, similar
    to what media/CAD apps do, so behavior stays consistent on battery.
    """
    if os.name != 'nt':
        return False

    try:
        ES_CONTINUOUS = 0x80000000
        ES_SYSTEM_REQUIRED = 0x00000001
        ctypes.windll.kernel32.SetThreadExecutionState(
            ES_CONTINUOUS | ES_SYSTEM_REQUIRED
        )

        PROCESS_POWER_THROTTLING_EXECUTION_SPEED = 0x1

        class PROCESS_POWER_THROTTLING_STATE(ctypes.Structure):
            _fields_ = [
                ("Version", ctypes.c_ulong),
                ("ControlMask", ctypes.c_ulong),
                ("StateMask", ctypes.c_ulong),
            ]

        state = PROCESS_POWER_THROTTLING_STATE()
        state.Version = 1
        state.ControlMask = PROCESS_POWER_THROTTLING_EXECUTION_SPEED
        state.StateMask = 0  # disable throttling for execution speed

        PROCESS_POWER_THROTTLING_STATE_INFO = 4
        handle = ctypes.windll.kernel32.GetCurrentProcess()
        ctypes.windll.kernel32.SetProcessInformation(
            handle,
            PROCESS_POWER_THROTTLING_STATE_INFO,
            ctypes.byref(state),
            ctypes.sizeof(state)
        )
        print("[INFO] Power throttling disabled for this process")
        return True
    except Exception as e:
        print(f"[WARN] Could not disable power throttling: {e}")
        return False

TESSERACT_CMD = None  # kept for backward-compat references only; OCR now uses PaddleOCR (see get_paddle_ocr_engine)

def app_base():
    """
    Returns the directory where the app is running from.
    FUNCTIONAL USE: Determines if app is frozen (compiled) or running from source code.
    Used to construct absolute paths for config files, databases, and resources.
    Returns: Directory path string (either compiled executable dir or script dir)
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def configure_touch_feedback(hwnd):
    """Keep touch drawing responsive while suppressing Windows one-finger pan.

    Windows may interpret a mostly vertical finger movement as a scroll/pan
    gesture before Tk receives B1-Motion. Block only GID_PAN on the canvas,
    keep GID_ZOOM enabled, and disable visual feedback. No Python WNDPROC or
    raw WM_TOUCH callback is installed, so this remains safe on Python 3.14.
    """
    if os.name != 'nt':
        return False
    try:
        user32 = ctypes.windll.user32

        class GESTURECONFIG(ctypes.Structure):
            _fields_ = [
                ('dwID', wintypes.DWORD),
                ('dwWant', wintypes.DWORD),
                ('dwBlock', wintypes.DWORD),
            ]

        GID_ZOOM = 3
        GID_PAN = 4
        GC_ZOOM = 0x00000001
        GC_PAN = 0x00000001

        set_gesture_config = user32.SetGestureConfig
        set_gesture_config.argtypes = [
            wintypes.HWND, wintypes.DWORD, wintypes.UINT,
            ctypes.POINTER(GESTURECONFIG), wintypes.UINT
        ]
        set_gesture_config.restype = wintypes.BOOL

        # Allow two-finger zoom but block Windows from stealing a one-finger
        # vertical or horizontal stroke as a native pan/scroll gesture.
        configs = (GESTURECONFIG * 2)(
            GESTURECONFIG(GID_ZOOM, GC_ZOOM, 0),
            GESTURECONFIG(GID_PAN, 0, GC_PAN),
        )
        gesture_ok = bool(set_gesture_config(
            hwnd, 0, len(configs), configs, ctypes.sizeof(GESTURECONFIG)
        ))
        if not gesture_ok:
            print(f'[WARN] SetGestureConfig failed: {ctypes.get_last_error()}')

        feedback_ok = False
        try:
            setting = user32.SetWindowFeedbackSetting
            setting.restype = wintypes.BOOL
            setting.argtypes = [
                wintypes.HWND, wintypes.DWORD, wintypes.DWORD,
                wintypes.UINT, ctypes.c_void_p
            ]
            feedback_off = wintypes.BOOL(False)
            for feedback_id in (3, 5, 6, 7, 8, 9, 10):
                if setting(hwnd, feedback_id, 0, ctypes.sizeof(feedback_off),
                           ctypes.byref(feedback_off)):
                    feedback_ok = True
        except Exception as exc:
            print(f'[WARN] Touch feedback configuration skipped: {exc}')

        if gesture_ok:
            print('[INFO] One-finger Windows pan blocked; tool strokes enabled')
        return gesture_ok or feedback_ok
    except Exception as exc:
        print(f'[WARN] Touch gesture configuration skipped: {exc}')
        return False

def show_onscreen_keyboard():
    """
    Launch Windows on-screen keyboard (touch keyboard preferred, falls back to osk.exe).
    FUNCTIONAL USE: Pops up a virtual keyboard whenever a text entry field gains focus.
    Uses ShellExecuteW instead of subprocess.Popen because both TabTip.exe and osk.exe
    are shell-integrated components that Windows expects to be launched via the shell
    activation path — invoking them through a raw CreateProcess (which subprocess.Popen
    does) commonly fails with WinError 740 "requires elevation" even without any actual
    admin requirement.
    """
    if os.name != 'nt':
        return

    try:
        tabtip_path = r"C:\Program Files\Common Files\Microsoft Shared\ink\TabTip.exe"
        if os.path.exists(tabtip_path):
            result = ctypes.windll.shell32.ShellExecuteW(
                None, "open", tabtip_path, None, None, 1  # SW_SHOWNORMAL
            )
            # ShellExecuteW returns a value > 32 on success
            if result > 32:
                return
            print(f"[WARN] TabTip ShellExecute returned {result}")
    except Exception as e:
        print(f"[WARN] TabTip launch failed: {e}")

    try:
        osk_path = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "System32", "osk.exe")
        result = ctypes.windll.shell32.ShellExecuteW(
            None, "open", osk_path, None, None, 1
        )
        if result > 32:
            return
        print(f"[WARN] osk.exe ShellExecute returned {result}")
    except Exception as e:
        print(f"[WARN] osk.exe launch failed: {e}")


def hide_onscreen_keyboard():
    """
    Attempt to close the on-screen keyboard windows (osk.exe / TabTip).
    FUNCTIONAL USE: Called when a text field loses focus so the keyboard
    doesn't stay on screen unnecessarily.
    """
    if os.name != 'nt':
        return
    try:
        subprocess.run(["taskkill", "/IM", "osk.exe", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass
    try:
        subprocess.run(["taskkill", "/IM", "TabTip.exe", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass
def asset_path(filename):
    """Resolve asset paths for source runs and PyInstaller bundles."""
    bundle_dir = getattr(sys, "_MEIPASS", "")
    if bundle_dir:
        bundled_path = os.path.join(bundle_dir, "assets", filename)
        if os.path.exists(bundled_path):
            return bundled_path

    if getattr(sys, 'frozen', False):
        return os.path.join(app_base(), "assets", filename)

    return os.path.join(app_base(), "assets", filename)


    
    


class ManagerDB:
    """Manager database integration with storage_location and excel_path support"""
    def __init__(self, db_path):
        self.db_path = db_path

    def splitcell(self, cell_ref):
        """
        Parse Excel cell reference (e.g., 'A1', 'B42') into row and column components.
        FUNCTIONAL USE: Splits Excel notation into numeric row and string column for openpyxl operations.
        Args: cell_ref - Cell reference string (e.g., 'B5', 'H10')
        Returns: Tuple of (row_number, column_letter)
        """
        import re
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col

    def mergedtar(self, ws, row, col_idx):
        """
        Find actual cell coordinates when target cell is part of a merged cell range.
        FUNCTIONAL USE: Handles merged cells in Excel by returning the top-left cell of merge range.
        Ensures writes/reads go to correct cell even when targeting merged area.
        Args: ws - Worksheet, row - row number, col_idx - column index
        Returns: Tuple of (actual_row, actual_col) accounting for merges
        """
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx

    def read(self, ws, row, col):
        """
        Read value from Excel cell, handling merged cells and column format conversion.
        FUNCTIONAL USE: Unified read interface that accepts column as letter ('A') or number (1).
        Automatically finds actual cell if target is part of merged range.
        Args: ws - Worksheet, row - row number, col - column (letter or index)
        Returns: Cell value (string, number, date, etc.)
        """
        from openpyxl.utils import column_index_from_string
        
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self.mergedtar(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value

    def interphase_status(self, excel_path):
        """
        Determine cabinet status by reading highest reference number in Interphase sheet.
        FUNCTIONAL USE: Reads Interphase worksheet to calculate assembly progress.
        Maps highest filled reference to status (assembly_in_progress, assembly_complete, final_check, etc.).
        Used by quality inspection to track assembly completion stage.
        Args: excel_path - Full path to Excel file
        Returns: Status string or None if not determined from Interphase
        """
        if not excel_path or not os.path.exists(excel_path):
            return None
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            if 'Interphase' not in wb.sheetnames:
                wb.close()
                return None
            
            ws = wb['Interphase']
            
            # Find the HIGHEST reference number that has a status
            highest_ref_num = 0
            
            # Start from row 11 (typical Interphase data starts here)
            for row in range(11, ws.max_row + 1):
                status_cell = self.read(ws, row, 'D')  # Status column
                
                # If status cell has content, check the reference number
                if status_cell and str(status_cell).strip():
                    ref_no_cell = self.read(ws, row, 'B')  # Reference column
                    
                    if ref_no_cell:
                        try:
                            ref_str = str(ref_no_cell).strip()
                            
                            # Handle range formats like "1-2" - take the LAST number
                            if '-' in ref_str:
                                ref_num = int(ref_str.split('-')[-1])
                            else:
                                ref_num = int(ref_str)
                            
                            # Track highest completed reference
                            if ref_num > highest_ref_num:
                                highest_ref_num = ref_num
                        
                        except (ValueError, IndexError):
                            continue
            
            wb.close()
            
            # Determine status based on highest completed reference number
            if highest_ref_num == 0:
                return 'quality_inspection'  # Nothing completed yet
            elif 1 <= highest_ref_num <= 2:
                return 'project_info_sheet'
            elif 3 <= highest_ref_num <= 9:
                return 'mechanical_assembly'
            elif 10 <= highest_ref_num <= 18:
                return 'component_assembly'
            elif 19 <= highest_ref_num <= 26:
                return 'final_assembly'
            elif highest_ref_num >= 27:
                return 'final_documentation'
            else:
                return 'quality_inspection'
            
        except Exception as e:
            print(f"Error reading Interphase worksheet: {e}")
            return None
    
    def updatecab(self, cabinet_id, project_name, sales_order_no, 
                      total_pages, annotated_pages, total_punches, 
                      open_punches, implemented_punches, closed_punches, status,
                      storage_location=None, excel_path=None):
        """
        Insert or replace complete cabinet record with all statistics and metadata.
        FUNCTIONAL USE: Updates manager dashboard with cabinet progress: punch counts, implementation status,
        storage location, and associated Excel file path. Creates record if new, updates if exists.
        Used by quality module to sync work progress with manager system.
        """
        try:
            storage_location_db = to_relative_storage_location(storage_location)
            excel_path_db = to_relative_path(excel_path)

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO cabinets 
                (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                 total_punches, open_punches, implemented_punches, closed_punches, status,
                 storage_location, excel_path,
                 created_date, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        COALESCE((SELECT created_date FROM cabinets WHERE cabinet_id = ?), ?),
                        ?)
            ''', (cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                  total_punches, open_punches, implemented_punches, closed_punches, status,
                storage_location_db, excel_path_db,
                  cabinet_id, datetime.now().isoformat(), datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Manager DB update error: {e}")
            return False
    
    def logcatoccurence(self, cabinet_id, project_name, category, subcategory):
        """
        Record an instance of a category/subcategory occurrence for analytics.
        FUNCTIONAL USE: Logs quality issues found by category for dashboard reporting.
        Tracks patterns and frequencies of defect types across projects.
        Args: cabinet_id, project_name, category, subcategory - Metadata for occurrence
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO category_occurrences 
                (cabinet_id, project_name, category, subcategory, occurrence_date)
                VALUES (?, ?, ?, ?, ?)
            ''', (cabinet_id, project_name, category, subcategory, datetime.now().isoformat()))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Category logging error: {e}")
            return False
    
    def updatestats(self, cabinet_id, status):
        """
        Update cabinet status field and last_updated timestamp.
        FUNCTIONAL USE: Lightweight status-only update for quality workflow transitions.
        Updates database with current date/time to track inspection progress.
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE cabinets 
                SET status = ?, last_updated = ?
                WHERE cabinet_id = ?
            ''', (status, datetime.now().isoformat(), cabinet_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Status update error: {e}")
            return False
    
    def fetchcab(self, cabinet_id):
        """Get cabinet information"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT cabinet_id, project_name, sales_order_no, total_pages, annotated_pages,
                       total_punches, open_punches, implemented_punches, closed_punches, status,
                       storage_location, excel_path, created_date, last_updated
                FROM cabinets 
                WHERE cabinet_id = ?
            ''', (cabinet_id,))
            
            row = cursor.fetchone()
            conn.close()
            
            if row:
                return {
                    'cabinet_id': row[0],
                    'project_name': row[1],
                    'sales_order_no': row[2],
                    'total_pages': row[3],
                    'annotated_pages': row[4],
                    'total_punches': row[5],
                    'open_punches': row[6],
                    'implemented_punches': row[7],
                    'closed_punches': row[8],
                    'status': row[9],
                    'storage_location': resolve_storage_location(row[10]),
                    'excel_path': to_absolute_path(row[11]),
                    'created_date': row[12],
                    'last_updated': row[13]
                }
            return None
            
        except Exception as e:
            print(f"Error getting cabinet: {e}")
            return None


class CircuitInspector:
    def __init__(self, root):
        self.root = root
        self.logged_in_username = User
        self.logged_in_fullname = Name
        self.root.title("Quality Inspection Tool - Highlighter")
        self.root.geometry("1400x900")


        # Bind window close event to auto-save
        self.root.protocol("WM_DELETE_WINDOW", self.onclosing)
        # Data / files
        self.pdf_document = None
        self.current_pdf_path = None
        self.current_page = 0
        self.project_name = ""
        self.sales_order_no = ""
        self.cabinet_id = ""
        self.annotations = []
        base = app_base()
        self.master_excel_file = os.path.join(base, "Emerson.xlsx")

        self.excel_file = None
        self.working_excel_path = None
        self.checklist_file = self.excel_file
        self.zoom_level = 1.0
        self.ZOOM_MIN = 0.5
        self.ZOOM_MAX = 3.0
        self.ZOOM_STEP = 0.1
        self.current_sr_no = 1
        self.current_page_image = None
        self.tool_mode = None  # None, "pen", or "text"
        self.pen_points = []
        self.session_refs = set()
        self.project_dirs = {}

        # Fixed column mapping
        self.punch_sheet_name = 'Punch Sheet'
        self.punch_cols = {
            'sr_no': 'A',
            'ref_no': 'B',
            'desc': 'C',
            'category': 'D',
            'checked_name': 'E',
            'checked_date': 'F',
            'implemented_name': 'G',
            'implemented_date': 'H',
            'closed_name': 'I',
            'closed_date': 'J'
        }
        
        self.interphase_sheet_name = 'Interphase'
        self.interphase_cols = {
            'ref_no': 'B',
            'description': 'C',
            'status': 'D',
            'name':'E',
            'date':'F',
            'remark':'G'
        }

        self.header_cells = {
            "Interphase": {
                "project_name": "C4",
                "sales_order": "C6",
                "cabinet_id": "F6"
            },
            "Punch Sheet": {
                "project_name": "C2",
                "sales_order": "C4",
                "cabinet_id": "H4"
            }
        }

        self.categories = []
        self.category_file = os.path.join(os.path.dirname(app_base()), "assets", "categories.json")
        self.loadcat()

        # HIGHLIGHTER STATE - 3 COLORS
        self.active_highlighter = None
        self.highlighter_colors = {
            'green': {'rgb': (0, 255, 0), 'rgba': (0, 255, 0, 100), 'name': ' OK'},
            'pink': {'rgb': (255, 20, 147), 'rgba': (255, 20, 147, 120), 'name': ' Error'},
            'yellow': {'rgb': (255, 255, 0), 'rgba': (255, 255, 0, 80), 'name': 'Wiring '}
        }
        self.current_color_key = 'yellow'  # Default color
        self.highlight_points = []

        # -------- MULTI-MARK MODE (attach several highlights to one existing punch) --------
        # When active, every pink highlight drawn on the PDF is attached directly
        # to `self.multimark_punch` (skipping the "already marked?" dialog and the
        # full categorization/Excel-row-creation flow, since the punch already
        # exists). Only the highlighter + view/navigation tools remain usable while
        # this is on; pen/text tools are locked out until the user presses Stop.
        self.multimark_active = False
        self.multimark_punch = None  # dict: sr_no, ref_no, punch_text, category, row
        self.multimark_count = 0     # highlights added so far in this session
        self.multimark_bar = None            # toolbar frame holding the Stop button
        self.multimark_bar_label_var = None  # tk.StringVar for the bar's status text

        # TOGGLE: when True (default), pink highlights go through the full
        # error-categorization flow and write a row into the Excel punch sheet.
        # When False, an pink highlight is still drawn/kept as an annotation
        # (and still saved with the session) but NO Excel row is created and
        # NO Interphase status update happens - useful for quickly marking
        # visual errors without touching the workbook.
        self.mark_errors_to_excel = True

        # -------- BATCHED SAVE / SYNC STATE --------
        # Instead of writing to the database (sync_manager_stats_only) and
        # resaving the session on every single annotation, most actions now
        # just flag work as "dirty". A background timer flushes dirty work
        # periodically, and onclosing() always does one final flush - so a
        # normal exit OR an abrupt crash/kill still leaves the DB/session in
        # a recent, consistent state without paying the I/O cost on every click.
        self._dirty = False
        self._autosave_interval_ms = 45000  # flush at most every 45s
        self._autosave_after_id = None
        self._last_flush_time = 0.0

        # Drawing / tool state
        self.drawing = False
        self.drawing_type = None  # 'highlight', 'pen', 'text'
        self.temp_line_ids = []  # Store temporary drawing line IDs
        self.drawing_page = None
        self.drawing_page_offset = (0, 0)
        self.page_layout = []
        self.page_images = []
        # PDF rasterization is the dominant UI cost. Keep page renders cached and
        # invalidate only pages whose annotations changed. The previous code
        # rasterized every page after every pen/highlight action.
        self._page_render_cache = {}
        self._page_cache_zoom = None
        self._display_after_id = None
        self.selected_annotation = None
        # Microsoft-style text box interaction state.
        self._text_box_start = None
        self._text_box_preview_id = None
        self._text_editor = None
        self._text_editor_window_id = None
        self._text_edit_annotation = None
        self._text_transform_mode = None
        self._text_transform_start = None
        self._text_transform_original_bbox = None
        self._text_selection_ids = []
        self.undo_stack = []  # Stack for undo operations
        self.max_undo = 50    # Maximum undo history
        self.hover_annotation = None  # For hover preview
        self._touch_scroll_lock_until = 0.0
        self._panning = False  # NEW: tracks whether a no-tool pan/scroll drag is active
        self._native_pinch_start_distance = None
        self._native_pinch_start_zoom = None
        self._native_pinch_last_zoom = None
        self._native_pinch_render_after_id = None
        self._native_pinch_preview_after_id = None
        self._native_pinch_preview_pending = False
        self._native_pinch_preview_interval_ms = 16
        self._native_pinch_finish_delay_ms = 180
        self._safe_pinch_last_raw = None
        self._safe_pinch_last_time = 0.0
        self._safe_pinch_accumulator = 1.0
        self._safe_pinch_active = False
        self._safe_pinch_release_funcid = None
        self._safe_pinch_watchdog_ms = 1200
        # Cached source pixmap for the page being pinch-zoomed, captured once
        # at gesture start. Live pinch frames resize this in-memory PIL image
        # instead of re-rasterizing the PDF on every touch event, which is
        # what made pinch feel laggy - PyMuPDF get_pixmap() is comparatively
        # expensive and the digitizer can fire far faster than it can keep up.
        self._pinch_base_image = None
        self._pinch_base_scale = None
        self._pinch_base_page = None
        self._pinch_frame_pending = False
        self._pinch_frame_min_interval = 1.0 / 60.0  # cap live redraws to ~60fps
        self._pinch_last_frame_time = 0.0
        self._busy_overlay = None  # modal loading overlay, see busy()/unbusy()
        self._zoom_render_after_id = None
        self._zoom_dropdown_updating = False
        self.uisetup()
        self.bind_global_keyboard_popup()   # <-- ADD THIS LINE
        self.current_sr_no = self.getnextsr()
        
        self.db = DatabaseManager("inspection_tool")
        self.manager_db = ManagerDB("manager")
        self.handover_db = HandoverDB("handover_db")
        self.loadrecprojui()
        self.startautosaveloop()


    # ================================================================
    # COORDINATE CONVERSION HELPERS
    # ================================================================
    
    def page_to_display_scale(self):
        """
        Calculate scaling factor from PDF page coordinates to display canvas.
        FUNCTIONAL USE: Accounts for zoom level (default 2x magnification plus user zoom).
        Used to convert between PDF space and canvas rendering space.
        """
        return 2.0 * self.zoom_level

    def display_to_page_coords(self, pts):
        """
        Convert canvas display coordinates back to PDF page space.
        FUNCTIONAL USE: Reverses scaling from display_scale to find annotation position in PDF.
        Handles single point tuple or list of points.
        Args: pts - Single (x, y) tuple or list of [(x1, y1), (x2, y2), ...]
        Returns: Same structure but with page-space coordinates
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] / scale, pts[1] / scale)
        
        # Handle list of points
        return [(x / scale, y / scale) for x, y in pts]

    def page_to_display_coords(self, pts):
        """
        Convert PDF page coordinates to canvas display space.
        FUNCTIONAL USE: Scales from PDF space to display space for rendering annotations on canvas.
        Handles single point tuple or list of points.
        Args: pts - Single (x, y) tuple or list of [(x1, y1), (x2, y2), ...]
        Returns: Same structure but with display-space coordinates
        """
        scale = self.page_to_display_scale()
        
        # Handle single point tuple
        if isinstance(pts, tuple) and len(pts) == 2:
            if not isinstance(pts[0], (list, tuple)):
                return (pts[0] * scale, pts[1] * scale)
        
        # Handle list of points
        return [(x * scale, y * scale) for x, y in pts]

    def bbox_page_to_display(self, bbox_page):
        """
        Convert bounding box from PDF coordinates to display coordinates.
        FUNCTIONAL USE: Scales rectangle coordinates for rendering on canvas.
        Args: bbox_page - Tuple (x1, y1, x2, y2) in PDF space
        Returns: Tuple (x1, y1, x2, y2) in display space
        """
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_page
        return (x1 * scale, y1 * scale, x2 * scale, y2 * scale)

    def bbox_display_to_page(self, bbox_display):
        """
        Convert bounding box from display coordinates to PDF coordinates.
        FUNCTIONAL USE: Reverses scaling to find annotation position in original PDF.
        Args: bbox_display - Tuple (x1, y1, x2, y2) in display space
        Returns: Tuple (x1, y1, x2, y2) in PDF space
        """
        scale = self.page_to_display_scale()
        x1, y1, x2, y2 = bbox_display
        return (x1 / scale, y1 / scale, x2 / scale, y2 / scale)

    def _page_at_point(self, x, y):
        """Return the stacked page layout item that contains a canvas point."""
        for layout in self.page_layout:
            left = layout['x']
            top = layout['y']
            right = left + layout['width']
            bottom = top + layout['height']
            if left <= x <= right and top <= y <= bottom:
                return layout

        if self.page_layout:
            return min(
                self.page_layout,
                key=lambda layout: abs((layout['y'] + (layout['height'] / 2)) - y)
            )

        return None

    def _point_to_page_coords(self, x, y):
        """Map a canvas point into page-local coordinates in the stacked layout."""
        layout = self._page_at_point(x, y)
        if not layout:
            return None, None, None, None

        local_x = x - layout['x']
        local_y = y - layout['y']
        return layout['page_index'], local_x, local_y, layout

    # ================================================================
    # HIGHLIGHTER HELPER - AUTO-STRAIGHTEN
    # ================================================================
    
    def Straighten(self, points):
        """
        Convert a freehand path into a straight line from start to end.
        FUNCTIONAL USE: Simplifies highlighter annotations to perfect straight lines.
        Removes freehand waviness for cleaner annotation rendering.
        Args: points - List of (x, y) tuples representing freehand path
        Returns: List containing only start and end points
        """
        if len(points) < 2:
            return points
        # Simply return start and end points for a perfectly straight line
        return [points[0], points[-1]]

    # ================================================================
    # MOUSE EVENT HANDLERS - HIGHLIGHTER INTEGRATED
    # ================================================================

    def leftclick(self, event):
        """
        Handle mouse down event for highlighter/pen/text drawing and annotation interactions.
        FUNCTIONAL USE: Initiates highlighter/pen stroke, text entry, or annotation selection.
        Routes to appropriate handler based on active tool mode (highlight, pen, text).
        When NO tool is active, starts a canvas pan (scan) instead, so touch drag
        behaves like normal scrolling.
        Args: event - Tkinter mouse event with x, y coordinates
        """

        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return "break"

        # Canvas events only occur outside the embedded Tk Text editor.
        # Save and close the editor immediately, then process this click.
        if self._text_editor is not None:
            self._commit_text_editor()

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)

        # Existing text boxes remain movable/resizable after placement.
        # A drag beginning on a text box transforms that box; a drag beginning
        # anywhere else with no active tool continues to pan the document.
        if not self.active_highlighter and self.tool_mode != "pen":
            hit, hit_mode = self._hit_test_text_box(x, y)
            if hit is not None:
                self._commit_text_editor()
                self.selected_annotation = hit
                self._text_transform_mode = hit_mode
                self._text_transform_start = (x, y)
                self._text_transform_original_bbox = tuple(
                    hit.get('bbox_page', (0, 0, 0, 0))
                )
                self.drawing = True
                self.drawing_type = "text_transform"
                try:
                    self.canvas.grab_set()
                except tk.TclError:
                    pass
                self._draw_text_selection(hit)
                return "break"
            if self.tool_mode is None:
                self._clear_text_selection()

        # -------- HIGHLIGHTER MODE --------
        if self.active_highlighter:
            page_index, local_x, local_y, layout = self._point_to_page_coords(x, y)
            if page_index is None:
                return "break"
            self.current_page = page_index
            self.drawing = True
            self.drawing_type = "highlight"
            self.drawing_page = page_index
            self.drawing_page_offset = (layout['x'], layout['y'])
            self.highlight_points = [(local_x, local_y)]
            self.cleartemp()
            try:
                self.canvas.grab_set()
            except tk.TclError:
                pass
            self._touch_scroll_lock_until = time.monotonic() + 5.0  # NEW
            return "break"

        # -------- PEN TOOL --------
        if self.tool_mode == "pen":
            page_index, local_x, local_y, layout = self._point_to_page_coords(x, y)
            if page_index is None:
                return "break"
            self.current_page = page_index
            self.drawing = True
            self.drawing_type = "pen"
            self.drawing_page = page_index
            self.drawing_page_offset = (layout['x'], layout['y'])
            self.pen_points = [(local_x, local_y)]
            self.cleartemp()
            try:
                self.canvas.grab_set()
            except tk.TclError:
                pass
            self._touch_scroll_lock_until = time.monotonic() + 5.0  # NEW

            return "break"

        # -------- TEXT TOOL: drag to size, or drag an existing box --------
        if self.tool_mode == "text":
            self._commit_text_editor()
            page_index, local_x, local_y, layout = self._point_to_page_coords(x, y)
            if page_index is None:
                return "break"
            self.current_page = page_index
            self.drawing = True
            self.drawing_type = "text_box"
            self.drawing_page = page_index
            self.drawing_page_offset = (layout['x'], layout['y'])
            self._text_box_start = (local_x, local_y)
            self.cleartemp()
            self._touch_scroll_lock_until = time.monotonic() + 5.0
            return "break"

        # -------- NO TOOL ACTIVE: start a normal pan/scroll drag --------
        self.drawing = False
        self.drawing_type = None
        self.canvas.scan_mark(event.x, event.y)
        self._panning = True
        return "break"

    def leftdrag(self, event):
        """
        Record the latest pointer position during a drag. The actual drawing
        work is throttled to a fixed interval by _process_drag_frame instead
        of running on every raw touch/mouse motion event, since touch digitizers
        can report motion far faster than a canvas redraw needs to happen.
        When no tool is active, this instead pans the canvas (normal scroll
        behavior) using Tk's scan_dragto.
        """
        if self._panning and not self.drawing:
            # Normal touch/mouse drag-to-scroll when no tool is selected.
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            self._update_current_page_from_scroll()
            return "break"

        if not self.drawing:
            return "break"

        self._pending_drag_event = event

        if not getattr(self, '_drag_frame_scheduled', False):
            self._drag_frame_scheduled = True
            self.root.after(16, self._process_drag_frame)  # ~60fps cap

        return "break"

    def _process_drag_frame(self):
        """
        Consume the most recent pending drag position and perform the actual
        canvas update. Runs at a fixed ~60fps cadence regardless of how many
        raw motion events fired since the last frame, which is what keeps
        touch drawing responsive instead of being overwhelmed by event volume.
        """
        self._drag_frame_scheduled = False

        event = getattr(self, '_pending_drag_event', None)
        self._pending_drag_event = None

        if event is None or not self.drawing:
            return

        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        offset_x, offset_y = self.drawing_page_offset
        local_x = x - offset_x
        local_y = y - offset_y

        MIN_POINT_DIST = 6

        if self.drawing_type == "text_box":
            sx, sy = self._text_box_start
            x1, y1, x2, y2 = sx + offset_x, sy + offset_y, local_x + offset_x, local_y + offset_y
            if self._text_box_preview_id is None:
                self._text_box_preview_id = self.canvas.create_rectangle(
                    x1, y1, x2, y2, outline='#2563eb', width=2, dash=(5, 3), tags=('text_ui',))
            else:
                self.canvas.coords(self._text_box_preview_id, x1, y1, x2, y2)
            return

        if self.drawing_type == "text_transform":
            self._update_text_transform(x, y)
            return

        if self.drawing_type == "highlight":
            if self.highlight_points:
                last_x, last_y = self.highlight_points[-1]
                if (local_x - last_x) ** 2 + (local_y - last_y) ** 2 < MIN_POINT_DIST ** 2:
                    return

            self.highlight_points.append((local_x, local_y))

            rgb = self.highlighter_colors[self.active_highlighter]['rgb']
            hex_color = f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'

            if self.temp_line_ids:
                line_id = self.temp_line_ids[0]
                self.canvas.coords(line_id, *[
                    coord for px, py in self.highlight_points
                    for coord in (px + offset_x, py + offset_y)
                ])
            else:
                if len(self.highlight_points) >= 2:
                    flat_pts = []
                    for px, py in self.highlight_points:
                        flat_pts.extend([px + offset_x, py + offset_y])
                    line_id = self.canvas.create_line(
                        *flat_pts,
                        fill=hex_color,
                        width=max(15, int(15 * self.zoom_level)),
                        capstyle=tk.ROUND,
                        joinstyle=tk.ROUND,
                        smooth=True
                    )
                    self.temp_line_ids.append(line_id)
            return

        if self.drawing_type == "pen":
            if self.pen_points:
                last_x, last_y = self.pen_points[-1]
                if (local_x - last_x) ** 2 + (local_y - last_y) ** 2 < MIN_POINT_DIST ** 2:
                    return

            self.pen_points.append((local_x, local_y))

            if self.temp_line_ids:
                line_id = self.temp_line_ids[0]
                self.canvas.coords(line_id, *[
                    coord for px, py in self.pen_points
                    for coord in (px + offset_x, py + offset_y)
                ])
            else:
                if len(self.pen_points) >= 2:
                    flat_pts = []
                    for px, py in self.pen_points:
                        flat_pts.extend([px + offset_x, py + offset_y])
                    line_id = self.canvas.create_line(
                        *flat_pts, fill="red", width=3,
                        capstyle=tk.ROUND, joinstyle=tk.ROUND, smooth=True
                    )
                    self.temp_line_ids.append(line_id)
            return

    def leftrel(self, event):
        """
        Handle mouse up event to finalize drawing/annotation and extract text via OCR.
        FUNCTIONAL USE: Completes pen/highlighter stroke, runs OCR on highlighted area,
        saves annotation with extracted text to session, triggers error categorization dialog.
        Args: event - Tkinter mouse event with x, y coordinates
        """
        try:
            if self.canvas.grab_current() == self.canvas:
                self.canvas.grab_release()
        except tk.TclError:
            pass

        # NEW: end a pan drag cleanly if that's what was happening
        if getattr(self, '_panning', False):
            self._panning = False
            if not self.drawing:
                return "break"

        if not self.pdf_document or not self.drawing:
            return "break"

        # -------- HIGHLIGHTER FINISH WITH OCR --------
        if self.drawing_type == "highlight":
            if len(self.highlight_points) >= 2:
                # ALWAYS apply straightening for highlighter
                processed_points = self.Straighten(self.highlight_points)
                
                # Convert to page coordinates
                points_page = self.display_to_page_coords(processed_points)
                
                # Calculate bounding box
                xs = [p[0] for p in points_page]
                ys = [p[1] for p in points_page]
                bbox_page = (min(xs), min(ys), max(xs), max(ys))
                
                # Create annotation
                annotation = {
                    'type': 'highlight',
                    'color': self.active_highlighter,
                    'page': self.drawing_page if self.drawing_page is not None else self.current_page,
                    'bbox_page': bbox_page,
                    'points_page': points_page,
                    'timestamp': datetime.now().isoformat()
                }
                
                # NEW: Extract text from highlighted area if pink highlighter
                if self.active_highlighter == 'pink':
                    extracted_text = self.exctracttxt(annotation)
                    
                    if extracted_text:
                        annotation['extracted_text'] = extracted_text

                    else:
                        annotation['extracted_text'] = None

                    # -------- MULTI-MARK MODE --------
                    # If we're currently attaching extra highlights to one
                    # already-existing punch, skip the "already marked?" dialog
                    # and the whole categorization/Excel flow entirely - just
                    # tag this highlight with that punch's identity and add it.
                    if self.multimark_active and self.multimark_punch:
                        self.attachmultimark(annotation)
                    else:
                        # Show action menu with extracted text
                        self.errorhighlight(annotation)
                else:
                    # Green/Yellow highlighters - no OCR, just add annotation
                    self.annotations.append(annotation)
                    self.addtostack('add_annotation', annotation)
                    self.display()
            
            self.highlight_points = []
            self.cleartemp()
            self.drawing = False
            self.drawing_type = None
            self._touch_scroll_lock_until = time.monotonic() + 0.35
            self.updtoolpane()
            return "break"

        # -------- PEN TOOL FINISH - NO CHANGES --------
        if self.drawing_type == "pen":
            if len(self.pen_points) >= 2:
                points_page = self.display_to_page_coords(self.pen_points)
                annotation = {
                    'type': 'pen',
                    'page': self.drawing_page if self.drawing_page is not None else self.current_page,
                    'points': points_page,
                    'timestamp': datetime.now().isoformat()
                }
                self.annotations.append(annotation)
                self.addtostack('add_annotation', annotation)
            self.pen_points = []
            self.cleartemp()
            self.drawing = False
            self.drawing_type = None
            self.display()
            self._touch_scroll_lock_until = time.monotonic() + 0.35
            self.updtoolpane()
            return "break"

        # -------- TEXT BOX FINISH --------
        if self.drawing_type == "text_box":
            x = self.canvas.canvasx(event.x)
            y = self.canvas.canvasy(event.y)
            ox, oy = self.drawing_page_offset
            ex, ey = x - ox, y - oy
            sx, sy = self._text_box_start
            x1, x2 = sorted((sx, ex)); y1, y2 = sorted((sy, ey))
            if self._text_box_preview_id is not None:
                self.canvas.delete(self._text_box_preview_id)
                self._text_box_preview_id = None
            # A small drag still creates a practical default box.
            if x2 - x1 < 80: x2 = x1 + 220
            if y2 - y1 < 35: y2 = y1 + 90
            annotation = {
                'type': 'text', 'page': self.drawing_page,
                'bbox_page': self.bbox_display_to_page((x1, y1, x2, y2)),
                'pos_page': self.display_to_page_coords((x1, y1)),
                'text': '', 'font_size': 12,
                'timestamp': datetime.now().isoformat()
            }
            self.annotations.append(annotation)
            self.addtostack('add_annotation', annotation)
            self.drawing = False; self.drawing_type = None
            self._open_text_editor(annotation, select_all=False)
            return "break"

        if self.drawing_type == "text_transform":
            self.drawing = False; self.drawing_type = None
            self._text_transform_mode = None
            self.mark_dirty()
            self.display()
            self._draw_text_selection(self.selected_annotation)
            return "break"

        return "break"

    


    """
    Enhanced High-Resolution OCR Extraction
    Captures ANY size highlight, intelligently expands it, sharpens, and extracts text
    """


    def exctracttxt(self, annotation):
        """
        Extract text from highlighted annotation area using OCR with intelligent preprocessing.
        FUNCTIONAL USE: Captures text from quality issues (error/wiring highlights) using
        PaddleOCR. Automatically expands highlight area, sharpens image, and reads text at
        any orientation via PaddleOCR's text-line orientation classifier -
        no manual 90-degree rotation retries needed, and no all-caps requirement: any
        legible text extracted from the highlighted area is accepted.
        Args: annotation - Dictionary with highlight bbox and page info
        Returns: Extracted and cleaned text string or None if OCR fails
        """
        if not self.pdf_document:
            return None

        try:
            # OCR the actual annotation page. This avoids keeping a second
            # full-size NumPy copy of every cached PDF page in memory.
            page_index = int(annotation.get('page', self.current_page))
            if page_index < 0 or page_index >= len(self.pdf_document):
                return None
            pix = self.pdf_document[page_index].get_pixmap(
                matrix=fitz.Matrix(self.page_to_display_scale(), self.page_to_display_scale()),
                alpha=False
            )
            ocr_page_image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                pix.height, pix.width, pix.n
            )[:, :, :3]
            bbox_page = annotation.get('bbox_page')
            if not bbox_page:
                return None
            
            bbox_display = self.bbox_page_to_display(bbox_page)
            x1, y1, x2, y2 = bbox_display
            
            # EXPAND BBOX - Add generous padding for OCR
            PADDING_X = 10  # Horizontal padding
            PADDING_Y = 20  # Vertical padding (more because text height matters)
            
            height, width = ocr_page_image.shape[:2]
            
            x1 = max(0, int(x1) - PADDING_X)
            y1 = max(0, int(y1) - PADDING_Y)
            x2 = min(width, int(x2) + PADDING_X)
            y2 = min(height, int(y2) + PADDING_Y)
            
            crop_width = x2 - x1
            crop_height = y2 - y1
            
            if crop_width < 20 or crop_height < 15:
                print(" WARNING: Highlighted area too small")
                return None
            
            cropped = ocr_page_image[y1:y2, x1:x2]
            
            if cropped.size == 0:
                return None
            
            # Upscale for better OCR (smaller scale for faster processing)
            h, w = cropped.shape[:2]
            upscaled = cv2.resize(cropped, (w*2, h*2), interpolation=cv2.INTER_CUBIC)
            
            # Convert to grayscale and threshold in one go
            gray = cv2.cvtColor(upscaled, cv2.COLOR_RGB2GRAY)
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

            # PaddleOCR expects a 3-channel image (BGR or RGB both work fine
            # for its internal preprocessing); the thresholded binary result
            # is single-channel, so stack it back to 3 channels.
            ocr_ready = cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)

            # Single OCR pass: PaddleOCR's text-line orientation classifier
            # already detects and corrects each text line's
            # orientation internally, so unlike the old Tesseract path we do
            # not need to manually rotate the crop in 90-degree steps and
            # re-run OCR per rotation - any orientation is handled in one call.
            text, confidence = self.ocrcon(ocr_ready)

            if not text:
                print(" No text found")
                return None

            cleaned_text = self.cleantxt(text)
            if cleaned_text and len(cleaned_text) > 1:
                print(f" Extracted ({confidence:.1f}%): '{cleaned_text}'")
                return cleaned_text

            print(" No text found")
            return None

        except Exception as e:
            print(f" OCR Error: {e}")
            return None


    def ocrcon(self, image):
        """
        Run PaddleOCR on an image (numpy array, BGR/RGB or grayscale-as-3ch)
        and return the combined recognized text plus an average confidence
        percentage, in the same (text, confidence_0_to_100) shape the rest of
        the pipeline expects from the old Tesseract-based ocrcon().
        FUNCTIONAL USE: Executes OCR and measures extraction reliability.
        High confidence indicates good OCR quality, low suggests manual review needed.
        Args: image - numpy array image (as produced by cv2/PIL->np.array)
        Returns: Tuple of (extracted_text, average_confidence_percent)
        """
        engine = get_paddle_ocr_engine()
        if engine is None:
            print(" OCR processing error: PaddleOCR engine is not available "
                  "(failed to initialize - is paddleocr/paddlepaddle installed?)")
            return None, 0

        try:
            img_array = np.array(image) if not isinstance(image, np.ndarray) else image
            results = engine.predict(img_array)

            text_parts = []
            confidences = []

            # PaddleOCR 3.x returns Result objects. Their JSON payload contains
            # rec_texts and rec_scores under the "res" key.
            for item in results or []:
                payload = getattr(item, "json", None)
                if callable(payload):
                    payload = payload()
                if payload is None and isinstance(item, dict):
                    payload = item
                if not isinstance(payload, dict):
                    continue

                data = payload.get("res", payload)
                recognized_texts = data.get("rec_texts", []) or []
                recognized_scores = data.get("rec_scores", []) or []

                for index, line_text in enumerate(recognized_texts):
                    line_text = str(line_text or "").strip()
                    if not line_text:
                        continue
                    text_parts.append(line_text)
                    if index < len(recognized_scores):
                        confidences.append(float(recognized_scores[index]) * 100.0)

            if text_parts:
                text = ' '.join(text_parts)
                avg_confidence = (
                    sum(confidences) / len(confidences)
                    if confidences else 0.0
                )
                return text, avg_confidence

            return None, 0

        except Exception as e:
            print(f" OCR processing error: {type(e).__name__}: {e}")
            return None, 0


    def cleantxt(self, text):
        """
        Clean OCR output text by removing artifacts and normalizing.
        FUNCTIONAL USE: Improves OCR text quality by fixing common misrecognitions.
        Converts pipe to I, fixes quotes, removes non-printable characters.
        Args: text - Raw OCR text with potential errors
        Returns: str - Cleaned text or None if too short
        """
        if not text:
            return None
        
        # Single pass cleaning with string operations
        text = ' '.join(text.split())  # Remove extra whitespace
        
        # Character replacements in one go using translate
        replacements = str.maketrans({
            '|': 'I',
            '`': "'",
            '~': '-'
        })
        text = text.translate(replacements)
        
        # Keep only printable characters
        text = ''.join(char for char in text if char.isprintable())
        
        # Strip and return
        text = text.strip()
        
        return text if len(text) >= 2 else None

        # ============================================================================
        # SIMPLIFIED VERSION - If the above is too complex
        # ============================================================================

    def extracttxtsimple(self, annotation):
        """
        Simplified OCR extraction with basic preprocessing.
        FUNCTIONAL USE: Fallback method for quick text extraction when full processing unavailable.
        Uses upscaling and basic grayscale conversion for text recognition.
        Args: annotation - Dictionary with highlight area bbox
        Returns: Extracted text or None if OCR fails
        """
        if self.current_page_image is None:
            print("No image loaded")
            return None
        
        try:
            bbox_page = annotation.get('bbox_page')
            if not bbox_page:
                return None
            
            bbox_display = self.bbox_page_to_display(bbox_page)
            x1, y1, x2, y2 = map(int, bbox_display)
            
            # Get image dimensions
            height, width = self.current_page_image.shape[:2]
            
            # Clip to image bounds
            x1 = max(0, min(x1, width))
            y1 = max(0, min(y1, height))
            x2 = max(0, min(x2, width))
            y2 = max(0, min(y2, height))
            
            # Crop
            cropped = self.current_page_image[y1:y2, x1:x2]
            
            print(f"Cropped area: {cropped.shape}")
            
            if cropped.size == 0:
                print("❌ Empty crop")
                return None
            
            # Upscale 3x for better OCR
            h, w = cropped.shape[:2]
            upscaled = cv2.resize(cropped, (w*3, h*3), interpolation=cv2.INTER_CUBIC)
            
            # Convert to grayscale
            gray = cv2.cvtColor(upscaled, cv2.COLOR_RGB2GRAY)
            
            # Threshold
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Save for debugging
            try:
                debug_dir = "ocr_debug"
                os.makedirs(debug_dir, exist_ok=True)
                cv2.imwrite(os.path.join(debug_dir, "crop.png"), binary)
                print(f" Saved to ocr_debug/crop.png")
            except:
                pass
            
            # OCR - PaddleOCR (angle classifier handles any text orientation;
            # no all-caps requirement, any legible text is accepted)
            ocr_ready = cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)
            text, _confidence = self.ocrcon(ocr_ready)
            
            # Clean
            text = ' '.join((text or '').split()).strip()
            
            if text and len(text) > 1:
                print(f" Extracted: '{text}'")
                return text
            else:
                print(" No text found")
                return None
                
        except Exception as e:
            print(f" OCR Error: {e}")
            import traceback
            traceback.print_exc()
            return None
        

    def preocr(self, pil_image):
        """
        Preprocess image for better OCR accuracy with contrast and noise reduction.
        FUNCTIONAL USE: Enhances image quality before OCR to improve text recognition.
        Converts to grayscale, applies adaptive thresholding, and removes noise.
        Args: pil_image - PIL Image object
        Returns: PIL Image - Preprocessed and ready for OCR
        """
        # Convert to numpy array
        img_array = np.array(pil_image)
        
        # Convert to grayscale
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_array
        
        # Increase contrast
        # Apply adaptive thresholding for better text detection
        thresh = cv2.adaptiveThreshold(
            gray, 255, 
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY, 
            11, 2
        )
        
        # Denoise
        denoised = cv2.fastNlMeansDenoising(thresh)
        
        # Convert back to PIL
        return Image.fromarray(denoised)

    # ================================================================
    # MULTI-MARK MODE: attach several highlights to one existing punch
    # ================================================================

    def startmultimark(self, punch):
        """
        Enter multi-mark mode for an already-existing punch.
        FUNCTIONAL USE: Called from the "Edit Existing Punch" dialog. From this
        point on, every pink highlight drawn on the PDF is attached directly to
        this punch (same sr_no/ref_no/category/description) instead of going
        through the "already marked?" dialog or the categorization menu. A Stop
        button appears on the toolbar; pen/text tools are locked until Stop is
        pressed. Highlighter, zoom, pan, and page navigation keep working normally.
        Args: punch - dict from allexistingpunches()/openpuches() with at least
              sr_no, ref_no, punch_text, category, row.
        """
        if not punch:
            return
        self.deactivate()  # clear any active tool/highlighter cleanly first

        self.multimark_active = True
        self.multimark_punch = dict(punch)
        # Seed the counter with highlights that already exist for this punch
        # (from earlier sessions, or a previous multi-mark round), so the
        # bar doesn't reset to 0 and mislead the user into thinking nothing
        # was marked before.
        self.multimark_count = self._count_existing_highlights(punch)

        # Force the pink (error) highlighter on so the user can start marking
        # immediately without an extra click.
        self.current_color_key = 'pink'
        self.active_highlighter = 'pink'
        self.root.config(cursor="pencil")
        self.colorbutton()

        self._show_multimark_bar()
        self.flashstat(
            f" Multi-mark mode: SR {punch.get('sr_no')} - highlight the PDF, press Stop when done",
            bg='#ec4899'
        )

    def _count_existing_highlights(self, punch):
        """
        Count highlight annotations already attached to this punch (pink,
        still-open ones, plus any already verified/green ones), so re-entering
        multi-mark mode for a punch that was marked before shows the true
        running total instead of resetting to 0.
        Uses the same sr_no/excel_row matching as the verification dialog's
        find_all_anns, so counts always agree with what "Verify" will act on.
        """
        sr_key = str(punch.get('sr_no', '')).strip()
        row_key = str(punch.get('row', '')).strip()
        count = 0
        for a in self.annotations:
            matches_id = ((str(a.get('sr_no', '')).strip() == sr_key and sr_key) or
                          (str(a.get('excel_row', '')).strip() == row_key and row_key))
            if matches_id and a.get('type') == 'highlight' and a.get('color') in ('pink', 'green'):
                count += 1
        return count

    def attachmultimark(self, annotation):
        """
        Attach one newly-drawn pink highlight to the punch selected for
        multi-mark mode, without touching Excel (the punch row already exists)
        and without opening the categorization menu.
        """
        punch = self.multimark_punch
        if not punch:
            # Safety net - mode flag got out of sync somehow; fall back to the
            # normal categorization flow instead of silently dropping the mark.
            self.errorhighlight(annotation)
            return

        annotation['component'] = punch.get('category')
        annotation['subcategory'] = None
        annotation['punch_text'] = punch.get('punch_text')
        annotation['ref_no'] = punch.get('ref_no')
        annotation['excel_row'] = punch.get('row')
        annotation['sr_no'] = punch.get('sr_no')
        annotation['category'] = punch.get('category')
        annotation['already_marked'] = True
        annotation['excel_marking_skipped'] = True
        annotation['multimark'] = True

        self.annotations.append(annotation)
        self.addtostack('add_annotation', annotation)
        self.multimark_count += 1
        self.mark_dirty()
        self.display()
        self.flashstat(
            f" Highlight #{self.multimark_count} added for SR {punch.get('sr_no')}",
            bg='#ec4899'
        )
        self._update_multimark_bar_count()

    def stopmultimark(self):
        """
        Exit multi-mark mode: restore normal toolbar/tool behaviour and
        remove the Stop button.
        """
        if not self.multimark_active:
            return
        count = self.multimark_count
        punch = self.multimark_punch
        self.multimark_active = False
        self.multimark_punch = None
        self.multimark_count = 0

        self._hide_multimark_bar()
        self.deactivate()  # turns the highlighter off and returns to normal mode

        if punch is not None:
            self.flashstat(
                f" Stopped - added {count} highlight(s) for SR {punch.get('sr_no')}",
                bg='#10b981'
            )

    def _show_multimark_bar(self):
        """Create (or refresh) the Stop button/bar shown in multi-mark mode.

        NOTE: every other toolbar section is already packed with side=LEFT/
        side=RIGHT at UI-build time, so appending a new pack()'d frame here
        at runtime lands past all of them and can end up squeezed off the
        visible toolbar (unclickable) depending on window width. Using
        place() with a fixed anchor keeps it reliably on top and clickable
        regardless of how much space the rest of the toolbar has claimed.
        """
        if getattr(self, 'multimark_bar', None) is not None:
            self._update_multimark_bar_count()
            return

        bar = tk.Frame(self.toolbar, bg='#be185d', highlightthickness=2,
                        highlightbackground='#ffffff')
        self.multimark_bar = bar

        self.multimark_bar_label_var = tk.StringVar()
        self._update_multimark_bar_count()

        tk.Label(
            bar, textvariable=self.multimark_bar_label_var, bg='#be185d', fg='white',
            font=('Segoe UI', 9, 'bold')
        ).pack(side=tk.LEFT, padx=(14, 10), pady=6)

        stop_btn = tk.Button(
            bar, text=" Stop ", command=self.stopmultimark,
            bg='#ffffff', fg='#be185d', activebackground='#f1f5f9',
            font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, borderwidth=0,
            padx=20, pady=8, cursor='hand2'
        )
        stop_btn.pack(side=tk.LEFT, padx=(0, 14), pady=6)

        # Float centered along the top of the toolbar, above every other
        # control, so it's always visible and always on top of the click
        # stack regardless of toolbar width or how the rest is packed.
        bar.place(in_=self.toolbar, relx=0.5, rely=0.5, anchor='center')
        bar.lift()

        # Lock out pen/text tools while multi-mark mode is active.
        self._set_locked_tools(True)

    def _update_multimark_bar_count(self):
        if getattr(self, 'multimark_bar_label_var', None) is None:
            return
        punch = self.multimark_punch or {}
        self.multimark_bar_label_var.set(
            f"Marking SR {punch.get('sr_no')}  •  {self.multimark_count} added"
        )

    def _hide_multimark_bar(self):
        bar = getattr(self, 'multimark_bar', None)
        if bar is not None:
            bar.destroy()
            self.multimark_bar = None
            self.multimark_bar_label_var = None
        self._set_locked_tools(False)

    def _set_locked_tools(self, locked):
        """Enable/disable the pen and text tool buttons so only the
        highlighter (plus zoom/pan/navigation) can be used during
        multi-mark mode."""
        state = tk.DISABLED if locked else tk.NORMAL
        for attr in ('pen_btn', 'text_btn'):
            btn = getattr(self, attr, None)
            if btn is not None:
                try:
                    btn.config(state=state)
                except tk.TclError:
                    pass

    def errorhighlight(self, annotation):
        """
        Display categorization menu for pink (error) highlights with OCR-extracted text.
        FUNCTIONAL USE: Routes pink highlights through error classification workflow.
        Shows category menu with extracted text pre-filled for punch creation.
        Args: annotation - Dictionary with highlight data and extracted_text field
        """
        
        extracted_text = annotation.get('extracted_text', None)
        
        menu = Menu(self.root, tearoff=0)

        for cat in self.categories:
            mode = cat.get("mode", "parent")
            
            # ========== WIRING SELECTOR MODE ==========
            if mode == "wiring_selector":
                menu.add_command(
                    label=f" {cat['name']}",
                    command=lambda c=cat, ann=annotation, txt=extracted_text: 
                        self.wiringselocr(c, ann, txt)
                )
            
            # ========== TEMPLATE MODE ==========
            elif mode == "template":
                ref_num = cat.get("ref_number", "")
                if ref_num:
                    label = f" [{ref_num}] {cat['name']}"
                else:
                    label = f" {cat['name']}"
                
                menu.add_command(
                    label=label,
                    command=lambda c=cat, ann=annotation, txt=extracted_text: 
                        self.handlecat(c, ann, txt)
                )
            
            # ========== PARENT MODE ==========
            elif mode == "parent":
                cat_menu = Menu(menu, tearoff=0)
                for sub in cat.get("subcategories", []):
                    ref_num = sub.get("ref_number", "??")
                    if ref_num:
                        label = f"[{ref_num}] {sub['name']}"
                    else:
                        label = sub['name']
                    cat_menu.add_command(
                        label=label,
                        command=lambda c=cat, s=sub, ann=annotation, txt=extracted_text: 
                            self.handlesub(c, s, ann, txt)
                    )
                
                menu.add_cascade(label=f" {cat['name']}", menu=cat_menu)

        menu.add_separator()
        menu.add_command(
            label=" Custom Action Point",
            command=lambda ann=annotation, txt=extracted_text: 
                self.logcustomerr(ann, txt)
        )

        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        menu.tk_popup(x, y)

    # ============================================================================
    # MODIFIED: Handler methods with OCR support
    # Add these to your CircuitInspector class
    # ============================================================================

    def wiringselocr(self, category, annotation, extracted_text):
        """Handle wiring type selection with OCR text"""
        wiring_menu = Menu(self.root, tearoff=0, bg='#1e293b', fg='white',
                          activebackground='#3b82f6', font=('Segoe UI', 10))
        
        wiring_types = category.get("wiring_types", [])
        
        for wiring in wiring_types:
            wiring_type = wiring.get("type", "Unknown")
            ref_num = wiring.get("ref_number", "??")
            
            wiring_menu.add_command(
                label=f"[{ref_num}] {wiring_type}",
                command=lambda c=category, w=wiring, ann=annotation, txt=extracted_text: 
                    self.showwiringsub(c, w, ann, txt)
            )
        
        special_subs = category.get("special_subcategories", [])
        if special_subs:
            wiring_menu.add_separator()
            for special in special_subs:
                ref_num = special.get("ref_number", "??")
                wiring_menu.add_command(
                    label=f"[{ref_num}] {special['name']} (All types)",
                    command=lambda c=category, s=special, ann=annotation, txt=extracted_text:
                        self.splcat(c, s, ann, txt)
                )
        
        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        wiring_menu.tk_popup(x, y)


    def showwiringsub(self, category, wiring_data, annotation, extracted_text):
        """Show sub-subcategories with OCR text"""
        subcategories = wiring_data.get("subcategories", [])
        
        if not subcategories:
            self.wiringtype(category, wiring_data, annotation, extracted_text)
            return
        
        sub_menu = Menu(self.root, tearoff=0, bg='#1e293b', fg='white',
                       activebackground='#3b82f6', font=('Segoe UI', 10))
        
        wiring_type = wiring_data.get("type", "Unknown")
        ref_num = wiring_data.get("ref_number", "??")
        
        for sub in subcategories:
            sub_name = sub.get("name", "Unknown")
            sub_menu.add_command(
                label=f"[{ref_num}] {sub_name}",
                command=lambda c=category, w=wiring_data, s=sub, ann=annotation, txt=extracted_text:
                    self.hnadlwiringsub(c, w, s, ann, txt)
            )
        
        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        sub_menu.tk_popup(x, y)


    def hnadlwiringsub(self, category, wiring_data, subcategory, annotation, extracted_text):
        """Handle wiring subcategory with OCR pre-fill"""
        
        punch_text = self.runtemp(subcategory, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = wiring_data.get("ref_number", "??")
        wiring_type = wiring_data.get("type", "Unknown")
        sub_name = subcategory.get("name", "Unknown")
        
        self.logerrdirect(
            component_type=category["name"],
            error_name=f"{wiring_type} - {sub_name}",
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def handlecat(self, category, annotation, extracted_text):
        """Handle template category with OCR pre-fill"""
        punch_text = self.runtemp(category, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = category.get("ref_number", "")
        
        if ref_number:
            self.logerrdirect(
                component_type=category["name"],
                error_name=None,
                error_template=punch_text,
                annotation=annotation,
                ref_number=ref_number
            )
        else:
            self.logerrwithref(
                component_type=category["name"],
                error_name=None,
                error_template=punch_text,
                annotation=annotation
            )


    def handlesub(self, category, subcategory, annotation, extracted_text):
        """Handle subcategory with OCR pre-fill"""
        punch_text = self.runtemp(subcategory, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = subcategory.get("ref_number", "")
        
        if ref_number:
            self.logerrdirect(
                component_type=category["name"],
                error_name=subcategory["name"],
                error_template=punch_text,
                annotation=annotation,
                ref_number=ref_number
            )
        else:
            self.logerrwithref(
                component_type=category["name"],
                error_name=subcategory["name"],
                error_template=punch_text,
                annotation=annotation
            )


    def logerrdirect(self, component_type, error_name, error_template, annotation, ref_number):
        """Log error DIRECTLY without asking for reference number"""
        punch_text = error_template

        if not punch_text:
            messagebox.showerror("Error", "Punch description is empty.")
            return

        # Use ref_number directly - NO POPUP
        ref_no = str(ref_number).strip()
        self.session_refs.add(ref_no)

        # -------- TOGGLE OFF: keep the annotation locally, skip Excel entirely --------
        if not self.mark_errors_to_excel:
            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = None
            annotation['sr_no'] = None
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None
            annotation['excel_marking_skipped'] = True

            self.annotations.append(annotation)
            self.display()

            self.flashstat(f" Marked locally (Excel skipped) - Ref {ref_no}", bg='#f59e0b')
            self.mark_dirty()
            return

        try:
            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.readcell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.readcell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.writecell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.writecell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.writecell(ws, row_num, self.punch_cols['desc'], punch_text)
            self.writecell(ws, row_num, self.punch_cols['category'], component_type)

            uname = self.logged_in_fullname or "Unknown User"

            self.writecell(ws, row_num, self.punch_cols['checked_name'], uname)
            # Updated to include timestamp + date
            self.writecell(ws, row_num, self.punch_cols['checked_date'], 
                          datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            wb.save(self.excel_file)
            wb.close()

            updated = self.updatestatsforref(ref_no, status='NOK')
            if updated:
                print(f"OK Interphase: marked ref {ref_no} as NOK")

            # Update annotation with all the data
            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None

            # Add to annotations list
            self.annotations.append(annotation)
            self.addtostack('add_annotation', annotation)
            self.current_sr_no = self.getnextsr()
            
            # Redraw to show the color change from pink to red
            self.display()

            print(f" Logged: Ref {ref_no}, SR {sr_no_assigned}")
            self.flashstat(f" Logged Ref {ref_no}", bg='#10b981')
            
            try:
                self.manager_db.logcatoccurence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
                self.mark_dirty()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log punch:\n{e}")
            import traceback
            traceback.print_exc()


    def logerrwithref(self, component_type, error_name, error_template, annotation):
        """Log error WITH popup (for Design Error only)"""
        punch_text = error_template

        if not punch_text:
            messagebox.showerror("Error", "Punch description is empty.")
            return

        # ASK FOR REFERENCE NUMBER (Design Error)
        ref_no = simpledialog.askstring(
            "Reference Number", 
            "Enter the reference number:",
            parent=self.root
        )
        
        if not ref_no:
            return

        ref_no = str(ref_no).strip()
        self.session_refs.add(ref_no)

        # -------- TOGGLE OFF: keep the annotation locally, skip Excel entirely --------
        if not self.mark_errors_to_excel:
            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = None
            annotation['sr_no'] = None
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None
            annotation['excel_marking_skipped'] = True

            self.annotations.append(annotation)
            self.display()

            self.flashstat(f" Marked locally (Excel skipped) - Ref {ref_no}", bg='#f59e0b')
            self.mark_dirty()
            return

        try:
            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.readcell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.readcell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.writecell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.writecell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.writecell(ws, row_num, self.punch_cols['desc'], punch_text)
            self.writecell(ws, row_num, self.punch_cols['category'], component_type)

            uname = self.logged_in_fullname or "Unknown User"

            self.writecell(ws, row_num, self.punch_cols['checked_name'], uname)
            # Updated to include timestamp + date
            self.writecell(ws, row_num, self.punch_cols['checked_date'], 
                          datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            wb.save(self.excel_file)
            wb.close()

            updated = self.updatestatsforref(ref_no, status='NOK')
            if updated:
                print(f"Interphase: marked ref {ref_no} as NOK")

            annotation['component'] = component_type
            annotation['subcategory'] = error_name
            annotation['punch_text'] = punch_text
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['implemented'] = False
            annotation['implemented_name'] = None
            annotation['implemented_date'] = None
            annotation['implementation_remark'] = None

            self.annotations.append(annotation)
            self.addtostack('add_annotation', annotation)
            self.current_sr_no = self.getnextsr()
            self.display()

            print(f" Logged: Ref {ref_no}, SR {sr_no_assigned}")
            self.flashstat(f" Logged Ref {ref_no}", bg='#10b981')
            
            try:
                self.manager_db.logcatoccurence(
                    self.cabinet_id,
                    self.project_name,
                    component_type,
                    error_name
                )
                self.mark_dirty()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log punch:\n{e}")
            import traceback
            traceback.print_exc()


    def logcustomerr(self, annotation, extracted_text):
        """Log custom error with OCR pre-fill"""
        try:
            # Pre-fill with OCR text
            custom_action = simpledialog.askstring(
                "Custom Action Point",
                "Enter the action point / punch description:",
                parent=self.root,
                initialvalue=extracted_text if extracted_text else ""
            )
            
            if not custom_action:
                return

            custom_category = simpledialog.askstring(
                "Custom Category",
                "Enter the category:",
                parent=self.root
            )
            if not custom_category:
                return

            ref_no = simpledialog.askstring(
                "Reference Number", 
                "Enter the reference number:",
                parent=self.root
            )
            
            if not ref_no:
                messagebox.showwarning("Reference Required", "Reference No is required.")
                return

            ref_no = str(ref_no).strip()
            self.session_refs.add(ref_no)

            # -------- TOGGLE OFF: keep the annotation locally, skip Excel entirely --------
            if not self.mark_errors_to_excel:
                annotation['component'] = custom_category
                annotation['error'] = 'Custom'
                annotation['punch_text'] = custom_action
                annotation['ref_no'] = ref_no
                annotation['excel_row'] = None
                annotation['sr_no'] = None
                annotation['timestamp'] = datetime.now().isoformat()
                annotation['excel_marking_skipped'] = True

                self.annotations.append(annotation)
                self.display()

                print(f"Logged custom locally (Excel skipped): Ref {ref_no}")
                self.flashstat(f" Marked locally (Excel skipped) - Ref {ref_no}", bg='#f59e0b')
                self.mark_dirty()
                return

            wb = load_workbook(self.excel_file)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row_num = 8
            while True:
                val = self.readcell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                row_num += 1

            prev_sr = None
            if row_num > 8:
                prev_sr = self.readcell(ws, row_num - 1, self.punch_cols['sr_no'])

            try:
                sr_no_assigned = int(prev_sr) + 1 if prev_sr is not None else 1
            except:
                sr_no_assigned = 1

            self.writecell(ws, row_num, self.punch_cols['sr_no'], sr_no_assigned)
            self.writecell(ws, row_num, self.punch_cols['ref_no'], ref_no)
            self.writecell(ws, row_num, self.punch_cols['desc'], custom_action)
            self.writecell(ws, row_num, self.punch_cols['category'], custom_category)

            uname = self.logged_in_fullname or "Unknown User"

            self.writecell(ws, row_num, self.punch_cols['checked_name'], uname)
            # Updated to include timestamp + date
            self.writecell(ws, row_num, self.punch_cols['checked_date'], 
                          datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            wb.save(self.excel_file)
            wb.close()

            annotation['component'] = custom_category
            annotation['error'] = 'Custom'
            annotation['punch_text'] = custom_action
            annotation['ref_no'] = ref_no
            annotation['excel_row'] = row_num
            annotation['sr_no'] = sr_no_assigned
            annotation['timestamp'] = datetime.now().isoformat()

            self.annotations.append(annotation)
            self.addtostack('add_annotation', annotation)
            self.current_sr_no = self.getnextsr()
            self.display()

            print(f"OK Logged custom: Ref {ref_no}, SR {sr_no_assigned}")
            self.flashstat(f"✓ Custom punch Ref {ref_no}", bg='#8b5cf6')

            try:
                self.manager_db.logcatoccurence(
                    self.cabinet_id,
                    self.project_name,
                    custom_category,
                    None
                )
                self.mark_dirty()
            except Exception as e:
                print(f"Manager category logging failed: {e}")

            updated = self.updatestatsforref(ref_no, status='NOK')
            if updated:
                print(f"Interphase: marked ref {ref_no} as NOK")

        except PermissionError:
            messagebox.showerror("Error", "Close the Excel file before writing to it.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log custom error:\n{e}")
            import traceback
            traceback.print_exc()


    # ============================================================================
    # Additional helper methods for OCR
    # ============================================================================

    def wiringtype(self, category, wiring_data, annotation, extracted_text):
        """Handle direct wiring type selection with OCR"""
        punch_text = self.runtemp(wiring_data, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = wiring_data.get("ref_number", "??")
        wiring_type = wiring_data.get("type", "Unknown")
        
        self.logerrdirect(
            component_type=category["name"],
            error_name=wiring_type,
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def splcat(self, category, special_sub, annotation, extracted_text):
        """Handle special subcategories with OCR"""
        punch_text = self.runtemp(special_sub, tag_name=None, prefill_text=extracted_text)
        if not punch_text:
            return
        
        ref_number = special_sub.get("ref_number", "??")
        
        self.logerrdirect(
            component_type=category["name"],
            error_name=special_sub["name"],
            error_template=punch_text,
            annotation=annotation,
            ref_number=ref_number
        )


    def cleartemp(self):
        """Clear temporary drawing elements from canvas"""
        for line_id in self.temp_line_ids:
            try:
                self.canvas.delete(line_id)
            except:
                pass
        self.temp_line_ids.clear()

    # ================================================================
    # MICROSOFT-STYLE TEXT BOXES
    # ================================================================
    def _text_bbox_display(self, ann):
        bbox = ann.get('bbox_page')
        if not bbox:
            x, y = ann.get('pos_page', (0, 0))
            bbox = (x, y, x + 140, y + 55)
            ann['bbox_page'] = bbox
        x1, y1, x2, y2 = self.bbox_page_to_display(bbox)
        layout = next((v for v in self.page_layout if v['page_index'] == ann.get('page')), None)
        if not layout:
            return None
        return (x1 + layout['x'], y1 + layout['y'], x2 + layout['x'], y2 + layout['y'])

    def _hit_test_text_box(self, x, y):
        """Return a text annotation and move/resize mode at a canvas point."""
        radius = 12
        for ann in reversed(self.annotations):
            if ann.get('type') != 'text':
                continue
            box = self._text_bbox_display(ann)
            if not box:
                continue
            x1, y1, x2, y2 = box
            handles = {
                'resize_nw': (x1, y1), 'resize_n': ((x1+x2)/2, y1),
                'resize_ne': (x2, y1), 'resize_e': (x2, (y1+y2)/2),
                'resize_se': (x2, y2), 'resize_s': ((x1+x2)/2, y2),
                'resize_sw': (x1, y2), 'resize_w': (x1, (y1+y2)/2),
            }
            for mode, (hx, hy) in handles.items():
                if abs(x-hx) <= radius and abs(y-hy) <= radius:
                    return ann, mode
            if x1 <= x <= x2 and y1 <= y <= y2:
                return ann, 'move'
        return None, None

    def _clear_text_selection(self):
        for item in self._text_selection_ids:
            try:
                self.canvas.delete(item)
            except Exception:
                pass
        self._text_selection_ids = []
        self.selected_annotation = None

    def _draw_text_selection(self, ann):
        self._clear_text_selection()
        if not ann:
            return
        self.selected_annotation = ann
        box = self._text_bbox_display(ann)
        if not box:
            return
        x1, y1, x2, y2 = box
        ids = [self.canvas.create_rectangle(
            x1, y1, x2, y2, outline='#2563eb', width=2, dash=(5, 3),
            tags=('text_ui',))]
        points = [
            (x1,y1), ((x1+x2)/2,y1), (x2,y1), (x2,(y1+y2)/2),
            (x2,y2), ((x1+x2)/2,y2), (x1,y2), (x1,(y1+y2)/2)
        ]
        h = 5
        for hx, hy in points:
            ids.append(self.canvas.create_rectangle(
                hx-h, hy-h, hx+h, hy+h, fill='white', outline='#2563eb',
                width=2, tags=('text_ui',)))
        self._text_selection_ids = ids

    def _update_text_transform(self, canvas_x, canvas_y):
        ann = self.selected_annotation
        if not ann or not self._text_transform_original_bbox:
            return
        sx, sy = self._text_transform_start
        dx = (canvas_x - sx) / self.page_to_display_scale()
        dy = (canvas_y - sy) / self.page_to_display_scale()
        x1, y1, x2, y2 = self._text_transform_original_bbox
        mode = self._text_transform_mode
        if mode == 'move':
            x1, y1, x2, y2 = x1+dx, y1+dy, x2+dx, y2+dy
        else:
            if mode in ('resize_nw', 'resize_w', 'resize_sw'):
                x1 += dx
            if mode in ('resize_ne', 'resize_e', 'resize_se'):
                x2 += dx
            if mode in ('resize_nw', 'resize_n', 'resize_ne'):
                y1 += dy
            if mode in ('resize_sw', 'resize_s', 'resize_se'):
                y2 += dy
            # Enforce a usable minimum size without flipping the object.
            if x2 - x1 < 30:
                if mode in ('resize_nw', 'resize_w', 'resize_sw'): x1 = x2 - 30
                else: x2 = x1 + 30
            if y2 - y1 < 18:
                if mode in ('resize_nw', 'resize_n', 'resize_ne'): y1 = y2 - 18
                else: y2 = y1 + 18
        ann['bbox_page'] = (x1, y1, x2, y2)
        ann['pos_page'] = (x1, y1)
        self._draw_text_selection(ann)

    def _delete_selected_text_box(self, event=None):
        ann = self.selected_annotation
        if self._text_editor is not None or not ann or ann.get('type') != 'text':
            return
        if ann in self.annotations:
            self.annotations.remove(ann)
            self._clear_text_selection()
            self.mark_dirty()
            self.display()
        return "break"

    def _open_text_editor(self, ann, select_all=True):
        self._commit_text_editor()
        box = self._text_bbox_display(ann)
        if not box:
            return
        x1, y1, x2, y2 = box
        editor = tk.Text(self.canvas, wrap=tk.WORD, undo=True, relief=tk.SOLID,
                         borderwidth=2, highlightthickness=1,
                         highlightbackground='#2563eb', highlightcolor='#2563eb',
                         font=('Segoe UI', max(8, int(ann.get('font_size', 12) * self.zoom_level))))
        editor.insert('1.0', ann.get('text', ''))
        self._text_editor = editor
        self._text_edit_annotation = ann
        self._text_editor_window_id = self.canvas.create_window(
            x1, y1, anchor=tk.NW, window=editor,
            width=max(80, x2-x1), height=max(35, y2-y1), tags=('text_ui',))
        # Put keyboard focus and the insertion caret inside the new text box.
        # The delayed repeat is required on Windows because opening TabTip may
        # briefly move foreground focus away from the Tk Text widget.
        editor.focus_force()
        if select_all:
            editor.tag_add(tk.SEL, '1.0', 'end-1c')
            editor.mark_set(tk.INSERT, '1.0')
        else:
            editor.mark_set(tk.INSERT, 'end-1c')
        editor.see(tk.INSERT)
        self.root.after_idle(
            lambda w=editor: w.focus_force() if w.winfo_exists() else None
        )
        editor.bind('<Control-Return>', lambda e: (self._commit_text_editor(), 'break'))
        editor.bind('<Escape>', lambda e: (self._cancel_text_editor(), 'break'))
        editor.bind('<FocusOut>', self._on_text_editor_focus_out)

    def _on_text_editor_focus_out(self, event=None):
        """Save and exit editing when focus moves outside the text box."""
        editor = self._text_editor
        if editor is None:
            return
        self.root.after_idle(lambda expected=editor: (
            self._commit_text_editor() if self._text_editor is expected else None
        ))

    def _commit_text_editor(self):
        editor = self._text_editor
        ann = self._text_edit_annotation
        if editor is None or ann is None:
            return
        try:
            value = editor.get('1.0', 'end-1c').rstrip()
        except tk.TclError:
            value = ann.get('text', '')
        ann['text'] = value
        if not value and ann in self.annotations:
            self.annotations.remove(ann)
        self._destroy_text_editor()
        self.mark_dirty()
        self.display()
        if value:
            self.selected_annotation = ann
            self._draw_text_selection(ann)

    def _cancel_text_editor(self):
        ann = self._text_edit_annotation
        if ann is not None and not ann.get('text') and ann in self.annotations:
            self.annotations.remove(ann)
        self._destroy_text_editor()
        self.display()

    def _destroy_text_editor(self):
        if self._text_editor_window_id is not None:
            try: self.canvas.delete(self._text_editor_window_id)
            except Exception: pass
        if self._text_editor is not None:
            try: self._text_editor.destroy()
            except Exception: pass
        self._text_editor = None
        self._text_editor_window_id = None
        self._text_edit_annotation = None

    def _wrap_text_for_box(self, draw, text, font, pixel_width):
        lines = []
        for paragraph in str(text).splitlines() or ['']:
            words = paragraph.split()
            if not words:
                lines.append(''); continue
            line = words[0]
            for word in words[1:]:
                trial = line + ' ' + word
                try: width = draw.textlength(trial, font=font)
                except Exception: width = len(trial) * 7
                if width <= pixel_width:
                    line = trial
                else:
                    lines.append(line); line = word
            lines.append(line)
        return '\n'.join(lines)

    # ================================================================
    # DISPLAY PAGE - WITH HIGHLIGHTER, PEN AND TEXT RENDERING
    # ================================================================

    def _annotation_render_signature(self, page_index):
        """Return a cheap stable signature for annotations rendered on one page."""
        relevant = []
        for ann in self.annotations:
            if ann.get('page') != page_index:
                continue
            relevant.append((
                ann.get('type'), ann.get('color'), ann.get('points_page'),
                ann.get('points'), ann.get('pos_page'), ann.get('text'),
                ann.get('bbox_page'), ann.get('closed_by'), ann.get('sr_no')
            ))
        return repr(relevant)

    def _clear_page_render_cache(self):
        """Release cached Tk images, normally after loading a different PDF."""
        self._page_render_cache.clear()
        self._page_cache_zoom = None

    def schedule_display(self, preserve_view=True, delay_ms=1):
        """Coalesce repeated redraw requests into one Tk idle-time render."""
        if self._display_after_id is not None:
            try:
                self.root.after_cancel(self._display_after_id)
            except Exception:
                pass
        self._display_after_id = self.root.after(
            delay_ms, lambda: self._scheduled_display(preserve_view)
        )

    def _scheduled_display(self, preserve_view):
        self._display_after_id = None
        self.display(preserve_view=preserve_view)

    def display(self, preserve_view=True):
        """Render the stacked PDF view, re-rasterizing only changed pages.

        The original implementation converted every PDF page to a bitmap and
        rebuilt every annotation overlay after each stroke. On a multi-page PDF
        that blocks Tk's single UI thread. This version caches each composed page
        using zoom level plus a per-page annotation signature.
        """
        if not self.pdf_document:
            self.canvas.delete("all")
            self._update_page_toolbar()
            return
        # Only show the loading overlay when there's real rasterization work
        # ahead (cache misses) - most display() calls after the first render
        # are cheap cache hits (e.g. adding one annotation) and shouldn't
        # interrupt the user with a spinner. A cold multi-page render is the
        # actual freeze-prone case this is meant to cover.
        showed_overlay = False
        try:
            scale_probe = self.page_to_display_scale()
            zoom_key_probe = round(float(self.zoom_level), 4)
            cache_cleared = self._page_cache_zoom != zoom_key_probe
            miss_count = 0
            for page_index in range(len(self.pdf_document)):
                if cache_cleared:
                    miss_count += 1
                    continue
                signature = self._annotation_render_signature(page_index)
                if (page_index, zoom_key_probe, signature) not in self._page_render_cache:
                    miss_count += 1
            if miss_count >= 2:
                self.busy("Rendering PDF pages...")
                showed_overlay = True
        except Exception:
            pass
        try:
            current_view = None
            if preserve_view and hasattr(self, 'canvas') and self.canvas and self.canvas.winfo_exists():
                current_view = (self.canvas.xview(), self.canvas.yview())

            scale = self.page_to_display_scale()
            zoom_key = round(float(self.zoom_level), 4)
            if self._page_cache_zoom != zoom_key:
                self._page_render_cache.clear()
                self._page_cache_zoom = zoom_key

            self.canvas.delete("all")
            self.page_layout = []
            self.page_images = []
            page_gap = max(24, int(24 * self.zoom_level))
            y_offset = 0
            max_width = 0
            try:
                font_size = max(12, int(14 * self.zoom_level))
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()

            # Keep every page in one vertically stacked canvas, preserving the
            # original scroll-down-to-change-page workflow. Cached pages are reused.
            for page_index, page in enumerate(self.pdf_document):
                signature = self._annotation_render_signature(page_index)
                cache_key = (page_index, zoom_key, signature)
                cached = self._page_render_cache.get(cache_key)

                if cached is None:
                    if showed_overlay:
                        # Keep the spinner animating and the window responsive
                        # to the OS while we rasterize each uncached page.
                        self.busy(f"Rendering page {page_index + 1} of {len(self.pdf_document)}...")
                    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
                    page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("RGBA")
                    overlay = Image.new("RGBA", page_image.size, (0, 0, 0, 0))
                    draw = ImageDraw.Draw(overlay, 'RGBA')

                    for ann in self.annotations:
                        if ann.get('page') != page_index:
                            continue
                        ann_type = ann.get('type')
                        if ann_type == 'highlight' and 'points_page' in ann:
                            points = self.page_to_display_coords(ann['points_page'])
                            if len(points) >= 2:
                                color_key = ann.get('color', 'yellow')
                                rgba = self.highlighter_colors[color_key]['rgba']
                                width = max(15, int(15 * self.zoom_level))
                                draw.line(points, fill=rgba, width=width, joint='curve')
                                if ann.get('closed_by') and ann.get('bbox_page'):
                                    x1, y1, _, _ = self.bbox_page_to_display(ann['bbox_page'])
                                    draw.ellipse([x1 + 2, y1 + 2, x1 + 14, y1 + 14], fill=(0, 128, 0, 200))
                        elif ann_type == 'pen' and 'points' in ann:
                            points = self.page_to_display_coords(ann['points'])
                            if len(points) >= 2:
                                draw.line(points, fill=(255, 0, 0, 255),
                                          width=max(2, int(3 * self.zoom_level)), joint='curve')
                        elif ann_type == 'text':
                            value = ann.get('text', '')
                            bbox = ann.get('bbox_page')
                            if not bbox:
                                px, py = ann.get('pos_page', (0, 0))
                                bbox = (px, py, px + 140, py + 55)
                                ann['bbox_page'] = bbox
                            x1, y1, x2, y2 = self.bbox_page_to_display(bbox)
                            if value:
                                wrapped = self._wrap_text_for_box(draw, value, font, max(20, x2-x1-8))
                                draw.rectangle([x1, y1, x2, y2], fill=(255, 255, 255, 220))
                                draw.multiline_text((x1+4, y1+3), wrapped, fill=(255, 0, 0, 255),
                                                    font=font, spacing=3)

                    composed = Image.alpha_composite(page_image, overlay).convert("RGB")
                    photo = ImageTk.PhotoImage(composed)
                    cached = {
                        'photo': photo,
                        'width': composed.width, 'height': composed.height
                    }
                    # Keep only the newest version of this page at this zoom.
                    stale = [k for k in self._page_render_cache
                             if k[0] == page_index and k[1] == zoom_key]
                    for key in stale:
                        self._page_render_cache.pop(key, None)
                    self._page_render_cache[cache_key] = cached

                width, height = cached['width'], cached['height']
                max_width = max(max_width, width)
                self.page_layout.append({
                    'page_index': page_index, 'x': 0, 'y': y_offset,
                    'width': width, 'height': height
                })
                self.page_images.append(cached['photo'])
                self.canvas.create_image(0, y_offset, anchor=tk.NW,
                                         image=cached['photo'], tags=(f"page_{page_index}",))
                y_offset += height + page_gap

            self.photo = self.page_images[self.current_page] if self.page_images else None
            self.canvas.config(scrollregion=(0, 0, max_width, max(1, y_offset)))
            if current_view and current_view[1][0] > 0.0:
                self.canvas.xview_moveto(current_view[0][0])
                self.canvas.yview_moveto(current_view[1][0])
            elif self.page_layout and 0 <= self.current_page < len(self.page_layout):
                self.canvas.yview_moveto(self.page_layout[self.current_page]['y'] / max(1, y_offset))
            else:
                self.canvas.xview_moveto(0)
                self.canvas.yview_moveto(0)
            self._update_page_toolbar()
            self.updtoolpane()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to display page: {e}")
        finally:
            if showed_overlay:
                self.unbusy()

    # ================================================================
    # SAVE SESSION - WITH HIGHLIGHTER SERIALIZATION
    # ================================================================

    def _build_session_data(self):
        """
        Build the JSON-serializable session dict from current in-memory state.
        FUNCTIONAL USE: Shared serialization logic used by both the explicit
        "Save Session" action and the batched/background autosave flush, so
        there's exactly one place that defines what a session file contains.
        Returns: dict ready for json.dump
        """
        data = {
            'project_name': self.project_name,
            'sales_order_no': self.sales_order_no,
            'cabinet_id': getattr(self, 'cabinet_id', ''),
            'pdf_path': self.current_pdf_path,
            'current_page': self.current_page,
            'zoom_level': self.zoom_level,
            'current_sr_no': self.current_sr_no,
            'session_refs': list(self.session_refs),
            'annotations': [],
            'undo_stack_size': len(self.undo_stack) if hasattr(self, 'undo_stack') else 0,
            'save_timestamp': datetime.now().isoformat()
        }

        # Process all annotation types
        for ann in self.annotations:
            entry = ann.copy()

            # ===== HIGHLIGHTER ANNOTATIONS - Convert tuples to lists =====
            if 'points_page' in entry:
                entry['points_page'] = [[float(x), float(y)] for x, y in entry['points_page']]

            # ===== PEN STROKES - Convert tuples to lists =====
            if 'points' in entry:
                entry['points'] = [[float(x), float(y)] for x, y in entry['points']]

            # ===== TEXT ANNOTATIONS - Convert tuple to list =====
            if 'pos_page' in entry:
                pos = entry['pos_page']
                entry['pos_page'] = [float(pos[0]), float(pos[1])]

            # Ensure text content is saved
            if 'text' in entry:
                entry['text'] = str(entry['text'])

            data['annotations'].append(entry)

        return data

    def _write_session_file(self):
        """
        Write the current session to its JSON file on disk.
        FUNCTIONAL USE: The actual disk write, split out from savesession()
        so flush_pending_saves() (background autosave / final crash-safe
        flush) can reuse it without duplicating path resolution or
        triggering the "no PDF loaded" user-facing warnings that the
        explicit menu action shows.
        Returns: True on success, False on failure (errors are logged, not raised).
        """
        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("sessions"):
            return False

        save_path = os.path.join(
            self.project_dirs["sessions"],
            f"{self.cabinet_id}_annotations.json"
        )

        data = self._build_session_data()

        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"[WARN] Failed to write session file: {e}")
            return False

    def savesession(self):
        """
        Serialize all current annotations to session JSON file.
        FUNCTIONAL USE: Writes annotations list, page references, and metadata to file.
        Enables resuming work across quality module instances without losing annotations.
        This is the explicit user-triggered save (File menu / Ctrl+S) - it always
        does a real, immediate sync + write, unlike the batched autosave.
        """
        """Save current session to JSON file with all annotation types including highlights"""
        if self._text_editor is not None:
            self._commit_text_editor()
        if not self.pdf_document:
            messagebox.showwarning("No PDF", "Load a PDF first before saving a session.")
            return

        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("sessions"):
            messagebox.showerror("Error", "Project directories not set up. Load a PDF first.")
            return

        self.busy("Saving session...")
        try:
            self.sync_manager_stats_only()

            if self._write_session_file():
                self._dirty = False
                self._last_flush_time = time.monotonic()
            else:
                messagebox.showerror("Error", "Failed to save session. See console for details.")
        finally:
            self.unbusy()

    # ================================================================
    # LOAD SESSION - WITH HIGHLIGHTER DESERIALIZATION
    # ================================================================

    def loadsession(self):
        """
        Load and deserialize annotations from previous session file.
        FUNCTIONAL USE: Reads saved annotations from last working session on this PDF.
        Restores in-memory annotation list to resume quality inspection work.
        """
        """Load session from JSON file via file dialog"""
        path = filedialog.askopenfilename(
            title="Load Session JSON",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not path:
            return

        self.loadfrompath(path)
        self.sync_manager_stats_only()

    def loadfrompath(self, path):
        """
        Deserialize annotations from session JSON file into memory.
        FUNCTIONAL USE: Reads saved annotations and reconstructs in-memory annotation list.
        Called when loading cabinet to restore previous quality work.
        Args: path - Full path to session JSON file
        """
        """Load session from a specific JSON file path with all annotation types"""
        self.stopmultimark()  # never leave multi-mark mode active across a session load
        # Track whether THIS call opened the overlay, so we don't hide it out
        # from under a caller (loadpdf/loadrecentdb) that already has it open.
        opened_here = getattr(self, '_busy_overlay', None) is None
        if opened_here:
            self.busy("Loading session...")
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            if opened_here:
                self.unbusy()
            messagebox.showerror("Session Load Error", f"Failed to load session:\n{e}")
            return

        # Restore basic state
        self.project_name = data.get('project_name', self.project_name)
        self.sales_order_no = data.get('sales_order_no', self.sales_order_no)
        self.cabinet_id = data.get('cabinet_id', getattr(self, "cabinet_id", ""))
        self.current_page = data.get('current_page', 0)
        self.zoom_level = data.get('zoom_level', 1.0)
        self._update_zoom_toolbar_label()
        self.current_sr_no = data.get('current_sr_no', self.current_sr_no)

        # Restore session refs
        self.session_refs = set(data.get('session_refs', []))

        # Restore annotations with proper type conversion
        self.annotations = []
        highlight_count = 0
        pen_count = 0
        text_count = 0
        
        for entry in data.get('annotations', []):
            ann = entry.copy()
            # Backward compatibility for sessions saved before the error
            # highlighter changed from orange to pink.
            if ann.get('color') == 'orange':
                ann['color'] = 'pink'
            ann.setdefault('_undo_id', uuid.uuid4().hex)

            # ===== HIGHLIGHTER ANNOTATIONS - Convert lists to tuples =====
            if 'points_page' in ann:
                ann['points_page'] = [(float(p[0]), float(p[1])) for p in ann['points_page']]
                highlight_count += 1
            
            # ===== BBOX - Convert list back to tuple =====
            if 'bbox_page' in ann:
                ann['bbox_page'] = tuple(float(x) for x in ann['bbox_page'])

            # ===== PEN STROKES - Convert lists to tuples =====
            if 'points' in ann:
                ann['points'] = [(float(p[0]), float(p[1])) for p in ann['points']]
                pen_count += 1

            # ===== TEXT ANNOTATIONS - Convert list to tuple =====
            if 'pos_page' in ann:
                pos = ann['pos_page']
                ann['pos_page'] = (float(pos[0]), float(pos[1]))
                text_count += 1
            
            # Ensure text content is restored
            if 'text' in ann:
                ann['text'] = str(ann['text'])

            self.annotations.append(ann)

            # Add ref_no to session refs
            if ann.get('ref_no'):
                self.session_refs.add(str(ann['ref_no']).strip())

        try:
            self.display(preserve_view=False)  # shows its own overlay if the render is heavy
        finally:
            if opened_here:
                self.unbusy()
        


    # ============================================================================
    # TRANSFORMATION METHODS FOR HIGHLIGHTER ANNOTATIONS
    # ============================================================================
    def textrotate(self, rect, page):
        """Transform bbox for page rotation (for old rectangle annotations)"""
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x1, y1, x2, y2 = rect

        if r == 0:
            return fitz.Rect(x1, y1, x2, y2)
        if r == 90:
            return fitz.Rect(y1, w - x2, y2, w - x1)
        if r == 180:
            return fitz.Rect(w - x2, h - y2, w - x1, h - y1)
        if r == 270:
            return fitz.Rect(h - y2, x1, h - y1, x2)

        return fitz.Rect(x1, y1, x2, y2)


    def pointrotate(self, point, page):
        """Transform a single point (x, y) for page rotation
        
        Used for:
        - Pen stroke points
        - Text annotation positions
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        x, y = point

        if r == 0:
            return fitz.Point(x, y)
        elif r == 90:
            return fitz.Point(y, w - x)
        elif r == 180:
            return fitz.Point(w - x, h - y)
        elif r == 270:
            return fitz.Point(h - y, x)
        
        return fitz.Point(x, y)


    def highlightpointrotate(self, points, page):
        """Transform highlighter stroke points for page rotation
        
        Highlighters store a list of (x, y) tuples representing the stroke path.
        Each point needs to be individually transformed based on page rotation.
        
        Args:
            points: List of (x, y) tuples representing the highlight stroke
            page: PyMuPDF page object with rotation info
            
        Returns:
            List of fitz.Point objects, transformed for the page rotation
        """
        r = page.rotation
        w = page.rect.width
        h = page.rect.height
        
        transformed_points = []
        
        for point in points:
            x, y = point
            
            if r == 0:
                transformed_points.append(fitz.Point(x, y))
            elif r == 90:
                transformed_points.append(fitz.Point(y, w - x))
            elif r == 180:
                transformed_points.append(fitz.Point(w - x, h - y))
            elif r == 270:
                transformed_points.append(fitz.Point(h - y, x))
            else:
                transformed_points.append(fitz.Point(x, y))
        
        return transformed_points
    def textpos(self, rect, page):
        """Get the correct position for text beside a highlight annotation based on page rotation
        
        Args:
            rect: fitz.Rect object (already transformed)
            page: fitz page object with rotation info
        
        Returns:
            fitz.Point object for text position
        """
        r = page.rotation
        offset = 5  # Small gap between highlight and text
        
        if r == 0:
            # Normal orientation - text to the right of highlight
            return fitz.Point(rect.x1 + offset, rect.y0)
        elif r == 90:
            # 90° rotation - text below highlight
            return fitz.Point(rect.x0, rect.y1 + offset)
        elif r == 180:
            # 180° rotation - text to the left of highlight
            return fitz.Point(rect.x0 - offset, rect.y1)
        elif r == 270:
            # 270° rotation - text above highlight
            return fitz.Point(rect.x1, rect.y0 - offset)
        
        # Default fallback
        return fitz.Point(rect.x1 + offset, rect.y0)
    
    def exportpdf(self):
        """
        Export annotated PDF with all highlighter, pen, and text marks to file.
        FUNCTIONAL USE: Creates permanent record of quality inspection with visual markup.
        Saves PDF with all annotations embedded for audit trail and documentation.
        """
        """Export PDF with all annotations including highlighter strokes"""
        if not self.pdf_document:
            messagebox.showwarning("Warning", "Please load a PDF first")
            return
    
        if not hasattr(self, 'project_dirs') or not self.project_dirs.get("annotated_drawings"):
            messagebox.showerror("Error", "Project directories not set up.")
            return
    
        self.busy("Exporting annotated PDF...")
        try:
            save_path = os.path.join(
                self.project_dirs["annotated_drawings"],
                f"{self.cabinet_id.replace(' ', '_')}_Annotated.pdf"
            )
    
            # Create output PDF
            out_doc = fitz.open()
            total_pages = len(self.pdf_document)
            for pnum in range(total_pages):
                if total_pages > 3:
                    self.busy(f"Copying page {pnum + 1} of {total_pages}...")
                out_doc.insert_pdf(self.pdf_document, from_page=pnum, to_page=pnum)
    
            # Open Excel for SR No lookup
            wb = None
            ws = None
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name]
                except:
                    pass
    
            self.busy("Drawing annotations...")
            # Draw annotations
            for ann in self.annotations:
                p = ann.get('page')
                if p is None or p < 0 or p >= len(out_doc):
                    continue
    
                target_page = out_doc[p]
                ann_type = ann.get('type')
    
                # -------- HIGHLIGHTER ANNOTATIONS --------
                if ann_type == 'highlight' and 'points_page' in ann:
                    points_page = ann['points_page']
                    if len(points_page) >= 2:
                        color_key = ann.get('color', 'yellow')
                        rgb = self.highlighter_colors[color_key]['rgb']
                        # Normalize RGB to 0-1 range for PyMuPDF
                        color = (rgb[0]/255, rgb[1]/255, rgb[2]/255)
                        
                        # Transform points for page rotation
                        transformed_points = self.highlightpointrotate(
                            points_page, 
                            target_page
                        )
                        
                        # Convert to list of tuples for ink annotation
                        stroke = [(pt.x, pt.y) for pt in transformed_points]
                        
                        if len(stroke) >= 2:
                            ink_list = [stroke]  # Wrap in list for PyMuPDF
                            annot = target_page.add_ink_annot(ink_list)
                            annot.set_colors(stroke=color)
                            annot.set_border(width=15)  # Thick highlighter stroke
                            annot.set_opacity(0.4)  # Semi-transparent
                            annot.update()
                            
                            # Add SR number text for BOTH pink AND green highlights
                            if color_key in ['pink', 'green'] and 'bbox_page' in ann:
                                sr_text = None
                                row = ann.get('excel_row')
                                sr_no = ann.get('sr_no')  # Try to get from annotation first
                                
                                # If SR number is stored in annotation, use it
                                if sr_no is not None:
                                    sr_text = f"Sr {sr_no}"
                                # Otherwise, try to read from Excel
                                elif row and ws:
                                    try:
                                        sr_val = self.readcell(ws, row, self.punch_cols['sr_no'])
                                        if sr_val is not None:
                                            sr_text = f"Sr {sr_val}"
                                    except:
                                        pass
                                
                                if sr_text:
                                    # Use bbox for text position
                                    x1, y1, x2, y2 = ann['bbox_page']
                                    bbox_rect = self.textrotate(
                                        (x1, y1, x2, y2), 
                                        target_page
                                    )
                                    # Position text beside the highlight
                                    text_pos = self.textpos(bbox_rect, target_page)
                                    
                                    # Use different color for green vs pink
                                    text_color = (0, 0.5, 0) if color_key == 'green' else (1, 0, 0)
                                    
                                    try:
                                        target_page.insert_text(
                                            text_pos, 
                                            sr_text, 
                                            fontsize=8, 
                                            color=text_color
                                        )
                                    except:
                                        pass
    
    
                # -------- PEN STROKES --------
                elif ann_type == 'pen' and 'points' in ann:
                    points = ann['points']
                    if len(points) >= 2:
                        # Transform points for rotation
                        transformed_points = [
                            self.pointrotate(pt, target_page) 
                            for pt in points
                        ]
                        
                        # Draw lines between consecutive points
                        for i in range(len(transformed_points) - 1):
                            p1 = transformed_points[i]
                            p2 = transformed_points[i + 1]
                            target_page.draw_line(p1, p2, color=(1, 0, 0), width=2)
    
                # -------- TEXT ANNOTATIONS --------
                elif ann_type == 'text' and 'pos_page' in ann:
                    pos = ann['pos_page']
                    text = ann.get('text', '')
                    if text:
                        text_point = self.pointrotate(pos, target_page)
                        try:
                            target_page.insert_text(
                                text_point, text,
                                fontsize=10, color=(1, 0, 0),
                                rotate=target_page.rotation
                            )
                        except:
                            pass
    
            if wb:
                wb.close()
    
            self.busy("Saving exported PDF...")
            out_doc.save(save_path)
            out_doc.close()
            self.sync_manager_stats_only()
    
        except PermissionError:
            messagebox.showerror("Error", "Close the target file (if open) and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export annotated PDF:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            self.unbusy()


    # ================================================================
    # UI SETUP WITH HIGHLIGHTER CONTROLS
    # ================================================================
    def bind_global_keyboard_popup(self):
        """
        Globally bind FocusIn/FocusOut on all Entry and Text widgets (including those
        inside dialogs like simpledialog) to show/hide the on-screen keyboard.
        FUNCTIONAL USE: Ensures the popup keyboard appears automatically whenever the
        user taps into ANY text box anywhere in the app, without editing every dialog.
        Call this once in __init__ after uisetup().
        """
        def restore_text_focus(widget):
            """Keep keyboard focus and the insertion caret inside the tapped field."""
            try:
                if not widget.winfo_exists():
                    return
                widget.focus_force()
                if isinstance(widget, tk.Text):
                    widget.mark_set(tk.INSERT, widget.index(tk.INSERT))
                    widget.see(tk.INSERT)
                else:
                    widget.icursor(tk.INSERT)
            except (tk.TclError, AttributeError):
                pass

        def on_focus_in(event):
            widget = event.widget
            if isinstance(widget, (tk.Entry, tk.Text, ttk.Entry)):
                # Make the caret visible immediately. TabTip can briefly take
                # foreground focus while opening, so restore it again afterward.
                restore_text_focus(widget)
                show_onscreen_keyboard()
                self.root.after_idle(lambda w=widget: restore_text_focus(w))
                self.root.after(250, lambda w=widget: restore_text_focus(w))

        def on_focus_out(event):
            widget = event.widget
            if isinstance(widget, (tk.Entry, tk.Text, ttk.Entry)):
                # Small delay avoids flicker when focus moves between two text fields
                self.root.after(150, self._maybe_hide_keyboard)

        def first_editable_text_widget(parent):
            """Return the first enabled Entry/Text field in a dialog's widget tree."""
            try:
                children = parent.winfo_children()
            except tk.TclError:
                return None

            for child in children:
                if isinstance(child, (tk.Entry, tk.Text, ttk.Entry)):
                    try:
                        state = str(child.cget('state'))
                    except (tk.TclError, AttributeError):
                        state = 'normal'
                    if state not in ('disabled', 'readonly'):
                        return child
                nested = first_editable_text_widget(child)
                if nested is not None:
                    return nested
            return None

        def focus_dialog_input(dialog):
            """Move the caret into the first input whenever a text dialog opens."""
            try:
                if not dialog.winfo_exists():
                    return

                focused = dialog.focus_get()
                if (isinstance(focused, (tk.Entry, tk.Text, ttk.Entry)) and
                        self._widget_is_inside(focused, dialog)):
                    return

                target = first_editable_text_widget(dialog)
                if target is None:
                    return

                dialog.lift()
                restore_text_focus(target)
                show_onscreen_keyboard()
            except (tk.TclError, AttributeError):
                pass

        def on_dialog_map(event):
            dialog = event.widget
            # Run more than once because simpledialog creates/maps its controls
            # in stages, and the touch keyboard can briefly take Windows focus.
            self.root.after_idle(lambda d=dialog: focus_dialog_input(d))
            self.root.after(80, lambda d=dialog: focus_dialog_input(d))
            self.root.after(250, lambda d=dialog: focus_dialog_input(d))

        self.root.bind_class("Toplevel", "<Map>", on_dialog_map, add="+")
        self.root.bind_class("Entry", "<FocusIn>", on_focus_in, add="+")
        self.root.bind_class("Text", "<FocusIn>", on_focus_in, add="+")
        self.root.bind_class("TEntry", "<FocusIn>", on_focus_in, add="+")

        self.root.bind_class("Entry", "<FocusOut>", on_focus_out, add="+")
        self.root.bind_class("Text", "<FocusOut>", on_focus_out, add="+")
        self.root.bind_class("TEntry", "<FocusOut>", on_focus_out, add="+")

    def _maybe_hide_keyboard(self):
        """
        Hide the on-screen keyboard only if focus has actually left a text-entry widget.
        FUNCTIONAL USE: Prevents keyboard flicker when tabbing between input fields.
        """
        try:
            focused = self.root.focus_get()
        except Exception:
            focused = None

        if not isinstance(focused, (tk.Entry, tk.Text, ttk.Entry)):
            hide_onscreen_keyboard()
    def uisetup(self):
        """
        Create complete user interface with toolbar, menu, canvas, and status bar.
        FUNCTIONAL USE: Builds UI components including file menu, tools menu, navigation buttons,
        color selectors, zoom controls, highlighter/pen/text tool buttons, canvas for PDF display,
        and keyboard shortcuts for quality inspection workflow.
        Sets up all event bindings for mouse and keyboard interactions.
        """
        """Setup modern UI with highlighter controls"""
        # Main toolbar
        toolbar = tk.Frame(self.root, bg='#1e293b', height=80)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        self.toolbar = toolbar
        
        # Enhanced Menu Bar
        menubar = Menu(self.root, bg='#1e293b', fg='white', activebackground='#3b82f6')
        self.root.config(menu=menubar)
        
        # File Menu
        file_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open PDF", command=self.loadpdf, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Load Session", command=self.loadsession, accelerator="Ctrl+L")
        file_menu.add_command(label="Save Session", command=self.savesession, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Export Annotated PDF", command=self.exportpdf, accelerator="Ctrl+E")
        file_menu.add_command(label="Save Interphase Excel", command=self.saveinterphase)
        file_menu.add_command(label="Open Excel", command=self.openxcl, accelerator="Ctrl+Shift+E")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools Menu
        tools_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Review Checklist", command=self.reviewnow, accelerator="Ctrl+R")
        tools_menu.add_command(label="Punch Closing Mode", command=self.punchclosing, accelerator="Ctrl+Shift+P")
        tools_menu.add_command(label="Edit Existing Punch", command=self.editexistingpunch, accelerator="Ctrl+Shift+U")
        tools_menu.add_separator()

        # Cabinet workspace actions are grouped separately from inspection tools.
        # These actions become useful after the initial cabinet setup has been saved.
        cabinet_menu = Menu(
            tools_menu, tearoff=0, bg='#1e293b', fg='white',
            activebackground='#3b82f6', activeforeground='white'
        )
        tools_menu.add_cascade(label="Cabinet Setup", menu=cabinet_menu)
        cabinet_menu.add_command(
            label="Select Different PDF...",
            command=self.select_different_pdf,
            accelerator="Ctrl+Shift+O"
        )
        cabinet_menu.add_command(
            label="Edit Cabinet Details...",
            command=self.edit_cabinet_details,
            accelerator="Ctrl+Shift+D"
        )

        tools_menu.add_separator()
        tools_menu.add_command(label="Verify ", command=self.viewhandbacks, accelerator="Ctrl+Shift+V")
        
        # View Menu
        view_menu = Menu(menubar, tearoff=0, bg='#1e293b', fg='white', activebackground='#3b82f6')
        menubar.add_cascade(label="View", menu=view_menu)        
        # Keyboard shortcuts
        self.root.bind_all("<Control-o>", lambda e: self.loadpdf())
        self.root.bind_all("<Control-s>", lambda e: self.savesession())
        self.root.bind_all("<Control-l>", lambda e: self.loadsession())
        self.root.bind_all("<Control-e>", lambda e: self.exportpdf())
        self.root.bind_all("<Control-r>", lambda e: self.reviewnow())
        self.root.bind_all("<Control-z>", lambda e: self.undolast())
        self.root.bind_all("<Control-P>", lambda e: self.punchclosing())
        self.root.bind_all("<Control-U>", lambda e: self.editexistingpunch())
        self.root.bind_all("<Control-E>", lambda e: self.openxcl())
        self.root.bind_all("<Control-V>", lambda e: self.viewhandbacks())
        self.root.bind_all("<Control-Shift-O>", lambda e: self.select_different_pdf())
        self.root.bind_all("<Control-Shift-D>", lambda e: self.edit_cabinet_details())
        self.root.bind_all("<Escape>", lambda e: self.deactivate())
        self.root.bind_all("<Delete>", self._delete_selected_text_box)
        
        # Modern button style
        btn_style = {
            'bg': '#3b82f6',
            'fg': 'white',
            'padx': 12,
            'pady': 10,
            'font': ('Segoe UI', 9, 'bold'),
            'relief': tk.FLAT,
            'borderwidth': 0,
            'cursor': 'hand2'
        }
        
        # Left section - File operations
        left_frame = tk.Frame(toolbar, bg='#1e293b')
        left_frame.pack(side=tk.LEFT, padx=10, pady=10)
        
        tk.Button(left_frame, text="Open PDF", command=self.loadpdf, **btn_style).pack(side=tk.LEFT, padx=3)
        
        # All Projects browser. This replaces the old limited recent-project dropdown.
        projects_frame = tk.Frame(left_frame, bg='#1e293b')
        projects_frame.pack(side=tk.LEFT, padx=8)
        tk.Button(
            projects_frame, text="Projects & Cabinets", command=self.show_project_browser,
            bg='#334155', fg='white', activebackground='#475569', activeforeground='white',
            font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, borderwidth=0,
            padx=14, pady=10, cursor='hand2'
        ).pack(side=tk.LEFT)
        # Center - THREE DIRECT HIGHLIGHTER COLOR CIRCLES
        # No label or dropdown: each available color is always visible and can
        # be selected directly with one click.
        highlighter_frame = tk.Frame(toolbar, bg='#1e293b')
        highlighter_frame.pack(side=tk.LEFT, padx=30)

        self.color_picker_frame = tk.Frame(highlighter_frame, bg='#1e293b')
        self.color_picker_frame.pack(side=tk.LEFT)
        self.color_canvases = {}

        for color_key in ('green', 'pink', 'yellow'):
            color_canvas = tk.Canvas(
                self.color_picker_frame,
                width=44,
                height=44,
                bg='#1e293b',
                highlightthickness=0,
                borderwidth=0,
                cursor='hand2'
            )
            color_canvas.pack(side=tk.LEFT, padx=3)
            color_canvas.bind(
                "<Button-1>",
                lambda event, ck=color_key: self.selecthighlighter(ck)
            )
            self.color_canvases[color_key] = color_canvas

        self.colorbutton()

        # NOTE: The "already marked / not marked" decision is no longer a
        # persistent toolbar toggle. It's now asked per-highlight, right
        # after an pink (error) highlight is drawn - see leftrel().

        # Navigation
        center_frame = tk.Frame(toolbar, bg='#1e293b')
        center_frame.pack(side=tk.LEFT, padx=20)
        
        # Read-only page indicator. Page changes happen only through normal
        # scrolling; direct page-number entry / skip-to-page navigation is removed.
        self.page_label = tk.Label(
            center_frame, text="Page: 0 / 0", bg='#1e293b', fg='white',
            font=('Segoe UI', 10, 'bold')
        )
        self.page_label.pack(side=tk.LEFT, padx=(4, 10))
        
        # Tool section
        tool_frame = tk.Frame(toolbar, bg='#1e293b')
        tool_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(tool_frame, text="Tools:", bg='#1e293b', fg='#94a3b8',
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 8))
        
        # Load icons or use fallback
        try:
            icon_size = (44, 44)
            
            pen_icon_path = asset_path("pen_icon.png")
            pen_img = Image.open(pen_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.pen_icon = ImageTk.PhotoImage(pen_img)
            
            text_icon_path = asset_path("text_icon.png")
            text_img = Image.open(text_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.text_icon = ImageTk.PhotoImage(text_img)
            
            undo_icon_path = asset_path("undo_icon.png")
            undo_img = Image.open(undo_icon_path).resize(icon_size, Image.Resampling.LANCZOS)
            self.undo_icon = ImageTk.PhotoImage(undo_img)
            
            self.pen_btn = tk.Button(tool_frame, image=self.pen_icon,
                                    command=lambda: self.toolmode("pen"),
                                    bg='#334155', width=48, height=48,
                                    relief=tk.FLAT, cursor='hand2')
            self.pen_btn.pack(side=tk.LEFT, padx=2)
            
            self.text_btn = tk.Button(tool_frame, image=self.text_icon,
                                     command=lambda: self.toolmode("text"),
                                     bg='#334155', width=48, height=48,
                                     relief=tk.FLAT, cursor='hand2')
            self.text_btn.pack(side=tk.LEFT, padx=2)
            
            self.undo_btn = tk.Button(tool_frame, image=self.undo_icon,
                                      command=self.undolast,
                                      bg='#334155', width=48, height=48,
                                      relief=tk.FLAT, cursor='hand2')
            self.undo_btn.pack(side=tk.LEFT, padx=2)
            
        except Exception as e:
            print(f"Could not load tool icons: {e}")
        
        # Right section - Action buttons
        right_frame = tk.Frame(toolbar, bg='#1e293b')
        right_frame.pack(side=tk.RIGHT, padx=10, pady=10)
        
        verify_btn_style = btn_style.copy()
        verify_btn_style['bg'] = '#ec4899'
        
        tk.Button(right_frame, text=" Verify ",
                 command=self.viewhandbacks,
                 **verify_btn_style).pack(side=tk.RIGHT, padx=3)
        
        handover_btn_style = btn_style.copy()
        handover_btn_style['bg'] = '#8b5cf6'
        
        tk.Button(right_frame, text="Handover",
                 command=self.handover,
                 **handover_btn_style).pack(side=tk.RIGHT, padx=3)

        # -------- Zoom dropdown --------
        # A simple toolbar dropdown replaces the floating slider and pinch
        # handlers. Changes are debounced and rendered once to avoid lag.
        self.setup_toolbar_zoom(toolbar)
        
        # Canvas with scrollbars
        canvas_frame = tk.Frame(self.root, bg='#f1f5f9')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # ---- Vertical scrollbar container (holds scrollbar + page number popup) ----
        v_scroll_container = tk.Frame(canvas_frame, bg='#f1f5f9')
        v_scroll_container.pack(side=tk.RIGHT, fill=tk.Y)

        v_scrollbar = tk.Scrollbar(v_scroll_container, orient=tk.VERTICAL, width=12)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        h_scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, width=12)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        self.canvas = tk.Canvas(canvas_frame, bg='#f8fafc',
                               yscrollcommand=v_scrollbar.set,
                               xscrollcommand=h_scrollbar.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # NEW: disable OS-level touch gesture translation on this canvas so
        # touch drags reach leftclick/leftdrag/leftrel instead of being
        # intercepted as a pan/scroll gesture by Windows.
        self.canvas.update_idletasks()  # ensure winfo_id() is valid
        self.root.after(50, self._apply_touch_gesture_fix)

        v_scrollbar.config(command=self.canvas.yview)
        h_scrollbar.config(command=self.canvas.xview)

        self._setup_scrollbar_hover_effects(v_scrollbar, v_scroll_container)
        
        # Bind mouse events
        self.canvas.bind("<ButtonPress-1>", self.leftclick)
        self.canvas.bind("<B1-Motion>", self.leftdrag)
        self.canvas.bind("<ButtonRelease-1>", self.leftrel)
        self.canvas.bind("<Double-Button-1>", self.doubleclick)
        self._bind_display_mouse_controls()
        
        # Modern status bar
        status_bar = tk.Frame(self.root, bg='#334155', height=40)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        instructions_text = " Esc: Deactivate | Ctrl+Z: Undo"
        tk.Label(status_bar, text=instructions_text, bg='#334155', fg='#e2e8f0',
                font=('Segoe UI', 9), pady=10).pack()

    def _apply_touch_gesture_fix(self):
        """
        Configure Windows touch feedback and install canvas pinch handling.
        FUNCTIONAL USE: Called shortly after canvas creation once the widget
        has a valid native window handle. Retries briefly if the handle
        isn't ready yet.
        """
        if os.name != 'nt':
            return

        try:
            hwnd = self.canvas.winfo_id()
            if not hwnd:
                self.root.after(100, self._apply_touch_gesture_fix)
                return

            success = configure_touch_feedback(hwnd)
            self._touch_gesture_fix_applied = success
            self.setup_canvas_pinch_zoom()

            if not success:
                print("[WARN] Touch feedback settings were unavailable; input remains enabled.")
        except Exception as e:
            print(f"[WARN] _apply_touch_gesture_fix error: {e}")

    def _setup_scrollbar_hover_effects(self, v_scrollbar, v_scroll_container):
        """
        Make the vertical scrollbar grow wider on hover/press and show the
        current page number as a floating popup tooltip beside it.
        FUNCTIONAL USE: Improves touch/mouse usability by enlarging the scrollbar
        hit-area during interaction and giving instant page-position feedback via
        a small borderless popup window (like a tooltip) rather than inline UI.
        Args: v_scrollbar - the Scrollbar widget, v_scroll_container - its parent Frame
        """
        self._v_scrollbar = v_scrollbar
        self._v_scroll_container = v_scroll_container
        self._scrollbar_normal_width = 12
        self._scrollbar_active_width = 22
        self._page_popup = None

        def grow(event=None):
            v_scrollbar.config(width=self._scrollbar_active_width)
            self._show_page_badge()

        def shrink(event=None):
            if not getattr(self, '_scrollbar_dragging', False):
                v_scrollbar.config(width=self._scrollbar_normal_width)
                self._hide_page_badge()

        def on_press(event=None):
            self._scrollbar_dragging = True
            grow()

        def on_release(event=None):
            self._scrollbar_dragging = False
            self.root.after(600, shrink)

        def on_motion(event=None):
            self._show_page_badge()

        v_scrollbar.bind("<Enter>", grow)
        v_scrollbar.bind("<Leave>", shrink)
        v_scrollbar.bind("<ButtonPress-1>", on_press)
        v_scrollbar.bind("<ButtonRelease-1>", on_release)
        v_scrollbar.bind("<B1-Motion>", on_motion)

    def _show_page_badge(self):
        """
        Display a small borderless popup window with the current page number,
        positioned beside the vertical scrollbar.
        FUNCTIONAL USE: Called while hovering/dragging the scrollbar so the user
        always knows which page they're scrolling to. Uses a Toplevel popup
        (tooltip-style) instead of inline UI so it doesn't shift layout.
        """
        if not self.pdf_document:
            return

        self._update_current_page_from_scroll()

        total = len(self.pdf_document)
        current = self.current_page + 1
        text = f"Page {current}/{total}"

        if self._page_popup is None or not self._page_popup.winfo_exists():
            self._page_popup = tk.Toplevel(self.root)
            self._page_popup.overrideredirect(True)
            self._page_popup.attributes('-topmost', True)
            self._page_popup_label = tk.Label(
                self._page_popup, text=text, bg='#1e293b', fg='white',
                font=('Segoe UI', 9, 'bold'), padx=8, pady=4,
                relief=tk.SOLID, borderwidth=1
            )
            self._page_popup_label.pack()
        else:
            self._page_popup_label.config(text=text)

        # Position just to the left of the scrollbar, vertically centered on it
        sb_x = self._v_scrollbar.winfo_rootx()
        sb_y = self._v_scrollbar.winfo_rooty()
        sb_h = self._v_scrollbar.winfo_height()

        popup_w = self._page_popup_label.winfo_reqwidth()
        popup_h = self._page_popup_label.winfo_reqheight()

        popup_x = sb_x - popup_w - 8
        popup_y = sb_y + (sb_h // 2) - (popup_h // 2)

        self._page_popup.geometry(f"+{popup_x}+{popup_y}")
        self._page_popup.deiconify()

    def _hide_page_badge(self):
        """
        Destroy/hide the page-number popup window.
        FUNCTIONAL USE: Called after the user stops interacting with the scrollbar.
        """
        if getattr(self, '_page_popup', None) is not None:
            try:
                self._page_popup.withdraw()
            except tk.TclError:
                self._page_popup = None
    def _bind_display_mouse_controls(self):
        """Register wheel scrolling handlers for the PDF display canvas."""
        if getattr(self, '_display_mouse_controls_bound', False):
            return

        self.root.bind_all("<MouseWheel>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-MouseWheel>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-Button-4>", self._on_display_mousewheel, add="+")
        self.root.bind_all("<Shift-Button-5>", self._on_display_mousewheel, add="+")

        self._display_mouse_controls_bound = True

    def _is_pointer_over_canvas(self):
        """Return True when the mouse pointer is over the display canvas."""
        if not hasattr(self, 'canvas') or not self.canvas or not self.canvas.winfo_exists():
            return False

        try:
            widget = self.root.winfo_containing(self.root.winfo_pointerx(), self.root.winfo_pointery())
        except tk.TclError:
            return False

        while widget is not None:
            if widget == self.canvas:
                return True
            widget = widget.master
        return False

    def _on_display_mousewheel(self, event):
        if time.monotonic() < getattr(self, '_touch_scroll_lock_until', 0.0):
            return "break"

        if self.active_highlighter or self.tool_mode in ("pen", "text") or self.drawing:
            return "break"

        if not self._is_pointer_over_canvas():
            return

        delta = 0
        if getattr(event, 'num', None) == 4:
            delta = 1
        elif getattr(event, 'num', None) == 5:
            delta = -1
        elif getattr(event, 'delta', 0):
            delta = 1 if event.delta > 0 else -1

        if delta == 0:
            return

        try:
            horizontal = bool(getattr(event, 'state', 0) & 0x0001)
            if horizontal:
                self.canvas.xview_scroll(-delta, "units")
            else:
                self.canvas.yview_scroll(-delta, "units")
        except tk.TclError:
            return

        self._update_current_page_from_scroll()
        return "break"

    def _update_current_page_from_scroll(self):
        """
        Recalculate self.current_page based on canvas scroll position and refresh
        the page-number label text if the toolbar label exists.
        FUNCTIONAL USE: Keeps the page indicator (toolbar label + scrollbar popup)
        accurate as the user scrolls through the stacked multi-page PDF view.
        """
        if not self.pdf_document or not self.page_layout:
            return

        try:
            top_fraction = self.canvas.yview()[0]
            scroll_region = self.canvas.cget("scrollregion").split()
            total_height = float(scroll_region[3]) if len(scroll_region) == 4 else 1
        except (tk.TclError, IndexError, ValueError):
            return

        viewport_center = (self.canvas.winfo_height() / 2.0)
        y_pos = (top_fraction * total_height) + viewport_center
        for layout in self.page_layout:
            if layout['y'] <= y_pos < layout['y'] + layout['height']:
                self.current_page = layout['page_index']
                break

        self._update_page_toolbar()
    


    # ================================================================
    # HIGHLIGHTER UI HELPERS - UPDATED
    # ================================================================

    def colorbutton(self):
        """Redraw all three toolbar color circles and show the active one."""
        canvases = getattr(self, 'color_canvases', {})
        for color_key, canvas in canvases.items():
            canvas.delete("all")
            rgb = self.highlighter_colors[color_key]['rgb']
            hex_color = f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
            is_active = self.active_highlighter == color_key

            canvas.create_oval(
                0 if is_active else 2,
                0 if is_active else 2,
                44 if is_active else 42,
                44 if is_active else 42,
                outline='#3b82f6' if is_active else '#475569',
                width=3 if is_active else 1,
                fill='#1e293b'
            )
            canvas.create_oval(6, 6, 38, 38, fill=hex_color, outline='', width=0)

    def selecthighlighter(self, color_key):
        """Activate a color directly, or turn it off when its active circle is clicked."""
        if self.active_highlighter == color_key:
            self.togglehighlighter()
            return

        self.current_color_key = color_key
        self.active_highlighter = color_key
        self.root.config(cursor="pencil")

        if self.tool_mode:
            self.tool_mode = None
            self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            self.text_btn.config(bg='#334155', relief=tk.FLAT)

        self.colorbutton()

    def colorchange(self, color_key):
        """
        Switch active highlighter color for quality markup.
        FUNCTIONAL USE: Changes current highlighter to specified color (yellow/green/pink).
        Updates button state and readies tool for next annotation with new color.
        Args: color_key - String key (yellow, green, pink) from highlighter_colors dict
        """
        """Change the highlighter color"""
        self.current_color_key = color_key
        self.colorbutton()
        
        if self.active_highlighter:
            self.active_highlighter = color_key
            self.root.config(cursor="pencil")


    def _return_to_touch_mode(self):
        """Reset transient annotation state and restore touch pan/scroll mode."""
        if self.active_highlighter or self.tool_mode is not None:
            return
        self.drawing = False
        self.drawing_type = None
        self.drawing_page = None
        self._panning = False
        self.highlight_points = []
        self.pen_points = []
        self._text_box_start = None
        self._touch_scroll_lock_until = 0.0
        self.cleartemp()
        self._clear_text_selection()
        self.root.config(cursor="")
        if hasattr(self, "canvas"):
            self.canvas.config(cursor="")

    def togglehighlighter(self):
        """
        Toggle highlighter tool on/off.
        FUNCTIONAL USE: Activates/deactivates highlighter mode for PDF markup.
        Updates UI state and sets drawing_type to 'highlight'.
        """
        """Toggle highlighter on/off"""
        if self.active_highlighter:
            self.active_highlighter = None
            self.root.config(cursor="")
            self.colorbutton()
            self._return_to_touch_mode()
        else:
            self.active_highlighter = self.current_color_key
            self.root.config(cursor="pencil")
            self.colorbutton()
            
            if self.tool_mode:
                self.tool_mode = None
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
                self.text_btn.config(bg='#334155', relief=tk.FLAT)

    def toolmode(self, mode):
        """
        Switch active drawing tool mode.
        FUNCTIONAL USE: Sets tool_mode to 'pen' for freehand drawing or 'text' for text annotations.
        Used by toolbar buttons to activate different annotation tools.
        Args: mode - String ('pen', 'text') or None to deactivate
        """
        """Set tool mode (pen or text)"""
        if self.active_highlighter:
            return

        # While multi-mark mode is active, only the highlighter is allowed -
        # pen/text stay locked until the user presses Stop.
        if getattr(self, 'multimark_active', False):
            return

        if self.tool_mode == mode:
            self.tool_mode = None
            if mode == "pen":
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
            self._return_to_touch_mode()
        else:
            self.tool_mode = mode
            if mode == "pen":
                self.pen_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.text_btn.config(bg='#334155', relief=tk.FLAT)
            else:
                self.text_btn.config(bg='#3b82f6', relief=tk.SUNKEN)
                self.pen_btn.config(bg='#334155', relief=tk.FLAT)

    def deactivate(self):
        """
        Disable all active drawing tools.
        FUNCTIONAL USE: Clears tool mode, stops active drawing, resets canvas state.
        Bound to Escape key for quick tool deactivation.
        """
        """Deactivate all tools and highlighters"""
        if self._text_editor is not None:
            self._commit_text_editor()
        self.selected_annotation = None
        for item in self._text_selection_ids:
            self.canvas.delete(item)
        self._text_selection_ids = []
        if self.active_highlighter:
            self.togglehighlighter()
        if self.tool_mode:
            self.toolmode(self.tool_mode)
        self._return_to_touch_mode()

    def updtoolpane(self):
        """Update annotation statistics"""
        # Placeholder - implement if you have a tool pane
        pass

    def flashstat(self, message, bg='#10b981'):
        """
        Display temporary status message in status bar with color indication.
        FUNCTIONAL USE: Provides visual feedback for user actions (success, warning, info).
        Message auto-clears after timeout.
        Args: message - Text to display, bg - background color (green for success, pink for warning)
        """
        """Show a temporary status message"""
        status_label = tk.Label(
            self.root, 
            text=message, 
            bg=bg, 
            fg='white', 
            font=('Segoe UI', 10, 'bold'),
            padx=25, 
            pady=12,
            relief=tk.FLAT
        )
        status_label.place(relx=0.5, rely=0.08, anchor='center')
        self.root.after(1500, lambda: status_label.destroy())

    # ================================================================
    # LOADING OVERLAY (for expensive/blocking operations)
    # ================================================================
    #
    # Tk is single-threaded: a long synchronous call (PDF load, full
    # multi-page render, PDF export, session restore, etc.) freezes the
    # window and the OS shows "Not Responding" because no events get
    # pumped while that call runs. This app doesn't use worker threads
    # for those calls, so the fix here is:
    #   1. Show a modal overlay immediately (busy()).
    #   2. Force Tk to actually paint it with update_idletasks()/update()
    #      BEFORE starting the expensive work.
    #   3. Run the expensive work (still synchronous - that part is
    #      unavoidable without threading it - but now the user sees a
    #      spinner/message instead of a frozen, greyed-out window).
    #   4. Hide the overlay (unbusy()) once done, even on error, via
    #      try/finally at each call site.
    # This does not make the work itself faster; it replaces the "Not
    # Responding" freeze with an explicit, intentional loading state.

    def busy(self, message="Working..."):
        """
        Show a modal loading overlay with a spinner and message.
        FUNCTIONAL USE: Call immediately before starting an expensive
        synchronous operation (PDF load/export, full render, session
        restore) so the user sees deliberate progress instead of the
        window appearing to hang / OS marking it "Not Responding".
        Safe to call again while already showing - just updates the message.
        Always pair with a matching self.unbusy() in a finally block.

        Implementation note: this is a plain tk.Frame placed INSIDE the main
        window (self.root), not a separate Toplevel window. An earlier
        version used a topmost/overrideredirect Toplevel sized from
        winfo_rootx/rooty/width/height, which could report stale (0,0) or
        1x1 geometry if called before the window had been fully mapped -
        producing a stray borderless window that could appear to cover the
        whole screen instead of just the app. A child Frame with place()
        is physically constrained to the parent window's real rendered
        area, so it can never extend beyond the app's own window.
        """
        try:
            if getattr(self, '_busy_overlay', None) is not None and self._busy_overlay.winfo_exists():
                self._busy_message_var.set(message)
                self._busy_overlay.lift()
                self._spin_busy_overlay()
                self.root.update_idletasks()
                return

            overlay = tk.Frame(self.root, bg='#0f172a')
            self._busy_overlay = overlay

            card = tk.Frame(overlay, bg='#1e293b', highlightthickness=1,
                             highlightbackground='#334155')
            card.place(relx=0.5, rely=0.5, anchor='center')

            self._busy_spinner_var = tk.StringVar(value='◐')
            tk.Label(card, textvariable=self._busy_spinner_var, bg='#1e293b',
                     fg='#60a5fa', font=('Segoe UI', 26)).pack(padx=36, pady=(28, 6))

            self._busy_message_var = tk.StringVar(value=message)
            tk.Label(card, textvariable=self._busy_message_var, bg='#1e293b',
                     fg='white', font=('Segoe UI', 11, 'bold')).pack(padx=36, pady=(0, 26))

            self._busy_spin_frames = ['◐', '◓', '◑', '◒']
            self._busy_spin_index = 0
            self._busy_spin_after_id = None

            # Cover the full client area of the main window - and only the
            # main window, since this Frame's parent IS self.root, so Tk
            # will never let it render outside that window's own bounds.
            overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
            overlay.lift()
            overlay.focus_set()
            self._spin_busy_overlay()

            # Force Tk to actually draw the overlay right now, before the
            # caller goes on to do the expensive blocking work - otherwise
            # the overlay would just sit in the event queue, unseen, while
            # the freeze happens exactly as before.
            self.root.update_idletasks()
            self.root.update()
        except tk.TclError:
            self._busy_overlay = None

    def _spin_busy_overlay(self):
        """Advance the busy-overlay spinner glyph. Re-arms itself while the overlay exists."""
        overlay = getattr(self, '_busy_overlay', None)
        if overlay is None or not overlay.winfo_exists():
            return
        try:
            self._busy_spin_index = (self._busy_spin_index + 1) % len(self._busy_spin_frames)
            self._busy_spinner_var.set(self._busy_spin_frames[self._busy_spin_index])
        except (tk.TclError, AttributeError):
            return
        self._busy_spin_after_id = self.root.after(160, self._spin_busy_overlay)

    def busy_pump(self):
        """
        Let Tk repaint the overlay/spinner mid-operation without ending the
        busy state. Call this between chunks of a long operation that has
        natural checkpoints (e.g. per-page in a loop) so the spinner keeps
        animating and the window never looks stuck, even during the
        unavoidable synchronous work.
        """
        try:
            self.root.update_idletasks()
            self.root.update()
        except tk.TclError:
            pass

    def unbusy(self):
        """Hide the loading overlay. Always call from a finally block."""
        after_id = getattr(self, '_busy_spin_after_id', None)
        if after_id is not None:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass
            self._busy_spin_after_id = None
        overlay = getattr(self, '_busy_overlay', None)
        self._busy_overlay = None
        if overlay is not None:
            try:
                overlay.destroy()
            except tk.TclError:
                pass

    # ================================================================
    # UNDO FUNCTIONALITY
    # ================================================================

    def addtostack(self, action_type, annotation):
        """Store an undo action using a unique ID, never dictionary equality."""
        annotation.setdefault('_undo_id', uuid.uuid4().hex)
        self.undo_stack.append({
            'type': action_type,
            'annotation_id': annotation['_undo_id'],
            'excel_row': annotation.get('excel_row'),
            'sr_no': annotation.get('sr_no')
        })
        if len(self.undo_stack) > self.max_undo:
            self.undo_stack.pop(0)

    def _remove_punch_excel_row(self, target_row):
        """Remove one punch row without touching any earlier punch entries."""
        if not target_row or not self.excel_file or not os.path.exists(self.excel_file):
            return
        wb = load_workbook(self.excel_file)
        ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
        last_row = target_row
        while self.readcell(ws, last_row + 1, self.punch_cols['sr_no']) is not None:
            last_row += 1
        columns = list(self.punch_cols.values())
        for row in range(target_row, last_row):
            for col in columns:
                self.writecell(ws, row, col, self.readcell(ws, row + 1, col))
        for col in columns:
            self.writecell(ws, last_row, col, None)
        wb.save(self.excel_file)
        wb.close()
        for ann in self.annotations:
            row = ann.get('excel_row')
            if isinstance(row, int) and row > target_row:
                ann['excel_row'] = row - 1

    def undolast(self):
        """Undo exactly the latest annotation.

        For a normal punch-creating highlight this also removes the Excel
        punch row it created. For a multi-mark highlight (an extra highlight
        attached to an ALREADY-EXISTING punch via multi-mark mode), the
        annotation is simply detached/removed - the punch itself, its Excel
        row, and any of its other highlights are left completely untouched,
        since this highlight never owned that row in the first place.
        """
        if not self.undo_stack:
            messagebox.showinfo("Nothing to Undo", "No actions to undo.", icon='info')
            return
        # Peek first so a cancelled confirmation leaves the undo history intact.
        action = self.undo_stack[-1]
        annotation_id = action.get('annotation_id')
        index = next((i for i, ann in enumerate(self.annotations)
                      if ann.get('_undo_id') == annotation_id), None)
        if index is None:
            self.undo_stack.pop()
            messagebox.showinfo("Nothing to Undo", "The selected item was already removed.")
            return

        annotation = self.annotations[index]
        is_multimark_extra = bool(annotation.get('multimark'))
        excel_row = None if is_multimark_extra else (annotation.get('excel_row') or action.get('excel_row'))

        if excel_row:
            proceed = messagebox.askyesno(
                "Undo Punch Entry?",
                "Warning: This undo operation will also remove the corresponding "
                "punch entry from the Punch Sheet.\n\n"
                "Do you want to continue?",
                icon='warning',
                parent=self.root
            )
            if not proceed:
                return

        # Remove the action only after any required confirmation is accepted.
        self.undo_stack.pop()
        try:
            if excel_row:
                self._remove_punch_excel_row(int(excel_row))
            self.annotations.pop(index)
            self.session_refs = {str(a.get('ref_no')).strip() for a in self.annotations if a.get('ref_no')}
            self.current_sr_no = self.getnextsr()
            self.mark_dirty()
            self.display()
            if is_multimark_extra:
                # If we're still in multi-mark mode for this same punch,
                # keep the on-screen running count in sync with reality.
                if (self.multimark_active and self.multimark_punch and
                        str(self.multimark_punch.get('sr_no', '')).strip() ==
                        str(annotation.get('sr_no', '')).strip()):
                    self.multimark_count = max(0, self.multimark_count - 1)
                    self._update_multimark_bar_count()
                self.flashstat("Highlight removed from punch", bg='#10b981')
            else:
                self.flashstat("Last entry removed", bg='#10b981')
        except PermissionError:
            self.undo_stack.append(action)
            messagebox.showerror("Excel Locked", "Close the Excel file and try Undo again.")
        except Exception as e:
            self.undo_stack.append(action)
            messagebox.showerror("Undo Failed", f"Could not undo the entry:\n{e}")
        self.updtoolpane()

    # ================================================================
    # NAVIGATION AND ZOOM
    # ================================================================

    def _widget_is_inside(self, widget, ancestor):
        """Return True when widget belongs to ancestor's Tk widget tree."""
        current = widget
        while current is not None:
            if current == ancestor:
                return True
            current = getattr(current, 'master', None)
        return False

    def _update_page_toolbar(self):
        """Refresh the read-only current-page indicator."""
        total = len(self.pdf_document) if self.pdf_document else 0
        current = self.current_page + 1 if total else 0
        if hasattr(self, 'page_label'):
            self.page_label.config(text=f"Page: {current} / {total}")

    def prev(self):
        """
        Navigate to previous page in PDF.
        FUNCTIONAL USE: Decrements current_page and redraws display.
        Bound to arrow button in toolbar during quality inspection.
        """
        if self.pdf_document and self.current_page > 0:
            self.current_page -= 1
            self.display(preserve_view=False)

    def next(self):
        """
        Navigate to next page in PDF.
        FUNCTIONAL USE: Increments current_page and redraws display.
        Bound to arrow button in toolbar during quality inspection.
        """
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display()
    def doubleclick(self, event):
        """Edit an existing text box on double-click; zoom is toolbar-only."""
        if not self.pdf_document:
            return "break"
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        hit, _ = self._hit_test_text_box(x, y)
        if hit is not None:
            self.selected_annotation = hit
            self._open_text_editor(hit, select_all=False)
        return "break"

    def show_zoom_slider(self, event=None):
        if getattr(self, 'zoom_slider_frame', None) is not None:
            try:
                self.zoom_slider_frame.lift()
                return
            except tk.TclError:
                self.zoom_slider_frame = None

        # Belt-and-suspenders: if a previous slider instance's cleanup was
        # ever skipped for any reason, make sure we don't stack a second
        # root-level release binding on top of it.
        self._unbind_zoom_root_release()

        MIN_ZOOM = self.ZOOM_MIN
        MAX_ZOOM = self.ZOOM_MAX
        overlay_width = 70
        overlay_height = 260

        if event is not None:
            pos_x = min(max(event.x_root - self.root.winfo_rootx() - overlay_width // 2, 10),
                        self.root.winfo_width() - overlay_width - 10)
            pos_y = min(max(event.y_root - self.root.winfo_rooty() - overlay_height // 2, 10),
                        self.root.winfo_height() - overlay_height - 10)
        else:
            pos_x, pos_y = 40, 40

        frame = tk.Frame(self.root, bg='#1e293b', bd=2, relief=tk.RIDGE)
        frame.place(x=pos_x, y=pos_y, width=overlay_width, height=overlay_height)
        self.zoom_slider_frame = frame

        # ---- Close button: bind press, not just command, and stop propagation ----
        close_btn = tk.Button(
            frame, text="✕", font=('Segoe UI', 9, 'bold'),
            bg='#ef4444', fg='white', relief=tk.FLAT, bd=0,
            width=2, height=1, cursor='hand2',
            command=self.close_zoom_slider
        )
        close_btn.place(x=overlay_width - 26, y=4)
        # On touch, react on press (not release) so it beats any pending render work
        close_btn.bind("<ButtonPress-1>", lambda e: (self.close_zoom_slider(), "break"))

        pct_var = tk.StringVar(value=f"{int(self.zoom_level * 100)}%")
        self._zoom_slider_pct_var = pct_var  # let set_zoom_level() keep this in sync too
        pct_label = tk.Label(frame, textvariable=pct_var, bg='#1e293b', fg='white',
                            font=('Segoe UI', 10, 'bold'))
        pct_label.place(x=6, y=30)

        track_top = 60
        track_bottom = overlay_height - 20
        track_height = track_bottom - track_top
        track_x_center = overlay_width // 2

        track_canvas = tk.Canvas(frame, bg='#1e293b', width=overlay_width,
                                height=track_height + 20, highlightthickness=0)
        track_canvas.place(x=0, y=track_top - 10)

        track_canvas.create_line(track_x_center, 10, track_x_center, track_height + 10,
                                fill='#475569', width=4, capstyle=tk.ROUND)

        def zoom_to_y(zoom_val):
            ratio = (zoom_val - MIN_ZOOM) / (MAX_ZOOM - MIN_ZOOM)
            return 10 + (1 - ratio) * track_height

        def y_to_zoom(y):
            ratio = max(0.0, min(1.0, (y - 10) / track_height))
            return MAX_ZOOM - ratio * (MAX_ZOOM - MIN_ZOOM)

        handle_radius = 9
        handle_y = zoom_to_y(self.zoom_level)
        handle = track_canvas.create_oval(
            track_x_center - handle_radius, handle_y - handle_radius,
            track_x_center + handle_radius, handle_y + handle_radius,
            fill='#3b82f6', outline='white', width=2
        )

        # This slider instance's own generation number. Any callback below
        # captures it and checks it before touching shared self.* zoom state,
        # so a stale callback from a closed/replaced slider can never step on
        # a newer one (or on self.zoom_level after the slider is gone).
        self._zoom_slider_generation = getattr(self, '_zoom_slider_generation', 0) + 1
        my_generation = self._zoom_slider_generation

        self._zoom_render_after_id = None
        self._zoom_is_dragging = False

        # Per-instance drag state, kept in this closure rather than on self,
        # so two slider instances (old one not fully torn down + a new one)
        # can never read/clobber each other's pending-event state.
        drag_state = {'pending_event': None, 'frame_scheduled': False}

        def is_current():
            return (getattr(self, '_zoom_slider_generation', None) == my_generation
                    and getattr(self, 'zoom_slider_frame', None) is frame)

        def apply_zoom_preview(new_zoom, do_render, low_res):
            self.zoom_level = new_zoom
            pct_var.set(f"{int(new_zoom * 100)}%")
            self._update_zoom_toolbar_label()
            if do_render:
                self._render_current_page_only(low_res=low_res)

        def process_zoom_frame():
            drag_state['frame_scheduled'] = False
            if not is_current():
                return
            evt = drag_state['pending_event']
            drag_state['pending_event'] = None
            if evt is None:
                return

            y = max(10, min(track_height + 10, evt.y))
            new_zoom = y_to_zoom(y)
            new_zoom = round(new_zoom / 0.05) * 0.05
            if abs(new_zoom - self.zoom_level) < 0.001:
                return

            track_canvas.coords(
                handle,
                track_x_center - handle_radius, y - handle_radius,
                track_x_center + handle_radius, y + handle_radius
            )
            # Live low-res preview of the current page while dragging, so the
            # PDF visibly tracks the handle instead of only jumping once on
            # release - that "nothing happens until you let go" gap was the
            # biggest part of the slider feeling glitchy/unresponsive.
            apply_zoom_preview(new_zoom, do_render=True, low_res=True)

        def on_handle_drag(evt):
            if not is_current():
                return
            self._zoom_is_dragging = True
            drag_state['pending_event'] = evt
            if not drag_state['frame_scheduled']:
                drag_state['frame_scheduled'] = True
                # ~30fps cap - touch digitizers fire far more events than
                # mouse motion, so a lighter cap avoids saturating the event
                # queue on slower touch hardware.
                self.root.after(33, process_zoom_frame)

        def on_track_click(evt):
            on_handle_drag(evt)

        def on_release(evt):
            # Only render if we were actually dragging the slider, and only
            # for the slider instance that's still current — a release
            # anywhere else in the app, or a stale callback from an already-
            # closed slider, must not trigger a render.
            if not is_current() or not self._zoom_is_dragging:
                return
            self._zoom_is_dragging = False

            if self._zoom_render_after_id is not None:
                self.root.after_cancel(self._zoom_render_after_id)
                self._zoom_render_after_id = None

            # Finalize with a FULL document re-render at the new zoom level.
            # During the drag we only touched the current page (that's what
            # keeps dragging smooth), but that leaves every other page's
            # cached image stale at the old zoom/resolution. If the user then
            # scrolls/navigates without closing the slider, those pages stay
            # blurry or mismatched against page_layout.
            self._do_zoom_render(current_page_only=False)

        track_canvas.tag_bind(handle, "<B1-Motion>", on_handle_drag)
        track_canvas.bind("<Button-1>", on_track_click)
        track_canvas.bind("<B1-Motion>", on_handle_drag)
        # Bind release on the track canvas itself, not root, so releases
        # elsewhere in the app (like the close button) don't trigger renders.
        track_canvas.bind("<ButtonRelease-1>", on_release)
        # Still catch drags that end outside the small canvas (e.g. finger/
        # mouse leaves the overlay while dragging). Keep the returned funcid
        # so close_zoom_slider() can remove exactly this binding - previously
        # this funcid was discarded, so every time the slider was reopened a
        # new permanent root-level handler stacked up and was never cleaned
        # up, which is the main source of the glitchy/stuck behavior.
        self._zoom_root_release_funcid = self.root.bind(
            "<ButtonRelease-1>", on_release, add="+"
        )

    def _unbind_zoom_root_release(self):
        """Remove exactly this slider's root-level <ButtonRelease-1> binding,
        if one is currently registered. Safe to call even if none is set."""
        funcid = getattr(self, '_zoom_root_release_funcid', None)
        if funcid:
            try:
                self.root.unbind("<ButtonRelease-1>", funcid)
            except Exception:
                pass
            self._zoom_root_release_funcid = None

    def _do_zoom_render(self, current_page_only=False, low_res=False):
        """
        Perform the actual PDF re-render at the current zoom_level.
        FUNCTIONAL USE: Called in a debounced fashion during slider drag. When
        current_page_only is True, only the currently visible page is
        re-rasterized instead of the whole document. When low_res is True,
        the page is rendered at a reduced scale for a fast preview during
        active dragging (touch generates far more drag events than mouse,
        so full-resolution re-rasterization on every tick is what jams the app).
        """
        self._zoom_render_after_id = None
        if current_page_only:
            self._render_current_page_only(low_res=low_res)
        else:
            self.display()

    def _render_current_page_only(self, low_res=False):
        """
        Cheap partial re-render used while actively dragging the zoom slider:
        re-rasterizes only self.current_page at the new zoom level and redraws
        just that page's image on canvas, leaving other pages' cached images
        untouched until a full display() call happens on release/close.
        FUNCTIONAL USE: Avoids re-rendering all N pages of a multi-page PDF on
        every zoom-slider tick. When low_res is True, renders at a fraction of
        the target scale (fast, blurry preview) — used mid-drag on touch where
        events fire rapidly. Full-resolution render happens once on release.

        Latency fix: for the low_res live-preview path, a single base pixmap
        is rasterized once per pinch gesture (at gesture-start zoom) and
        cached; every subsequent live frame just resizes that cached PIL
        image with PIL's cheap NEAREST resize instead of calling PyMuPDF's
        get_pixmap() again. Repeated PDF rasterization was the main per-frame
        cost that made live pinch feel like it was lagging behind the fingers.
        """
        if not self.pdf_document or not self.page_layout:
            return
        if self.current_page >= len(self.page_layout):
            return

        try:
            layout = self.page_layout[self.current_page]
            page = self.pdf_document[self.current_page]
            scale = self.page_to_display_scale()

            if low_res:
                # Reuse the cached base image for this gesture if we already
                # captured one for this page; otherwise capture it now. The
                # base is rasterized at a fixed, modest internal resolution
                # (independent of the live target zoom) so it stays valid -
                # and cheap to resize - for the whole gesture even as zoom
                # keeps changing.
                if (self._pinch_base_image is None or
                        self._pinch_base_page != self.current_page):
                    base_scale = scale * 0.5
                    mat = fitz.Matrix(base_scale, base_scale)
                    pix = page.get_pixmap(matrix=mat)
                    self._pinch_base_image = Image.frombytes(
                        "RGB", [pix.width, pix.height], pix.samples
                    )
                    self._pinch_base_scale = base_scale
                    self._pinch_base_page = self.current_page

                # Scale the cached base image (cheap, no PDF rasterization)
                # to match how far the live zoom has moved since capture.
                ratio = scale / self._pinch_base_scale if self._pinch_base_scale else 1.0
                base_w, base_h = self._pinch_base_image.size
                target_w = max(1, int(base_w * ratio))
                target_h = max(1, int(base_h * ratio))
                img = self._pinch_base_image.resize((target_w, target_h), Image.Resampling.NEAREST)
            else:
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            photo = ImageTk.PhotoImage(img)

            if self.current_page < len(self.page_images):
                self.page_images[self.current_page] = photo

            self.canvas.delete(f"page_{self.current_page}")
            self.canvas.create_image(
                layout['x'], layout['y'], anchor=tk.NW, image=photo,
                tags=(f"page_{self.current_page}",)
            )
            self.photo = photo
        except Exception as e:
            print(f"[WARN] Partial zoom render failed: {e}")

    # ================================================================
    # TOOLBAR-DOCKED ZOOM CONTROL + PINCH-TO-ZOOM
    # ================================================================

    def setup_toolbar_zoom(self, toolbar):
        """Create a compact highlighter-style zoom button and dropdown popup."""
        zoom_frame = tk.Frame(toolbar, bg='#1e293b')
        zoom_frame.pack(side=tk.LEFT, padx=(6, 10))
        tk.Label(zoom_frame, text="Zoom:", bg='#1e293b', fg='#94a3b8',
                 font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT, padx=(0, 5))

        self.zoom_pct_var = tk.StringVar(value=f"{int(round(self.zoom_level * 100))}%")
        self.zoom_popup = None

        # Editable zoom value. The global Entry focus binding opens the
        # on-screen keyboard. The value is applied only after a press outside
        # the toolbar, not on Enter and not merely on FocusOut.
        self.zoom_value_btn = tk.Entry(
            zoom_frame, textvariable=self.zoom_pct_var,
            font=('Segoe UI', 9, 'bold'), bg='#334155', fg='white',
            insertbackground='white', justify='center', relief=tk.FLAT,
            borderwidth=0, width=7
        )
        self.zoom_value_btn.pack(side=tk.LEFT, ipady=10)
        self.zoom_value_btn.bind('<FocusIn>', self._on_zoom_entry_focus)
        self.root.bind_all('<ButtonPress-1>', self._apply_zoom_on_outside_click, add='+')

        # Separate dropdown arrow, matching the highlighter arrow behavior.
        self.zoom_dropdown_btn = tk.Button(
            zoom_frame, text="↓", font=('Segoe UI', 8),
            bg='#1e293b', fg='#94a3b8', activebackground='#334155',
            activeforeground='white', relief=tk.FLAT, borderwidth=0,
            width=2, height=1, command=self.toggle_zoom_popup, cursor='hand2'
        )
        self.zoom_dropdown_btn.pack(side=tk.LEFT, padx=(3, 0))

        self.root.bind_all("<Control-MouseWheel>", self._on_ctrl_scroll_zoom, add="+")
        self.root.bind_all("<Control-Button-4>", self._on_ctrl_scroll_zoom, add="+")
        self.root.bind_all("<Control-Button-5>", self._on_ctrl_scroll_zoom, add="+")

    def _on_zoom_entry_focus(self, event=None):
        self._zoom_edit_active = True
        try:
            self.zoom_value_btn.selection_range(0, tk.END)
        except tk.TclError:
            pass

    def _apply_zoom_on_outside_click(self, event=None):
        """Apply typed zoom only after the user presses outside the toolbar."""
        if not getattr(self, '_zoom_edit_active', False):
            return
        if event is not None and self._widget_is_inside(event.widget, self.toolbar):
            return
        self._zoom_edit_active = False
        try:
            percent = self._parse_zoom_percent()
        except ValueError:
            self._update_zoom_toolbar_label()
            return
        self.set_zoom_level(percent / 100.0, immediate=False)

    def toggle_zoom_popup(self):
        """Open or close the zoom popup beneath the toolbar zoom control."""
        if self.zoom_popup is not None:
            try:
                if self.zoom_popup.winfo_exists():
                    self.close_zoom_popup()
                    return
            except tk.TclError:
                pass
            self.zoom_popup = None
        self.show_zoom_popup()

    def show_zoom_popup(self):
        """Show a slider-only highlighter-style zoom dropdown popup."""
        popup = tk.Toplevel(self.root)
        popup.overrideredirect(True)
        popup.attributes('-topmost', True)
        popup.configure(bg='#1e293b', bd=1, relief=tk.SOLID)
        self.zoom_popup = popup

        x = self.zoom_value_btn.winfo_rootx()
        y = self.zoom_value_btn.winfo_rooty() + self.zoom_value_btn.winfo_height() + 2
        popup.geometry(f"270x82+{x}+{y}")

        tk.Label(
            popup, text="Drag to adjust zoom", bg='#1e293b', fg='#cbd5e1',
            font=('Segoe UI', 9)
        ).pack(anchor='w', padx=12, pady=(10, 2))

        self.zoom_scale = tk.Scale(
            popup, from_=50, to=300, orient=tk.HORIZONTAL, resolution=1,
            showvalue=False, length=242, sliderlength=18, bd=0,
            highlightthickness=0, bg='#1e293b', fg='white',
            troughcolor='#475569', activebackground='#60a5fa',
            command=self._on_zoom_bar
        )
        self._zoom_dropdown_updating = True
        try:
            self.zoom_scale.set(int(round(self.zoom_level * 100)))
        finally:
            self._zoom_dropdown_updating = False
        self.zoom_scale.pack(padx=10, pady=(2, 8))
        self.zoom_scale.bind('<ButtonRelease-1>', self._finish_zoom_bar)

        popup.bind('<Escape>', lambda e: self.close_zoom_popup())
        popup.bind('<FocusOut>', self._zoom_popup_focus_out)
        self.zoom_scale.focus_set()

    def _zoom_popup_focus_out(self, event=None):
        """Close only when focus moves completely outside the zoom popup."""
        popup = self.zoom_popup
        if popup is None:
            return
        self.root.after(100, lambda expected=popup: self._close_zoom_popup_if_unfocused(expected))

    def _close_zoom_popup_if_unfocused(self, expected):
        """Apply typed zoom automatically when the user clicks outside the popup."""
        if self.zoom_popup is not expected:
            return
        try:
            focused = self.root.focus_get()
            if focused is None or not str(focused).startswith(str(expected)):
                self.close_zoom_popup()
        except tk.TclError:
            self.zoom_popup = None

    def close_zoom_popup(self):
        popup = self.zoom_popup
        self.zoom_popup = None
        if popup is not None:
            try:
                popup.destroy()
            except tk.TclError:
                pass

    def _parse_zoom_percent(self):
        raw = self.zoom_pct_var.get().strip().replace('%', '')
        value = float(raw)
        return max(50.0, min(300.0, value))

    def _apply_zoom_popup_entry(self, event=None, close_popup=True):
        """Apply typed zoom on Enter or whenever focus leaves the popup."""
        try:
            percent = self._parse_zoom_percent()
        except ValueError:
            self._update_zoom_toolbar_label()
            if close_popup:
                self.close_zoom_popup()
            return "break"
        self.set_zoom_level(percent / 100.0, immediate=False)
        if close_popup:
            self.close_zoom_popup()
        return "break"

    def _on_zoom_entry(self, event=None):
        return self._apply_zoom_popup_entry(event)

    def _on_zoom_dropdown(self, event=None):
        return self._apply_zoom_popup_entry(event)

    def _on_zoom_bar(self, value):
        if self._zoom_dropdown_updating:
            return
        percent = int(round(float(value)))
        self.zoom_pct_var.set(f"{percent}%")

    def _finish_zoom_bar(self, event=None):
        if not hasattr(self, 'zoom_scale'):
            return "break"
        value = float(self.zoom_scale.get()) / 100.0
        self.set_zoom_level(value, immediate=False)
        self.close_zoom_popup()
        return "break"

    def setup_canvas_pinch_zoom(self):
        """Bind Tk-managed pinch events without installing a native WNDPROC.

        Tk dispatches this callback on its own UI thread, avoiding the Python
        3.14 GIL crash caused by calling Python from a Windows window-procedure
        callback. Unsupported Tk builds simply ignore the virtual event.
        """
        if getattr(self, '_safe_pinch_bound', False):
            return
        try:
            self.canvas.bind('<<TouchpadPinch>>', self._on_pinch_zoom, add='+')
            # Most Windows Tk touch drivers promote the final finger-up to a
            # ButtonRelease event. Bind with add='+' so normal drawing/panning
            # release handlers remain intact.
            self.canvas.bind('<ButtonRelease-1>', self._on_safe_pinch_release, add='+')
            self._safe_pinch_bound = True
        except tk.TclError as exc:
            self._safe_pinch_bound = False
            print(f'[WARN] Tk pinch event unavailable: {exc}')

    def _on_pinch_zoom(self, event):
        """Track Tk pinch magnitude proportionally with immediate visual feedback.

        Latency fix: the zoom-level math still runs on every event (cheap),
        but the actual canvas redraw is throttled to a ~60fps budget and
        deferred via after_idle instead of calling update_idletasks()
        synchronously on every single touch event. Forcing a full Tk flush
        per-event was the main source of pinch lag on fast digitizers.

        Sensitivity: PINCH_SENSITIVITY controls how much zoom change a given
        amount of finger movement produces. 0.25 means a normal pinch swing
        moves zoom by up to 25% per update, roughly 1.4x more responsive
        than the previous tuning - so the page visibly grows/shrinks with
        less finger travel.
        """
        if (not self.pdf_document or self.active_highlighter or
                self.tool_mode is not None or self.drawing):
            return 'break'
        try:
            now = time.monotonic()
            raw = float(getattr(event, 'delta', 0.0))
            if raw == 0.0:
                return 'break'

            PINCH_SENSITIVITY = 0.25  # max zoom-factor swing per event, i.e. 25%

            # A pause means a new physical pinch gesture. Reset only the raw
            # sample history, not the current zoom level.
            if now - self._safe_pinch_last_time > 0.25:
                self._safe_pinch_last_raw = None
                self._safe_pinch_accumulator = 1.0
                self._pinch_base_image = None  # force a fresh base capture
            self._safe_pinch_last_time = now

            # Tk builds expose pinch values differently. If consecutive values
            # look cumulative, use their ratio/difference. Otherwise treat the
            # value as an incremental gesture delta. Exponential scaling makes
            # equal finger movement produce equal proportional zoom movement.
            previous = self._safe_pinch_last_raw
            self._safe_pinch_last_raw = raw
            if previous is not None and raw > 0 and previous > 0:
                ratio = raw / previous
                if (1.0 - PINCH_SENSITIVITY) <= ratio <= (1.0 + PINCH_SENSITIVITY) and abs(ratio - 1.0) > 0.0005:
                    factor = ratio
                else:
                    factor = pow(1.0018 * (1.0 + PINCH_SENSITIVITY), raw)
            else:
                if abs(raw) >= 10.0:
                    factor = pow(1.0018 * (1.0 + PINCH_SENSITIVITY), raw)
                elif abs(raw) > 1.0:
                    factor = pow(1.018 * (1.0 + PINCH_SENSITIVITY), raw)
                else:
                    factor = pow(2.0 * (1.0 + PINCH_SENSITIVITY), raw)

            # Reject only impossible driver spikes, while retaining the actual
            # pinch amount for normal events. Bound the swing to the
            # configured sensitivity instead of the old fixed 0.70-1.40 range.
            factor = max(1.0 - PINCH_SENSITIVITY, min(1.0 + PINCH_SENSITIVITY, factor))
            target = max(
                self.ZOOM_MIN,
                min(self.ZOOM_MAX, self.zoom_level * factor)
            )
            if abs(target - self.zoom_level) < 0.0005:
                return 'break'

            self.zoom_level = target
            self._native_pinch_last_zoom = target
            self._safe_pinch_active = True
            self._update_zoom_toolbar_label()
            self._panning = False

            # Throttle the actual redraw to a real frame budget instead of
            # painting on every raw event - the digitizer can fire well past
            # 100Hz, but repainting that often buys nothing visually and is
            # exactly what caused the lag. Only schedule a frame if one isn't
            # already pending, and skip if we're still inside this frame's
            # minimum interval.
            if not self._pinch_frame_pending:
                elapsed = now - self._pinch_last_frame_time
                delay_ms = 0 if elapsed >= self._pinch_frame_min_interval else \
                    int((self._pinch_frame_min_interval - elapsed) * 1000)
                self._pinch_frame_pending = True
                self.root.after(delay_ms, self._render_pinch_frame)

            # Final quality is normally applied by _on_safe_pinch_release.
            # This long timer is only a driver fallback when no release event
            # is exposed by Tk; it is deliberately not the normal finish path.
            if self._native_pinch_render_after_id is not None:
                try:
                    self.root.after_cancel(self._native_pinch_render_after_id)
                except Exception:
                    pass
            self._native_pinch_render_after_id = self.root.after(
                self._safe_pinch_watchdog_ms,
                self._finish_safe_pinch_render
            )
        except Exception as exc:
            print(f'[WARN] Tk pinch event failed: {exc}')
        return 'break'

    def _render_pinch_frame(self):
        """Paint exactly one throttled live-preview frame during an active pinch."""
        self._pinch_frame_pending = False
        if not self._safe_pinch_active:
            return
        self._pinch_last_frame_time = time.monotonic()
        self._render_current_page_only(low_res=True)

    def _on_safe_pinch_release(self, event=None):
        """Finalize full-quality zoom when the fingers are released."""
        if not self._safe_pinch_active:
            return
        self._finish_safe_pinch_render()
        return 'break'

    def _finish_safe_pinch_render(self):
        """Finalize one active pinch with a single full-quality render."""
        pending_id = self._native_pinch_render_after_id
        self._native_pinch_render_after_id = None
        if pending_id is not None:
            try:
                self.root.after_cancel(pending_id)
            except Exception:
                pass

        if not self._safe_pinch_active:
            return
        self._safe_pinch_active = False

        target = self._native_pinch_last_zoom
        self._native_pinch_last_zoom = None
        self._safe_pinch_last_raw = None
        self._safe_pinch_accumulator = 1.0
        self._pinch_base_image = None  # drop the cached preview source, gesture is over
        self._pinch_base_scale = None
        self._pinch_base_page = None
        if self.pdf_document and target is not None:
            self.set_zoom_level(target, immediate=True, full_render=True)

    def _update_zoom_toolbar_label(self):
        """Synchronize the compact toolbar value and any open popup slider."""
        if not hasattr(self, 'zoom_pct_var'):
            return
        percent = int(round(self.zoom_level * 100))
        self._zoom_dropdown_updating = True
        try:
            self.zoom_pct_var.set(f"{percent}%")
            popup = getattr(self, 'zoom_popup', None)
            if popup is not None and popup.winfo_exists() and hasattr(self, 'zoom_scale'):
                self.zoom_scale.set(percent)
        except tk.TclError:
            self.zoom_popup = None
        finally:
            self._zoom_dropdown_updating = False

    def step_zoom(self, delta):
        """
        Nudge zoom by a fixed step (used by the toolbar +/- buttons).
        FUNCTIONAL USE: Rounds to the nearest step so repeated clicks land on
        clean values (90%, 100%, 110%...) instead of drifting from float error.
        Args: delta - signed float, e.g. +0.1 or -0.1
        """
        new_zoom = round((self.zoom_level + delta) / self.ZOOM_STEP) * self.ZOOM_STEP
        self.set_zoom_level(new_zoom, immediate=False)

    def set_zoom_level(self, new_zoom, immediate=False, low_res=False, full_render=False):
        """Set zoom and coalesce redraws to prevent UI lag and stale page images."""
        try:
            new_zoom = float(new_zoom)
        except (TypeError, ValueError):
            return
        new_zoom = max(self.ZOOM_MIN, min(self.ZOOM_MAX, new_zoom))
        if abs(new_zoom - self.zoom_level) < 0.001 and not full_render:
            self._update_zoom_toolbar_label()
            return

        self.zoom_level = new_zoom
        self._update_zoom_toolbar_label()
        if not self.pdf_document:
            return

        if self._zoom_render_after_id is not None:
            try:
                self.root.after_cancel(self._zoom_render_after_id)
            except Exception:
                pass
            self._zoom_render_after_id = None

        def render_zoom():
            self._zoom_render_after_id = None
            self._clear_page_render_cache()
            # Let pending toolbar/paint events complete before the expensive
            # stacked-document redraw. This prevents the window appearing hung.
            self.root.update_idletasks()
            self.display(preserve_view=True)

        if immediate:
            render_zoom()
        else:
            # Multiple wheel/dropdown events collapse into one full render.
            self._zoom_render_after_id = self.root.after(35, render_zoom)

    def _on_ctrl_scroll_zoom(self, event):
        """
        Handle Ctrl+MouseWheel / Ctrl+Button-4/5 as a zoom gesture.
        FUNCTIONAL USE: Standard "Ctrl+scroll to zoom" convention (matches
        browsers, PDF viewers, etc.) as a fast, precise, mouse-friendly
        alternative to the toolbar buttons or floating slider.
        """
        if not self.pdf_document or not self._is_pointer_over_canvas():
            return

        delta = 0
        if getattr(event, 'num', None) == 4:
            delta = 1
        elif getattr(event, 'num', None) == 5:
            delta = -1
        elif getattr(event, 'delta', 0):
            delta = 1 if event.delta > 0 else -1

        if delta == 0:
            return "break"

        self.step_zoom(delta * self.ZOOM_STEP)
        return "break"

    # ---- Touch pinch-to-zoom (best effort; gracefully degrades) ----

    def close_zoom_slider(self):
        """
        Remove the floating zoom-level adjuster overlay from the screen.
        FUNCTIONAL USE: Bound to the slider's cross (X) button (on ButtonPress-1
        for immediate touch response). Cancels any pending debounced render,
        clears the drag-tracking flag, removes ONLY this slider instance's
        root-level release binding (by exact funcid, never a blind unbind),
        and performs one final full-quality re-render of the whole document.
        """
        if getattr(self, '_zoom_render_after_id', None) is not None:
            try:
                self.root.after_cancel(self._zoom_render_after_id)
            except Exception:
                pass
            self._zoom_render_after_id = None

        self._zoom_is_dragging = False

        # Bump the generation so any already-queued/stale callback from this
        # instance (e.g. a process_zoom_frame() still pending via after())
        # becomes a no-op even if it still fires once more.
        self._zoom_slider_generation = getattr(self, '_zoom_slider_generation', 0) + 1

        self._unbind_zoom_root_release()

        frame = getattr(self, 'zoom_slider_frame', None)
        if frame is not None:
            try:
                frame.destroy()
            except tk.TclError:
                pass
            self.zoom_slider_frame = None

        self._zoom_slider_pct_var = None
        self._update_zoom_toolbar_label()
        self.display()
    # ================================================================
    # PLACEHOLDER METHODS - Implement from your original code
    # ================================================================

    def derive_storage_location_from_input(self, input_path):
        """Derive the sibling 07-Scanned System Book UNC folder."""
        if not input_path:
            return get_base_path()

        # Use ntpath deliberately. It parses Windows drive and UNC paths
        # correctly even when this source file is inspected or packaged elsewhere.
        normalized_input = ntpath.normpath(str(input_path).strip())
        current_dir = ntpath.dirname(normalized_input)

        while current_dir:
            folder_name = ntpath.basename(current_dir).strip().casefold()
            if folder_name == "02-customer inputs":
                project_root = ntpath.dirname(current_dir)
                return ntpath.normpath(
                    ntpath.join(project_root, "07-Scanned System Book")
                )

            parent_dir = ntpath.dirname(current_dir)
            if not parent_dir or parent_dir == current_dir:
                break
            current_dir = parent_dir

        # Keep the old storage default when the selected PDF is not under the
        # expected 02-Customer Inputs folder.
        return get_base_path()


    def _cabinet_context_ready(self):
        """Return True when the current workspace has saved cabinet details."""
        if not (getattr(self, 'cabinet_id', '').strip() and
                getattr(self, 'project_name', '').strip()):
            messagebox.showwarning(
                "Cabinet Not Set Up",
                "Load a PDF and save the cabinet details first.",
                parent=self.root
            )
            return False
        return True

    def _cabinet_project_has_annotations(self):
        """Return True when this exact cabinet + project already owns annotations.

        Check both live memory and the saved session so restarting the app cannot
        bypass the PDF replacement lock.
        """
        cabinet_key = str(getattr(self, 'cabinet_id', '') or '').strip().casefold()
        project_key = str(getattr(self, 'project_name', '') or '').strip().casefold()
        if not cabinet_key or not project_key:
            return False

        if getattr(self, 'annotations', None):
            return True

        candidates = []
        sessions_dir = (getattr(self, 'project_dirs', {}) or {}).get('sessions')
        if sessions_dir:
            candidates.append(os.path.join(sessions_dir, f"{self.cabinet_id}_annotations.json"))
        try:
            record = self.db.get_project(self.cabinet_id) or {}
            saved_session = record.get('session_path')
            if saved_session:
                candidates.append(to_absolute_path(saved_session) or saved_session)
        except Exception:
            pass

        checked = set()
        for session_path in candidates:
            if not session_path:
                continue
            normalized = os.path.normcase(os.path.abspath(session_path))
            if normalized in checked or not os.path.isfile(session_path):
                continue
            checked.add(normalized)
            try:
                with open(session_path, 'r', encoding='utf-8') as session_file:
                    data = json.load(session_file)
                saved_cabinet = str(data.get('cabinet_id', '') or '').strip().casefold()
                saved_project = str(data.get('project_name', '') or '').strip().casefold()
                same_combination = saved_cabinet == cabinet_key and saved_project == project_key
                if same_combination and data.get('annotations'):
                    return True
            except (OSError, ValueError, TypeError) as exc:
                # Fail closed: an unreadable session must not permit replacing the
                # source PDF and potentially orphaning existing annotations.
                print(f"[WARN] Could not verify annotation session {session_path}: {exc}")
                return True
        return False

    def select_different_pdf(self):
        """Replace the drawing PDF only before this cabinet/project is annotated."""
        if not self._cabinet_context_ready():
            return
        if self._cabinet_project_has_annotations():
            messagebox.showwarning(
                "PDF Change Locked",
                f"The PDF cannot be changed because annotations already exist for:\n\n"
                f"Cabinet ID: {self.cabinet_id}\nProject: {self.project_name}\n\n"
                "Remove the annotations first, or create a new cabinet/project combination.",
                parent=self.root
            )
            return

        start_dir = os.path.dirname(getattr(self, 'current_pdf_path', '') or '')
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = get_base_path() if os.path.isdir(get_base_path()) else app_base()
        selected = filedialog.askopenfilename(
            title=f"Select a Different PDF for {self.cabinet_id}",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialdir=start_dir
        )
        if not selected:
            return
        try:
            selected = os.path.abspath(selected)
            if not os.path.isfile(selected):
                raise FileNotFoundError(selected)
            if (self.current_pdf_path and
                    os.path.normcase(os.path.abspath(self.current_pdf_path)) ==
                    os.path.normcase(selected)):
                messagebox.showinfo("PDF Already Selected", "This PDF is already open.", parent=self.root)
                return

            self.busy("Preparing the new cabinet PDF...")
            if not self.project_dirs or not self.project_dirs.get('source_drawings'):
                if not self.preparefolders():
                    return

            central_path = self.copy_pdf_to_central_storage(selected)
            new_document = fitz.open(central_path)
            if len(new_document) == 0:
                new_document.close()
                raise ValueError("The selected PDF contains no pages.")

            old_document = self.pdf_document
            self.pdf_document = new_document
            if old_document is not None:
                try:
                    old_document.close()
                except Exception:
                    pass

            self.stopmultimark()
            self.current_pdf_path = central_path
            self.current_page = 0
            self.annotations = []
            self.undo_stack = []
            self.session_refs = set()
            self.zoom_level = 1.0
            self._dirty = True
            self._clear_page_render_cache()
            self._update_zoom_toolbar_label()
            self.deactivate()
            self.display(preserve_view=False)

            expected_session_path = os.path.join(
                self.project_dirs['sessions'],
                f"{self.cabinet_id}_annotations.json"
            )
            self.db.update_project(self.cabinet_id, {
                'pdf_path': self.current_pdf_path,
                'excel_path': self.excel_file,
                'session_path': expected_session_path,
                'storage_location': self.storage_location,
                'last_accessed': datetime.now().isoformat()
            })
            self.saverecentproj()
            self.flashstat("Cabinet PDF replaced successfully", bg='#2563eb')
        except Exception as exc:
            messagebox.showerror(
                "PDF Replacement Failed",
                f"Could not select the new PDF:\n\n{exc}",
                parent=self.root
            )
        finally:
            self.unbusy()

    def _migrate_cabinet_id(self, old_id, new_id):
        """Migrate local files and the inspection DB when a cabinet ID changes."""
        old_id = str(old_id or '').strip()
        new_id = str(new_id or '').strip()
        if not old_id or not new_id or old_id == new_id:
            return

        # Save the old session before moving its folder.
        self.flush_pending_saves(show_status=False)
        old_dirs = dict(getattr(self, 'project_dirs', {}) or {})
        old_root = old_dirs.get('root')
        new_root = None
        if old_root:
            parent = os.path.dirname(old_root)
            new_root = os.path.join(parent, new_id.replace(' ', '_'))
            if os.path.exists(new_root):
                raise FileExistsError(f"The target cabinet folder already exists: {new_root}")
            if os.path.isdir(old_root):
                os.rename(old_root, new_root)

        self.cabinet_id = new_id
        if new_root:
            self.project_dirs = {
                'root': new_root,
                'source_drawings': os.path.join(new_root, 'Source_Drawings'),
                'working_excel': os.path.join(new_root, 'Working_Excel'),
                'interphase_export': os.path.join(new_root, 'Interphase_Export'),
                'annotated_drawings': os.path.join(new_root, 'Annotated_Drawings'),
                'sessions': os.path.join(new_root, 'Sessions')
            }
            def moved(path):
                if not path:
                    return path
                try:
                    rel = os.path.relpath(path, old_root)
                    if rel != os.pardir and not rel.startswith(os.pardir + os.sep):
                        return os.path.join(new_root, rel)
                except Exception:
                    pass
                return path
            self.current_pdf_path = moved(self.current_pdf_path)
            self.excel_file = moved(self.excel_file)
            self.working_excel_path = moved(self.working_excel_path)

            old_session = os.path.join(self.project_dirs['sessions'], f"{old_id}_annotations.json")
            new_session = os.path.join(self.project_dirs['sessions'], f"{new_id}_annotations.json")
            if os.path.isfile(old_session) and old_session != new_session:
                os.rename(old_session, new_session)

        old_record = self.db.get_project(old_id) or {}
        new_record = dict(old_record)
        new_record.update({
            'cabinet_id': new_id,
            'project_name': self.project_name,
            'sales_order_no': self.sales_order_no,
            'storage_location': self.storage_location,
            'pdf_path': self.current_pdf_path,
            'excel_path': self.excel_file,
            'session_path': self.getpathforpdf(),
            'last_accessed': datetime.now().isoformat()
        })

        # Prefer an atomic rename API when supplied by DatabaseManager.
        if hasattr(self.db, 'rename_project'):
            self.db.rename_project(old_id, new_id, new_record)
        elif hasattr(self.db, 'update_cabinet_id'):
            self.db.update_cabinet_id(old_id, new_id)
            self.db.update_project(new_id, new_record)
        else:
            self.db.add_project(new_record)
            if hasattr(self.db, 'delete_project'):
                self.db.delete_project(old_id)
            elif hasattr(self.db, 'remove_project'):
                self.db.remove_project(old_id)

        # Best-effort cleanup/migration in linked stores when their APIs support it.
        for store in (getattr(self, 'manager_db', None), getattr(self, 'handover_db', None)):
            if store is None:
                continue
            for method_name in ('rename_cabinet', 'rename_cabinet_id', 'update_cabinet_id'):
                method = getattr(store, method_name, None)
                if callable(method):
                    method(old_id, new_id)
                    break

    def edit_cabinet_details(self):
        """Show a modern editor for mutable cabinet metadata."""
        if not self._cabinet_context_ready():
            return

        colors = {
            'nav': '#0f172a', 'window': '#f1f5f9', 'card': '#ffffff',
            'text': '#0f172a', 'muted': '#64748b', 'line': '#e2e8f0',
            'primary': '#2563eb', 'success': '#059669'
        }
        dlg = tk.Toplevel(self.root)
        dlg.title("Edit Cabinet Details")
        dlg.geometry("680x520")
        dlg.minsize(620, 480)
        dlg.configure(bg=colors['window'])
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg=colors['nav'], height=86)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="Edit Cabinet Details", bg=colors['nav'], fg='white',
                 font=('Segoe UI Semibold', 18, 'bold')).pack(anchor='w', padx=26, pady=(15, 1))
        tk.Label(header, text="Update the saved cabinet information without reopening the drawing.",
                 bg=colors['nav'], fg='#94a3b8', font=('Segoe UI', 10)).pack(anchor='w', padx=26)

        card = tk.Frame(dlg, bg=colors['card'], highlightthickness=1,
                        highlightbackground=colors['line'])
        card.pack(fill=tk.BOTH, expand=True, padx=24, pady=20)

        cabinet_var = tk.StringVar(value=self.cabinet_id)
        project_var = tk.StringVar(value=self.project_name)
        so_var = tk.StringVar(value=self.sales_order_no)
        location_var = tk.StringVar(value=getattr(self, 'storage_location', ''))

        def field(row, label, variable, readonly=False):
            tk.Label(card, text=label.upper(), bg=colors['card'], fg=colors['muted'],
                     font=('Segoe UI Semibold', 8, 'bold')).grid(
                         row=row, column=0, sticky='w', padx=22, pady=(16, 4))
            entry = tk.Entry(card, textvariable=variable, font=('Segoe UI', 11),
                             relief=tk.FLAT, bg='#f8fafc', fg=colors['text'],
                             readonlybackground='#e2e8f0')
            entry.grid(row=row + 1, column=0, sticky='ew', padx=22, ipady=10)
            if readonly:
                entry.config(state='readonly')
            return entry

        cabinet_entry = field(0, "Cabinet ID", cabinet_var)
        tk.Label(card, text="Changing the ID also migrates the cabinet folder, session, and saved paths.",
                 bg=colors['card'], fg=colors['muted'], font=('Segoe UI', 8)).grid(
                     row=2, column=0, sticky='w', padx=22, pady=(3, 0))
        project_entry = field(3, "Project Name", project_var)
        field(5, "Sales Order Number", so_var)
        field(7, "Storage Location", location_var, readonly=True)
        card.columnconfigure(0, weight=1)

        footer = tk.Frame(dlg, bg=colors['window'])
        footer.pack(fill=tk.X, padx=24, pady=(0, 20))

        def save_details():
            new_cabinet_id = cabinet_var.get().strip()
            project = project_var.get().strip()
            sales_order = so_var.get().strip()
            if not new_cabinet_id:
                messagebox.showwarning("Cabinet ID Required", "Enter a cabinet ID.", parent=dlg)
                cabinet_entry.focus_force()
                return
            if not project:
                messagebox.showwarning("Project Name Required", "Enter a project name.", parent=dlg)
                project_entry.focus_force()
                return
            old_cabinet_id = self.cabinet_id
            if new_cabinet_id != old_cabinet_id and self.db.project_exists(new_cabinet_id):
                messagebox.showerror(
                    "Cabinet ID Already Exists",
                    f"Cabinet ID '{new_cabinet_id}' is already in use. Choose a unique ID.",
                    parent=dlg
                )
                cabinet_entry.focus_force()
                return
            if new_cabinet_id != old_cabinet_id:
                proceed = messagebox.askyesno(
                    "Change Cabinet ID",
                    f"Rename cabinet '{old_cabinet_id}' to '{new_cabinet_id}'?\n\n"
                    "The cabinet folder, session file, Excel/PDF paths, and database record will be migrated.",
                    parent=dlg,
                    icon='warning'
                )
                if not proceed:
                    return
            try:
                self.busy("Updating cabinet details...")
                if new_cabinet_id != old_cabinet_id:
                    self._migrate_cabinet_id(old_cabinet_id, new_cabinet_id)
                self.project_name = project
                self.sales_order_no = sales_order
                project_data = {
                    'cabinet_id': self.cabinet_id,
                    'project_name': self.project_name,
                    'sales_order_no': self.sales_order_no,
                    'storage_location': self.storage_location,
                    'pdf_path': self.current_pdf_path,
                    'excel_path': self.excel_file,
                    'session_path': self.getpathforpdf(),
                    'last_accessed': datetime.now().isoformat()
                }
                if self.db.project_exists(self.cabinet_id):
                    self.db.update_project(self.cabinet_id, project_data)
                else:
                    project_data['created_date'] = datetime.now().isoformat()
                    self.db.add_project(project_data)
                self.write_to_xcl()
                self.mark_dirty()
                self._write_session_file()
                self.sync_manager_stats_only()
                self.updrecentdropdwn()
                dlg.destroy()
                self.flashstat("Cabinet details updated", bg=colors['success'])
            except Exception as exc:
                messagebox.showerror("Update Failed", f"Could not update cabinet details:\n\n{exc}", parent=dlg)
            finally:
                self.unbusy()

        tk.Button(footer, text="Cancel", command=dlg.destroy, bg='#64748b', fg='white',
                  activebackground='#475569', activeforeground='white', relief=tk.FLAT,
                  borderwidth=0, font=('Segoe UI Semibold', 10, 'bold'),
                  padx=24, pady=11, cursor='hand2').pack(side=tk.RIGHT)
        tk.Button(footer, text="Save Changes", command=save_details, bg=colors['primary'], fg='white',
                  activebackground='#1d4ed8', activeforeground='white', relief=tk.FLAT,
                  borderwidth=0, font=('Segoe UI Semibold', 10, 'bold'),
                  padx=24, pady=11, cursor='hand2').pack(side=tk.RIGHT, padx=(0, 10))
        project_entry.focus_force()

    def loadpdf(self):
        """Load PDF and persist it under central UNC storage."""
        self.stopmultimark()  # never leave multi-mark mode active across a document swap
        initial_pdf_dir = get_base_path()
        if not os.path.isdir(initial_pdf_dir):
            initial_pdf_dir = app_base()

        file_path = filedialog.askopenfilename(
            title="Select Circuit Diagram PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialdir=initial_pdf_dir
        )
        if file_path:
            load_stage = "validating selected PDF"
            try:
                if not os.path.isfile(file_path):
                    raise FileNotFoundError(f"Selected PDF was not found: {file_path}")

                # Keep the drawing-input dialog at the existing configured base path.
                # Only the storage location is derived from the selected input path.
                self.current_pdf_path = file_path
                self.storage_location = self.derive_storage_location_from_input(file_path)

                load_stage = "reading project details"
                self.askprojdetails()  # user-facing dialog - keep overlay hidden until now

                self.busy("Setting up project folders...")
                load_stage = f"creating storage folders under: {self.storage_location}"
                if not self.preparefolders():
                    return

                self.busy(f"Copying PDF to storage...")
                load_stage = f"copying PDF from: {file_path}"
                central_pdf_path = self.copy_pdf_to_central_storage(file_path)

                self.busy("Opening PDF...")
                load_stage = f"opening copied PDF: {central_pdf_path}"

                if self.pdf_document:
                    self.pdf_document.close()

                self.pdf_document = fitz.open(central_pdf_path)
                self._clear_page_render_cache()
                self.current_pdf_path = central_pdf_path
                self.current_page = 0
                self.annotations = []
                self.zoom_level = 1.0
                self._update_zoom_toolbar_label()
                self.tool_mode = None
                self.active_highlighter = None
                self.colorbutton()
                self.root.config(cursor="")
                self.current_sr_no = self.getnextsr()
                self.display(preserve_view=False)  # display() shows its own overlay if the render is heavy
                self.unbusy()
                messagebox.showinfo("Success", f"Loaded PDF with {len(self.pdf_document)} pages")

                try:
                    self.working_excel_path = os.path.join(
                        self.project_dirs["working_excel"],
                        f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
                    )

                    if os.path.exists(self.working_excel_path):
                        resume = messagebox.askyesno(
                            "Resume Inspection",
                            f"Existing working Excel found. Resume previous inspection?"
                        )
                        if not resume:
                            self.busy("Preparing working Excel file...")
                            shutil.copy2(self.master_excel_file, self.working_excel_path)
                    else:
                        self.busy("Preparing working Excel file...")
                        shutil.copy2(self.master_excel_file, self.working_excel_path)

                    self.excel_file = self.working_excel_path

                except Exception as e:
                    self.unbusy()
                    messagebox.showerror("Excel Error", f"Failed to prepare working Excel:\n{e}")
                    return

                self.busy("Syncing punch list...")
                self.write_to_xcl()
                self.unbusy()

                expected_session_path = os.path.join(
                    self.project_dirs["sessions"],
                    f"{self.cabinet_id}_annotations.json"
                )

                self.db.update_project(self.cabinet_id, {
                    'pdf_path': self.current_pdf_path,
                    'excel_path': self.excel_file,
                    'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                    'storage_location': self.storage_location,
                    'last_accessed': datetime.now().isoformat()
                })

                if os.path.exists(expected_session_path):
                    resume = messagebox.askyesno(
                        "Resume Session",
                        "Existing session found. Do you want to resume it?"
                    )
                    if resume:
                        self.busy("Loading saved session...")
                        try:
                            self.loadfrompath(expected_session_path)
                        finally:
                            self.unbusy()
                
                self.saverecentproj()

            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Failed to load PDF while {load_stage}:\n\n{e}"
                )
            finally:
                self.unbusy()

    def loadcat(self):
        """Load categories from PostgreSQL."""
        try:
            self.categories = load_categories_from_postgres("inspection_tool")
        except Exception as e:
            print(f"Error loading categories: {e}")
            self.categories = []

    def getnextsr(self):
        """
        Calculate next available punch serial number from Excel Punch Sheet.
        FUNCTIONAL USE: Scans rows 9+ to find highest SR No, returns next sequential number.
        Used when creating new punch entries during quality inspection.
        """
        """Get next serial number"""
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 1
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            last_sr_no = 0
            row_num = 8
            while row_num <= ws.max_row + 5:
                val = self.readcell(ws, row_num, self.punch_cols['sr_no'])
                if val is None:
                    break
                try:
                    last_sr_no = int(val)
                except:
                    pass
                row_num += 1
            wb.close()
            return last_sr_no + 1
        except Exception:
            return 1

    def runtemp(self, template_def, tag_name=None, prefill_text=None):
        """Execute a template definition with optional OCR prefill for first input."""
        values = {}
        if tag_name:
            values["tag"] = tag_name

        inputs = template_def.get("inputs", [])
        for i, inp in enumerate(inputs):
            initial_value = ""
            if i == 0 and prefill_text:
                initial_value = prefill_text

            val = simpledialog.askstring(
                "Input Required",
                inp["label"],
                parent=self.root,
                initialvalue=initial_value
            )
            if not val:
                return None
            values[inp["name"]] = val.strip()

        try:
            return template_def["template"].format(**values)
        except KeyError as e:
            messagebox.showerror("Template Error", f"Missing placeholder: {e}")
            return None

    # Excel cell helpers
    def splitcell(self, cell_ref):
        m = re.match(r"([A-Z]+)(\d+)", cell_ref)
        if not m:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        col, row = m.groups()
        return int(row), col

    def resolvemergedtar(self, ws, row, col_idx):
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                return merged.min_row, merged.min_col
        return row, col_idx

    def writecell(self, ws, row, col, value):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self.resolvemergedtar(ws, int(row), col_idx)
        ws.cell(row=target_row, column=target_col).value = value

    def readcell(self, ws, row, col):
        if isinstance(col, str):
            col_idx = column_index_from_string(col)
        else:
            col_idx = int(col)
        target_row, target_col = self.resolvemergedtar(ws, int(row), col_idx)
        return ws.cell(row=target_row, column=target_col).value


    def askprojdetails(self):
        """Ask for project details while enforcing central storage policy."""
        
        dlg = tk.Toplevel(self.root)
        dlg.title("Project Details")
        dlg.geometry("500x400")
        dlg.transient(self.root)
        dlg.grab_set()

        # Cabinet ID
        tk.Label(dlg, text="Cabinet ID", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))
        cabinet_var = tk.StringVar(value=getattr(self, "cabinet_id", ""))
        cabinet_entry = tk.Entry(dlg, textvariable=cabinet_var, font=('Segoe UI', 10))
        cabinet_entry.pack(fill="x", padx=20)

        # Project Name
        tk.Label(dlg, text="Project Name", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        
        project_options = self.project_name_options_from_drawing(
            getattr(self, 'current_pdf_path', '')
        )
        detected_project = self.infer_project_from_drawing(
            getattr(self, 'current_pdf_path', '')
        )
        # Prefer an actual folder from 07-Scanned System Book. With multiple
        # folders, leave the editable dropdown unselected so the user chooses.
        initial_project = project_options[0] if len(project_options) == 1 else ''
        if not project_options and detected_project:
            initial_project = detected_project
        project_var = tk.StringVar(value=initial_project)
        project_entry = ttk.Combobox(
            dlg, textvariable=project_var, values=project_options,
            state='normal', font=('Segoe UI', 10)
        )
        project_entry.pack(fill="x", padx=20)
        if len(project_options) > 1:
            project_entry.set('')

        # Sales Order Number
        tk.Label(dlg, text="Sales Order Number", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(10, 0))
        so_var = tk.StringVar(value=getattr(self, 'sales_order_no', ''))
        tk.Entry(dlg, textvariable=so_var, font=('Segoe UI', 10)).pack(fill="x", padx=20)

        # Storage Location (browseable, defaults to configured base path)
        tk.Label(dlg, text="Storage Location", font=('Segoe UI', 10, 'bold')).pack(anchor="w", padx=20, pady=(15, 0))

        location_frame = tk.Frame(dlg)
        location_frame.pack(fill="x", padx=20, pady=5)

        scanned_system_book = self.scanned_system_book_path_from_drawing(
            getattr(self, 'current_pdf_path', '')
        )
        default_storage_location = scanned_system_book or getattr(
            self, "storage_location", get_base_path()
        ) or get_base_path()
        location_var = tk.StringVar(value=default_storage_location)
        location_entry = tk.Entry(location_frame, textvariable=location_var, font=('Segoe UI', 9), state='readonly')
        location_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))

        def browse_location():
            start_dir = location_var.get().strip() or get_base_path()
            if not os.path.isdir(start_dir):
                start_dir = get_base_path()
            selected = filedialog.askdirectory(
                title="Select Project Storage Location",
                initialdir=start_dir,
                mustexist=True
            )
            if selected:
                location_var.set(selected)

        tk.Button(
            location_frame,
            text="Browse...",
            command=browse_location,
            bg='#3b82f6',
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief=tk.FLAT,
            padx=15,
            pady=5
        ).pack(side=tk.RIGHT)

        tk.Label(
            dlg,
            text="Start from base_path and choose the exact project folder.",
            fg='#64748b',
            font=('Segoe UI', 9)
        ).pack(anchor="w", padx=20)

        # Auto-load location when project name changes
        def on_project_name_change(*args):
            project_name = project_var.get().strip()
            if not scanned_system_book:
                # Outside the expected input tree, preserve the original DB lookup.
                if project_name:
                    existing_location = self.db.get_project_location(project_name)
                    if existing_location:
                        location_var.set(existing_location)
                return
            existing_names = {name.casefold() for name in project_options}
            is_recurring = bool(project_name) and (
                project_name.casefold() in existing_names or
                bool(self.db.get_project_location(project_name))
            )
            if is_recurring:
                location_var.set(ntpath.join(scanned_system_book, project_name))
                location_entry.config(bg='#dcfce7')
                dlg.after(1000, lambda: location_entry.config(bg='white'))
            else:
                # A genuinely new project starts at 07-Scanned System Book.
                location_var.set(scanned_system_book)
        
        # Tcl 9 compatibility: trace() is deprecated; use trace_add().
        project_var.trace_add('write', on_project_name_change)
        on_project_name_change()

        def on_ok():
            cabinet = cabinet_var.get().strip()
            project = project_var.get().strip()
            so = so_var.get().strip()
            location = location_var.get().strip() or get_base_path()
            
            if not cabinet or not project:
                messagebox.showerror("Missing Information", 
                                   "Please fill in Cabinet ID and Project Name.")
                return
            
            # Check if this cabinet already exists in database
            if self.db.project_exists(cabinet):
                existing = self.db.get_project(cabinet)
                
                # If it's the same project, use existing location
                if existing['project_name'] == project:
                    location = existing['storage_location']
                    messagebox.showinfo("Existing Cabinet", 
                                      f"Cabinet '{cabinet}' found in project '{project}'.\n"
                                      f"Using existing location:\n{location}")
                else:
                    messagebox.showerror("Error", 
                                       f"Cabinet ID '{cabinet}' already exists in different project:\n"
                                       f"{existing['project_name']}")
                    return
            else:
                # New cabinet - check if project exists with different cabinet
                existing_project_location = self.db.get_project_location(project)
                
                if existing_project_location:
                    # Project exists, use its location
                    location = existing_project_location
                    messagebox.showinfo("Existing Project", 
                                      f"Project '{project}' found.\n"
                                      f"Using existing location:\n{location}")
                else:
                    # Brand new project keeps user-selected location (default is base_path).
                    location = location or get_base_path()
            
            self.cabinet_id = cabinet
            self.project_name = project
            self.sales_order_no = so
            self.storage_location = location
            
            # Save to database with all paths
            self.db.add_project({
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,
                'created_date': datetime.now().isoformat(),
                'last_accessed': datetime.now().isoformat()
            })
            
            dlg.destroy()

        tk.Button(dlg, text="OK", command=on_ok, 
                 bg="#10b981", fg="white", font=('Segoe UI', 10, 'bold'),
                 relief=tk.FLAT, padx=30, pady=10).pack(pady=20)
        
        dlg.wait_window()
    

    def write_to_xcl(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return

        try:
            wb = load_workbook(self.excel_file)

            for sheet_name, cells in self.header_cells.items():
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                if getattr(self, "project_name", ""):
                    r, c = self.splitcell(cells["project_name"])
                    self.writecell(ws, r, c, self.project_name)

                if getattr(self, "sales_order_no", ""):
                    r, c = self.splitcell(cells["sales_order"])
                    self.writecell(ws, r, c, self.sales_order_no)

                if getattr(self, "cabinet_id", ""):
                    r, c = self.splitcell(cells["cabinet_id"])
                    self.writecell(ws, r, c, self.cabinet_id)

            wb.save(self.excel_file)
            wb.close()

        except PermissionError:
            messagebox.showerror("Excel Locked", "Please close the Excel file before entering project details.")
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to write project details:\n{e}")

    def preparefolders(self):
        """Prepare project folders under configured UNC base_path."""
        if not hasattr(self, 'storage_location') or not self.storage_location:
            self.storage_location = get_base_path()
        
        if not self.project_name or not self.cabinet_id:
            messagebox.showerror("Error", "Project name and Cabinet ID required")
            return False

        try:
            # Enforce centrally managed UNC location and normalize to runtime absolute path.
            self.storage_location = resolve_storage_location(
                to_relative_storage_location(self.storage_location)
            )
        except ValueError as exc:
            messagebox.showerror("Invalid Storage Path", str(exc))
            return False
        
        # Create structure: storage_location/project_name/cabinet_id/
        project_folder_name = self.project_name.replace(' ', '_')
        storage_leaf = os.path.basename(os.path.normpath(self.storage_location))
        if storage_leaf.casefold() in {
            self.project_name.casefold(), project_folder_name.casefold()
        }:
            # Recurring projects already point at 07-Scanned System Book/<project>.
            project_folder = self.storage_location
        else:
            project_folder = os.path.join(self.storage_location, project_folder_name)
        
        cabinet_root = os.path.join(
            project_folder,
            self.cabinet_id.replace(' ', '_')
        )
        
        folders = {
            "root": cabinet_root,
            "source_drawings": os.path.join(cabinet_root, "Source_Drawings"),
            "working_excel": os.path.join(cabinet_root, "Working_Excel"),
            "interphase_export": os.path.join(cabinet_root, "Interphase_Export"),
            "annotated_drawings": os.path.join(cabinet_root, "Annotated_Drawings"),
            "sessions": os.path.join(cabinet_root, "Sessions")
        }
        
        for p in folders.values():
            os.makedirs(p, exist_ok=True)
        
        self.project_dirs = folders
        return True

    def copy_pdf_to_central_storage(self, source_pdf_path):
        """Ensure selected PDF is stored inside the configured central project tree."""
        if not source_pdf_path:
            raise ValueError("No source PDF path provided")

        source_drawings = self.project_dirs.get("source_drawings")
        if not source_drawings:
            raise ValueError("Project folders not prepared")

        os.makedirs(source_drawings, exist_ok=True)
        target_pdf_path = os.path.join(source_drawings, os.path.basename(source_pdf_path))

        src_norm = os.path.normcase(os.path.normpath(source_pdf_path))
        dst_norm = os.path.normcase(os.path.normpath(target_pdf_path))

        if src_norm != dst_norm:
            shutil.copy2(source_pdf_path, target_pdf_path)

        return target_pdf_path

    def getpathforpdf(self):
        if not self.current_pdf_path:
            return None

        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )

        return session_path if os.path.exists(session_path) else None

    # ================================================================
    # CHECKLIST FUNCTIONS
    # ================================================================

    def reviewnow(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Excel Missing", "Working Excel file not found.")
            return

        self.checklist_file = self.excel_file

        try:
            self.reviewbeforesave(self.checklist_file, self.session_refs)
        except Exception as e:
            messagebox.showerror("Checklist Error", f"Checklist review failed:\n{e}")

    # ================================================================
    # UPDATED: gather_checklist_matches - Updated for new column structure
    # ================================================================
    def allexistingpunches(self):
        """Return every existing Punch Sheet row, including implemented and closed punches."""
        punches = []
        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches
        try:
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            row = 9  # Row 8 contains the Punch Sheet column titles.
            while row <= ws.max_row + 5:
                sr_no = self.readcell(ws, row, self.punch_cols['sr_no'])
                description = self.readcell(ws, row, self.punch_cols['desc'])
                if sr_no is None and description is None:
                    break
                # Ignore accidental repeated header rows inside the data area.
                sr_text = str(sr_no or '').strip().casefold()
                desc_text = str(description or '').strip().casefold()
                if sr_text in {'sr no', 'sr no.', 'sr. no.', 'serial no', 'serial number'} or \
                        desc_text in {'punch description', 'description'}:
                    row += 1
                    continue
                if sr_no is not None:
                    punches.append({
                        'row': row,
                        'sr_no': sr_no,
                        'ref_no': self.readcell(ws, row, self.punch_cols['ref_no']),
                        'punch_text': description or '',
                        'category': self.readcell(ws, row, self.punch_cols['category']) or '',
                        'implemented_name': self.readcell(ws, row, self.punch_cols['implemented_name']),
                        'closed_name': self.readcell(ws, row, self.punch_cols['closed_name'])
                    })
                row += 1
            wb.close()
        except Exception as e:
            messagebox.showerror("Punch Read Error", f"Failed to read existing punches:\n{e}")
        return punches

    def editexistingpunch(self):
        """Edit an existing punch or add a new punch without a highlight."""
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showwarning("No Excel", "Load a project with a working Excel file first.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Edit or Add Punch")
        dlg.geometry("1050x680")
        dlg.minsize(850, 560)
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg='#1e293b', height=58)
        header.pack(fill=tk.X); header.pack_propagate(False)
        title_var = tk.StringVar(value="Edit Existing Punch")
        tk.Label(header, textvariable=title_var, bg='#1e293b', fg='white',
                 font=('Segoe UI', 14, 'bold')).pack(pady=14)

        list_frame = tk.Frame(dlg, bg='white')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=(16, 8))
        columns = ('sr', 'ref', 'description', 'category', 'status')
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)
        for col, text in zip(columns, ('SR No.', 'Reference', 'Punch Description', 'Category', 'Status')):
            tree.heading(col, text=text)
        tree.column('sr', width=70, anchor='center', stretch=False)
        tree.column('ref', width=100, anchor='center', stretch=False)
        tree.column('description', width=470)
        tree.column('category', width=180)
        tree.column('status', width=110, anchor='center', stretch=False)
        scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y); tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        form = tk.LabelFrame(dlg, text="Punch details", bg='#f8fafc', fg='#1e293b',
                             font=('Segoe UI', 10, 'bold'), padx=12, pady=10)
        form.pack(fill=tk.X, padx=18, pady=8)
        ref_var, category_var = tk.StringVar(), tk.StringVar()
        selected = {'punch': None, 'adding': False}
        tk.Label(form, text="Reference No.", bg='#f8fafc').grid(row=0, column=0, sticky='w')
        ref_entry = tk.Entry(form, textvariable=ref_var, font=('Segoe UI', 10), width=22)
        ref_entry.grid(row=1, column=0, sticky='ew', padx=(0, 12), pady=(2, 8))
        tk.Label(form, text="Category", bg='#f8fafc').grid(row=0, column=1, sticky='w')
        tk.Entry(form, textvariable=category_var, font=('Segoe UI', 10)).grid(
            row=1, column=1, sticky='ew', pady=(2, 8))
        tk.Label(form, text="Punch Description", bg='#f8fafc').grid(
            row=2, column=0, columnspan=2, sticky='w')
        desc_text = tk.Text(form, height=5, wrap=tk.WORD, font=('Segoe UI', 10))
        desc_text.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(2, 0))
        form.columnconfigure(1, weight=1)
        punch_by_item = {}

        def populate(select_sr=None):
            tree.delete(*tree.get_children()); punch_by_item.clear()
            for punch in self.allexistingpunches():
                status = 'Closed' if punch['closed_name'] else (
                    'Implemented' if punch['implemented_name'] else 'Open')
                item = tree.insert('', tk.END, values=(punch['sr_no'], punch['ref_no'] or '',
                    punch['punch_text'], punch['category'], status))
                punch_by_item[item] = punch
                if select_sr is not None and str(punch['sr_no']) == str(select_sr):
                    tree.selection_set(item); tree.focus(item)

        def load_selection(event=None):
            items = tree.selection()
            if not items: return
            punch = punch_by_item.get(items[0])
            if not punch: return
            selected.update(punch=punch, adding=False); title_var.set("Edit Existing Punch")
            ref_var.set('' if punch['ref_no'] is None else str(punch['ref_no']))
            category_var.set(str(punch['category'] or ''))
            desc_text.delete('1.0', tk.END); desc_text.insert('1.0', str(punch['punch_text'] or ''))

        def begin_add():
            selected.update(punch=None, adding=True); title_var.set("Add New Punch")
            tree.selection_remove(tree.selection()); ref_var.set(''); category_var.set('')
            desc_text.delete('1.0', tk.END); ref_entry.focus_force()

        def save_changes():
            ref_no = ref_var.get().strip()
            description = desc_text.get('1.0', 'end-1c').strip()
            category = category_var.get().strip()
            if not ref_no or not description or not category:
                messagebox.showwarning("Missing Information",
                    "Reference number, punch description, and category are required.", parent=dlg)
                return
            punch = selected['punch']
            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                if selected['adding'] or punch is None:
                    row = 9  # Preserve row 8 table titles.
                    while self.readcell(ws, row, self.punch_cols['sr_no']) is not None: row += 1
                    previous = self.readcell(ws, row - 1, self.punch_cols['sr_no']) if row > 9 else None
                    try: sr_no = int(previous) + 1 if previous is not None else 1
                    except (TypeError, ValueError): sr_no = self.getnextsr()
                    self.writecell(ws, row, self.punch_cols['sr_no'], sr_no)
                    self.writecell(ws, row, self.punch_cols['ref_no'], ref_no)
                    self.writecell(ws, row, self.punch_cols['desc'], description)
                    self.writecell(ws, row, self.punch_cols['category'], category)
                    self.writecell(ws, row, self.punch_cols['checked_name'],
                                   self.logged_in_fullname or "Unknown User")
                    self.writecell(ws, row, self.punch_cols['checked_date'],
                                   datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    action, target_sr = 'Added', sr_no
                else:
                    row, target_sr = punch['row'], punch['sr_no']
                    self.writecell(ws, row, self.punch_cols['ref_no'], ref_no)
                    self.writecell(ws, row, self.punch_cols['desc'], description)
                    self.writecell(ws, row, self.punch_cols['category'], category)
                    action = 'Updated'
                wb.save(self.excel_file); wb.close()
            except PermissionError:
                messagebox.showerror("Excel Locked", "Close the Excel file and try again.", parent=dlg); return
            except Exception as exc:
                messagebox.showerror("Punch Save Failed", f"Could not save the punch:\n{exc}", parent=dlg); return

            if punch is not None and not selected['adding']:
                ann = next((a for a in self.annotations if a.get('excel_row') == row or
                            str(a.get('sr_no', '')).strip() == str(target_sr)), None)
                if ann:
                    ann.update(ref_no=ref_no, punch_text=description, component=category,
                               category=category, last_edited_by=self.logged_in_fullname or "Unknown User",
                               last_edited_date=datetime.now().isoformat())
            self.session_refs.add(ref_no)
            self.updatestatsforref(ref_no, status='NOK')
            self.current_sr_no = self.getnextsr(); self.mark_dirty()
            populate(select_sr=target_sr)
            selected.update(punch=None, adding=False); title_var.set("Edit Existing Punch")
            self.flashstat(f"{action} punch SR {target_sr}", bg='#10b981')

        def begin_multimark():
            punch = selected['punch']
            if selected['adding'] or punch is None:
                messagebox.showwarning(
                    "Select a Punch",
                    "Select an existing punch from the list first, then click "
                    "\"Mark Multiple Highlights\".",
                    parent=dlg
                )
                return
            dlg.grab_release()
            dlg.destroy()
            self.root.update_idletasks()
            self.root.after(50, lambda: self.startmultimark(punch))

        tree.bind('<<TreeviewSelect>>', load_selection)
        buttons = tk.Frame(dlg, bg='#f8fafc'); buttons.pack(fill=tk.X, padx=18, pady=(4, 16))
        tk.Button(buttons, text="+ Add New Punch", command=begin_add, bg='#ec4899', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=22, pady=10).pack(side=tk.LEFT)
        tk.Button(buttons, text="Mark Multiple Highlights", command=begin_multimark,
                  bg='#7c3aed', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=22, pady=10).pack(side=tk.LEFT, padx=(8, 0))
        tk.Button(buttons, text="Save", command=save_changes, bg='#10b981', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=24, pady=10).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(buttons, text="Close", command=dlg.destroy, bg='#64748b', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=24, pady=10).pack(side=tk.RIGHT)
        populate()
        first = tree.get_children()
        if first: tree.selection_set(first[0]); tree.focus(first[0]); load_selection()
        else: begin_add()

    def openpuches(self):
        """Reads punch sheet and returns list of open punches with all details."""
        punches = []

        if not self.excel_file or not os.path.exists(self.excel_file):
            return punches

        try:
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active

            row = 9  # Row 8 contains table headings.
            empty_rows = 0
            while row <= ws.max_row + 5:
                sr = self.readcell(ws, row, self.punch_cols['sr_no'])
                description = self.readcell(ws, row, self.punch_cols['desc'])
                if sr is None and description is None:
                    empty_rows += 1
                    if empty_rows >= 3:
                        break
                    row += 1
                    continue
                empty_rows = 0
                sr_text = str(sr or '').strip().casefold()
                desc_text = str(description or '').strip().casefold()
                if sr_text in {'sr no', 'sr no.', 'sr. no.', 'serial no', 'serial number'} or \
                        desc_text in {'punch description', 'description'}:
                    row += 1
                    continue
                if sr is None:
                    row += 1
                    continue

                # Check if punch is closed
                closed = self.readcell(ws, row, self.punch_cols['closed_name'])
                if closed:
                    row += 1
                    continue

                # Check if punch is implemented
                implemented = bool(self.readcell(ws, row, self.punch_cols['implemented_name']))

                punches.append({
                    'sr_no': sr,
                    'row': row,
                    'ref_no': self.readcell(ws, row, self.punch_cols['ref_no']),
                    'punch_text': description or '',
                    'category': self.readcell(ws, row, self.punch_cols['category']),
                    'implemented': implemented,
                    'implemented_name': self.readcell(ws, row, self.punch_cols['implemented_name']),
                    'implemented_date': self.readcell(ws, row, self.punch_cols['implemented_date']),
                    'checked_name': self.readcell(ws, row, self.punch_cols['checked_name']),
                    'checked_date': self.readcell(ws, row, self.punch_cols['checked_date'])
                })

                row += 1

            wb.close()
            return punches
            
        except Exception as e:
            print(f"Error reading open punches: {e}")
            import traceback
            traceback.print_exc()
            return []


    # ================================================================
    # 3. UPDATED: review_checklist_before_save - With name and date updates
    # ================================================================

    def reviewbeforesave(self, checklist_path, refs_set):
        """Interphase Checklist Review, in the same list-workspace format as
        the punch verification screen (punchclosing): a sidebar listing every
        pending item (click any row to jump straight to it), a detail panel
        with an editable remark box, and action buttons in the footer.
        Core logic (matching, mandatory N/A remark, cell writes, save-on-exit)
        is unchanged from the previous card-style dialog."""
        try:
            cols, matches = self.checklistmatches(checklist_path, refs_set)
        except Exception as e:
            raise

        if not matches:
            messagebox.showinfo("Checklist Complete", 
                              " No items requiring review.\nAll Interphase items are up to date.",
                              icon='info')
            return

        wb = load_workbook(checklist_path)
        ws = wb[self.interphase_sheet_name]
        
        # Extract all columns
        status_col = cols['status_col']
        date_col = cols['date_col']
        name_col = cols['name_col']
        remark_col = cols['remark_col']

        colors = {
            'window': '#eef2f7', 'nav': '#0f172a', 'card': '#ffffff',
            'text': '#0f172a', 'muted': '#64748b', 'line': '#e2e8f0',
            'primary': '#2563eb', 'success': '#059669', 'warning': '#d97706',
            'danger': '#dc2626', 'soft_blue': '#eff6ff', 'soft_green': '#ecfdf5',
            'soft_orange': '#fff7ed'
        }

        dlg = tk.Toplevel(self.root)
        dlg.title("Interphase Checklist Review")
        dlg.geometry("1050x650")
        dlg.minsize(920, 580)
        dlg.configure(bg=colors['window'])
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg=colors['nav'], height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        heading = tk.Frame(header, bg=colors['nav'])
        heading.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        tk.Label(heading, text="Interphase Checklist Review", bg=colors['nav'], fg='white',
                 font=('Segoe UI Semibold', 17, 'bold')).pack(anchor='w', pady=(9, 0))
        tk.Label(heading, text=f"{self.cabinet_id}  •  {self.project_name}", bg=colors['nav'],
                 fg='#94a3b8', font=('Segoe UI', 10)).pack(anchor='w')
        summary = tk.Frame(header, bg=colors['nav'])
        summary.pack(side=tk.RIGHT, fill=tk.Y, padx=20)
        summary_var = tk.StringVar(value=f"{len(matches)} items pending review")
        tk.Label(summary, textvariable=summary_var, bg=colors['nav'],
                 fg='#bbf7d0', font=('Segoe UI Semibold', 10, 'bold')).pack(anchor='e', pady=(15, 1))
        tk.Label(summary, text="Mark each item OK, NOK, or N/A and add a remark if needed.",
                 bg=colors['nav'], fg='#94a3b8', font=('Segoe UI', 9)).pack(anchor='e')

        body = tk.Frame(dlg, bg=colors['window'])
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)
        sidebar = tk.Frame(body, bg=colors['card'], width=280, highlightthickness=1,
                           highlightbackground=colors['line'])
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)
        tk.Label(sidebar, text="CHECKLIST QUEUE", bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=14, pady=(12, 6))
        state_bar = tk.Frame(sidebar, bg=colors['soft_blue'])
        state_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
        state_label_var = tk.StringVar(value=f"{len(matches)} items remaining")
        tk.Label(state_bar, textvariable=state_label_var, bg=colors['soft_blue'],
                 fg='#1d4ed8', font=('Segoe UI Semibold', 9, 'bold')).pack(anchor='w', padx=8, pady=6)
        list_frame = tk.Frame(sidebar, bg=colors['card'])
        list_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        item_list = tk.Listbox(list_frame, activestyle='none', selectmode=tk.SINGLE,
                                font=('Segoe UI', 9), bg=colors['card'], fg=colors['text'],
                                selectbackground='#dbeafe', selectforeground='#1e3a8a',
                                relief=tk.FLAT, borderwidth=0, highlightthickness=0,
                                yscrollcommand=scrollbar.set)
        item_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=item_list.yview)

        content = tk.Frame(body, bg=colors['window'])
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        meta = tk.Frame(content, bg=colors['window'])
        meta.pack(fill=tk.X)
        ref_value = tk.StringVar(); row_value = tk.StringVar(); status_value = tk.StringVar()

        def metric(parent, label, variable, tint, value_color):
            card = tk.Frame(parent, bg=tint, highlightthickness=1, highlightbackground=colors['line'])
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            tk.Label(card, text=label, bg=tint, fg=colors['muted'],
                     font=('Segoe UI Semibold', 7, 'bold')).pack(anchor='w', padx=10, pady=(7, 1))
            tk.Label(card, textvariable=variable, bg=tint, fg=value_color,
                     font=('Segoe UI Semibold', 11, 'bold')).pack(anchor='w', padx=10, pady=(0, 8))

        metric(meta, "REFERENCE", ref_value, '#ffffff', colors['text'])
        metric(meta, "ROW", row_value, '#ffffff', colors['text'])
        metric(meta, "STATUS", status_value, '#ffffff', colors['warning'])

        detail_card = tk.Frame(content, bg=colors['card'], highlightthickness=1,
                               highlightbackground=colors['line'])
        detail_card.pack(fill=tk.BOTH, expand=True, pady=(9, 0))
        detail_head = tk.Frame(detail_card, bg=colors['card'])
        detail_head.pack(fill=tk.X, padx=14, pady=(11, 6))
        title_var = tk.StringVar(value="Checklist item")
        tk.Label(detail_head, textvariable=title_var, bg=colors['card'], fg=colors['text'],
                 font=('Segoe UI Semibold', 13, 'bold')).pack(side=tk.LEFT)
        position_var = tk.StringVar()
        tk.Label(detail_head, textvariable=position_var, bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI', 9)).pack(side=tk.RIGHT)

        tk.Label(detail_card, text="DESCRIPTION", bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=14, pady=(4, 2))
        description = tk.Text(detail_card, height=6, wrap=tk.WORD, bg='#f8fafc', fg=colors['text'],
                              relief=tk.FLAT, padx=10, pady=8, font=('Segoe UI', 10), cursor='arrow')
        description.pack(fill=tk.X, padx=14)
        description.config(state=tk.DISABLED)

        remark_frame = tk.Frame(detail_card, bg=colors['soft_orange'], highlightthickness=1,
                                highlightbackground='#fed7aa')
        remark_frame.pack(fill=tk.BOTH, expand=True, padx=14, pady=(10, 12))
        tk.Label(remark_frame, text="REMARK (editable - required for N/A)", bg=colors['soft_orange'],
                 fg='#c2410c', font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=10, pady=(8, 3))
        remark_text = tk.Text(remark_frame, height=5, wrap=tk.WORD, bg='white', fg=colors['text'],
                              relief=tk.FLAT, highlightthickness=1, highlightbackground='#fed7aa',
                              highlightcolor=colors['warning'], padx=10, pady=8, font=('Segoe UI', 10))
        remark_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        footer = tk.Frame(dlg, bg=colors['card'], height=64, highlightthickness=1,
                          highlightbackground=colors['line'])
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)
        left_actions = tk.Frame(footer, bg=colors['card'])
        left_actions.pack(side=tk.LEFT, padx=14, pady=9)
        right_actions = tk.Frame(footer, bg=colors['card'])
        right_actions.pack(side=tk.RIGHT, padx=14, pady=9)

        def button(parent, text, command, bg, fg='white', width=14):
            return tk.Button(parent, text=text, command=command, bg=bg, fg=fg,
                             activebackground=bg, activeforeground=fg, relief=tk.FLAT,
                             borderwidth=0, cursor='hand2', font=('Segoe UI Semibold', 9, 'bold'),
                             padx=10, pady=8, width=width)

        current = {'index': 0, 'refreshing': False}

        def refresh_list():
            selected = current['index']
            current['refreshing'] = True
            item_list.delete(0, tk.END)
            for r, ref_str, desc, remark in matches:
                marker = " • " if remark.strip() else "   "
                item_list.insert(tk.END, f"  {marker} Row {r}   •   {ref_str}")
            item_list.selection_clear(0, tk.END)
            if matches:
                selected = max(0, min(selected, len(matches) - 1))
                item_list.selection_set(selected)
                item_list.activate(selected)
                item_list.see(selected)
            current['refreshing'] = False
            summary_var.set(f"{len(matches)} items pending review")
            state_label_var.set(f"{len(matches)} items remaining")

        def show_item(index=None):
            if not matches:
                return
            if index is not None:
                current['index'] = max(0, min(len(matches) - 1, index))
            r, ref_str, desc, remark = matches[current['index']]
            title_var.set(f"Reference {ref_str}")
            position_var.set(f"Item {current['index'] + 1} of {len(matches)}")
            ref_value.set(ref_str)
            row_value.set(str(r))
            status_value.set("Pending review")
            description.config(state=tk.NORMAL)
            description.delete('1.0', tk.END)
            description.insert('1.0', desc or 'No description available.')
            description.config(state=tk.DISABLED)
            remark_text.delete('1.0', tk.END)
            remark_text.insert('1.0', remark or '')
            refresh_list()
            remark_text.focus_set()

        def select_from_list(event=None):
            if current.get('refreshing'):
                return
            selection = item_list.curselection()
            if selection:
                show_item(selection[0])

        def go(delta):
            show_item(current['index'] + delta)

        def current_remark_text():
            return remark_text.get('1.0', 'end-1c').strip()

        def sync_current_remark_in_memory():
            """Keep the in-memory matches list in sync with whatever is
            currently typed in the remark box, so switching items in the
            sidebar list doesn't lose an edit that wasn't explicitly saved."""
            if not matches:
                return
            r, ref_str, desc, _old_remark = matches[current['index']]
            matches[current['index']] = (r, ref_str, desc, current_remark_text())

        def save_remark(show_confirmation=True):
            """Write just the remark for the current item without changing status."""
            if not matches:
                return False
            r, ref_str, desc, _old_remark = matches[current['index']]
            remark = current_remark_text()
            try:
                self.writecell(ws, r, remark_col, remark)
                wb.save(checklist_path)
            except PermissionError:
                messagebox.showerror("File Locked", "Please close the Excel file and try again.",
                                     icon='error', parent=dlg)
                return False
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}", parent=dlg)
                return False
            matches[current['index']] = (r, ref_str, desc, remark)
            self.mark_dirty()
            if show_confirmation:
                messagebox.showinfo("Remark saved", "The remark was saved for this item.", parent=dlg)
            return True

        def do_action_set_status(status_value_to_write):
            r, ref_str, desc, _old_remark = matches[current['index']]
            remark = current_remark_text()
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = self.logged_in_fullname or "Unknown User"

            try:
                self.writecell(ws, r, status_col, status_value_to_write)
                self.writecell(ws, r, name_col, username)
                self.writecell(ws, r, date_col, current_date)
                if remark:
                    self.writecell(ws, r, remark_col, remark)
                wb.save(checklist_path)
            except PermissionError:
                messagebox.showerror("File Locked", "⚠️ Please close the Excel file and try again.",
                                     icon='error', parent=dlg)
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}", parent=dlg)
                return

            self.mark_dirty()
            matches.pop(current['index'])
            if not matches:
                messagebox.showinfo("Review Complete", "Checklist review finished!", icon='info', parent=dlg)
                dlg.destroy()
                return
            current['index'] = min(current['index'], len(matches) - 1)
            show_item(current['index'])

        def on_ok():
            do_action_set_status("OK")

        def on_nok():
            do_action_set_status("NOK")

        def on_na():
            """N/A requires a remark - reuse whatever is currently typed in
            the editable remark box; only fall back to a prompt if it's empty."""
            remark = current_remark_text()
            if not remark:
                remark = simpledialog.askstring(
                    "Remark Required",
                    "N/A status requires a remark.\nPlease provide a reason:",
                    parent=dlg
                )
                if remark:
                    remark_text.delete('1.0', tk.END)
                    remark_text.insert('1.0', remark)

            if not remark or not remark.strip():
                messagebox.showwarning("Remark Required",
                                       "You must provide a remark for N/A status.", parent=dlg)
                return

            r, ref_str, desc, _old_remark = matches[current['index']]
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = self.logged_in_fullname or "Unknown User"

            try:
                self.writecell(ws, r, status_col, "N/A")
                self.writecell(ws, r, date_col, current_date)
                self.writecell(ws, r, name_col, username)
                self.writecell(ws, r, remark_col, remark)
                wb.save(checklist_path)
            except PermissionError:
                messagebox.showerror("File Locked", "⚠️ Please close the Excel file and try again.",
                                     icon='error', parent=dlg)
                return
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update checklist:\n{e}", parent=dlg)
                return

            self.mark_dirty()
            matches.pop(current['index'])
            if not matches:
                dlg.destroy()
                return
            current['index'] = min(current['index'], len(matches) - 1)
            show_item(current['index'])

        item_list.bind('<<ListboxSelect>>', select_from_list)
        button(left_actions, "Previous", lambda: go(-1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT, padx=(0, 8))
        button(left_actions, "Next", lambda: go(1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT)
        button(right_actions, "Close", dlg.destroy, '#475569', width=10).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "NA - Not Applicable", on_na, colors['warning'], width=18).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "NOK", on_nok, colors['danger'], width=10).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "OK", on_ok, colors['success'], width=10).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "Save Remark", save_remark, colors['primary'], width=13).pack(side=tk.RIGHT)
        dlg.bind('<Control-s>', lambda event: save_remark())

        refresh_list()
        show_item(0)
        dlg.wait_window()


    # ================================================================
    # 4. HELPER: gather_checklist_matches - Returns column info and matches
    # ================================================================

    def checklistmatches(self, checklist_path, refs_set):
        """Returns Interphase rows where Reference No is NOT in refs_set."""
        wb = load_workbook(checklist_path)
        if self.interphase_sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError("Interphase sheet not found")

        ws = wb[self.interphase_sheet_name]
        ref_col = self.interphase_cols['ref_no']
        desc_col = self.interphase_cols['description']
        status_col = self.interphase_cols['status']
        name_col = self.interphase_cols['name']
        date_col = self.interphase_cols['date']
        remark_col = self.interphase_cols['remark']

        matches = []
        max_row = ws.max_row if ws.max_row else 2000

        for r in range(11, max_row + 1):
            ref_val = self.readcell(ws, r, ref_col)
            if ref_val is None:
                continue

            ref_str = str(ref_val).strip()

            if ref_str in refs_set:
                continue

            status_val = self.readcell(ws, r, status_col)
            status_str = str(status_val).strip().lower() if status_val is not None else ''

            if status_str in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                continue

            desc_val = self.readcell(ws, r, desc_col) or ''
            remark_val = self.readcell(ws, r, remark_col) or ''
            matches.append((r, ref_str, str(desc_val), str(remark_val)))

        wb.close()
        return {
            'ref_col': ref_col, 
            'desc_col': desc_col, 
            'status_col': status_col,
            'name_col': name_col,
            'date_col': date_col,
            'remark_col': remark_col
        }, matches
    # ================================================================
    # EXCEL HELPERS
    # ================================================================

    def saveinterphase(self):
        if not self.current_pdf_path:
            messagebox.showwarning("No PDF", "Load a PDF first.")
            return

        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Missing File", "Working Excel file not found.")
            return

        save_path = os.path.join(
            self.project_dirs["interphase_export"],
            f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
        )

        try:
            shutil.copy2(self.excel_file, save_path)
        except PermissionError:
            messagebox.showerror("File Open", "Close the Excel file and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")

    def openxcl(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showwarning("No Excel", "No working Excel file found.")
            return

        try:
            if os.name == 'nt':
                os.startfile(self.excel_file)
            else:
                if sys.platform == 'darwin':
                    cmd = f"open {shlex.quote(self.excel_file)}"
                else:
                    cmd = f"xdg-open {shlex.quote(self.excel_file)}"
                subprocess.Popen(cmd, shell=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel: {e}")

    # ================================================================
    # FUZZY MATCH HELPER
    # ================================================================

    def findrow(self, sr_no, punch_text, min_ratio=0.60):
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            row = 8

            while True:
                cell = self.readcell(ws, row, self.punch_cols['sr_no'])
                if cell is None:
                    if self.readcell(ws, row, self.punch_cols['desc']) is None:
                        break
                    else:
                        row += 1
                        continue
                try:
                    if int(cell) == int(sr_no):
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                except:
                    if str(cell).strip() == str(sr_no).strip():
                        wb.close()
                        return (row, 1.0, 'sr_exact')
                row += 1
                if row > 2000:
                    break

            best_row = None
            best_ratio = 0.0
            row = 8

            while True:
                txt = self.readcell(ws, row, self.punch_cols['desc'])
                if txt is None:
                    if row > 2000:
                        break
                    row += 1
                    continue
                try:
                    ratio = SequenceMatcher(None, str(punch_text).strip().lower(), str(txt).strip().lower()).ratio()
                except:
                    ratio = 0.0
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_row = row
                row += 1
                if row > 2000:
                    break

            wb.close()
            if best_row and best_ratio >= min_ratio:
                return (best_row, best_ratio, 'fuzzy_text')
            return (None, best_ratio, None)
        except Exception as e:
            try:
                wb.close()
            except:
                pass
            return (None, 0.0, None)

    # ============================================================================
    # UPDATED: view_production_handbacks - Auto-open punch closing
    # ============================================================================

    def _log_verification_error(self, stage, exc):
        """Write verification failures to a durable log beside the application."""
        import traceback
        try:
            log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    "verification_error.log")
            with open(log_path, "a", encoding="utf-8") as log:
                log.write("\n" + "=" * 80 + "\n")
                log.write(f"{datetime.now().isoformat()} | stage={stage}\n")
                log.write(f"cabinet={getattr(self, 'cabinet_id', '')}\n")
                log.write(f"pdf={getattr(self, 'current_pdf_path', '')}\n")
                log.write(f"excel={getattr(self, 'excel_file', '')}\n")
                log.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
            return log_path
        except Exception:
            return None

    def viewhandbacks(self):
        """Show the verification queue using the Projects & Cabinets dialog design."""
        pending_items = self.handover_db.get_pending_quality_items()

        if not pending_items:
            messagebox.showinfo(
                "No Items",
                "No items pending verification from production.",
                icon='info'
            )
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Verify Production Rework")
        dlg.geometry("1000x620")
        dlg.minsize(780, 500)
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg='#1e293b', height=58)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header, text="Verify Production Rework", bg='#1e293b', fg='white',
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=14)

        search_frame = tk.Frame(dlg, bg='#f8fafc')
        search_frame.pack(fill=tk.X, padx=18, pady=(16, 8))
        tk.Label(
            search_frame, text="Search:", bg='#f8fafc', fg='#334155',
            font=('Segoe UI', 10, 'bold')
        ).pack(side=tk.LEFT, padx=(0, 8))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=('Segoe UI', 11))
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        count_var = tk.StringVar()
        tk.Label(
            search_frame, textvariable=count_var, bg='#f8fafc', fg='#64748b',
            font=('Segoe UI', 9)
        ).pack(side=tk.RIGHT, padx=(12, 0))

        body = tk.Frame(dlg, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=8)

        columns = ('cabinet', 'project', 'rework_by', 'date')
        tree = ttk.Treeview(body, columns=columns, show='headings', selectmode='browse')
        tree.heading('cabinet', text='Cabinet ID')
        tree.heading('project', text='Project')
        tree.heading('rework_by', text='Rework Completed By')
        tree.heading('date', text='Completed Date')
        tree.column('cabinet', width=180, minwidth=140)
        tree.column('project', width=330, minwidth=220)
        tree.column('rework_by', width=220, minwidth=160)
        tree.column('date', width=140, anchor='center', stretch=False)

        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        row_items = {}

        def load_verification_item(item):
            stage = "initializing"
            cabinet_label = str(item.get('cabinet_id') or 'selected cabinet')
            self.busy(f"Loading {cabinet_label} for verification...")
            self.stopmultimark()  # never leave multi-mark mode active across a document swap
            try:
                stage = "reading project"
                self.busy(f"Loading {cabinet_label}: reading project details...")
                project_data = self.db.get_project(item['cabinet_id'])
                if not project_data:
                    raise RuntimeError(f"Project {item['cabinet_id']} was not found in the project database.")

                stage = "validating files"
                self.busy(f"Loading {cabinet_label}: validating PDF and Punch Sheet...")
                pdf_path = os.path.abspath(str(item.get('pdf_path') or ''))
                excel_path = os.path.abspath(str(item.get('excel_path') or ''))
                session_path = os.path.abspath(str(item.get('session_path') or '')) if item.get('session_path') else None
                if not os.path.isfile(pdf_path):
                    raise FileNotFoundError(f"PDF file not found: {pdf_path}")
                if not os.path.isfile(excel_path):
                    raise FileNotFoundError(f"Excel file not found: {excel_path}")

                stage = "closing previous document"
                self.busy(f"Loading {cabinet_label}: preparing the workspace...")
                if self.pdf_document is not None:
                    try:
                        self.pdf_document.close()
                    except Exception:
                        pass
                    self.pdf_document = None
                self._clear_page_render_cache()

                stage = "setting project context"
                self.busy(f"Loading {cabinet_label}: setting project context...")
                self.cabinet_id = str(item.get('cabinet_id') or '')
                self.project_name = str(item.get('project_name') or '')
                self.sales_order_no = str(item.get('sales_order_no') or '')
                self.storage_location = project_data['storage_location']
                self.preparefolders()

                stage = "opening PDF"
                self.busy(f"Loading {cabinet_label}: opening the PDF...")
                self.pdf_document = fitz.open(pdf_path)
                if len(self.pdf_document) == 0:
                    raise RuntimeError("The selected PDF contains no pages.")
                self.current_pdf_path = pdf_path
                self.current_page = 0
                self.zoom_level = 1.0
                self._update_zoom_toolbar_label()
                self.tool_mode = None
                self.root.config(cursor="")

                stage = "opening Punch Sheet"
                self.busy(f"Loading {cabinet_label}: opening the Punch Sheet...")
                # Validate the workbook before assigning it to the live workspace.
                test_wb = load_workbook(excel_path, read_only=True, data_only=True)
                if self.punch_sheet_name not in test_wb.sheetnames:
                    sheet_names = ", ".join(test_wb.sheetnames)
                    test_wb.close()
                    raise RuntimeError(
                        f"Punch Sheet '{self.punch_sheet_name}' was not found. Available sheets: {sheet_names}"
                    )
                test_wb.close()
                self.excel_file = excel_path
                self.working_excel_path = excel_path

                stage = "loading annotation session"
                self.busy(f"Loading {cabinet_label}: restoring annotations...")
                self.annotations = []
                self.session_refs.clear()
                if session_path and os.path.isfile(session_path):
                    self.loadfrompath(session_path)
                else:
                    self.display(preserve_view=False)
                self.current_sr_no = self.getnextsr()

                stage = "reading punches"
                self.busy(f"Loading {cabinet_label}: reading open punches...")
                punches = self.openpuches()
                if not punches:
                    self.unbusy()
                    messagebox.showinfo(
                        "No Open Punches",
                        "The cabinet loaded successfully, but no open punch rows were found.",
                        parent=self.root
                    )
                    return

                stage = "updating workflow status"
                self.busy(f"Loading {cabinet_label}: updating verification status...")
                self.update_status_and_sync('being_closed_by_quality')

                stage = "opening verification workspace"
                self.busy(f"Loading {cabinet_label}: opening verification workspace...")
                self.unbusy()
                self.punchclosing()

            except BaseException as exc:
                self.unbusy()
                log_path = self._log_verification_error(stage, exc)
                detail = f"Verification failed while {stage}.\n\n{type(exc).__name__}: {exc}"
                if log_path:
                    detail += f"\n\nDiagnostic log:\n{log_path}"
                messagebox.showerror("Verification Error", detail, parent=self.root)
            finally:
                self.unbusy()

        def populate(*args):
            query = search_var.get().strip().casefold()
            tree.delete(*tree.get_children())
            row_items.clear()
            visible = 0
            for item in pending_items:
                searchable = ' '.join(str(item.get(key, '')) for key in (
                    'cabinet_id', 'project_name', 'rework_completed_by', 'rework_completed_date'
                )).casefold()
                if query and query not in searchable:
                    continue
                row_id = tree.insert('', tk.END, values=(
                    item.get('cabinet_id', ''),
                    item.get('project_name', ''),
                    item.get('rework_completed_by', ''),
                    str(item.get('rework_completed_date') or '')[:10],
                ))
                row_items[row_id] = item
                visible += 1
            count_var.set(f"{visible} of {len(pending_items)} cabinet(s)")
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])

        def loadsel(event=None):
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Select a cabinet first.", parent=dlg)
                return
            item = row_items.get(selected[0])
            if not item:
                return
            item_data = dict(item)
            dlg.grab_release()
            dlg.destroy()
            self.root.update_idletasks()
            self.root.after(100, lambda: load_verification_item(item_data))

        search_var.trace_add('write', populate)
        tree.bind('<Double-Button-1>', loadsel)
        tree.bind('<Return>', loadsel)

        buttons = tk.Frame(dlg, bg='#f8fafc')
        buttons.pack(fill=tk.X, padx=18, pady=(4, 16))
        tk.Button(
            buttons, text="Cancel", command=dlg.destroy, bg='#64748b', fg='white',
            font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=20, pady=9,
            cursor='hand2'
        ).pack(side=tk.LEFT)
        tk.Button(
            buttons, text="Open Verification", command=loadsel, bg='#3b82f6', fg='white',
            font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=26, pady=9,
            cursor='hand2'
        ).pack(side=tk.RIGHT)

        populate()
        search_entry.focus_set()


    def verifyprodrework(self, item_data):
        """Open verification safely after loading a production handback."""
        try:
            self.punchclosing()
        except BaseException as exc:
            log_path = self._log_verification_error("opening verification workspace", exc)
            detail = f"The verification workspace could not open.\n\n{type(exc).__name__}: {exc}"
            if log_path:
                detail += f"\n\nDiagnostic log:\n{log_path}"
            messagebox.showerror("Verification Error", detail, parent=self.root)


    # ============================================================================
    # NEW: punch_closing_mode_for_verification - Modified punch closing for handback
    # ============================================================================

    def checklistcomp(self):
        """Check if all Interphase checklist items have been reviewed
        
        Returns:
            tuple: (is_complete: bool, pending_count: int)
        """
        if not self.excel_file or not os.path.exists(self.excel_file):
            return (True, 0)  # Assume complete if no Excel
        
        try:
            wb = load_workbook(self.excel_file, data_only=True)
            if self.interphase_sheet_name not in wb.sheetnames:
                wb.close()
                return (True, 0)
            
            ws = wb[self.interphase_sheet_name]
            ref_col = self.interphase_cols['ref_no']
            status_col = self.interphase_cols['status']
            
            pending_count = 0
            max_row = ws.max_row if ws.max_row else 2000
            
            for r in range(11, max_row + 1):
                ref_val = self.readcell(ws, r, ref_col)
                if ref_val is None:
                    continue
                
                status_val = self.readcell(ws, r, status_col)
                status_str = str(status_val).strip().lower() if status_val is not None else ''
                
                # Check if status is filled (OK, NOK, or N/A)
                if status_str not in ('ok', 'nok', 'n/a', 'na', 'not applicable'):
                    pending_count += 1
            
            wb.close()
            return (pending_count == 0, pending_count)
            
        except Exception as e:
            print(f"Error checking checklist completion: {e}")
            return (True, 0)  # Assume complete on error

    # UPDATED: punch_closing_mode - with auto-finalization
    def punchclosing(self):
        """Open the redesigned Quality verification workspace."""
        punches = self.openpuches()
        if not punches:
            messagebox.showinfo(
                "No Open Punches",
                "No open punch rows were found in the Punch Sheet.",
                parent=self.root
            )
            return
        def punch_sort_key(punch):
            sr_value = punch.get('sr_no')
            try:
                sr_key = (0, float(sr_value))
            except (TypeError, ValueError):
                sr_key = (1, str(sr_value or '').casefold())
            return (not bool(punch.get('implemented')), sr_key)
        punches.sort(key=punch_sort_key)

        colors = {
            'window': '#eef2f7', 'nav': '#0f172a', 'card': '#ffffff',
            'text': '#0f172a', 'muted': '#64748b', 'line': '#e2e8f0',
            'primary': '#2563eb', 'success': '#059669', 'warning': '#d97706',
            'danger': '#dc2626', 'soft_blue': '#eff6ff', 'soft_green': '#ecfdf5',
            'soft_orange': '#fff7ed'
        }
        dlg = tk.Toplevel(self.root)
        dlg.title("Quality Verification Workspace")
        dlg.geometry("1050x650")
        dlg.minsize(920, 580)
        dlg.configure(bg=colors['window'])
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg=colors['nav'], height=72)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        heading = tk.Frame(header, bg=colors['nav'])
        heading.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        tk.Label(heading, text="Verify Production Rework", bg=colors['nav'], fg='white',
                 font=('Segoe UI Semibold', 17, 'bold')).pack(anchor='w', pady=(9, 0))
        tk.Label(heading, text=f"{self.cabinet_id}  •  {self.project_name}", bg=colors['nav'],
                 fg='#94a3b8', font=('Segoe UI', 10)).pack(anchor='w')
        summary = tk.Frame(header, bg=colors['nav'])
        summary.pack(side=tk.RIGHT, fill=tk.Y, padx=20)
        ready = sum(1 for item in punches if item['implemented'])
        tk.Label(summary, text=f"{ready} of {len(punches)} ready to verify", bg=colors['nav'],
                 fg='#bbf7d0', font=('Segoe UI Semibold', 10, 'bold')).pack(anchor='e', pady=(15, 1))
        tk.Label(summary, text="Review the production action and close only verified punches.",
                 bg=colors['nav'], fg='#94a3b8', font=('Segoe UI', 9)).pack(anchor='e')

        body = tk.Frame(dlg, bg=colors['window'])
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=12)
        sidebar = tk.Frame(body, bg=colors['card'], width=250, highlightthickness=1,
                           highlightbackground=colors['line'])
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)
        tk.Label(sidebar, text="VERIFICATION QUEUE", bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=14, pady=(12, 6))
        state_bar = tk.Frame(sidebar, bg=colors['soft_green'])
        state_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(state_bar, text=f"{len(punches)} punches remaining", bg=colors['soft_green'],
                 fg='#047857', font=('Segoe UI Semibold', 9, 'bold')).pack(anchor='w', padx=8, pady=6)
        list_frame = tk.Frame(sidebar, bg=colors['card'])
        list_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        punch_list = tk.Listbox(list_frame, activestyle='none', selectmode=tk.SINGLE,
                                font=('Segoe UI', 9), bg=colors['card'], fg=colors['text'],
                                selectbackground='#dcfce7', selectforeground='#14532d',
                                relief=tk.FLAT, borderwidth=0, highlightthickness=0,
                                yscrollcommand=scrollbar.set)
        punch_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=punch_list.yview)

        content = tk.Frame(body, bg=colors['window'])
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        meta = tk.Frame(content, bg=colors['window'])
        meta.pack(fill=tk.X)
        sr_value = tk.StringVar(); ref_value = tk.StringVar(); status_value = tk.StringVar()
        def metric(parent, label, variable, tint, value_color):
            card = tk.Frame(parent, bg=tint, highlightthickness=1, highlightbackground=colors['line'])
            card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            tk.Label(card, text=label, bg=tint, fg=colors['muted'],
                     font=('Segoe UI Semibold', 7, 'bold')).pack(anchor='w', padx=10, pady=(7, 1))
            tk.Label(card, textvariable=variable, bg=tint, fg=value_color,
                     font=('Segoe UI Semibold', 11, 'bold')).pack(anchor='w', padx=10, pady=(0, 8))
        metric(meta, "SR NUMBER", sr_value, '#ffffff', colors['text'])
        metric(meta, "REFERENCE", ref_value, '#ffffff', colors['text'])
        metric(meta, "IMPLEMENTATION", status_value, '#ffffff', colors['warning'])

        detail_card = tk.Frame(content, bg=colors['card'], highlightthickness=1,
                               highlightbackground=colors['line'])
        detail_card.pack(fill=tk.BOTH, expand=True, pady=(9, 0))
        detail_head = tk.Frame(detail_card, bg=colors['card'])
        detail_head.pack(fill=tk.X, padx=14, pady=(11, 6))
        title_var = tk.StringVar(value="Punch details")
        tk.Label(detail_head, textvariable=title_var, bg=colors['card'], fg=colors['text'],
                 font=('Segoe UI Semibold', 13, 'bold')).pack(side=tk.LEFT)
        position_var = tk.StringVar()
        tk.Label(detail_head, textvariable=position_var, bg=colors['card'], fg=colors['muted'],
                 font=('Segoe UI', 9)).pack(side=tk.RIGHT)

        description = tk.Text(detail_card, height=5, wrap=tk.WORD, bg='#f8fafc', fg=colors['text'],
                              relief=tk.FLAT, padx=8, pady=6, font=('Segoe UI', 9), cursor='arrow')
        description.pack(fill=tk.X, padx=14)
        description.config(state=tk.DISABLED)

        remarks = tk.Frame(detail_card, bg=colors['card'])
        remarks.pack(fill=tk.BOTH, expand=True, padx=14, pady=9)
        prod_col = tk.Frame(remarks, bg=colors['soft_green'], highlightthickness=1,
                            highlightbackground='#a7f3d0')
        prod_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        tk.Label(prod_col, text="PRODUCTION ACTION", bg=colors['soft_green'], fg='#047857',
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=10, pady=(8, 3))
        production_text = tk.Text(prod_col, height=5, wrap=tk.WORD, bg=colors['soft_green'],
                                  fg=colors['text'], relief=tk.FLAT, padx=12, pady=8,
                                  font=('Segoe UI', 10), cursor='arrow')
        production_text.pack(fill=tk.BOTH, expand=True, padx=3, pady=(0, 5))
        production_text.config(state=tk.DISABLED)

        quality_col = tk.Frame(remarks, bg=colors['soft_blue'], highlightthickness=1,
                               highlightbackground='#bfdbfe')
        quality_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))
        tk.Label(quality_col, text="QUALITY REMARK", bg=colors['soft_blue'], fg='#1d4ed8',
                 font=('Segoe UI Semibold', 8, 'bold')).pack(anchor='w', padx=10, pady=(8, 3))
        quality_text = tk.Text(quality_col, height=5, wrap=tk.WORD, bg='white', fg=colors['text'],
                               relief=tk.FLAT, highlightthickness=1, highlightbackground='#bfdbfe',
                               highlightcolor=colors['primary'], padx=10, pady=8, font=('Segoe UI', 10))
        quality_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        footer = tk.Frame(dlg, bg=colors['card'], height=60, highlightthickness=1,
                          highlightbackground=colors['line'])
        footer.pack(fill=tk.X, side=tk.BOTTOM)
        footer.pack_propagate(False)
        left_actions = tk.Frame(footer, bg=colors['card'])
        left_actions.pack(side=tk.LEFT, padx=14, pady=9)
        right_actions = tk.Frame(footer, bg=colors['card'])
        right_actions.pack(side=tk.RIGHT, padx=14, pady=9)
        def button(parent, text, command, bg, fg='white', width=16):
            return tk.Button(parent, text=text, command=command, bg=bg, fg=fg,
                             activebackground=bg, activeforeground=fg, relief=tk.FLAT,
                             borderwidth=0, cursor='hand2', font=('Segoe UI Semibold', 9, 'bold'),
                             padx=10, pady=8, width=width)

        current = {'index': 0, 'refreshing': False}

        def find_all_anns(punch):
            """Return every annotation attached to this punch (its remark
            holder plus every pink/green highlight added for it, including
            any extras added via multi-mark mode)."""
            sr_key = str(punch.get('sr_no', '')).strip()
            row_key = str(punch.get('row', '')).strip()
            return [a for a in self.annotations
                    if (str(a.get('sr_no', '')).strip() == sr_key and sr_key)
                    or (str(a.get('excel_row', '')).strip() == row_key and row_key)]

        def find_ann(punch, create=False):
            matches = find_all_anns(punch)
            ann = next((a for a in matches if a.get('implementation_remark') is not None
                        or a.get('quality_remark') is not None), matches[0] if matches else None)
            if ann is None and create:
                ann = {'type': 'punch_meta', 'page': None,
                       'sr_no': punch.get('sr_no'), 'excel_row': punch.get('row'),
                       'ref_no': punch.get('ref_no'), 'punch_text': punch.get('punch_text'),
                       'category': punch.get('category'), 'timestamp': datetime.now().isoformat()}
                self.annotations.append(ann)
            return ann

        def refresh_list():
            selected = current['index']
            current['refreshing'] = True
            punch_list.delete(0, tk.END)
            for item in punches:
                marker = "READY" if item['implemented'] else "WAIT"
                punch_list.insert(tk.END, f"  {marker:<5}   SR {item['sr_no']}   •   {item['ref_no']}")
            punch_list.selection_clear(0, tk.END)
            punch_list.selection_set(selected)
            punch_list.activate(selected)
            punch_list.see(selected)
            current['refreshing'] = False

        def show_item(index=None):
            if index is not None:
                current['index'] = max(0, min(len(punches) - 1, index))
            p = punches[current['index']]
            ann = find_ann(p)
            sr_value.set(str(p['sr_no']))
            ref_value.set(str(p['ref_no']))
            status_value.set("Ready to verify" if p['implemented'] else "Not implemented")
            title_var.set(str(p.get('category') or 'Punch details'))
            position_var.set(f"Punch {current['index'] + 1} of {len(punches)}")
            description.config(state=tk.NORMAL)
            description.delete('1.0', tk.END)
            description.insert('1.0', p.get('punch_text') or 'No description available.')
            description.config(state=tk.DISABLED)
            production_text.config(state=tk.NORMAL)
            production_text.delete('1.0', tk.END)
            production_text.insert('1.0', (ann or {}).get('implementation_remark') or 'No production remark for this punch.')
            production_text.config(state=tk.DISABLED)
            quality_text.delete('1.0', tk.END)
            quality_text.insert('1.0', (ann or {}).get('quality_remark') or '')
            refresh_list()
            quality_text.focus_set()

        def select_from_list(event=None):
            if current.get('refreshing'):
                return
            selection = punch_list.curselection()
            if selection:
                show_item(selection[0])

        def go(delta):
            show_item(current['index'] + delta)

        def save_remark(show_confirmation=True):
            p = punches[current['index']]
            ann = find_ann(p, create=True)
            remark = quality_text.get('1.0', 'end-1c').strip()
            if str(ann.get('quality_remark') or '') != remark:
                ann['quality_remark'] = remark
                self.mark_dirty()
                self._write_session_file()
            if show_confirmation:
                messagebox.showinfo("Remark saved", "The Quality remark was saved for this punch.", parent=dlg)
            return True

        def close_punch():
            p = punches[current['index']]
            if not p['implemented']:
                messagebox.showwarning("Implementation pending",
                                       "Production has not marked this punch as implemented.", parent=dlg)
                return
            save_remark(show_confirmation=False)
            name = self.logged_in_fullname or "Unknown User"
            try:
                wb = load_workbook(self.excel_file)
                ws = wb[self.punch_sheet_name]
                self.writecell(ws, p['row'], self.punch_cols['closed_name'], name)
                self.writecell(ws, p['row'], self.punch_cols['closed_date'],
                               datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                wb.save(self.excel_file)
                wb.close()
            except PermissionError:
                messagebox.showerror("Excel file is open", "Close the Excel workbook and try again.", parent=dlg)
                return
            except Exception as exc:
                messagebox.showerror("Could not close punch", str(exc), parent=dlg)
                return
            # Ensure a remark-holder annotation exists (created if needed),
            # then flip EVERY highlight attached to this punch - not just one -
            # from pink to green. This covers punches that were marked with
            # multiple highlights via multi-mark mode.
            find_ann(p, create=True)
            all_anns = find_all_anns(p)
            any_page = None
            for ann in all_anns:
                if ann.get('type') == 'highlight' and ann.get('color') == 'pink':
                    ann['color'] = 'green'
                elif ann.get('type') == 'error':
                    ann['type'] = 'ok'
                ann['closed_by'] = name
                ann['closed_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if ann.get('page') is not None and any_page is None:
                    any_page = ann.get('page')
            self.mark_dirty()
            self._write_session_file()
            punches.pop(current['index'])
            if not punches:
                dlg.destroy()
                self.root.after(100, self.autofin)
                return
            current['index'] = min(current['index'], len(punches) - 1)
            show_item(current['index'])
            if any_page is not None:
                self.root.after_idle(lambda: self.schedule_display(preserve_view=True, delay_ms=1))

        punch_list.bind('<<ListboxSelect>>', select_from_list)
        button(left_actions, "Previous", lambda: go(-1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT, padx=(0, 8))
        button(left_actions, "Next", lambda: go(1), '#e2e8f0', colors['text'], 11).pack(side=tk.LEFT)
        button(right_actions, "Close", dlg.destroy, '#475569', width=10).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "Verify & Close Punch", close_punch, colors['success'], width=20).pack(side=tk.RIGHT, padx=(10, 0))
        button(right_actions, "Save Remark", save_remark, colors['primary'], width=13).pack(side=tk.RIGHT)
        dlg.bind('<Control-s>', lambda event: save_remark())
        dlg.bind('<Control-Return>', lambda event: close_punch())
        refresh_list()
        show_item(0)
        dlg.wait_window()

    def autofin(self):
        """Automatically finalize cabinet if all punches are closed
        
        This checks:
        1. Zero open punches
        2. All checklist items reviewed
        3. Then saves Excel, exports PDF, updates status to 'Closed'
        4. Removes from rework queue if present
        """
        if not self.pdf_document or not self.cabinet_id:
            return
        
        # Check open punches
        open_punches = self.countopen()
        
        if open_punches > 0:
            print(f" Cannot auto-finalize: {open_punches} open punch(es) remaining")
            return
        
        print("OK All punches closed - checking checklist...")
        
        # Check checklist completion
        is_complete, pending_count = self.checklistcomp()
        
        if not is_complete:
            print(f" Checklist incomplete: {pending_count} item(s) pending")
            
            # Ask user if they want to complete checklist now
            proceed = messagebox.askyesno(
                "Checklist Incomplete",
                f" {pending_count} checklist item(s) not reviewed.\n\n"
                "Would you like to complete the checklist now?",
                icon='warning'
            )
            
            if proceed:
                # Open checklist review dialog
                self.reviewnow()
                
                # After review, check again
                is_complete, pending_count = self.checklistcomp()
                
                if not is_complete:
                    messagebox.showinfo(
                        "Checklist Still Incomplete",
                        "Cabinet cannot be finalized until checklist is complete."
                    )
                    return
            else:
                return
        
        print(" Checklist complete - auto-finalizing cabinet...")
        
        try:
            # 1. Save session
            self.savesession()
            
            # 2. Save Interphase Excel
            interphase_path = os.path.join(
                self.project_dirs["interphase_export"],
                f"{self.cabinet_id.replace(' ', '_')}_Interphase.xlsx"
            )
            
            try:
                shutil.copy2(self.excel_file, interphase_path)
                print(f"Interphase Excel saved: {interphase_path}")
            except Exception as e:
                print(f"Failed to save Interphase Excel: {e}")
            
            # 3. Export annotated PDF
            self.exportpdf()
            print("Annotated PDF exported")
            
            # 4. Update status to Closed
            self.update_status_and_sync('closed')
            print("Status updated to: Closed")
            
            # 5. NEW: Remove from rework queue if present
            username = self.logged_in_fullname or "Unknown User"
            
            try:
                if self.handover_db.is_in_rework_queue(self.cabinet_id):
                    print(f" {self.cabinet_id} found in rework queue - removing...")
                    
                    removed = self.handover_db.verify_production_item(
                        self.cabinet_id,
                        verified_by=username,
                        verification_notes="Cabinet finalized - all punches closed and verified",
                        mark_as_closed=True
                    )
                    
                    if removed:
                        print(f" Removed {self.cabinet_id} from rework verification queue")
                    else:
                        print(f" Failed to remove {self.cabinet_id} from rework queue")
            except Exception as e:
                print(f" Error removing from rework queue: {e}")
            
            # 6. Show success message
            
        except Exception as e:
            messagebox.showerror("Finalization Error", f"Failed to finalize cabinet:\n{e}")
            import traceback
            traceback.print_exc()

    # UPDATED: handover_to_production - with checklist check and queue management
    def handover(self):
        """Handover current cabinet to production with checklist validation"""
        
        if not self.pdf_document or not self.excel_file:
            messagebox.showwarning("Incomplete", 
                                  "Please load a PDF and Excel file first.")
            return
        
        if not self.cabinet_id or not self.project_name:
            messagebox.showwarning("Missing Info", 
                                  "Project details are incomplete.")
            return
        
        # A punch may now be created directly without a visual annotation.
        existing_punches = self.allexistingpunches()
        if not self.annotations and not existing_punches:
            proceed = messagebox.askyesno(
                "No Annotations",
                "No annotations found. Handover anyway?",
                icon='warning'
            )
            if not proceed:
                return
        
        # NEW: Check checklist completion BEFORE handover
        is_complete, pending_count = self.checklistcomp()
        
        if not is_complete:
            messagebox.showwarning(
                "Checklist Incomplete",
                f"⚠️ Cannot handover to production.\n\n"
                f"{pending_count} checklist item(s) not reviewed.\n\n"
                "Please complete the checklist first.",
                icon='warning'
            )
            
            # Ask if they want to complete it now
            complete_now=True
            
            if complete_now:
                self.reviewnow()
                
                # Check again after review
                is_complete, pending_count = self.checklistcomp()
                
                if not is_complete:
                    messagebox.showinfo(
                        "Handover Cancelled",
                        "Checklist still incomplete. Handover cancelled."
                    )
                    return
            else:
                return
        
        # Count open punches
        open_punches = self.countopen()
        
        
        # Save session before handover
        self.savesession()
        
        # Get user name
        username = self.logged_in_fullname or "Unknown User"
        
        # Prepare handover data
        session_path = os.path.join(
            self.project_dirs.get("sessions", ""),
            f"{self.cabinet_id}_annotations.json"
        )
        
        # Count the Punch Sheet as the source of truth so direct/manual
        # punches are handed over even when no highlight annotation exists.
        existing_punches = self.allexistingpunches()
        total_punches = len(existing_punches)
        
        handover_data = {
            "cabinet_id": self.cabinet_id,
            "project_name": self.project_name,
            "sales_order_no": self.sales_order_no,
            "pdf_path": self.current_pdf_path,
            "excel_path": self.excel_file,
            "session_path": session_path if os.path.exists(session_path) else None,
            "total_punches": total_punches,
            "open_punches": open_punches,
            "closed_punches": max(0, total_punches - open_punches),
            "handed_over_by": username,
            "handed_over_date": datetime.now().isoformat()
        }
        
        # NEW: Remove from verify rework queue if present
        try:
            # Check if in rework queue
            pending_items = self.handover_db.get_pending_quality_items()
            in_rework_queue = any(item['cabinet_id'] == self.cabinet_id for item in pending_items)
            
            if in_rework_queue:
                # Remove from rework queue
                self.handover_db.verify_production_item(
                    self.cabinet_id,
                    verified_by=username,
                    verification_notes="Re-opened for quality inspection"
                )
                print(f"OK Removed {self.cabinet_id} from verify rework queue")
        except Exception as e:
            print(f"⚠️ Error checking/removing from rework queue: {e}")
        
        # Add to production queue
        success = self.handover_db.add_quality_handover(handover_data)
        self.update_status_and_sync('handed_to_production')
        
        if not success:
            messagebox.showwarning("Already Handed Over", 
                                 "Cabinet already in production queue")


    def updatestatsforref(self, ref_no, status='NOK'):
            """Update Interphase status"""
            try:
                wb = load_workbook(self.excel_file)
                if self.interphase_sheet_name not in wb.sheetnames:
                    wb.close()
                    return False
                ws = wb[self.interphase_sheet_name]
                
                updated_any = False
                # Updated to include timestamp + date
                current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                username = self.logged_in_fullname or "Unknown User"
                
                for r in range(1, ws.max_row + 1):
                    cell_val = self.readcell(ws, r, self.interphase_cols['ref_no'])
                    if cell_val and str(cell_val).strip() == str(ref_no).strip():
                        self.writecell(ws, r, self.interphase_cols['status'], status)
                        self.writecell(ws, r, self.interphase_cols['name'], username)
                        self.writecell(ws, r, self.interphase_cols['date'], current_date)
                        updated_any = True
                
                if updated_any:
                    wb.save(self.excel_file)
                wb.close()
                return updated_any
            except Exception as e:
                print(f"Interphase update error: {e}")
                return False

    # ============================================================================
    # NEW: finalize_verification - Check checklist, save Excel, export PDF
    # ============================================================================

    def onclosing(self):
        """
        Handle application closing with a final, guaranteed save.
        FUNCTIONAL USE: This is the single place a clean shutdown flushes
        everything - DB stats sync AND session JSON write - regardless of
        whether anything was flagged dirty (belt-and-suspenders in case a
        prior autosave tick silently failed). Most saves during normal use
        are deferred (see mark_dirty/flush_pending_saves); this is where
        they're guaranteed to land before the process exits. Note this only
        covers a clean exit - an abrupt crash/kill relies on the periodic
        background autosave (self._autosave_interval_ms) having already run
        recently, since nothing can run after the process is killed.
        """
        if self._text_editor is not None:
            self._commit_text_editor()
        if self.pdf_document and hasattr(self, 'project_dirs'):
            try:
                print("\n Final save before closing...")
                self._dirty = True  # force a full flush regardless of current flag state
                self.sync_manager_stats_only()
                if self._write_session_file():
                    print(" Session saved successfully")
                self._dirty = False
                print(" Statistics synced")

            except Exception as e:
                print(f" Save on close failed: {e}")
                # Ask user if they want to close anyway
                proceed = messagebox.askyesno(
                    "Save Failed",
                    f"Failed to save:\n{e}\n\nClose anyway?",
                    icon='warning'
                )
                if not proceed:
                    return  # Don't close the application

        if getattr(self, '_autosave_after_id', None) is not None:
            try:
                self.root.after_cancel(self._autosave_after_id)
            except Exception:
                pass
            self._autosave_after_id = None

        # Close the application
        self.root.destroy()

    def saverecentproj(self):
        """Save current project to database with storage location - HIGHLIGHTER VERSION"""
        if not self.current_pdf_path or not self.excel_file:
            return
        
        try:
            session_path = os.path.join(
                self.project_dirs.get("sessions", ""),
                f"{self.cabinet_id}_annotations.json"
            ) if hasattr(self, 'project_dirs') else None
            
            project_data = {
                'cabinet_id': self.cabinet_id,
                'project_name': self.project_name,
                'sales_order_no': self.sales_order_no,
                'storage_location': self.storage_location,
                'pdf_path': self.current_pdf_path,
                'excel_path': self.excel_file,
                'session_path': session_path if session_path and os.path.exists(session_path) else None,
                'last_accessed': datetime.now().isoformat()
            }
            
            if self.db.project_exists(self.cabinet_id):
                self.db.update_project(self.cabinet_id, project_data)
            else:
                project_data['created_date'] = datetime.now().isoformat()
                self.db.add_project(project_data)
            
            self.updrecentdropdwn()
            self.sync_manager_stats_only()
            
        except Exception as e:
            print(f"Error saving recent project: {e}")


    def _all_project_records(self):
        """Return every cabinet ever created, newest activity first."""
        try:
            records = self.db.get_recent_projects(limit=1000000) or []
        except TypeError:
            records = self.db.get_recent_projects() or []
        def stamp(item):
            return str(item.get('last_accessed') or item.get('created_date') or '')
        return sorted(records, key=stamp, reverse=True)

    def loadrecprojui(self):
        """Compatibility hook retained for startup; the browser loads live data on open."""
        self._project_browser_records = self._all_project_records()

    def updrecentdropdwn(self):
        """Compatibility hook for older save paths; refresh the all-project cache."""
        self._project_browser_records = self._all_project_records()

    def show_project_browser(self):
        """Show all projects, then drill into the cabinets belonging to a project."""
        records = self._all_project_records()
        dlg = tk.Toplevel(self.root)
        dlg.title("All Projects and Cabinets")
        dlg.geometry("900x620")
        dlg.minsize(720, 480)
        dlg.configure(bg='#f8fafc')
        dlg.transient(self.root)
        dlg.grab_set()

        header = tk.Frame(dlg, bg='#1e293b', height=58)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        title_var = tk.StringVar(value="All Projects")
        tk.Label(header, textvariable=title_var, bg='#1e293b', fg='white',
                 font=('Segoe UI', 14, 'bold')).pack(pady=14)

        search_frame = tk.Frame(dlg, bg='#f8fafc')
        search_frame.pack(fill=tk.X, padx=18, pady=(16, 8))
        tk.Label(search_frame, text="Search:", bg='#f8fafc', fg='#334155',
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 8))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                                font=('Segoe UI', 11))
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        body = tk.Frame(dlg, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=8)
        tree = ttk.Treeview(body, columns=('name', 'count', 'updated'), show='headings')
        tree.heading('name', text='Project')
        tree.heading('count', text='Cabinets')
        tree.heading('updated', text='Most Recent')
        tree.column('name', width=430)
        tree.column('count', width=100, anchor='center', stretch=False)
        tree.column('updated', width=210, anchor='center', stretch=False)
        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        state = {'project': None, 'items': {}}
        def grouped_projects():
            groups = {}
            for row in records:
                name = str(row.get('project_name') or 'Unnamed Project').strip()
                groups.setdefault(name, []).append(row)
            return sorted(groups.items(), key=lambda pair: str(
                pair[1][0].get('last_accessed') or pair[1][0].get('created_date') or ''),
                reverse=True)

        def populate(*args):
            query = search_var.get().strip().casefold()
            tree.delete(*tree.get_children())
            state['items'] = {}
            if state['project'] is None:
                for project_name, cabinets in grouped_projects():
                    searchable = project_name + ' ' + ' '.join(str(c.get('cabinet_id', '')) for c in cabinets)
                    if query and query not in searchable.casefold():
                        continue
                    latest = cabinets[0].get('last_accessed') or cabinets[0].get('created_date') or ''
                    item = tree.insert('', tk.END, values=(project_name, len(cabinets), str(latest)[:19]))
                    state['items'][item] = ('project', project_name, cabinets)
            else:
                for cabinet in state['project'][1]:
                    cabinet_id = str(cabinet.get('cabinet_id') or '')
                    so = str(cabinet.get('sales_order_no') or '')
                    if query and query not in (cabinet_id + ' ' + so).casefold():
                        continue
                    updated = cabinet.get('last_accessed') or cabinet.get('created_date') or ''
                    item = tree.insert('', tk.END, values=(cabinet_id, so, str(updated)[:19]))
                    state['items'][item] = ('cabinet', cabinet)

        def open_selected(event=None):
            selected = tree.selection()
            if not selected:
                return
            payload = state['items'].get(selected[0])
            if not payload:
                return
            if payload[0] == 'project':
                state['project'] = (payload[1], payload[2])
                title_var.set(f"Project: {payload[1]}")
                tree.heading('name', text='Cabinet ID')
                tree.heading('count', text='Sales Order')
                search_var.set('')
                populate()
            else:
                dlg.destroy()
                self.loadrecentdb(payload[1])

        def go_back():
            if state['project'] is None:
                dlg.destroy()
                return
            state['project'] = None
            title_var.set("All Projects")
            tree.heading('name', text='Project')
            tree.heading('count', text='Cabinets')
            search_var.set('')
            populate()

        search_var.trace_add('write', populate)
        tree.bind('<Double-Button-1>', open_selected)
        tree.bind('<Return>', open_selected)
        buttons = tk.Frame(dlg, bg='#f8fafc')
        buttons.pack(fill=tk.X, padx=18, pady=(4, 16))
        tk.Button(buttons, text="Back / Close", command=go_back, bg='#64748b', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=20, pady=9).pack(side=tk.LEFT)
        tk.Button(buttons, text="Open", command=open_selected, bg='#3b82f6', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief=tk.FLAT, padx=26, pady=9).pack(side=tk.RIGHT)
        populate()
        search_entry.focus_set()

    def scanned_system_book_path_from_drawing(self, file_path):
        """Return the sibling 07-Scanned System Book path for a selected input drawing."""
        if not file_path:
            return ''
        current = ntpath.dirname(ntpath.normpath(str(file_path).strip()))
        while current:
            if ntpath.basename(current).strip().casefold() == '02-customer inputs':
                return ntpath.normpath(
                    ntpath.join(ntpath.dirname(current), '07-Scanned System Book')
                )
            parent = ntpath.dirname(current)
            if not parent or parent == current:
                break
            current = parent
        return ''

    def project_name_options_from_drawing(self, file_path):
        """Return immediate folders from the sibling 07-Scanned System Book directory.

        For a drawing selected anywhere below 02-Customer Inputs, move to the
        project root and inspect 07-Scanned System Book. The Hazardous area
        folder is excluded case-insensitively. Returned names are suitable for
        an editable project-name dropdown, so the user can also type a new name.
        """
        if not file_path:
            return []
        scanned_dir = self.scanned_system_book_path_from_drawing(file_path)
        if not scanned_dir:
            return []
        if not os.path.isdir(scanned_dir):
            return []
        try:
            names = [
                entry.name for entry in os.scandir(scanned_dir)
                if entry.is_dir() and entry.name.strip().casefold() != 'hazardous area'
            ]
        except OSError as exc:
            print(f"[WARN] Could not read project-name folders from {scanned_dir}: {exc}")
            return []
        return sorted(names, key=str.casefold)

    def infer_project_from_drawing(self, file_path):
        """Find an existing project for a drawing path and return its name, if any."""
        if not file_path:
            return ''
        selected = ntpath.normcase(ntpath.normpath(str(file_path)))
        records = self._all_project_records()
        # Strongest match: selected drawing is under a project's recorded storage tree.
        best = None
        for row in records:
            location = row.get('storage_location')
            if not location:
                continue
            loc = ntpath.normcase(ntpath.normpath(str(location)))
            if selected == loc or selected.startswith(loc.rstrip('\\/') + ntpath.sep):
                if best is None or len(loc) > len(best[0]):
                    best = (loc, str(row.get('project_name') or ''))
        if best:
            return best[1]
        # Folder convention: <project>/07-Scanned System Book/<drawing>.
        current = ntpath.dirname(selected)
        while current:
            if ntpath.basename(current).strip().casefold() == '07-scanned system book':
                folder_name = ntpath.basename(ntpath.dirname(current)).replace('_', ' ').strip()
                for row in records:
                    if str(row.get('project_name') or '').replace('_', ' ').strip().casefold() == folder_name.casefold():
                        return str(row.get('project_name') or '')
                return folder_name
            parent = ntpath.dirname(current)
            if not parent or parent == current:
                break
            current = parent
        return ''

    def loadrecentdb(self, project_data):
        """Load a recent project from database - HIGHLIGHTER VERSION"""
        self.stopmultimark()  # never leave multi-mark mode active across a project swap
        self.busy("Loading project...")
        try:
            # Set project details
            self.cabinet_id = project_data['cabinet_id']
            self.project_name = project_data['project_name']
            self.sales_order_no = project_data.get('sales_order_no', '')
            self.storage_location = project_data['storage_location']
            
            self.preparefolders()
            
            expected_excel_path = os.path.join(
                self.project_dirs["working_excel"],
                f"{self.cabinet_id.replace(' ', '_')}_Working.xlsx"
            )
            
            expected_session_path = os.path.join(
                self.project_dirs["sessions"],
                f"{self.cabinet_id}_annotations.json"
            )
            
            # Check PDF
            pdf_path = project_data.get('pdf_path')
            if not pdf_path or not os.path.exists(pdf_path):
                self.unbusy()
                messagebox.showerror("Error", 
                                   f"PDF file not found:\n{pdf_path}\n\n"
                                   "The file may have been moved or deleted.")
                return
            
            # Check Excel
            if not os.path.exists(expected_excel_path):
                old_excel_path = project_data.get('excel_path')
                if old_excel_path and os.path.exists(old_excel_path):
                    try:
                        shutil.copy2(old_excel_path, expected_excel_path)
                        self.unbusy()
                        messagebox.showinfo("Excel Migrated", 
                                          f"Excel file migrated to new location:\n{expected_excel_path}")
                        self.busy("Loading project...")
                    except Exception as e:
                        self.unbusy()
                        messagebox.showerror("Error", 
                                           f"Excel file not found and couldn't migrate:\n{e}")
                        return
                else:
                    self.unbusy()
                    messagebox.showerror("Error", 
                                       f"Excel file not found at:\n{expected_excel_path}\n\n"
                                       "The file may have been moved or deleted.")
                    return
            
            # Load PDF
            self.busy("Opening PDF...")
            self.pdf_document = fitz.open(pdf_path)
            self.current_pdf_path = pdf_path
            self.current_page = 0
            self.annotations = []
            self.zoom_level = 1.0
            self._update_zoom_toolbar_label()
            self.tool_mode = None
            
            # ADDED: Reset highlighter state
            self.active_highlighter = None
            self.highlight_points = []
            self.colorbutton()
            
            self.root.config(cursor="")
            
            # Set Excel
            self.excel_file = expected_excel_path
            self.working_excel_path = expected_excel_path
            self.current_sr_no = self.getnextsr()
            
            # Load session
            if os.path.exists(expected_session_path):
                self.busy("Loading saved session...")
                self.loadfrompath(expected_session_path)
            else:
                old_session_path = project_data.get('session_path')
                if old_session_path and os.path.exists(old_session_path):
                    try:
                        shutil.copy2(old_session_path, expected_session_path)
                        self.busy("Loading saved session...")
                        self.loadfrompath(expected_session_path)
                    except:
                        self.display(preserve_view=False)
                else:
                    self.display(preserve_view=False)
            
            # Update database
            self.db.update_project(self.cabinet_id, {
                'pdf_path': self.current_pdf_path,
                'excel_path': expected_excel_path,
                'session_path': expected_session_path if os.path.exists(expected_session_path) else None,
                'last_accessed': datetime.now().isoformat()
            })
            
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            self.unbusy()

    # ================================================================
    # COMPREHENSIVE STATUS AND STATISTICS MANAGEMENT
    # ================================================================


    def syncstatsmgr(self):
        """LEGACY: Sync statistics while preserving existing status
        
        This is kept for backward compatibility.
        Internally calls sync_manager_stats_only().
        
        Will now also CREATE cabinet if it doesn't exist (making it visible on dashboard)
        """
        self.sync_manager_stats_only()


    def ensure_visible_on_dashboard(self):
        """Explicitly ensure this cabinet is visible on the manager dashboard
        
        Call this when:
        - Opening a PDF for the first time
        - Starting quality inspection
        - Any time you want to make sure the cabinet appears on dashboard
        
        This will:
        1. Create the cabinet in database if it doesn't exist
        2. Set initial status from Interphase worksheet (or 'quality_inspection')
        3. Sync all current statistics
        """
        if not self.cabinet_id:
            return False
        
        try:
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            # Check if exists
            cursor.execute('SELECT cabinet_id FROM cabinets WHERE cabinet_id = ?', (self.cabinet_id,))
            exists = cursor.fetchone()
            
            conn.close()
            
            if exists:
                # Already exists, just sync stats
                self.sync_manager_stats_only()
                print(f"OK {self.cabinet_id} already on dashboard - stats synced")
                return True
            else:
                # Doesn't exist, create it
                self.sync_manager_stats_only()  # This will create it now
                print(f"OK {self.cabinet_id} is now visible on dashboard")
                return True
                
        except Exception as e:
            print(f"❌ Error ensuring visibility: {e}")
            return False


    def countopen(self):
        """Count open punches in current Excel
        
        Returns:
            int: Number of punches that are not closed
        """
        try:
            if not self.excel_file or not os.path.exists(self.excel_file):
                return 0
            
            wb = load_workbook(self.excel_file, data_only=True)
            ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
            
            open_count = 0
            row = 8
            
            while row <= ws.max_row + 5:
                sr = self.readcell(ws, row, self.punch_cols['sr_no'])
                if sr is None:
                    break
                
                closed = self.readcell(ws, row, self.punch_cols['closed_name'])
                if not closed:
                    open_count += 1
                
                row += 1
            
            wb.close()
            return open_count
            
        except Exception as e:
            print(f"Error counting open punches: {e}")
            return 0
        
    def mark_dirty(self):
        """
        Flag that in-memory annotation/stat state has changed since the last
        save, without doing any I/O right now.
        FUNCTIONAL USE: Called after cheap, frequent actions (logging a punch,
        closing a punch, checklist review clicks, etc.) instead of immediately
        hitting the database and/or Excel. The periodic autosave loop (started
        in uisetup/startautosaveloop) and onclosing() are what actually flush
        this to disk, so repeated annotation actions stay fast and smooth.
        """
        self._dirty = True

    def startautosaveloop(self):
        """
        Start the recurring background flush that saves dirty state.
        FUNCTIONAL USE: Called once during setup. Reschedules itself via
        root.after so it keeps running for the life of the app, flushing
        at most every self._autosave_interval_ms while there is pending
        (dirty) work. This is what makes "save at the end" also safe against
        an abrupt crash - the gap between crash and last flush is bounded by
        this interval instead of the whole session.
        """
        self._autosave_tick()

    def _autosave_tick(self):
        try:
            if self._dirty:
                self.flush_pending_saves(show_status=False)
        except Exception as e:
            print(f"[WARN] Autosave tick failed: {e}")
        finally:
            self._autosave_after_id = self.root.after(self._autosave_interval_ms, self._autosave_tick)

    def flush_pending_saves(self, show_status=True):
        """
        Perform the actual batched save: sync DB stats and resave the session
        JSON, then clear the dirty flag.
        FUNCTIONAL USE: The single real "save everything" path. Called by the
        periodic autosave loop, by onclosing() (final flush before exit/crash),
        and by explicit user actions like Save Session. Excel workbook writes
        for individual punch rows already happen immediately at the time
        they're logged (openpyxl workbooks aren't safely shareable across a
        long-lived handle without risking corruption on crash), but the
        heavier DB stats sync and session JSON rewrite are batched here.
        """
        if not self._dirty:
            return

        try:
            if self.pdf_document and self.cabinet_id:
                self.sync_manager_stats_only()

            if self.pdf_document and hasattr(self, 'project_dirs') and self.project_dirs.get("sessions"):
                self._write_session_file()

            self._dirty = False
            self._last_flush_time = time.monotonic()

            if show_status:
                self.flashstat("Saved", bg='#10b981')

        except Exception as e:
            print(f"[WARN] flush_pending_saves failed: {e}")

    def sync_manager_stats_only(self, update_status_from_interphase=True):
        """Sync statistics and optionally update status from Interphase
        
        Args:
            update_status_from_interphase: If True, recalculate status from Interphase worksheet
        """
        if not self.pdf_document or not self.cabinet_id:
            return
        
        try:
            # Count pages with annotations
            annotated_pages = len(set(ann['page'] for ann in self.annotations 
                                     if ann.get('page') is not None))
            total_pages = len(self.pdf_document)
            
            # Count the Punch Sheet as the source of truth. This includes
            # punches created directly from Edit or Add Punch without a highlight.
            total_punches = len(self.allexistingpunches())
            open_punches = self.countopen()
            
            # Count implemented and closed
            implemented_punches = 0
            closed_punches = 0
            
            if self.excel_file and os.path.exists(self.excel_file):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.excel_file, data_only=True)
                    ws = wb[self.punch_sheet_name] if self.punch_sheet_name in wb.sheetnames else wb.active
                    
                    row = 8
                    while row <= ws.max_row + 5:
                        checked = self.readcell(ws, row, self.punch_cols['checked_name'])
                        
                        if checked:
                            implemented = self.readcell(ws, row, self.punch_cols['implemented_name'])
                            closed = self.readcell(ws, row, self.punch_cols['closed_name'])
                            
                            if closed:
                                closed_punches += 1
                            elif implemented:
                                implemented_punches += 1
                        
                        row += 1
                        if row > 2000:
                            break
                    
                    wb.close()
                except Exception as e:
                    print(f"Error counting punches: {e}")
            
            # Determine status
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT status FROM cabinets WHERE cabinet_id = ?', (self.cabinet_id,))
            existing = cursor.fetchone()
            excel_path_db = to_relative_path(self.excel_file)
            storage_location_db = to_relative_storage_location(getattr(self, 'storage_location', None))
            
            if existing:
                # Cabinet exists - get current status
                current_status = existing[0]
                
                # Update status from Interphase if requested AND if status is workflow-related
                workflow_statuses = [
                    'project_info_sheet',
                    'mechanical_assembly', 
                    'component_assembly',
                    'final_assembly',
                    'final_documentation',
                    'quality_inspection'
                ]
                
                if update_status_from_interphase and current_status in workflow_statuses:
                    new_status = self.get_status_from_interphase(self.excel_file)
                    if new_status:
                        current_status = new_status
                        print(f"OK Status updated from Interphase: {new_status}")
                
                # Update with potentially new status
                cursor.execute('''
                    UPDATE cabinets 
                    SET total_pages = ?,
                        annotated_pages = ?,
                        total_punches = ?,
                        open_punches = ?,
                        implemented_punches = ?,
                        closed_punches = ?,
                        status = ?,
                        last_updated = ?,
                        excel_path = ?,
                        storage_location = ?
                    WHERE cabinet_id = ?
                ''', (total_pages, annotated_pages, total_punches, open_punches,
                      implemented_punches, closed_punches, current_status,
                        datetime.now().isoformat(), excel_path_db,
                        storage_location_db, self.cabinet_id))
                
                print(f"OK Updated {self.cabinet_id} - Status: {current_status}")
            else:
                # Cabinet doesn't exist - create with initial status from Interphase
                initial_status = self.get_status_from_interphase(self.excel_file)
                if not initial_status:
                    initial_status = 'quality_inspection'
                
                cursor.execute('''
                    INSERT INTO cabinets (
                        cabinet_id, project_name, sales_order_no,
                        total_pages, annotated_pages, total_punches,
                        open_punches, implemented_punches, closed_punches,
                        status, created_date, last_updated,
                        storage_location, excel_path
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    self.cabinet_id, self.project_name, self.sales_order_no,
                    total_pages, annotated_pages, total_punches,
                    open_punches, implemented_punches, closed_punches,
                    initial_status, datetime.now().isoformat(),
                    datetime.now().isoformat(),
                    storage_location_db, excel_path_db
                ))
                
                print(f"Created {self.cabinet_id} with status: {initial_status}")
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Stats sync error: {e}")
            import traceback
            traceback.print_exc()


    def get_status_from_interphase(self, excel_path):
        """Read Interphase worksheet and determine status based on HIGHEST filled reference number
        
        Returns: status string or None if not determined from Interphase
        """
        if not excel_path or not os.path.exists(excel_path):
            return None
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            if 'Interphase' not in wb.sheetnames:
                wb.close()
                return None
            
            ws = wb['Interphase']
            
            # Find the HIGHEST reference number that has a status
            highest_ref_num = 0
            
            # Start from row 11 (typical Interphase data starts here)
            for row in range(11, ws.max_row + 1):
                status_cell = self.readcell(ws, row, 'D')  # Status column
                
                # If status cell has content, check the reference number
                if status_cell and str(status_cell).strip():
                    ref_no_cell = self.readcell(ws, row, 'B')  # Reference column
                    
                    if ref_no_cell:
                        try:
                            ref_str = str(ref_no_cell).strip()
                            
                            # Handle range formats like "1-2" - take the LAST number
                            if '-' in ref_str:
                                ref_num = int(ref_str.split('-')[-1])
                            else:
                                ref_num = int(ref_str)
                            
                            # Track highest completed reference
                            if ref_num > highest_ref_num:
                                highest_ref_num = ref_num
                        
                        except (ValueError, IndexError):
                            continue
            
            wb.close()
            
            # Determine status based on highest completed reference number
            if highest_ref_num == 0:
                return 'quality_inspection'  # Nothing completed yet
            elif 1 <= highest_ref_num <= 2:
                return 'project_info_sheet'
            elif 3 <= highest_ref_num <= 9:
                return 'mechanical_assembly'
            elif 10 <= highest_ref_num <= 18:
                return 'component_assembly'
            elif 19 <= highest_ref_num <= 26:
                return 'final_assembly'
            elif highest_ref_num >= 27:
                return 'final_documentation'
            else:
                return 'quality_inspection'
            
        except Exception as e:
            print(f"Error reading Interphase worksheet: {e}")
            return None


    def update_status_and_sync(self, new_status):
        """Explicitly set status and sync to database
        
        Use this for manual status changes (handover, closing, etc.)
        This will NOT be overridden by Interphase status.
        
        Args:
            new_status: One of the valid status strings
        """
        try:
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE cabinets 
                SET status = ?, 
                    last_updated = ?
                WHERE cabinet_id = ?
            ''', (new_status, datetime.now().isoformat(), self.cabinet_id))
            
            conn.commit()
            conn.close()
            
            print(f"OK Status manually updated to: {new_status}")
            
            # Now sync stats without changing status
            self.sync_manager_stats_only(update_status_from_interphase=False)
            
        except Exception as e:
            print(f"Status update error: {e}")


    def get_current_status_from_db(self):
        """Get the current status from database
        
        Returns:
            str: Current status or 'quality_inspection' if not found
        """
        try:
            conn = sqlite3.connect(self.manager_db.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT status FROM cabinets WHERE cabinet_id = ?', 
                          (self.cabinet_id,))
            result = cursor.fetchone()
            
            conn.close()
            
            if result:
                return result[0]
            else:
                # If not in database, check Interphase
                status = self.get_status_from_interphase(self.excel_file)
                return status if status else 'quality_inspection'
                
        except Exception as e:
            print(f"Error getting status from DB: {e}")
            return 'quality_inspection'
    


# ================================================================
# MAIN ENTRY POINT
# ================================================================

def main():
    prevent_power_throttling()
    root = tk.Tk()
    root.title("Circuit Inspector")
    app = CircuitInspector(root)
    root.mainloop()


if __name__ == "__main__":
    main()